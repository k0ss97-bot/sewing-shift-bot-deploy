import hashlib
import hmac
import io
import json
import logging
import os
import base64
import threading
import time
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import parse_qsl, quote, urlparse

from database import (
    add_fabric_receipt,
    add_edit_log,
    add_feedback_entry,
    add_operation,
    add_cutting_layout,
    add_warehouse_stock,
    adjust_fabric_stock,
    adjust_warehouse_stock,
    admin_close_shift,
    assign_production_task,
    assign_route_batch,
    cancel_production_task,
    cancel_route_batch_and_restore_inputs,
    close_shift,
    complete_route_batch_step_atomic,
    create_cutting_contour_batch_for_task,
    create_production_task,
    create_route_batch,
    create_route_batches_with_inputs,
    create_shift,
    delete_shift_by_id,
    ensure_admin_employee,
    get_active_operations,
    get_active_cutting_batch_product_names,
    get_active_production_tasks,
    get_active_production_tasks_for_contours,
    get_active_route_batches,
    get_all_product_colors,
    get_all_employees,
    get_all_user_accounts,
    get_all_operations,
    get_cutting_batches_for_cutting,
    get_cutting_batches_for_formation,
    get_cutting_batches_for_layout,
    get_cutting_batch_result_rows,
    get_cutting_batch_owner,
    get_employee_by_telegram_id,
    get_employee_by_id,
    get_employee_route_batches,
    get_employee_period_operation_totals,
    get_employee_period_summary,
    get_employee_shifts_by_period,
    get_employees_by_status,
    get_feedback_entries,
    get_feedback_entries_by_shift,
    get_fabric_stock_rows,
    get_fabric_stock_rows_with_ids,
    get_open_shifts,
    get_open_shift_for_today,
    get_pending_employees,
    get_period_employee_summary,
    get_period_employee_production_performance,
    get_period_fabric_movement_rows,
    get_period_operation_rows,
    get_period_production_task_item_rows,
    get_period_production_trace_event_rows,
    get_period_route_batch_input_rows,
    get_period_route_batch_input_lot_rows,
    get_period_route_batch_rows,
    get_period_shift_details,
    get_period_warehouse_movement_rows,
    get_product_colors,
    get_product_sizes,
    get_recent_shifts,
    get_production_task_attachment,
    get_route_batch_by_id,
    get_route_batch_by_trace_code,
    get_route_batch_defect_rows,
    get_route_batch_defect_photo,
    get_route_batch_defects,
    get_route_batch_input_lots,
    get_route_batch_inputs,
    get_route_batch_passport,
    get_production_task_by_id,
    get_production_task_fabric_rolls,
    get_production_task_fabric_defects,
    get_shift_for_today,
    get_shift_report,
    get_warehouse_stock_by_id,
    get_warehouse_stock_rows,
    get_today_shifts,
    hide_operation,
    has_route_completion_request,
    local_now,
    local_today,
    mark_cutting_batch_formed,
    reject_production_task_fabric_rolls,
    route_steps_from_snapshot,
    restore_operation,
    set_route_batch_work_state,
    update_cutting_batch_progress,
    update_employee_access_status,
    update_employee_position,
    update_employee_role,
    update_operation_field,
)
from catalog import PREPARATION_OPERATION_OPTIONS, format_color_label
from miniapp_auth import parse_auth_token
from miniapp_assets import MINIAPP_HTML
from route_maps import CUTTING_ROUTE, PRODUCT_ROUTE_MAPS
from webapp_auth import (
    WebRegistrationError,
    authenticate_web_credentials,
    build_clear_cookies,
    build_session_cookie,
    change_web_password,
    create_web_session,
    get_web_account_profiles_by_telegram_ids,
    get_web_session,
    init_web_auth,
    register_web_account,
    revoke_web_session,
    revoke_web_sessions_for_telegram_id,
    session_token_from_cookie,
)
from webapp_pwa import app_shell_revision, inject_pwa_markup, send_pwa_resource


AUTH_MAX_AGE_SECONDS = 24 * 60 * 60
MAX_ATTACHMENT_BYTES = 10 * 1024 * 1024
MAX_DEFECT_PHOTO_BYTES = 2 * 1024 * 1024
MAX_REQUEST_BODY_BYTES = 15 * 1024 * 1024
MAX_CONCURRENT_REQUESTS = 32
REQUEST_SOCKET_TIMEOUT_SECONDS = 30
CONTENT_SECURITY_POLICY = (
    "default-src 'self'; "
    "base-uri 'self'; "
    "connect-src 'self'; "
    "font-src 'self' data:; "
    "form-action 'self'; "
    "frame-ancestors 'self'; "
    "img-src 'self' data: blob:; "
    "media-src 'self' blob:; "
    "object-src 'none'; "
    "script-src 'self' 'unsafe-inline'; "
    "style-src 'self' 'unsafe-inline'; "
    "worker-src 'self' blob:"
)
ROUTES_MINIAPP_ENABLED = False
CUTTING_CONTOUR_OPERATION = "Нанесение контуров лекал на ткань"
CUTTING_LAYOUT_OPERATION = "Формирование настила"
CUTTING_CUT_OPERATION = "Раскрой"
CUTTING_FORM_OPERATION = "Формирование готового кроя"
DEFECT_REASONS = [
    "Ошибка раскроя",
    "Повреждение ткани",
    "Неверный размер",
    "Неверный цвет",
    "Пропуск операции",
    "Дефект строчки",
    "Загрязнение",
    "Повреждение фурнитуры",
    "Другое",
]
DEFECT_DISPOSITIONS = ["Списать", "На переделку", "Уточнить"]
PWA_HTML = inject_pwa_markup(MINIAPP_HTML)
PWA_SHELL_REVISION = app_shell_revision(PWA_HTML)


def get_admin_ids():
    admin_ids = []
    raw_admin_ids = os.getenv("ADMIN_IDS", "").replace("ADMIN_IDS=", "")

    for raw_admin_id in raw_admin_ids.split(","):
        raw_admin_id = raw_admin_id.strip()

        if not raw_admin_id:
            continue

        try:
            admin_ids.append(int(raw_admin_id))
        except ValueError:
            logging.warning("Invalid ADMIN_IDS item ignored: %s", raw_admin_id)

    return admin_ids


def is_admin(telegram_id: int):
    employee = get_employee_by_telegram_id(telegram_id)
    if employee is not None:
        return employee[4] == "admin" and employee[5] == "active"
    return telegram_id in get_admin_ids()


def get_employee_for_access(telegram_id: int):
    employee = get_employee_by_telegram_id(telegram_id)
    if employee is not None:
        return employee
    if telegram_id in get_admin_ids():
        return ensure_admin_employee(telegram_id)
    return None


def format_minutes(total_minutes: int | None):
    if total_minutes is None:
        return ""

    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours}:{minutes:02d}"


def parse_telegram_init_data(init_data: str, bot_token: str):
    if not init_data or not bot_token:
        return None

    parsed_items = dict(parse_qsl(init_data, keep_blank_values=True))
    received_hash = parsed_items.pop("hash", "")

    if not received_hash:
        return None

    data_check_string = "\n".join(
        f"{key}={value}" for key, value in sorted(parsed_items.items())
    )
    secret_key = hmac.new(
        b"WebAppData",
        bot_token.encode("utf-8"),
        hashlib.sha256,
    ).digest()
    calculated_hash = hmac.new(
        secret_key,
        data_check_string.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()

    if not hmac.compare_digest(calculated_hash, received_hash):
        return None

    try:
        auth_date = int(parsed_items.get("auth_date", "0"))
    except ValueError:
        return None

    now_epoch = time.time()
    if auth_date <= 0 or auth_date > now_epoch + 60 or now_epoch - auth_date > AUTH_MAX_AGE_SECONDS:
        return None

    try:
        return json.loads(parsed_items.get("user", "{}"))
    except json.JSONDecodeError:
        return None


def authenticate_payload(payload: dict, bot_token: str, debug: bool):
    if debug and payload.get("telegram_id"):
        try:
            return {"id": int(payload["telegram_id"])}
        except (TypeError, ValueError):
            return None

    telegram_user = parse_telegram_init_data(payload.get("initData", ""), bot_token)

    if telegram_user:
        return telegram_user

    return parse_auth_token(payload.get("authToken", ""), bot_token)


def shift_to_dict(shift):
    if not shift:
        return None

    return {
        "id": shift[0],
        "date": shift[2],
        "start_time": shift[3],
        "end_time": shift[4],
        "status": shift[5],
    }


def employee_to_dict(employee):
    if not employee:
        return None

    return {
        "id": employee[0],
        "telegram_id": employee[1],
        "full_name": employee[2],
        "position": employee[3],
        "role": employee[4],
        "status": employee[5],
    }


def get_shift_state(telegram_id: int, message: str = ""):
    employee = get_employee_for_access(telegram_id)

    if employee is None:
        return {
            "ok": False,
            "code": "not_registered",
            "message": (
                "Вы не зарегистрированы или ваша запись не перенесена в новую базу. "
                f"Ваш Telegram ID: {telegram_id}. Вернитесь в бот и нажмите /start."
            ),
            "employee": None,
            "shift": None,
        }

    employee_data = employee_to_dict(employee)

    if employee[5] != "active":
        return {
            "ok": False,
            "code": employee[5],
            "message": "Профиль ещё не активен. Дождитесь подтверждения администратора.",
            "employee": employee_data,
            "shift": None,
        }

    open_shift = get_open_shift_for_today(employee[0])
    today_shift = open_shift or get_shift_for_today(employee[0])
    shift_data = shift_to_dict(today_shift)

    if today_shift and today_shift[5] == "closed":
        shift_data["total_minutes_text"] = ""

    return {
        "ok": True,
        "code": "ok",
        "message": message,
        "employee": employee_data,
        "shift": shift_data,
        "has_open_shift": open_shift is not None,
    }


def open_shift_for_telegram(telegram_id: int):
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return get_shift_state(telegram_id)

    open_shift = get_open_shift_for_today(employee[0])

    if open_shift is not None:
        return get_shift_state(telegram_id, "Смена уже открыта.")

    today_shift = get_shift_for_today(employee[0])

    if today_shift is not None and today_shift[5] == "closed":
        return get_shift_state(telegram_id, "Сегодня смена уже закрыта.")

    shift = create_shift(employee[0])
    add_edit_log(
        telegram_id,
        "employee",
        "Открыл смену из миниаппа",
        "shift",
        shift["id"],
        f"Дата: {shift['shift_date']}, начало: {shift['start_time']}",
    )
    return get_shift_state(telegram_id, "Смена открыта.")


def close_shift_for_telegram(telegram_id: int):
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return get_shift_state(telegram_id)

    open_shift = get_open_shift_for_today(employee[0])

    if open_shift is None:
        return get_shift_state(telegram_id, "У вас нет открытой смены.")

    result = close_shift(open_shift[0])

    if result is None:
        return get_shift_state(telegram_id, "Не удалось закрыть смену.")

    add_edit_log(
        telegram_id,
        "employee",
        "Закрыл смену из миниаппа",
        "shift",
        open_shift[0],
        f"Окончание: {result['end_time']}, отработано: {format_minutes(result['total_minutes'])}",
    )

    response = get_shift_state(
        telegram_id,
        f"Смена закрыта. Отработано: {format_minutes(result['total_minutes'])}.",
    )

    if response.get("shift"):
        response["shift"]["total_minutes"] = result["total_minutes"]
        response["shift"]["total_minutes_text"] = format_minutes(result["total_minutes"])

    return response


def operation_row_to_dict(row):
    operation_name, product_size, product_color, quantity, unit = row

    return {
        "operation_name": operation_name,
        "product_size": product_size,
        "product_color": format_color_label(product_color),
        "quantity": quantity,
        "unit": unit,
    }


def feedback_row_to_dict(row):
    feedback_date, feedback_time, full_name, position, category, message, shift_id = row

    return {
        "date": feedback_date,
        "time": feedback_time,
        "employee": full_name,
        "position": position,
        "category": category,
        "message": message,
        "shift_id": shift_id,
    }


def get_current_report_for_telegram(telegram_id: int):
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {
            "ok": False,
            "message": "Нет активного профиля.",
            "shift": None,
            "operations": [],
            "feedback": [],
        }

    shift = get_open_shift_for_today(employee[0]) or get_shift_for_today(employee[0])

    if shift is None:
        return {
            "ok": True,
            "message": "За сегодня отчёта пока нет.",
            "shift": None,
            "operations": [],
            "feedback": [],
        }

    return {
        "ok": True,
        "message": "",
        "shift": shift_to_dict(shift),
        "operations": [operation_row_to_dict(row) for row in get_shift_report(shift[0])],
        "feedback": [feedback_row_to_dict(row) for row in get_feedback_entries_by_shift(shift[0])],
    }


def submit_feedback_for_telegram(telegram_id: int, category: str, message: str):
    category = category.strip()
    message = message.strip()

    if category not in {"Производство", "Бытовое"}:
        return {"ok": False, "message": "Выберите раздел обратной связи."}

    if not message:
        return {"ok": False, "message": "Введите сообщение."}

    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля."}

    shift = get_open_shift_for_today(employee[0]) or get_shift_for_today(employee[0])
    shift_id = shift[0] if shift else None
    feedback_id = add_feedback_entry(employee[0], shift_id, category, message)

    if feedback_id is None:
        return {"ok": False, "message": "Не удалось сохранить сообщение."}

    add_edit_log(
        telegram_id,
        "admin" if is_admin(telegram_id) else "employee",
        "Отправил обратную связь из миниаппа",
        "feedback",
        feedback_id,
        f"{category}: {message}",
    )

    return {
        "ok": True,
        "message": "Сообщение отправлено.",
        "report": get_current_report_for_telegram(telegram_id),
    }


def get_route_step(product_name: str, step_index: int):
    steps = PRODUCT_ROUTE_MAPS.get(product_name, [])

    if step_index < 0 or step_index >= len(steps):
        return None

    return steps[step_index]


def get_route_steps_for_batch(batch: dict):
    return route_steps_from_snapshot(
        batch.get("route_snapshot") or "",
        batch.get("product_name") or "",
    )


def get_route_step_for_batch(batch: dict, step_index: int | None = None):
    steps = get_route_steps_for_batch(batch)
    index = batch.get("route_step_index", 0) if step_index is None else step_index
    if index < 0 or index >= len(steps):
        return None
    return steps[index]


def get_route_previous_status(batch: dict):
    if batch["status"] == "done":
        steps = get_route_steps_for_batch(batch)
        return steps[-1]["status_after"] if steps else "Процесс завершён"

    step_index = batch["route_step_index"]

    if step_index == 0:
        return "Партия создана"

    previous_step = get_route_step_for_batch(batch, step_index - 1)
    return previous_step["status_after"] if previous_step else "Партия в работе"


def route_batch_identity(batch: dict):
    return (
        f"{batch['product_name']} / "
        f"{batch['product_size']} / "
        f"{format_color_label(batch['product_color'])} / "
        f"{batch['quantity']} шт"
    )


def route_batch_assignee_to_dict(batch: dict):
    employee_id = batch.get("assigned_employee_id")

    if not employee_id:
        return None

    employee = get_employee_by_id(employee_id)

    if not employee:
        return {
            "id": employee_id,
            "full_name": "Сотрудник",
            "position": "",
        }

    return employee_to_dict(employee)


def can_employee_work_route_step(employee, current_step: dict | None):
    return bool(employee and current_step and employee[3] == current_step["position"])


def get_preparation_folder(operation_name: str):
    return PREPARATION_OPERATION_OPTIONS.get(operation_name, {}).get("folder", "")


def is_elastic_preparation_operation(operation_name: str):
    return get_preparation_folder(operation_name) == "Нарезание резинки"


def route_step_category(route_step: dict | None):
    if not route_step:
        return ""

    position = route_step.get("position", "")
    operation = route_step.get("operation", "")
    operation_lower = operation.lower()
    preparation_folder = get_preparation_folder(operation)

    if position == "Раскройщик":
        return "Раскрой"

    if position == "Упаковщик":
        if preparation_folder in {"Нарезание резинки", "Нарезание дублерина", "Дублирование"}:
            return "Подготовка"
        if "подготов" in operation_lower:
            return "Подготовка"
        if "упаков" in operation_lower or "склад" in operation_lower:
            return "Упаковка"
        if "вто" in operation_lower or "заутюж" in operation_lower or "проклей" in operation_lower:
            return "ВТО"
        return "Подготовка"

    if position == "Швея":
        if operation == "Сшивание резинок в кольцо":
            return "Подготовка"
        if "подготов" in operation_lower or "стачивание" in operation_lower:
            return "Подготовка"
        if "оверлок" in operation_lower or "сборка" in operation_lower:
            return "Оверлок"
        return "Прямострочка"

    return position


def should_auto_create_next_preparation_task(current_step: dict | None, next_step: dict | None):
    if not current_step or not next_step:
        return False

    return (
        is_elastic_preparation_operation(current_step.get("operation", ""))
        and next_step.get("position") == "Швея"
        and next_step.get("operation") == "Сшивание резинок в кольцо"
    )


def should_skip_cut_output_step(route_step: dict):
    operation_name = route_step.get("operation", "")
    preparation_folder = get_preparation_folder(operation_name)
    return preparation_folder in {"Нарезание резинки", "Нарезание дублерина"} or operation_name == "Сшивание резинок в кольцо"


def accepted_stock_stages_for_route_step(product_name: str, step_index: int):
    steps = PRODUCT_ROUTE_MAPS.get(product_name, [])

    if step_index < len(CUTTING_ROUTE) or step_index >= len(steps):
        return []

    accepted_stages = []

    if step_index > len(CUTTING_ROUTE):
        accepted_stages.append(steps[step_index - 1]["status_after"])

    first_cut_output_step = next(
        (
            index
            for index, route_step in enumerate(steps[len(CUTTING_ROUTE):], start=len(CUTTING_ROUTE))
            if not should_skip_cut_output_step(route_step)
        ),
        None,
    )

    if step_index == first_cut_output_step:
        accepted_stages.append("Раскроенные")

    return list(dict.fromkeys(accepted_stages))


def stock_size_tokens(value: str):
    size_text = str(value or "").split("(", 1)[0]
    return {item.strip() for item in size_text.split(",") if item.strip()}


def stock_sizes_overlap(first_value: str, second_value: str):
    first_sizes = stock_size_tokens(first_value)
    second_sizes = stock_size_tokens(second_value)
    return bool(first_sizes and second_sizes and first_sizes.intersection(second_sizes))


def route_map_to_dict(product_name: str):
    return {
        "product_name": product_name,
        "sizes": get_product_sizes(product_name),
        "colors": [format_color_label(color) for color in get_product_colors(product_name)],
        "raw_colors": get_product_colors(product_name),
        "steps": [
            {
                "number": index,
                "position": route_step["position"],
                "operation": route_step["operation"],
                "status_after": route_step["status_after"],
                "category": route_step_category(route_step),
                "accepted_stock_stages": accepted_stock_stages_for_route_step(product_name, index - 1),
                "requires_all_stock_stages": len(accepted_stock_stages_for_route_step(product_name, index - 1)) > 1,
            }
            for index, route_step in enumerate(PRODUCT_ROUTE_MAPS[product_name], start=1)
        ],
    }


def route_task_to_dict(batch: dict, current_step: dict, viewer_employee=None):
    assignee = route_batch_assignee_to_dict(batch)
    viewer_employee_id = viewer_employee[0] if viewer_employee else None
    is_assigned_to_viewer = bool(viewer_employee_id and batch.get("assigned_employee_id") == viewer_employee_id)
    is_free = not batch.get("assigned_employee_id")
    is_done = batch.get("status") == "done"
    work_state = "done" if is_done else (batch.get("work_state") or ("free" if is_free else "in_work"))
    status_text = {
        "free": "Свободно",
        "in_work": "В работе",
        "paused": "Пауза",
        "blocked": "Заблокировано",
        "done": "Завершено",
        "cancelled": "Отменено",
    }.get(work_state, "В работе")

    return {
        "id": batch["id"],
        "identity": route_batch_identity(batch),
        "product_name": batch["product_name"],
        "product_size": batch["product_size"],
        "product_color": format_color_label(batch["product_color"]),
        "quantity": batch["quantity"],
        "current_status": get_route_previous_status(batch),
        "position": current_step["position"],
        "operation": current_step["operation"],
        "status_after": current_step["status_after"],
        "category": route_step_category(current_step),
        "assigned_employee_id": batch.get("assigned_employee_id"),
        "assigned_employee": assignee,
        "assigned_employee_name": assignee["full_name"] if assignee else "",
        "assigned_at": batch.get("assigned_at") or "",
        "good_quantity": batch.get("good_quantity") or 0,
        "defect_quantity": batch.get("defect_quantity") or 0,
        "priority": batch.get("priority") or "normal",
        "due_date": batch.get("due_date") or "",
        "parent_batch_id": batch.get("parent_batch_id"),
        "trace_code": batch.get("trace_code") or f"RB-{batch['id']:06d}",
        "route_version": batch.get("route_version") or "",
        "work_state": work_state,
        "blocked_reason": batch.get("blocked_reason") or "",
        "paused_at": batch.get("paused_at") or "",
        "last_activity_at": batch.get("last_activity_at") or batch.get("updated_at") or "",
        "handover_count": int(batch.get("handover_count") or 0),
        "defects": get_route_batch_defects(batch["id"]),
        "inputs": [
            {
                **input_row,
                "product_color_label": format_color_label(input_row["product_color"]),
                "quantity_text": format_number(input_row["quantity"]),
            }
            for input_row in get_route_batch_inputs(batch["id"])
        ],
        "input_lots": get_route_batch_input_lots(batch["id"]),
        "work_status": work_state,
        "status_text": status_text,
        "is_assigned_to_me": is_assigned_to_viewer,
        "can_take": is_free and work_state == "free",
        "can_complete": is_assigned_to_viewer and work_state == "in_work",
        "can_pause": is_assigned_to_viewer and work_state == "in_work",
        "can_resume": is_assigned_to_viewer and work_state in {"paused", "blocked"},
        "can_release": is_assigned_to_viewer and not is_done,
    }


def get_route_tasks_for_telegram(telegram_id: int):
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля.", "tasks": []}

    tasks = []

    for batch in get_active_route_batches():
        current_step = get_route_step_for_batch(batch)

        if current_step is None:
            continue

        if is_admin(telegram_id) or can_employee_work_route_step(employee, current_step):
            tasks.append(route_task_to_dict(batch, current_step, employee))

    return {"ok": True, "tasks": tasks}


def get_completed_route_tasks_for_telegram(telegram_id: int):
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return []

    tasks = []

    for batch in get_employee_route_batches(employee[0], "done"):
        step_index = max(0, batch["route_step_index"] - 1)
        completed_step = get_route_step_for_batch(batch, step_index)

        if completed_step is None:
            continue

        tasks.append(route_task_to_dict(batch, completed_step, employee))

    return tasks


def start_route_task_for_telegram(telegram_id: int, batch_id: int):
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля."}

    if is_admin(telegram_id):
        return {"ok": False, "message": "Администратор не выполняет производственные задания."}

    if get_open_shift_for_today(employee[0]) is None:
        return {"ok": False, "message": "Откройте смену перед выбором задания."}

    batch = get_route_batch_by_id(batch_id)

    if batch is None or batch["status"] != "active":
        return {"ok": False, "message": "Это задание уже недоступно."}

    current_step = get_route_step_for_batch(batch)

    if not can_employee_work_route_step(employee, current_step):
        return {"ok": False, "message": "Это задание доступно другой должности."}

    if batch.get("assigned_employee_id") and batch["assigned_employee_id"] != employee[0]:
        assignee = route_batch_assignee_to_dict(batch)
        assignee_name = assignee["full_name"] if assignee else "другого сотрудника"
        return {"ok": False, "message": f"Задание уже в работе у {assignee_name}."}

    assigned_batch = assign_route_batch(batch_id, employee[0])

    if assigned_batch is None:
        return {"ok": False, "message": "Не удалось взять задание в работу."}

    add_edit_log(
        telegram_id,
        "employee",
        "Взял задание в работу из миниаппа",
        "route_batch",
        batch_id,
        route_batch_identity(assigned_batch),
    )

    return {
        "ok": True,
        "message": "Задание взято в работу.",
        "tasks": get_route_tasks_for_telegram(telegram_id)["tasks"],
        "completed_tasks": get_completed_route_tasks_for_telegram(telegram_id),
    }


def create_route_batch_for_telegram(telegram_id: int, payload: dict):
    if not ROUTES_MINIAPP_ENABLED:
        return {"ok": False, "message": "Создание маршрутов в мини-приложении отключено."}

    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля."}

    product_name = (payload.get("product_name") or "").strip()
    product_size = (payload.get("product_size") or "").strip()
    product_color = (payload.get("product_color") or "").strip()

    try:
        quantity = int(payload.get("quantity") or 0)
    except (TypeError, ValueError):
        quantity = 0

    if product_name not in PRODUCT_ROUTE_MAPS:
        return {"ok": False, "message": "Выберите изделие."}

    if quantity <= 0:
        return {"ok": False, "message": "Введите количество больше 0."}

    sizes = get_product_sizes(product_name)
    colors = get_product_colors(product_name)

    if sizes and product_size not in sizes:
        return {"ok": False, "message": "Выберите размер из списка."}

    if colors and product_color not in colors:
        return {"ok": False, "message": "Выберите цвет из списка."}

    batch = create_route_batch(
        product_name,
        product_size or "без размера",
        product_color or "без цвета",
        quantity,
        employee[0],
    )

    if batch is None:
        return {"ok": False, "message": "Не удалось создать партию."}

    current_step = get_route_step_for_batch(batch)
    add_edit_log(
        telegram_id,
        "admin" if is_admin(telegram_id) else "employee",
        "Создал задание по операции из миниаппа",
        "route_batch",
        batch["id"],
        route_batch_identity(batch),
    )

    return {
        "ok": True,
        "message": f"Задание #{batch['id']} создано.",
        "batch": route_task_to_dict(batch, current_step, employee),
        "tasks": get_route_tasks_for_telegram(telegram_id)["tasks"],
    }


def normalize_defect_photo(payload: dict):
    photo = payload.get("defect_photo") or {}
    if not isinstance(photo, dict) or not photo:
        return None, ""

    file_name = str(photo.get("file_name") or "defect-photo.jpg").strip()
    mime_type = str(photo.get("mime_type") or "").strip().lower()
    content_base64 = str(photo.get("content_base64") or "").strip()
    if not content_base64:
        return None, ""
    if mime_type not in {"image/jpeg", "image/png", "image/webp"}:
        return None, "Фото брака должно быть в формате JPG, PNG или WebP."
    if content_base64.startswith("data:"):
        content_base64 = content_base64.split(",", 1)[-1]
    try:
        raw_content = base64.b64decode(content_base64, validate=True)
    except (TypeError, ValueError):
        return None, "Не удалось прочитать фотографию брака."
    if not raw_content or len(raw_content) > MAX_DEFECT_PHOTO_BYTES:
        return None, "Фотография брака должна быть не больше 2 МБ."
    return {
        "file_name": file_name[:160],
        "mime_type": mime_type,
        "content_base64": content_base64,
    }, ""


def complete_route_task_for_telegram(telegram_id: int, batch_id: int, payload: dict | None = None):
    payload = payload or {}
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля."}

    if is_admin(telegram_id):
        return {"ok": False, "message": "Администратор не выполняет производственные задания."}

    request_id = str(payload.get("request_id") or "").strip()
    if request_id and has_route_completion_request(batch_id, employee[0], request_id):
        return {
            "ok": True,
            "message": "Задание уже было синхронизировано. Данные обновлены.",
            "tasks": get_route_tasks_for_telegram(telegram_id)["tasks"],
            "completed_tasks": get_completed_route_tasks_for_telegram(telegram_id),
            "production": get_production_state_for_telegram(telegram_id),
            "quality": {
                "defect_reasons": DEFECT_REASONS,
                "defect_dispositions": DEFECT_DISPOSITIONS,
            },
        }

    open_shift = get_open_shift_for_today(employee[0])
    if open_shift is None:
        return {"ok": False, "message": "Откройте смену перед выполнением задания."}

    batch = get_route_batch_by_id(batch_id)

    if batch is None or batch["status"] != "active":
        return {"ok": False, "message": "Эта партия уже недоступна."}

    current_step = get_route_step_for_batch(batch)

    if not can_employee_work_route_step(employee, current_step):
        return {"ok": False, "message": "Это задание сейчас доступно другой должности."}

    if batch.get("assigned_employee_id") != employee[0]:
        assignee = route_batch_assignee_to_dict(batch)
        if assignee:
            return {"ok": False, "message": f"Задание в работе у {assignee['full_name']}."}
        return {"ok": False, "message": "Сначала выберите задание во вкладке «Задания»."}

    work_state = batch.get("work_state") or "in_work"
    if work_state in {"paused", "blocked"}:
        return {"ok": False, "message": "Сначала возобновите задание."}

    try:
        good_quantity = int(payload.get("good_quantity") if payload.get("good_quantity") not in (None, "") else batch["quantity"])
        defect_quantity = int(payload.get("defect_quantity") or 0)
    except (TypeError, ValueError):
        return {"ok": False, "message": "Количество годного и брака должно быть целым числом."}

    if good_quantity < 0 or defect_quantity < 0:
        return {"ok": False, "message": "Количество не может быть отрицательным."}

    if good_quantity + defect_quantity != batch["quantity"]:
        return {"ok": False, "message": "Распределите всё количество задания между годным и браком."}

    defect_reason = str(payload.get("defect_reason") or "").strip()
    defect_disposition = str(payload.get("defect_disposition") or "").strip()
    defect_comment = str(payload.get("defect_comment") or "").strip()

    if defect_quantity > 0:
        if defect_reason not in DEFECT_REASONS:
            return {"ok": False, "message": "Выберите причину брака."}
        if defect_disposition not in DEFECT_DISPOSITIONS:
            return {"ok": False, "message": "Выберите решение по браку."}

    defect_photo, photo_error = normalize_defect_photo(payload)
    if photo_error:
        return {"ok": False, "message": photo_error}

    next_step_index = batch["route_step_index"] + 1
    next_step = get_route_step_for_batch(batch, next_step_index)
    item_type = "semifinished" if next_step else "finished"
    ready_for_position = next_step["position"] if next_step else "Склад"
    stage_name = current_step["status_after"]

    operation_row = next(
        (
            row for row in get_active_operations(position=current_step["position"])
            if row[2] == current_step["operation"]
        ),
        None,
    )
    completion = complete_route_batch_step_atomic(
        batch["id"],
        employee[0],
        batch["route_step_index"],
        current_step["operation"],
        current_step["position"],
        next_step_index,
        good_quantity,
        defect_quantity,
        item_type,
        stage_name,
        ready_for_position,
        auto_create_next=should_auto_create_next_preparation_task(current_step, next_step),
        defect_reason=defect_reason,
        defect_disposition=defect_disposition,
        defect_comment=defect_comment,
        defect_photo_name=(defect_photo or {}).get("file_name", ""),
        defect_photo_mime_type=(defect_photo or {}).get("mime_type", ""),
        defect_photo_base64=(defect_photo or {}).get("content_base64", ""),
        shift_id=open_shift[0] if open_shift else None,
        operation_id=operation_row[0] if operation_row else None,
        request_id=request_id,
    )

    if completion is None:
        return {"ok": False, "message": "Задание уже завершено или данные изменились. Обновите список."}

    updated_batch = completion["batch"]
    auto_batch = completion["auto_batch"]
    rework_batch = completion["rework_batch"]

    add_edit_log(
        telegram_id,
        "admin" if is_admin(telegram_id) else "employee",
        "Завершил задание по операции из миниаппа",
        "route_batch",
        batch["id"],
        f"{route_batch_identity(batch)}. Этап: {current_step['operation']}",
    )

    if completion.get("replayed"):
        message = "Задание уже было синхронизировано. Данные обновлены."
    elif rework_batch:
        message = f"Этап завершён. Брак: {defect_quantity} шт. Создано задание на переделку #{rework_batch['id']}."
    elif auto_batch:
        message = f"Этап завершён. Следующее задание создано для должности {next_step['position']}."
    elif next_step:
        message = f"Этап завершён. Результат на складе для должности {next_step['position']}."
    else:
        message = "Этап завершён. Готовая продукция принята на склад."

    return {
        "ok": True,
        "message": message,
        "tasks": get_route_tasks_for_telegram(telegram_id)["tasks"],
        "completed_tasks": get_completed_route_tasks_for_telegram(telegram_id),
        "production": get_production_state_for_telegram(telegram_id),
        "quality": {
            "defect_reasons": DEFECT_REASONS,
            "defect_dispositions": DEFECT_DISPOSITIONS,
        },
    }


def route_task_work_action_for_telegram(telegram_id: int, batch_id: int, payload: dict):
    employee = get_employee_for_access(telegram_id)
    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля."}

    action = str(payload.get("action") or "").strip().lower()
    reason = str(payload.get("reason") or "").strip()
    if action not in {"pause", "block", "resume", "release"}:
        return {"ok": False, "message": "Неизвестное действие с заданием."}
    if action in {"block", "release"} and not reason:
        return {"ok": False, "message": "Укажите причину."}
    if action == "resume" and not is_admin(telegram_id) and get_open_shift_for_today(employee[0]) is None:
        return {"ok": False, "message": "Откройте смену перед продолжением задания."}

    batch = get_route_batch_by_id(batch_id)
    if batch is None or batch["status"] != "active":
        return {"ok": False, "message": "Задание уже недоступно."}
    current_step = get_route_step_for_batch(batch)
    if not is_admin(telegram_id) and not can_employee_work_route_step(employee, current_step):
        return {"ok": False, "message": "Это задание доступно другой должности."}

    updated_batch = set_route_batch_work_state(
        batch_id,
        employee[0],
        action,
        reason,
        force=is_admin(telegram_id),
    )
    if updated_batch is None:
        return {"ok": False, "message": "Не удалось изменить состояние задания. Обновите список."}

    messages = {
        "pause": "Задание поставлено на паузу.",
        "block": "Проблема зафиксирована. Администратор увидит блокировку.",
        "resume": "Работа по заданию продолжена.",
        "release": "Задание передано обратно в свободные.",
    }
    add_edit_log(
        telegram_id,
        "admin" if is_admin(telegram_id) else "employee",
        messages[action],
        "route_batch",
        batch_id,
        reason or route_batch_identity(updated_batch),
    )
    return {
        "ok": True,
        "message": messages[action],
        "batch": route_task_to_dict(updated_batch, get_route_step_for_batch(updated_batch), employee),
        "tasks": get_route_tasks_for_telegram(telegram_id).get("tasks", []),
        "completed_tasks": get_completed_route_tasks_for_telegram(telegram_id),
    }


def can_access_route_batch(telegram_id: int, batch: dict | None):
    if batch is None:
        return False
    if is_admin(telegram_id):
        return True
    employee = get_employee_for_access(telegram_id)
    if employee is None or employee[5] != "active":
        return False
    if batch.get("assigned_employee_id") == employee[0]:
        return True
    return batch.get("status") == "active" and can_employee_work_route_step(employee, get_route_step_for_batch(batch))


def get_route_passport_for_telegram(telegram_id: int, batch_id: int):
    batch = get_route_batch_by_id(batch_id)
    if not can_access_route_batch(telegram_id, batch):
        return {"ok": False, "message": "Нет доступа к паспорту партии."}
    passport = get_route_batch_passport(batch_id)
    if passport is None:
        return {"ok": False, "message": "Паспорт партии не найден."}

    event_labels = {
        "task_created": "Создано задание на раскрой",
        "materials_reserved": "Материал зарезервирован",
        "task_started": "Задание взято в работу",
        "cutting_contours_done": "Контуры нанесены",
        "cutting_layout_done": "Настил сформирован",
        "cutting_progress": "Раскрой обновлён",
        "cutting_output_received": "Крой принят на склад",
        "cutting_formed": "Раскрой завершён",
        "batch_created": "Создано производственное задание",
        "inputs_reserved": "Компоненты зарезервированы",
        "task_paused": "Работа приостановлена",
        "task_blocked": "Зафиксирована блокировка",
        "task_resumed": "Работа продолжена",
        "task_released": "Задание передано",
        "operation_completed": "Операция завершена",
        "task_cancelled": "Задание отменено",
    }
    passport["events"] = [
        {**event, "event_text": event_labels.get(event["event_type"], event["event_type"])}
        for event in passport.get("events", [])
    ]
    passport["batches"] = [
        {
            **item,
            "product_color_label": format_color_label(item.get("product_color") or ""),
            "assignee": route_batch_assignee_to_dict(item),
            "step": get_route_step_for_batch(
                item,
                max(0, item["route_step_index"] - 1) if item.get("status") == "done" else item["route_step_index"],
            ),
        }
        for item in passport.get("batches", [])
    ]
    return {"ok": True, "passport": passport}


def lookup_route_trace_for_telegram(telegram_id: int, trace_code: str):
    batch = get_route_batch_by_trace_code(trace_code)
    if not can_access_route_batch(telegram_id, batch):
        return {"ok": False, "message": "Партия не найдена или недоступна."}
    employee = get_employee_for_access(telegram_id)
    step_index = max(0, batch["route_step_index"] - 1) if batch["status"] == "done" else batch["route_step_index"]
    return {
        "ok": True,
        "batch": route_task_to_dict(batch, get_route_step_for_batch(batch, step_index), employee),
    }


def get_route_catalog():
    return [route_map_to_dict(product_name) for product_name in PRODUCT_ROUTE_MAPS]


def get_routes_payload(telegram_id: int):
    return {
        "enabled": ROUTES_MINIAPP_ENABLED,
        "catalog": get_route_catalog(),
        "tasks": get_route_tasks_for_telegram(telegram_id).get("tasks", []),
        "completed_tasks": get_completed_route_tasks_for_telegram(telegram_id),
    }


def sort_size_key(value: str):
    value_text = str(value)

    if value_text.isdigit():
        return (0, int(value_text))

    return (1, value_text)


def split_group_concat(value: str | None, *, sort_sizes: bool = False):
    items = [item for item in (value or "").split(",") if item]

    if sort_sizes:
        return sorted(items, key=sort_size_key)

    return sorted(items)


def format_number(value: float):
    if float(value).is_integer():
        return str(int(value))

    return str(value).rstrip("0").rstrip(".")


def parse_positive_int(value):
    try:
        parsed = int(str(value or "").strip())
    except (TypeError, ValueError):
        return None

    if parsed <= 0:
        return None

    return parsed


def parse_task_planning(payload: dict):
    priority = str(payload.get("priority") or "normal").strip()
    due_date = str(payload.get("due_date") or "").strip()

    if priority not in {"low", "normal", "high", "urgent"}:
        return None, None, "Выберите приоритет задания."

    if due_date:
        try:
            time.strptime(due_date, "%Y-%m-%d")
        except ValueError:
            return None, None, "Укажите срок в формате даты."

    return priority, due_date, ""


def infer_attachment_mime_type(file_name: str, mime_type: str):
    lower_name = file_name.lower()

    if mime_type and mime_type != "application/octet-stream":
        return mime_type

    if lower_name.endswith(".pdf"):
        return "application/pdf"

    if lower_name.endswith(".docx"):
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

    if lower_name.endswith(".doc"):
        return "application/msword"

    if lower_name.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if lower_name.endswith(".xls"):
        return "application/vnd.ms-excel"

    return mime_type


def normalize_task_attachment(payload: dict):
    attachment = payload.get("attachment") or {}

    if not attachment:
        return None

    file_name = str(attachment.get("file_name") or "").strip()
    mime_type = str(attachment.get("mime_type") or "").strip()
    content_base64 = str(attachment.get("content_base64") or "").strip()
    allowed_extensions = (".pdf", ".doc", ".docx", ".xls", ".xlsx")
    allowed_mime_prefixes = (
        "application/pdf",
        "application/msword",
        "application/vnd.openxmlformats-officedocument",
        "application/vnd.ms-excel",
    )

    if not file_name or not content_base64:
        return None

    if len(content_base64) > ((MAX_ATTACHMENT_BYTES + 2) // 3) * 4 + 8:
        return {"error": "Файл должен быть не больше 10 МБ."}

    try:
        decoded_content = base64.b64decode(content_base64, validate=True)
    except (ValueError, TypeError):
        return {"error": "Файл повреждён или передан в неверном формате."}

    if len(decoded_content) > MAX_ATTACHMENT_BYTES:
        return {"error": "Файл должен быть не больше 10 МБ."}

    if not file_name.lower().endswith(allowed_extensions):
        return {"error": "Можно прикрепить только Word, Excel или PDF."}

    mime_type = infer_attachment_mime_type(file_name, mime_type)

    if not mime_type:
        return {"error": "Формат файла должен быть Word, Excel или PDF."}

    if not any(mime_type.startswith(prefix) for prefix in allowed_mime_prefixes):
        return {"error": "Формат файла должен быть Word, Excel или PDF."}

    return {
        "file_name": file_name,
        "mime_type": mime_type,
        "content_base64": content_base64,
    }


def get_order_color_options():
    colors = []

    for color in get_all_product_colors():
        if color not in colors:
            colors.append(color)

    for _material_name, product_color, _quantity, _unit, _updated_at in get_fabric_stock_rows():
        if product_color not in colors:
            colors.append(product_color)

    return colors


def task_fabric_rolls_to_dict(task_id: int | None):
    if not task_id:
        return []

    defects = get_production_task_fabric_defects(task_id)
    rejected_by_item = {}
    for defect in defects:
        key = (defect["material_name"], defect["product_color"])
        rejected_by_item[key] = rejected_by_item.get(key, 0) + int(defect["quantity"] or 0)

    return [
        {
            "material_name": material_name,
            "product_color": product_color,
            "product_color_label": format_color_label(product_color),
            "rolls": rolls,
            "rejected_rolls": rejected_by_item.get((material_name, product_color), 0),
            "available_rolls": max(0, int(rolls or 0) - rejected_by_item.get((material_name, product_color), 0)),
            "defects": [
                {
                    **defect,
                    "product_color_label": format_color_label(defect["product_color"]),
                }
                for defect in defects
                if defect["material_name"] == material_name and defect["product_color"] == product_color
            ],
        }
        for material_name, product_color, rolls in get_production_task_fabric_rolls(task_id)
    ]


def task_attachment_to_dict(task_id: int | None):
    if not task_id:
        return None

    attachment = get_production_task_attachment(task_id)

    if attachment is None:
        return None

    return {
        "task_id": task_id,
        "file_name": attachment["file_name"],
        "mime_type": attachment["mime_type"],
    }


def production_catalog_to_dict():
    catalog = []

    for product_name in PRODUCT_ROUTE_MAPS:
        colors = get_product_colors(product_name)
        catalog.append(
            {
                "product_name": product_name,
                "sizes": get_product_sizes(product_name),
                "colors": colors,
                "color_labels": [format_color_label(color) for color in colors],
            }
        )

    return catalog


def production_task_to_dict(row, viewer_employee=None):
    (
        task_id,
        product_name,
        status,
        created_at,
        sizes_text,
        colors_text,
        priority,
        due_date,
        assigned_employee_id,
        assigned_at,
    ) = row
    colors = split_group_concat(colors_text)
    fabric_rolls = task_fabric_rolls_to_dict(task_id)
    attachment = task_attachment_to_dict(task_id)
    assigned_employee = get_employee_by_id(assigned_employee_id) if assigned_employee_id else None
    assignee = employee_to_dict(assigned_employee) if assigned_employee else None
    viewer_employee_id = viewer_employee[0] if viewer_employee else None
    is_assigned_to_viewer = bool(viewer_employee_id and assigned_employee_id == viewer_employee_id)
    is_free = not assigned_employee_id
    process_status_text = {
        "active": "ожидает контуров",
        "contours_done": "контуры нанесены",
        "in_cutting": "в раскрое",
        "formed": "готовый крой сформирован",
        "cancelled": "отменено",
    }.get(status, status)

    return {
        "id": task_id,
        "product_name": product_name,
        "status": status,
        "status_text": "Свободно" if is_free else "В работе",
        "process_status_text": process_status_text,
        "work_status": "free" if is_free else "in_work",
        "created_at": created_at,
        "priority": priority or "normal",
        "due_date": due_date or "",
        "sizes": split_group_concat(sizes_text, sort_sizes=True),
        "colors": colors,
        "color_labels": [format_color_label(color) for color in colors],
        "fabric_rolls": fabric_rolls,
        "attachment": attachment,
        "assigned_employee_id": assigned_employee_id,
        "assigned_employee": assignee,
        "assigned_employee_name": assignee["full_name"] if assignee else "",
        "assigned_at": assigned_at or "",
        "is_assigned_to_me": is_assigned_to_viewer,
        "can_take": is_free,
    }


def fabric_stock_to_dict(row):
    stock_id, material_name, product_color, quantity, unit, updated_at = row

    return {
        "id": stock_id,
        "material_name": material_name,
        "product_color": product_color,
        "product_color_label": format_color_label(product_color),
        "quantity": quantity,
        "quantity_text": format_number(quantity),
        "unit": unit,
        "updated_at": updated_at,
    }


def warehouse_stock_to_dict(row: dict):
    item_type_text = {
        "semifinished": "Полуфабрикаты",
        "finished": "Готовая продукция",
    }.get(row["item_type"], row["item_type"])

    return {
        **row,
        "item_type_text": item_type_text,
        "product_color_label": format_color_label(row["product_color"]),
        "quantity_text": format_number(row["quantity"]),
        "title": f"{row['product_name']} — {row['stage_name']}",
    }


def get_cutting_operation(product_name: str, operation_name: str):
    operations = get_active_operations("Раскройщик", product_name, "Раскрой изделий")

    for operation_id, _number, name, unit in operations:
        if name == operation_name:
            return {
                "id": operation_id,
                "name": f"{product_name}: {name}",
                "base_name": name,
                "unit": unit,
            }

    return None


def get_cutting_contour_operation(product_name: str):
    return get_cutting_operation(product_name, CUTTING_CONTOUR_OPERATION)


def cutting_product_names():
    product_names = list(PRODUCT_ROUTE_MAPS)

    for product_name in get_active_cutting_batch_product_names():
        if product_name not in product_names:
            product_names.append(product_name)

    return product_names


def cutting_task_assignment_fields(task_id: int | None, viewer_employee=None, batch_id: int | None = None):
    task = get_production_task_by_id(task_id) if task_id else None
    legacy_owner = get_cutting_batch_owner(batch_id) if not task and batch_id else None
    assigned_employee_id = task.get("assigned_employee_id") if task else (legacy_owner.get("employee_id") if legacy_owner else None)
    assigned_employee = get_employee_by_id(assigned_employee_id) if assigned_employee_id else None
    assignee = employee_to_dict(assigned_employee) if assigned_employee else None
    viewer_employee_id = viewer_employee[0] if viewer_employee else None
    is_assigned_to_viewer = bool(viewer_employee_id and assigned_employee_id == viewer_employee_id)
    is_free = not assigned_employee_id
    return {
        "assigned_employee_id": assigned_employee_id,
        "assigned_employee": assignee,
        "assigned_employee_name": assignee["full_name"] if assignee else "",
        "assigned_at": task.get("assigned_at") if task else (legacy_owner.get("assigned_at") if legacy_owner else ""),
        "work_status": "free" if is_free else "in_work",
        "status_text": "Свободно" if is_free else "В работе",
        "is_assigned_to_me": is_assigned_to_viewer,
        "can_take": is_free,
    }


def cutting_stage_task_to_dict(stage: str, row, viewer_employee=None):
    if stage == "contours":
        task = production_task_to_dict(row, viewer_employee)
        return {
            **task,
            "task_kind": "cutting_stage",
            "category": "Раскрой",
            "stage": "contours",
            "stage_title": "Нанесение контуров лекал на ткань",
            "source_id": task["id"],
            "next_action": "Начать контуры",
        }

    if stage == "layout":
        batch_id, product_name, contour_date, production_task_id, employee_name, sizes_text, colors_text = row
        colors = split_group_concat(colors_text)
        return {
            **cutting_task_assignment_fields(production_task_id, viewer_employee, batch_id),
            "id": batch_id,
            "source_id": batch_id,
            "production_task_id": production_task_id,
            "task_kind": "cutting_stage",
            "category": "Раскрой",
            "stage": "layout",
            "stage_title": "Формирование настила",
            "next_action": "Сохранить настил",
            "product_name": product_name,
            "status": "contours_done",
            "process_status_text": "контуры нанесены",
            "created_at": contour_date,
            "employee": employee_name or "",
            "sizes_text": sizes_text or "",
            "colors": colors,
            "color_labels": [format_color_label(color) for color in colors],
            "fabric_rolls": task_fabric_rolls_to_dict(production_task_id),
            "attachment": task_attachment_to_dict(production_task_id),
        }

    if stage == "cutting":
        batch_id, product_name, contour_date, layout_date, progress, production_task_id, colors_text = row
        return {
            **cutting_task_assignment_fields(production_task_id, viewer_employee, batch_id),
            "id": batch_id,
            "source_id": batch_id,
            "production_task_id": production_task_id,
            "task_kind": "cutting_stage",
            "category": "Раскрой",
            "stage": "cutting",
            "stage_title": "Раскрой",
            "next_action": "Обновить процент",
            "product_name": product_name,
            "status": "in_cutting" if progress else "layout_done",
            "process_status_text": f"готовность {progress or 0}%",
            "created_at": layout_date or contour_date,
            "progress": progress or 0,
            "colors_text": colors_text or "",
            "fabric_rolls": task_fabric_rolls_to_dict(production_task_id),
            "attachment": task_attachment_to_dict(production_task_id),
        }

    batch_id, product_name, contour_date, layout_date, progress, production_task_id = row
    return {
        **cutting_task_assignment_fields(production_task_id, viewer_employee, batch_id),
        "id": batch_id,
        "source_id": batch_id,
        "production_task_id": production_task_id,
        "task_kind": "cutting_stage",
        "category": "Раскрой",
        "stage": "formation",
        "stage_title": "Формирование готового кроя",
        "next_action": "Принять крой на склад",
        "product_name": product_name,
        "status": "cutting_done",
        "process_status_text": "раскрой завершён",
        "created_at": layout_date or contour_date,
        "progress": progress or 100,
        "fabric_rolls": task_fabric_rolls_to_dict(production_task_id),
        "attachment": task_attachment_to_dict(production_task_id),
    }


def get_cutting_stage_tasks(employee):
    if not employee or employee[5] != "active" or employee[3] != "Раскройщик":
        return []

    tasks = [
        cutting_stage_task_to_dict("contours", row, employee)
        for row in get_active_production_tasks_for_contours()
    ]

    for product_name in cutting_product_names():
        tasks.extend(cutting_stage_task_to_dict("layout", row, employee) for row in get_cutting_batches_for_layout(product_name))
        tasks.extend(cutting_stage_task_to_dict("cutting", row, employee) for row in get_cutting_batches_for_cutting(product_name))
        tasks.extend(cutting_stage_task_to_dict("formation", row, employee) for row in get_cutting_batches_for_formation(product_name))

    return tasks


def start_cutting_task_for_telegram(telegram_id: int, task_id: int):
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля."}

    if employee[3] != "Раскройщик":
        return {"ok": False, "message": "Задания на раскрой доступны раскройщику."}

    if get_open_shift_for_today(employee[0]) is None:
        return {"ok": False, "message": "Откройте смену перед выбором задания."}

    task = get_production_task_by_id(task_id)

    if task is None or task["status"] not in {"active", "contours_done", "in_cutting"}:
        return {"ok": False, "message": "Задание уже недоступно."}

    if task.get("assigned_employee_id") not in (None, employee[0]):
        assignee = get_employee_by_id(task["assigned_employee_id"])
        assignee_name = assignee[1] if assignee else "другого сотрудника"
        return {"ok": False, "message": f"Задание уже в работе у {assignee_name}."}

    assigned_task = assign_production_task(task_id, employee[0])

    if assigned_task is None:
        current_task = get_production_task_by_id(task_id)
        assignee = get_employee_by_id(current_task.get("assigned_employee_id")) if current_task else None
        assignee_name = assignee[1] if assignee else "другого сотрудника"
        return {"ok": False, "message": f"Задание уже в работе у {assignee_name}."}

    add_edit_log(
        telegram_id,
        "employee",
        "Взял задание на раскрой в работу из миниаппа",
        "production_task",
        task_id,
        assigned_task["product_name"],
    )

    return {
        "ok": True,
        "message": "Задание взято в работу.",
        "production": get_production_state_for_telegram(telegram_id),
    }


def cutting_task_access_error(employee, task_id: int | None, batch_id: int | None = None):
    task = get_production_task_by_id(task_id) if task_id else None

    if task is None and batch_id:
        legacy_owner = get_cutting_batch_owner(batch_id)
        if legacy_owner and legacy_owner.get("employee_id") == employee[0]:
            return ""
        if legacy_owner and legacy_owner.get("employee_id"):
            assignee = get_employee_by_id(legacy_owner["employee_id"])
            assignee_name = assignee[1] if assignee else "другого сотрудника"
            return f"Задание в работе у {assignee_name}."

    if task is None or task["status"] not in {"active", "contours_done", "in_cutting"}:
        return "Задание уже недоступно."

    assigned_employee_id = task.get("assigned_employee_id")

    if assigned_employee_id is None:
        return "Сначала выберите задание во вкладке «Задания»."

    if assigned_employee_id != employee[0]:
        assignee = get_employee_by_id(assigned_employee_id)
        assignee_name = assignee[1] if assignee else "другого сотрудника"
        return f"Задание в работе у {assignee_name}."

    return ""


def get_production_state_for_telegram(telegram_id: int):
    employee = get_employee_for_access(telegram_id)
    is_admin_user = is_admin(telegram_id)
    can_work_contours = bool(employee and employee[5] == "active" and employee[3] == "Раскройщик")
    order_colors = get_order_color_options()
    warehouse_rows = [warehouse_stock_to_dict(row) for row in get_warehouse_stock_rows()]

    return {
        "catalog": production_catalog_to_dict(),
        "order_colors": order_colors,
        "order_color_labels": [format_color_label(color) for color in order_colors],
        "fabric_stock": [fabric_stock_to_dict(row) for row in get_fabric_stock_rows_with_ids()] if is_admin_user else [],
        "warehouse_stock": warehouse_rows if is_admin_user else [],
        "tasks": [production_task_to_dict(row, employee) for row in get_active_production_tasks()] if is_admin_user else [],
        "cutting_tasks": get_cutting_stage_tasks(employee),
        "contour_tasks": [
            production_task_to_dict(row, employee)
            for row in get_active_production_tasks_for_contours()
        ] if can_work_contours else [],
        "can_admin": is_admin_user,
        "can_contours": can_work_contours,
    }


def add_fabric_receipt_for_telegram(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    material_name = (payload.get("material_name") or "").strip()
    product_color = (payload.get("product_color") or "").strip()
    quantity = parse_positive_int(payload.get("quantity"))

    if not material_name:
        return {"ok": False, "message": "Введите материал."}

    if not product_color:
        return {"ok": False, "message": "Выберите цвет."}

    if quantity is None:
        return {"ok": False, "message": "Введите количество рулонов больше 0."}

    employee = get_employee_for_access(telegram_id)
    row = add_fabric_receipt(
        material_name,
        product_color,
        quantity,
        employee[0] if employee else None,
        unit="рул",
    )

    if row is None:
        return {"ok": False, "message": "Не удалось сохранить приход ткани."}

    stock_id, stock_material, stock_color, total_quantity, unit, _updated_at = row
    add_edit_log(
        telegram_id,
        "admin",
        "Добавил приход ткани из миниаппа",
        "fabric_stock",
        stock_id,
        f"{stock_material}, {stock_color}: +{format_number(quantity)} {unit}, остаток {format_number(total_quantity)} {unit}",
    )

    return {
        "ok": True,
        "message": "Приход ткани сохранён.",
        "production": get_production_state_for_telegram(telegram_id),
    }


def adjust_stock_for_telegram(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    stock_kind = (payload.get("stock_kind") or "").strip()
    reason = (payload.get("reason") or "").strip()
    try:
        stock_id = int(payload.get("stock_id") or 0)
        new_quantity = int(payload.get("quantity"))
    except (TypeError, ValueError):
        return {"ok": False, "message": "Остаток должен быть целым числом от 0."}

    if stock_id <= 0 or new_quantity < 0:
        return {"ok": False, "message": "Остаток должен быть целым числом от 0."}
    if not reason:
        return {"ok": False, "message": "Укажите причину корректировки."}

    employee = get_employee_for_access(telegram_id)
    employee_id = employee[0] if employee else None

    if stock_kind == "fabric":
        row = adjust_fabric_stock(stock_id, new_quantity, employee_id, reason)
        entity_type = "fabric_stock"
    elif stock_kind == "warehouse":
        row = adjust_warehouse_stock(stock_id, new_quantity, employee_id, reason)
        entity_type = "warehouse_stock"
    else:
        return {"ok": False, "message": "Неизвестный вид складского остатка."}

    if row is None:
        return {
            "ok": False,
            "message": "Не удалось скорректировать остаток. Обновите склад и проверьте новое количество.",
        }

    add_edit_log(
        telegram_id,
        "admin",
        "Скорректировал складской остаток",
        entity_type,
        stock_id,
        f"Новый остаток: {new_quantity}; причина: {reason}",
    )
    return {
        "ok": True,
        "message": "Остаток скорректирован, причина записана в журнал движений.",
        "production": get_production_state_for_telegram(telegram_id),
    }


def reject_fabric_rolls_for_telegram(telegram_id: int, payload: dict):
    employee = get_employee_for_access(telegram_id)
    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля."}
    if employee[3] != "Раскройщик":
        return {"ok": False, "message": "Списывать рулоны в брак может только раскройщик."}
    if get_open_shift_for_today(employee[0]) is None:
        return {"ok": False, "message": "Откройте смену перед списанием рулонов в брак."}

    product_color = (payload.get("product_color") or "").strip()
    comment = (payload.get("comment") or "").strip()
    quantity = parse_positive_int(payload.get("quantity"))
    try:
        task_id = int(payload.get("task_id") or 0)
    except (TypeError, ValueError):
        task_id = 0

    if task_id <= 0 or not product_color:
        return {"ok": False, "message": "Выберите рулоны из текущего задания."}
    if quantity is None:
        return {"ok": False, "message": "Количество брака должно быть больше 0."}
    if not comment:
        return {"ok": False, "message": "Добавьте комментарий к браку рулонов."}

    result = reject_production_task_fabric_rolls(
        task_id,
        employee[0],
        product_color,
        quantity,
        comment,
    )
    if result is None:
        return {
            "ok": False,
            "message": "Не удалось списать рулоны. Проверьте доступное количество и что задание находится у вас в работе.",
        }

    add_edit_log(
        telegram_id,
        "employee",
        "Списал рулоны ткани в брак",
        "production_task",
        task_id,
        f"{format_color_label(product_color)}: {quantity} рул.; {comment}",
    )
    return {
        "ok": True,
        "message": f"В брак списано: {quantity} рул.",
        "production": get_production_state_for_telegram(telegram_id),
    }


def create_production_task_for_telegram(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    product_name = (payload.get("product_name") or "").strip()
    material_name = (payload.get("material_name") or "Ткань").strip() or "Ткань"
    sizes = [str(size).strip() for size in payload.get("sizes", []) if str(size).strip()]
    colors = [str(color).strip() for color in payload.get("colors", []) if str(color).strip()]
    raw_fabric_rolls = payload.get("fabric_rolls") or {}
    priority, due_date, planning_error = parse_task_planning(payload)

    if planning_error:
        return {"ok": False, "message": planning_error}

    if product_name not in PRODUCT_ROUTE_MAPS:
        return {"ok": False, "message": "Выберите изделие."}

    allowed_sizes = set(get_product_sizes(product_name))
    allowed_colors = set(get_order_color_options())

    if not sizes or any(size not in allowed_sizes for size in sizes):
        return {"ok": False, "message": "Выберите размеры из списка."}

    if not colors or any(color not in allowed_colors for color in colors):
        return {"ok": False, "message": "Выберите цвета из списка."}

    if material_name != "Ткань":
        return {"ok": False, "message": "Пока для раскроя доступен только материал: Ткань."}

    fabric_rolls = {}

    for color in colors:
        rolls = parse_positive_int(raw_fabric_rolls.get(color))

        if rolls is None:
            return {"ok": False, "message": f"Укажите количество рулонов для цвета {format_color_label(color)}."}

        fabric_rolls[color] = rolls

    attachment = normalize_task_attachment(payload)

    if attachment and attachment.get("error"):
        return {"ok": False, "message": attachment["error"]}

    employee = get_employee_for_access(telegram_id)
    task = create_production_task(
        product_name,
        sizes,
        colors,
        employee[0] if employee else None,
        f"Материал: {material_name}; рулоны: " + ", ".join(f"{format_color_label(color)} — {rolls} рул." for color, rolls in fabric_rolls.items()),
        fabric_rolls=fabric_rolls,
        material_name=material_name,
        attachment=attachment,
        priority=priority,
        due_date=due_date,
    )

    if task is None:
        return {"ok": False, "message": "Не удалось создать задание. Проверьте остатки рулонов на складе."}

    add_edit_log(
        telegram_id,
        "admin",
        "Создал производственное задание из миниаппа",
        "production_task",
        task["id"],
        f"{task['product_name']}; размеры: {', '.join(sizes)}; рулоны: {fabric_rolls}",
    )

    return {
        "ok": True,
        "message": f"Задание #{task['id']} создано.",
        "production": get_production_state_for_telegram(telegram_id),
    }


def create_order_task_for_telegram(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    product_name = (payload.get("product_name") or "").strip()
    task_type = (payload.get("task_type") or "cutting").strip()
    material_name = (payload.get("material_name") or "").strip()
    sizes = [str(size).strip() for size in payload.get("sizes", []) if str(size).strip()]
    colors = [str(color).strip() for color in payload.get("colors", []) if str(color).strip()]
    raw_fabric_rolls = payload.get("fabric_rolls") or {}
    priority, due_date, planning_error = parse_task_planning(payload)

    if planning_error:
        return {"ok": False, "message": planning_error}

    if product_name not in PRODUCT_ROUTE_MAPS:
        return {"ok": False, "message": "Выберите изделие."}

    if task_type not in {"cutting", "route"}:
        return {"ok": False, "message": "Выберите тип задания."}

    employee = get_employee_for_access(telegram_id)
    employee_id = employee[0] if employee else None

    if task_type == "cutting":
        allowed_sizes = set(get_product_sizes(product_name))
        allowed_colors = set(get_order_color_options())

        if not sizes or any(size not in allowed_sizes for size in sizes):
            return {"ok": False, "message": "Выберите размеры из списка."}

        if not colors or any(color not in allowed_colors for color in colors):
            return {"ok": False, "message": "Выберите цвета из списка."}

        material_name = material_name or "Ткань"

        if material_name != "Ткань":
            return {"ok": False, "message": "Пока для раскроя доступен только материал: Ткань."}

        fabric_rolls = {}

        for color in colors:
            rolls = parse_positive_int(raw_fabric_rolls.get(color))

            if rolls is None:
                return {"ok": False, "message": f"Укажите количество рулонов для цвета {format_color_label(color)}."}

            fabric_rolls[color] = rolls

        attachment = normalize_task_attachment(payload)

        if attachment and attachment.get("error"):
            return {"ok": False, "message": attachment["error"]}

        task = create_production_task(
            product_name,
            sizes,
            colors,
            employee_id,
            f"Материал: {material_name}; рулоны: " + ", ".join(f"{format_color_label(color)} — {rolls} рул." for color, rolls in fabric_rolls.items()),
            fabric_rolls=fabric_rolls,
            material_name=material_name,
            attachment=attachment,
            priority=priority,
            due_date=due_date,
        )

        if task is None:
            return {"ok": False, "message": "Не удалось создать задание на раскрой. Проверьте остатки рулонов на складе."}

        add_edit_log(
            telegram_id,
            "admin",
            "Создал задание на раскрой из миниаппа",
            "production_task",
            task["id"],
            f"{task['product_name']}; материал: {material_name}; размеры: {', '.join(sizes)}; рулоны: {fabric_rolls}",
        )

        return {
            "ok": True,
            "message": f"Задание на раскрой #{task['id']} создано.",
            "production": get_production_state_for_telegram(telegram_id),
            "routes": get_routes_payload(telegram_id),
        }

    try:
        route_step_index = int(payload.get("route_step_index"))
    except (TypeError, ValueError):
        route_step_index = -1

    route_step = get_route_step(product_name, route_step_index)

    if route_step is None:
        return {"ok": False, "message": "Выберите операцию."}

    if route_step_index < len(CUTTING_ROUTE):
        return {"ok": False, "message": "Для раскроя используйте тип задания «Раскрой»."}

    accepted_stock_stages = accepted_stock_stages_for_route_step(product_name, route_step_index)

    raw_stock_items = payload.get("stock_items") or []

    if not raw_stock_items:
        return {"ok": False, "message": "Выберите полуфабрикат со склада."}

    selected_stock_rows = []
    selected_stock_ids = set()

    for item in raw_stock_items:
        try:
            stock_id = int(item.get("stock_id") or 0)
            quantity = int(item.get("quantity") or 0)
        except (AttributeError, TypeError, ValueError):
            return {"ok": False, "message": "Количество по складу должно быть целым числом."}

        if quantity <= 0:
            continue

        if stock_id in selected_stock_ids:
            return {"ok": False, "message": "Один складской остаток выбран несколько раз."}

        stock_row = get_warehouse_stock_by_id(stock_id)

        if stock_row is None or stock_row["quantity"] <= 0:
            return {"ok": False, "message": "Один из полуфабрикатов уже недоступен."}

        if stock_row["item_type"] != "semifinished":
            return {"ok": False, "message": "Для задания выберите полуфабрикат."}

        if stock_row["product_name"] != product_name:
            return {"ok": False, "message": "Полуфабрикат не соответствует изделию."}

        if stock_row["ready_for_position"] != route_step["position"]:
            return {"ok": False, "message": f"Этот полуфабрикат не для должности {route_step['position']}."}

        if stock_row["stage_name"] not in accepted_stock_stages:
            return {
                "ok": False,
                "message": "Этот полуфабрикат получен после другой операции. Сначала завершите предыдущую или смежную операцию.",
            }

        if quantity > stock_row["quantity"]:
            return {"ok": False, "message": "Количество больше остатка на складе."}

        selected_stock_ids.add(stock_id)
        selected_stock_rows.append({**stock_row, "selected_quantity": quantity})

    if not selected_stock_rows:
        return {"ok": False, "message": "Введите количество хотя бы по одной строке склада."}

    primary_stage = "Раскроенные" if "Раскроенные" in accepted_stock_stages else accepted_stock_stages[0]
    component_stages = [stage for stage in accepted_stock_stages if stage != primary_stage]
    primary_rows = [row for row in selected_stock_rows if row["stage_name"] == primary_stage]

    if not primary_rows:
        return {"ok": False, "message": f"Выберите основной полуфабрикат: {primary_stage}."}

    component_budgets = {row["id"]: row["selected_quantity"] for row in selected_stock_rows}
    input_groups = []

    for primary_row in primary_rows:
        group_inputs = [
            {
                "stock_id": primary_row["id"],
                "quantity": primary_row["selected_quantity"],
                "input_role": "main",
            }
        ]

        for component_stage in component_stages:
            required_quantity = primary_row["selected_quantity"]
            candidates = [
                row
                for row in selected_stock_rows
                if row["stage_name"] == component_stage
                and stock_sizes_overlap(primary_row["product_size"], row["product_size"])
                and component_budgets.get(row["id"], 0) > 0
            ]

            for component_row in candidates:
                available_quantity = component_budgets[component_row["id"]]
                allocated_quantity = min(required_quantity, available_quantity)

                if allocated_quantity <= 0:
                    continue

                group_inputs.append(
                    {
                        "stock_id": component_row["id"],
                        "quantity": allocated_quantity,
                        "input_role": "component",
                    }
                )
                component_budgets[component_row["id"]] -= allocated_quantity
                required_quantity -= allocated_quantity

                if required_quantity == 0:
                    break

            if required_quantity > 0:
                return {
                    "ok": False,
                    "message": f"Для размера {primary_row['product_size']} не хватает компонента «{component_stage}».",
                }

        input_groups.append((primary_row, group_inputs))

    created_batches = create_route_batches_with_inputs(
        product_name,
        employee_id,
        route_step_index,
        [
            {
                "product_size": primary_row["product_size"],
                "product_color": primary_row["product_color"],
                "quantity": primary_row["selected_quantity"],
                "input_items": group_inputs,
            }
            for primary_row, group_inputs in input_groups
        ],
        priority=priority,
        due_date=due_date,
    )

    if not created_batches:
        return {"ok": False, "message": "Не удалось одновременно списать все компоненты задания."}

    add_edit_log(
        telegram_id,
        "admin",
        "Создал задания по операции из миниаппа",
        "route_batch",
        created_batches[0]["id"],
        (
            f"{product_name}; {route_step['position']} — {route_step['operation']}; "
            f"входов на задание: {len(accepted_stock_stages)}; партий: {len(created_batches)}"
        ),
    )

    return {
        "ok": True,
        "message": f"Создано заданий: {len(created_batches)}.",
        "production": get_production_state_for_telegram(telegram_id),
        "routes": get_routes_payload(telegram_id),
    }


def delete_order_task_for_telegram(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    task_kind = (payload.get("task_kind") or "").strip()

    try:
        task_id = int(payload.get("task_id") or 0)
    except (TypeError, ValueError):
        task_id = 0

    if task_id <= 0:
        return {"ok": False, "message": "Выберите задание."}

    employee = get_employee_for_access(telegram_id)
    employee_id = employee[0] if employee else None

    if task_kind == "production":
        task = get_production_task_by_id(task_id)

        if task is None:
            return {"ok": False, "message": "Задание не найдено."}

        cancelled_task = cancel_production_task(task_id, employee_id)

        if cancelled_task is None:
            return {"ok": False, "message": "Это задание уже нельзя удалить."}

        add_edit_log(
            telegram_id,
            "admin",
            "Удалил производственное задание из миниаппа",
            "production_task",
            task_id,
            f"{task['product_name']}; прежний статус: {task['status']}",
        )
        message = f"Задание #{task_id} удалено."
    elif task_kind == "route":
        batch = get_route_batch_by_id(task_id)

        if batch is None:
            return {"ok": False, "message": "Задание не найдено."}

        cancelled_batch = cancel_route_batch_and_restore_inputs(task_id, employee_id)

        if cancelled_batch is None:
            return {"ok": False, "message": "Это задание уже нельзя удалить."}

        add_edit_log(
            telegram_id,
            "admin",
            "Удалил задание по операции из миниаппа",
            "route_batch",
            task_id,
            route_batch_identity(batch),
        )
        message = f"Задание #{task_id} удалено."
    else:
        return {"ok": False, "message": "Неизвестный тип задания."}

    return {
        "ok": True,
        "message": message,
        "production": get_production_state_for_telegram(telegram_id),
        "routes": get_routes_payload(telegram_id),
    }


def submit_production_contours_for_telegram(telegram_id: int, payload: dict):
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля."}

    if employee[3] != "Раскройщик" and not is_admin(telegram_id):
        return {"ok": False, "message": "Задания на раскрой доступны раскройщику."}

    shift = get_open_shift_for_today(employee[0])

    if shift is None:
        return {"ok": False, "message": "Откройте смену перед выполнением задания."}

    try:
        task_id = int(payload.get("task_id") or 0)
    except (TypeError, ValueError):
        task_id = 0

    task_rows = [row for row in get_active_production_tasks_for_contours() if row[0] == task_id]

    if not task_rows:
        return {"ok": False, "message": "Задание уже недоступно."}

    access_error = cutting_task_access_error(employee, task_id)
    if access_error:
        return {"ok": False, "message": access_error}

    task = production_task_to_dict(task_rows[0], employee)
    operation = get_cutting_contour_operation(task["product_name"])

    if operation is None:
        return {"ok": False, "message": "Для изделия не найдена операция контуров."}

    raw_quantities = payload.get("quantities") or {}
    matrix_quantities = {}

    for color in task["colors"]:
        for product_size in task["sizes"]:
            key = f"{product_size}|{color}"
            try:
                quantity = int(raw_quantities.get(key) or 0)
            except (TypeError, ValueError):
                return {"ok": False, "message": "Количество должно быть целым числом."}

            if quantity < 0:
                return {"ok": False, "message": "Количество не может быть отрицательным."}

            matrix_quantities[(product_size, color)] = quantity

    positive_rows = [
        (product_size, color, quantity)
        for (product_size, color), quantity in matrix_quantities.items()
        if quantity > 0
    ]

    if not positive_rows:
        return {"ok": False, "message": "Укажите количество хотя бы в одной строке."}

    batch_id = create_cutting_contour_batch_for_task(
        task_id,
        task["product_name"],
        shift[0],
        employee[0],
        operation["id"],
        matrix_quantities,
    )

    if batch_id is None:
        return {"ok": False, "message": "Не удалось создать партию раскроя."}

    add_edit_log(
        telegram_id,
        "employee",
        "Выполнил контуры по производственному заданию из миниаппа",
        "production_task",
        task_id,
        f"Партия раскроя {batch_id}, строк добавлено: {len(positive_rows)}",
    )

    return {
        "ok": True,
        "message": f"Контуры сохранены. Партия раскроя #{batch_id}.",
        "production": get_production_state_for_telegram(telegram_id),
    }


def parse_cutting_sizes_text(value: str):
    sizes = []

    for item in (value or "").split(","):
        size = item.split(" - ", 1)[0].strip()

        if size:
            sizes.append(size)

    return sizes


def submit_cutting_stage_for_telegram(telegram_id: int, payload: dict):
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {"ok": False, "message": "Нет активного профиля."}

    if employee[3] != "Раскройщик" and not is_admin(telegram_id):
        return {"ok": False, "message": "Этапы раскроя доступны раскройщику."}

    shift = get_open_shift_for_today(employee[0])

    if shift is None:
        return {"ok": False, "message": "Откройте смену перед выполнением задания."}

    stage = (payload.get("stage") or "").strip()

    if stage == "contours":
        return submit_production_contours_for_telegram(telegram_id, payload)

    try:
        batch_id = int(payload.get("batch_id") or 0)
    except (TypeError, ValueError):
        batch_id = 0

    if batch_id <= 0:
        return {"ok": False, "message": "Выберите задание."}

    if stage == "layout":
        batch_row = None

        for product_name in cutting_product_names():
            for row in get_cutting_batches_for_layout(product_name):
                if row[0] == batch_id:
                    batch_row = row
                    break
            if batch_row:
                break

        if batch_row is None:
            return {"ok": False, "message": "Сначала нужно выполнить нанесение контуров."}

        access_error = cutting_task_access_error(employee, batch_row[3], batch_id)
        if access_error:
            return {"ok": False, "message": access_error}

        operation = get_cutting_operation(batch_row[1], CUTTING_LAYOUT_OPERATION)

        if operation is None:
            return {"ok": False, "message": "Для изделия не найдена операция настила."}

        raw_layers = payload.get("color_layers") or {}
        color_layers = {}
        colors = split_group_concat(batch_row[6])

        for color in colors:
            try:
                layers = int(raw_layers.get(color) or 0)
            except (TypeError, ValueError):
                return {"ok": False, "message": "Количество слоёв должно быть целым числом."}

            if layers < 0:
                return {"ok": False, "message": "Количество слоёв не может быть отрицательным."}

            if layers > 0:
                color_layers[color] = layers

        if not color_layers:
            return {"ok": False, "message": "Укажите слои хотя бы по одному цвету."}

        if not add_cutting_layout(batch_id, shift[0], employee[0], operation["id"], color_layers):
            return {"ok": False, "message": "Настил уже недоступен."}

        added_count = len(parse_cutting_sizes_text(batch_row[5])) * len(color_layers)

        add_edit_log(
            telegram_id,
            "employee",
            "Выполнил формирование настила из миниаппа",
            "cutting_batch",
            batch_id,
            f"Цветов: {len(color_layers)}, строк отчёта: {added_count}",
        )

        return {
            "ok": True,
            "message": "Настил сохранён. Следующий этап: раскрой.",
            "production": get_production_state_for_telegram(telegram_id),
        }

    if stage == "cutting":
        batch_row = None

        for product_name in cutting_product_names():
            for row in get_cutting_batches_for_cutting(product_name):
                if row[0] == batch_id:
                    batch_row = row
                    break
            if batch_row:
                break

        if batch_row is None:
            return {"ok": False, "message": "Сначала нужно выполнить формирование настила."}

        access_error = cutting_task_access_error(employee, batch_row[5], batch_id)
        if access_error:
            return {"ok": False, "message": access_error}

        operation = get_cutting_operation(batch_row[1], CUTTING_CUT_OPERATION)

        if operation is None:
            return {"ok": False, "message": "Для изделия не найдена операция раскроя."}

        try:
            progress = int(payload.get("progress") or 0)
        except (TypeError, ValueError):
            progress = 0

        if progress not in {25, 50, 75, 100}:
            return {"ok": False, "message": "Выберите готовность: 25, 50, 75 или 100%."}

        if not update_cutting_batch_progress(batch_id, shift[0], employee[0], operation["id"], progress):
            return {"ok": False, "message": "Раскрой уже недоступен."}

        add_edit_log(
            telegram_id,
            "employee",
            "Обновил процент раскроя из миниаппа",
            "cutting_batch",
            batch_id,
            f"{progress}%",
        )

        message = "Раскрой завершён. Следующий этап: формирование готового кроя." if progress == 100 else "Готовность раскроя сохранена."

        return {
            "ok": True,
            "message": message,
            "production": get_production_state_for_telegram(telegram_id),
        }

    if stage == "formation":
        batch_row = None

        for product_name in cutting_product_names():
            for row in get_cutting_batches_for_formation(product_name):
                if row[0] == batch_id:
                    batch_row = row
                    break
            if batch_row:
                break

        if batch_row is None:
            return {"ok": False, "message": "Сначала нужно завершить раскрой на 100%."}

        access_error = cutting_task_access_error(employee, batch_row[5], batch_id)
        if access_error:
            return {"ok": False, "message": access_error}

        operation = get_cutting_operation(batch_row[1], CUTTING_FORM_OPERATION)

        if operation is None:
            return {"ok": False, "message": "Для изделия не найдена операция формирования готового кроя."}

        result_rows = get_cutting_batch_result_rows(batch_id)

        if not mark_cutting_batch_formed(batch_id, shift[0], employee[0], operation["id"]):
            return {"ok": False, "message": "Готовый крой уже недоступен."}

        add_edit_log(
            telegram_id,
            "employee",
            "Сформировал готовый крой из миниаппа",
            "cutting_batch",
            batch_id,
            f"На склад передано строк: {len([row for row in result_rows if row[2] > 0])}",
        )

        return {
            "ok": True,
            "message": "Готовый крой принят на склад.",
            "production": get_production_state_for_telegram(telegram_id),
        }

    return {"ok": False, "message": "Неизвестный этап."}


ADMIN_MENU = [
    {
        "id": "requests",
        "title": "Заявки",
        "buttons": ["Заявки"],
    },
    {
        "id": "reports",
        "title": "Отчёты",
        "buttons": [
            "Отчёт за сегодня",
            "Отчёт за период",
            "Отчёт по сотруднику",
            "Править отчёт",
        ],
    },
    {
        "id": "shifts",
        "title": "Смены",
        "buttons": ["Открытые смены", "Последние смены", "Удалить смену"],
    },
    {
        "id": "employees",
        "title": "Сотрудники",
        "buttons": [
            "Список сотрудников",
            "Активные сотрудники",
            "Неактивные сотрудники",
            "Активировать сотрудника",
            "Отключить сотрудника",
            "Сменить должность",
        ],
    },
    {
        "id": "operations",
        "title": "Операции",
        "buttons": [
            "Список операций",
            "Добавить операцию",
            "Изменить операцию",
            "Скрыть операцию",
            "Вернуть операцию",
        ],
    },
]

POSITIONS = ["Швея", "Упаковщик", "Раскройщик", "Ремонт"]


def clean_date(value: str | None, fallback: str):
    value = (value or "").strip()

    if len(value) == 10 and value[4] == "-" and value[7] == "-":
        return value

    return fallback


def month_bounds():
    today = local_today()
    return today.replace(day=1).isoformat(), today.isoformat()


def quarter_bounds():
    today = local_today()
    quarter_start_month = ((today.month - 1) // 3) * 3 + 1
    return today.replace(month=quarter_start_month, day=1).isoformat(), today.isoformat()


def employee_admin_to_dict(employee, web_profiles=None):
    employee_id, full_name, position, telegram_id_value, employee_status = employee
    profile = (web_profiles or {}).get(int(telegram_id_value), {})

    return {
        "id": employee_id,
        "full_name": full_name,
        "position": position or "-",
        "telegram_id": telegram_id_value,
        "status": employee_status,
        "email": profile.get("email", ""),
        "phone": profile.get("phone", ""),
        "is_web_account": bool(profile),
    }


def user_account_admin_to_dict(employee, web_profiles=None):
    employee_id, full_name, position, telegram_id_value, employee_status, role, registered_at = employee
    profile = (web_profiles or {}).get(int(telegram_id_value), {})

    return {
        "id": employee_id,
        "full_name": full_name,
        "position": position or "-",
        "telegram_id": telegram_id_value,
        "status": employee_status,
        "role": role or "employee",
        "registered_at": registered_at or "",
        "email": profile.get("email", ""),
        "phone": profile.get("phone", ""),
        "is_web_account": bool(profile),
    }


def pending_employee_to_dict(employee, web_profiles=None):
    employee_id, full_name, position, telegram_id_value, registered_at = employee
    profile = (web_profiles or {}).get(int(telegram_id_value), {})

    return {
        "id": employee_id,
        "full_name": full_name,
        "position": position or "-",
        "telegram_id": telegram_id_value,
        "registered_at": registered_at,
        "email": profile.get("email", ""),
        "phone": profile.get("phone", ""),
        "is_web_account": bool(profile),
    }


def open_shift_to_dict(shift):
    shift_id, full_name, shift_date, start_time = shift

    return {
        "id": shift_id,
        "employee": full_name,
        "date": shift_date,
        "start_time": start_time,
    }


def recent_shift_to_dict(shift):
    shift_id, full_name, shift_date, start_time, end_time, shift_status, operation_count = shift

    return {
        "id": shift_id,
        "employee": full_name,
        "date": shift_date,
        "start_time": start_time,
        "end_time": end_time,
        "status": shift_status,
        "operation_count": operation_count,
    }


def shift_detail_to_dict(row):
    if len(row) == 7:
        shift_id, full_name, shift_date, start_time, end_time, total_minutes, shift_status = row
    else:
        shift_id = None
        shift_date, full_name, start_time, end_time, total_minutes, shift_status = row

    return {
        "id": shift_id,
        "date": shift_date,
        "employee": full_name,
        "start_time": start_time,
        "end_time": end_time,
        "total_minutes": total_minutes,
        "total_time": format_minutes(total_minutes or 0),
        "status": shift_status,
    }


def employee_shift_to_dict(row):
    shift_id, shift_date, start_time, end_time, total_minutes, shift_status = row

    return {
        "id": shift_id,
        "date": shift_date,
        "start_time": start_time,
        "end_time": end_time,
        "total_minutes": total_minutes,
        "total_time": format_minutes(total_minutes or 0),
        "status": shift_status,
    }


def employee_summary_to_dict(row):
    employee_id, full_name, shift_count, total_minutes = row[:4]

    return {
        "employee_id": employee_id,
        "full_name": full_name,
        "shift_count": shift_count,
        "total_minutes": total_minutes,
        "total_time": format_minutes(total_minutes or 0),
    }


def operation_summary_to_dict(row):
    (
        shift_date,
        full_name,
        work_group,
        operation_name,
        product_size,
        product_color,
        total_quantity,
        unit,
    ) = row

    return {
        "date": shift_date,
        "employee": full_name,
        "group": work_group,
        "operation": operation_name,
        "size": product_size,
        "color": format_color_label(product_color),
        "quantity": total_quantity,
        "unit": unit,
    }


def minutes_to_excel_time(total_minutes):
    if total_minutes is None:
        return ""

    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours}:{minutes:02d}"


def style_excel_sheet(sheet):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="F1E1D6")
    header_font = Font(bold=True)

    if sheet.max_row == 0:
        return

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = min(max_length + 3, 45)

    sheet.freeze_panes = "A2"

    if not sheet.auto_filter.ref:
        sheet.auto_filter.ref = sheet.dimensions


def append_excel_total_row(sheet, label: str, value, label_column: int = 1, value_column: int = 2):
    from openpyxl.styles import Font, PatternFill

    row_number = sheet.max_row + 1
    sheet.cell(row=row_number, column=label_column, value=label)
    sheet.cell(row=row_number, column=value_column, value=value)

    for cell in sheet[row_number]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF6EE")


def set_excel_filter_range(sheet, last_data_row: int | None = None):
    from openpyxl.utils import get_column_letter

    if sheet.max_row == 0 or sheet.max_column == 0:
        return

    end_row = last_data_row if last_data_row is not None else sheet.max_row
    end_column = get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = f"A1:{end_column}{max(1, end_row)}"


def append_excel_filtered_total_row(sheet, label: str, label_column: int, value_column: int):
    from openpyxl.utils import get_column_letter

    last_data_row = sheet.max_row

    if last_data_row < 2:
        value = 0
    else:
        value_letter = get_column_letter(value_column)
        value = f"=SUBTOTAL(109,{value_letter}2:{value_letter}{last_data_row})"

    append_excel_total_row(sheet, label, value, label_column=label_column, value_column=value_column)
    set_excel_filter_range(sheet, last_data_row=last_data_row)


def append_excel_operation_rows(sheet, operation_rows):
    sheet.append(["Дата", "Сотрудник", "Группа", "Операция", "Размер", "Цвет", "Количество", "Ед."])

    for row in operation_rows:
        shift_date, full_name, work_group, operation_name, product_size, product_color, quantity, unit = row
        sheet.append([
            shift_date,
            full_name,
            work_group,
            operation_name,
            product_size,
            format_color_label(product_color),
            quantity,
            unit,
        ])

    append_excel_filtered_total_row(sheet, "Итого по фильтру", label_column=6, value_column=7)


def append_excel_feedback_rows(sheet, feedback_rows):
    sheet.append(["Дата", "Время", "Сотрудник", "Должность", "Раздел", "Сообщение", "ID смены"])

    for row in feedback_rows:
        feedback_date, feedback_time, full_name, position, category, message_text, shift_id = row
        sheet.append([
            feedback_date,
            feedback_time,
            full_name,
            position,
            category,
            message_text,
            shift_id or "",
        ])


def append_excel_shift_rows(sheet, shift_rows, employee_name: str | None = None):
    sheet.append(["Дата", "Сотрудник", "Пришёл", "Ушёл", "Отработано", "Статус"])

    for row in shift_rows:
        if employee_name is None:
            shift_date, full_name, start_time, end_time, total_minutes, status = row[-6:]
        else:
            _, shift_date, start_time, end_time, total_minutes, status = row
            full_name = employee_name

        status_text = "Закрыта" if status == "closed" else "Открыта"
        sheet.append([
            shift_date,
            full_name,
            start_time,
            end_time or "",
            minutes_to_excel_time(total_minutes),
            status_text,
        ])


def route_task_excel_step(row):
    step_index = int(row[5] or 0)

    if row[6] == "done":
        step_index = max(0, step_index - 1)

    batch = get_route_batch_by_id(row[0])
    return (get_route_step_for_batch(batch, step_index) if batch else get_route_step(row[1], step_index)) or {}


def append_excel_route_task_rows(sheet, rows):
    statuses = {"active": "В работе", "done": "Завершено", "cancelled": "Удалено"}
    sheet.append([
        "ID задания", "Создано", "Назначено", "Завершено", "Изделие", "Операция",
        "Должность", "Размер", "Цвет", "План", "Годное", "Брак",
        "Выполнение, %", "Брак, %", "Статус", "Сотрудник",
        "Приоритет", "Срок", "Исходное задание", "Код партии", "Версия маршрута",
        "Состояние работы", "Причина остановки", "Передач",
    ])

    for row in rows:
        step = route_task_excel_step(row)
        batch = get_route_batch_by_id(row[0]) or {}
        plan = int(row[4] or 0)
        good = int(row[14] or 0)
        defect = int(row[15] or 0)
        sheet.append([
            row[0], row[7] or "", row[10] or "", row[9] or "", row[1],
            step.get("operation", ""), step.get("position", row[13] or ""), row[2],
            format_color_label(row[3]), plan, good, defect,
            round(good * 100 / plan, 1) if plan else 0,
            round(defect * 100 / plan, 1) if plan else 0,
            statuses.get(row[6], row[6]), row[12] or "Свободно",
            row[16] or "normal", row[17] or "", row[18] or "",
            batch.get("trace_code") or f"RB-{row[0]:06d}", batch.get("route_version") or "",
            batch.get("work_state") or "", batch.get("blocked_reason") or "",
            int(batch.get("handover_count") or 0),
        ])

    set_excel_filter_range(sheet)


def append_excel_route_input_rows(sheet, rows):
    roles = {"main": "Основной", "component": "Компонент"}
    types = {"semifinished": "Полуфабрикат", "finished": "Готовая продукция"}
    sheet.append([
        "ID задания", "Роль", "ID остатка", "Тип", "Номенклатура", "Размер",
        "Цвет", "Этап", "Для должности", "Количество", "Ед.", "Статус задания",
        "Создано", "Сотрудник",
    ])

    for row in rows:
        sheet.append([
            row[0], roles.get(row[1], row[1]), row[3], types.get(row[4], row[4]), row[5],
            row[6], format_color_label(row[7]), row[8], row[9], row[2], row[10],
            row[11], row[12], row[13] or "Свободно",
        ])

    set_excel_filter_range(sheet)


def append_excel_route_input_lot_rows(sheet, rows):
    roles = {"main": "Основной", "component": "Компонент"}
    types = {"semifinished": "Полуфабрикат", "finished": "Готовая продукция"}
    sheet.append([
        "ID задания", "Код задания", "Роль", "Количество", "ID складской партии",
        "Код складской партии", "Источник", "ID источника", "Тип", "Номенклатура",
        "Размер", "Цвет", "Этап", "Для должности", "Создано", "Сотрудник",
    ])
    for row in rows:
        sheet.append([
            row[0], row[1] or f"RB-{row[0]:06d}", roles.get(row[2], row[2]), row[3],
            row[4], row[5], row[6], row[7] or "", types.get(row[8], row[8]), row[9],
            row[10], format_color_label(row[11]), row[12], row[13], row[14], row[15] or "Свободно",
        ])
    set_excel_filter_range(sheet)


def append_excel_trace_event_rows(sheet, rows):
    event_labels = {
        "task_created": "Создано задание на раскрой",
        "materials_reserved": "Материал зарезервирован",
        "task_started": "Задание взято в работу",
        "cutting_contours_done": "Контуры нанесены",
        "cutting_layout_done": "Настил сформирован",
        "cutting_progress": "Прогресс раскроя",
        "cutting_output_received": "Крой принят на склад",
        "cutting_formed": "Раскрой завершён",
        "batch_created": "Создано задание операции",
        "inputs_reserved": "Компоненты зарезервированы",
        "task_paused": "Пауза",
        "task_blocked": "Блокировка",
        "task_resumed": "Продолжено",
        "task_released": "Передано",
        "operation_completed": "Операция завершена",
        "task_cancelled": "Задание отменено",
    }
    sheet.append([
        "ID события", "Дата и время", "Событие", "ID задания", "Код партии",
        "ID задания раскроя", "ID партии раскроя", "Сотрудник", "ID смены",
        "Операция", "Должность", "Количество", "Годное", "Брак", "Причина", "Подробности",
    ])
    for row in rows:
        sheet.append([
            row[0], row[1], event_labels.get(row[2], row[2]), row[3] or "", row[4] or "",
            row[5] or "", row[6] or "", row[8] or "Система", row[9] or "",
            row[10] or "", row[11] or "", row[12], row[13], row[14], row[15] or "", row[16] or "",
        ])
    set_excel_filter_range(sheet)


def append_excel_defect_rows(sheet, rows):
    sheet.append([
        "ID записи", "ID задания", "Дата", "Сотрудник", "Изделие", "Операция",
        "Размер", "Цвет", "Брак", "Причина", "Решение", "Комментарий",
        "Задание на переделку",
    ])

    for row in rows:
        sheet.append([
            row[0], row[1], row[2], row[14] or "", row[3], row[6], row[4],
            format_color_label(row[5]), row[8], row[9], row[10], row[11] or "", row[12] or "",
        ])

    set_excel_filter_range(sheet)


def append_excel_production_task_rows(sheet, rows):
    statuses = {
        "active": "Контуры", "contours_done": "Настил", "in_cutting": "Раскрой",
        "cut_done": "Формирование кроя", "formed": "Завершено", "cancelled": "Удалено",
    }
    sheet.append([
        "ID задания", "Создано", "Завершено", "Изделие", "Размер", "Цвет",
        "Контуры, шт", "Сформировано, шт", "Статус",
        "Приоритет", "Срок",
    ])

    for row in rows:
        sheet.append([
            row[0], row[3], row[4] or "", row[1], row[5], format_color_label(row[6]),
            int(row[7] or 0), int(row[8] or 0), statuses.get(row[2], row[2]),
            row[9] or "normal", row[10] or "",
        ])

    set_excel_filter_range(sheet)


def append_excel_warehouse_movement_rows(sheet, rows):
    types = {"semifinished": "Полуфабрикат", "finished": "Готовая продукция"}
    sheet.append([
        "ID движения", "Дата и время", "ID остатка", "Тип", "Номенклатура", "Размер",
        "Цвет", "Этап", "Для должности", "Изменение", "Ед.", "Вид движения",
        "Источник", "ID источника", "Сотрудник", "Причина корректировки",
    ])

    for row in rows:
        sheet.append([
            row[0], row[2], row[1], types.get(row[3], row[3]), row[4], row[5],
            format_color_label(row[6]), row[7], row[8], row[9], row[10], row[11],
            row[12], row[13] or "", row[14] or "Система", row[15] or "",
        ])

    set_excel_filter_range(sheet)


def append_excel_fabric_movement_rows(sheet, rows):
    sheet.append([
        "ID движения", "Дата и время", "Материал", "Цвет", "Изменение", "Ед.",
        "Вид движения", "Комментарий", "Сотрудник",
    ])

    for row in rows:
        sheet.append([
            row[0], row[1], row[2], format_color_label(row[3]), row[4], row[5],
            row[6], row[7] or "", row[8] or "Система",
        ])

    set_excel_filter_range(sheet)


def append_excel_warehouse_balance_rows(sheet, rows):
    types = {"semifinished": "Полуфабрикат", "finished": "Готовая продукция"}
    sheet.append([
        "ID остатка", "Тип", "Номенклатура", "Размер", "Цвет", "Этап",
        "Для должности", "Количество", "Ед.", "Обновлено",
    ])
    for row in rows:
        sheet.append([
            row["id"], types.get(row["item_type"], row["item_type"]), row["product_name"],
            row["product_size"], format_color_label(row["product_color"]), row["stage_name"],
            row["ready_for_position"], row["quantity"], row["unit"], row["updated_at"],
        ])
    set_excel_filter_range(sheet)


def append_excel_fabric_balance_rows(sheet, rows):
    sheet.append(["Материал", "Цвет", "Остаток", "Ед.", "Обновлено"])
    for material, color, quantity, unit, updated_at in rows:
        sheet.append([material, format_color_label(color), quantity, unit, updated_at])
    set_excel_filter_range(sheet)


def append_excel_wip_rows(sheet, rows):
    sheet.append(["Этап", "Активных заданий", "Количество, шт", "Свободно", "Просрочено"])
    for row in rows:
        sheet.append([row["stage"], row["tasks"], row["quantity"], row["free"], row["overdue"]])
    set_excel_filter_range(sheet)


def append_excel_alert_rows(sheet, rows):
    sheet.append(["Тип", "Событие", "Подробности"])
    type_labels = {
        "overdue": "Просрочено", "free": "Свободно", "defect": "Брак",
        "blocked": "Заблокировано", "paused": "Пауза", "stalled": "Нет движения",
    }
    for row in rows:
        sheet.append([type_labels.get(row["type"], row["type"]), row["title"], row["detail"]])
    set_excel_filter_range(sheet)


def workbook_to_bytes(workbook):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def create_period_excel_bytes(start_date: str, end_date: str):
    from openpyxl import Workbook

    employee_summary = get_period_employee_summary(start_date, end_date)
    operation_rows = get_period_operation_rows(start_date, end_date)
    shift_rows = get_period_shift_details(start_date, end_date)
    feedback_rows = get_feedback_entries(start_date, end_date)
    route_rows = get_period_route_batch_rows(start_date, end_date)
    defect_rows = get_route_batch_defect_rows(start_date, end_date)
    route_input_rows = get_period_route_batch_input_rows(start_date, end_date)
    route_input_lot_rows = get_period_route_batch_input_lot_rows(start_date, end_date)
    trace_event_rows = get_period_production_trace_event_rows(start_date, end_date)
    production_rows = get_period_production_task_item_rows(start_date, end_date)
    warehouse_movement_rows = get_period_warehouse_movement_rows(start_date, end_date)
    fabric_movement_rows = get_period_fabric_movement_rows(start_date, end_date)
    production_control = get_production_control_payload(start_date, end_date)

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    summary_sheet.append(["Сотрудник", "Смен", "Часы"])
    total_shifts = 0
    total_minutes = 0

    for employee_id, full_name, shift_count, employee_minutes in employee_summary:
        total_shifts += shift_count or 0
        total_minutes += employee_minutes or 0
        summary_sheet.append([full_name, shift_count, minutes_to_excel_time(employee_minutes)])

    append_excel_total_row(summary_sheet, "Итого", total_shifts, label_column=1, value_column=2)
    summary_sheet.cell(row=summary_sheet.max_row, column=3, value=minutes_to_excel_time(total_minutes))
    summary_sheet.append([])
    summary_sheet.append(["Производственный показатель", "Значение"])
    summary_sheet.append(["Маршрутных заданий", len(route_rows)])
    summary_sheet.append(["План, шт", production_control["plan"]])
    summary_sheet.append(["Годная продукция, шт", production_control["fact"]])
    summary_sheet.append(["Брак, шт", production_control["defect_quantity"]])
    summary_sheet.append(["Заданий раскроя", len({row[0] for row in production_rows})])
    summary_sheet.append(["FPY, %", production_control["fpy"]])
    summary_sheet.append(["Средний cycle time, мин", production_control["average_cycle_minutes"]])
    summary_sheet.append(["Средний lead time, мин", production_control["average_lead_minutes"]])
    summary_sheet.append(["Выполнено в срок, %", production_control["schedule_adherence"]])
    summary_sheet.append(["WIP активных заданий, шт", production_control["active_quantity"]])
    summary_sheet.append(["Полуфабрикаты на складе, шт", production_control["semifinished_quantity"]])
    summary_sheet.append(["Период", f"{start_date} — {end_date}"])

    shifts_sheet = workbook.create_sheet("Смены по дням", 1)
    append_excel_shift_rows(shifts_sheet, shift_rows)

    operations_sheet = workbook.create_sheet("Операции")
    append_excel_operation_rows(operations_sheet, operation_rows)

    for group_name in ("Раскрой", "Пошив", "Упаковка"):
        group_sheet = workbook.create_sheet(group_name)
        append_excel_operation_rows(
            group_sheet,
            [row for row in operation_rows if row[2] == group_name],
        )

    production_sheet = workbook.create_sheet("Задания раскроя")
    append_excel_production_task_rows(production_sheet, production_rows)

    route_sheet = workbook.create_sheet("Маршрутные задания")
    append_excel_route_task_rows(route_sheet, route_rows)

    route_inputs_sheet = workbook.create_sheet("Входы заданий")
    append_excel_route_input_rows(route_inputs_sheet, route_input_rows)

    route_input_lots_sheet = workbook.create_sheet("Партии компонентов")
    append_excel_route_input_lot_rows(route_input_lots_sheet, route_input_lot_rows)

    trace_sheet = workbook.create_sheet("Хронология партий")
    append_excel_trace_event_rows(trace_sheet, trace_event_rows)

    defects_sheet = workbook.create_sheet("Брак")
    append_excel_defect_rows(defects_sheet, defect_rows)

    warehouse_movements_sheet = workbook.create_sheet("Движения склада")
    append_excel_warehouse_movement_rows(warehouse_movements_sheet, warehouse_movement_rows)

    fabric_movements_sheet = workbook.create_sheet("Движения материалов")
    append_excel_fabric_movement_rows(fabric_movements_sheet, fabric_movement_rows)

    warehouse_balance_sheet = workbook.create_sheet("Остатки склада")
    append_excel_warehouse_balance_rows(warehouse_balance_sheet, get_warehouse_stock_rows())

    fabric_balance_sheet = workbook.create_sheet("Остатки материалов")
    append_excel_fabric_balance_rows(fabric_balance_sheet, get_fabric_stock_rows())

    wip_sheet = workbook.create_sheet("WIP по этапам")
    append_excel_wip_rows(wip_sheet, production_control["stages"])

    alerts_sheet = workbook.create_sheet("Отклонения")
    append_excel_alert_rows(alerts_sheet, production_control["alerts"])

    feedback_sheet = workbook.create_sheet("Обратная связь")
    append_excel_feedback_rows(feedback_sheet, feedback_rows)

    for sheet in workbook.worksheets:
        style_excel_sheet(sheet)

    return workbook_to_bytes(workbook)


def create_employee_excel_bytes(employee_id: int, start_date: str, end_date: str):
    from openpyxl import Workbook

    employee_summary = get_employee_period_summary(employee_id, start_date, end_date)

    if employee_summary is None:
        return None, None

    _, full_name, position, shift_count, total_minutes = employee_summary
    operation_rows = [
        row for row in get_period_operation_rows(start_date, end_date)
        if row[1] == full_name
    ]
    shift_rows = get_employee_shifts_by_period(employee_id, start_date, end_date)
    feedback_rows = get_feedback_entries(start_date, end_date, employee_id=employee_id)
    route_rows = get_period_route_batch_rows(start_date, end_date, employee_id=employee_id)
    defect_rows = get_route_batch_defect_rows(start_date, end_date, employee_id=employee_id)
    route_input_rows = get_period_route_batch_input_rows(start_date, end_date, employee_id=employee_id)
    route_input_lot_rows = get_period_route_batch_input_lot_rows(start_date, end_date, employee_id=employee_id)
    trace_event_rows = get_period_production_trace_event_rows(start_date, end_date, employee_id=employee_id)
    warehouse_movement_rows = get_period_warehouse_movement_rows(start_date, end_date, employee_id=employee_id)
    fabric_movement_rows = get_period_fabric_movement_rows(start_date, end_date, employee_id=employee_id)
    employee_defect_quantity = sum(int(row[8] or 0) for row in defect_rows)
    employee_good_quantity = sum(int(row[14] or 0) for row in route_rows if date_in_period(row[9], start_date, end_date))

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Итог"
    summary_sheet.append(["Показатель", "Значение"])
    summary_sheet.append(["Сотрудник", full_name])
    summary_sheet.append(["Должность", position or ""])
    summary_sheet.append(["Период", f"{start_date} — {end_date}"])
    summary_sheet.append(["Отработано смен", shift_count])
    summary_sheet.append(["Отработано часов", minutes_to_excel_time(total_minutes)])
    summary_sheet.append(["Маршрутных заданий", len(route_rows)])
    summary_sheet.append([
        "План, шт",
        sum(int(row[4] or 0) for row in route_rows if row[6] != "cancelled" and date_in_period(row[7], start_date, end_date)),
    ])
    summary_sheet.append(["Годная продукция, шт", employee_good_quantity])
    summary_sheet.append(["Брак, шт", employee_defect_quantity])
    summary_sheet.append([
        "FPY, %",
        round(employee_good_quantity * 100 / (employee_good_quantity + employee_defect_quantity), 1)
        if employee_good_quantity + employee_defect_quantity else 0,
    ])

    shifts_sheet = workbook.create_sheet("Смены по дням", 1)
    append_excel_shift_rows(shifts_sheet, shift_rows, employee_name=full_name)

    operations_sheet = workbook.create_sheet("Операции")
    append_excel_operation_rows(operations_sheet, operation_rows)

    route_sheet = workbook.create_sheet("Маршрутные задания")
    append_excel_route_task_rows(route_sheet, route_rows)

    route_inputs_sheet = workbook.create_sheet("Входы заданий")
    append_excel_route_input_rows(route_inputs_sheet, route_input_rows)

    route_input_lots_sheet = workbook.create_sheet("Партии компонентов")
    append_excel_route_input_lot_rows(route_input_lots_sheet, route_input_lot_rows)

    trace_sheet = workbook.create_sheet("Хронология партий")
    append_excel_trace_event_rows(trace_sheet, trace_event_rows)

    defects_sheet = workbook.create_sheet("Брак")
    append_excel_defect_rows(defects_sheet, defect_rows)

    warehouse_movements_sheet = workbook.create_sheet("Движения склада")
    append_excel_warehouse_movement_rows(warehouse_movements_sheet, warehouse_movement_rows)

    fabric_movements_sheet = workbook.create_sheet("Движения материалов")
    append_excel_fabric_movement_rows(fabric_movements_sheet, fabric_movement_rows)

    feedback_sheet = workbook.create_sheet("Обратная связь")
    append_excel_feedback_rows(feedback_sheet, feedback_rows)

    for sheet in workbook.worksheets:
        style_excel_sheet(sheet)

    return workbook_to_bytes(workbook), full_name


def employee_operation_total_to_dict(row):
    operation_name, total_quantity, unit = row

    return {
        "operation": operation_name,
        "quantity": total_quantity,
        "unit": unit,
    }


def operation_admin_to_dict(row):
    operation_id, number, name, position, operation_group, folder, unit, is_active = row

    return {
        "id": operation_id,
        "number": number,
        "name": name,
        "position": position or "-",
        "group": operation_group or "-",
        "folder": folder or "-",
        "unit": unit or "шт",
        "active": bool(is_active),
    }


def quantity_as_float(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def quantity_text(value):
    return format_number(quantity_as_float(value))


def date_in_period(value: str | None, start_date: str, end_date: str):
    return bool(value and start_date <= value[:10] <= end_date)


def minutes_between(start_value: str | None, end_value: str | None):
    if not start_value or not end_value:
        return None

    try:
        delta = datetime.fromisoformat(end_value) - datetime.fromisoformat(start_value)
    except ValueError:
        return None

    return max(0, round(delta.total_seconds() / 60, 1))


def production_control_route_row_to_dict(row):
    step_index = int(row[5] or 0)
    if row[6] == "done":
        step_index = max(0, step_index - 1)
    batch = get_route_batch_by_id(row[0])
    step = get_route_step_for_batch(batch, step_index) if batch else get_route_step(row[1], step_index)
    step = step or {}
    cycle_minutes = minutes_between(row[10], row[9])
    lead_minutes = minutes_between(row[7], row[9])
    due_date = row[17] or ""
    completed_date = (row[9] or "")[:10]

    return {
        "id": row[0],
        "task_kind": "route",
        "product": row[1],
        "size": row[2],
        "color": format_color_label(row[3]),
        "quantity": int(row[4] or 0),
        "status": row[6],
        "status_text": {
            "active": "В работе" if row[11] else "Свободно",
            "done": "Завершено",
            "cancelled": "Отменено",
        }.get(row[6], row[6]),
        "operation": step.get("operation", "Операция"),
        "position": step.get("position", row[13] or ""),
        "stage": f"{route_step_category(step)} · {step.get('position', '')}".strip(" ·"),
        "employee": row[12] or "",
        "good_quantity": int(row[14] or 0),
        "defect_quantity": int(row[15] or 0),
        "priority": row[16] or "normal",
        "due_date": due_date,
        "created_at": row[7] or "",
        "assigned_at": row[10] or "",
        "completed_at": row[9] or "",
        "cycle_minutes": cycle_minutes,
        "lead_minutes": lead_minutes,
        "on_time": completed_date <= due_date if due_date and completed_date else None,
        "parent_batch_id": row[18],
        "trace_code": (batch or {}).get("trace_code") or f"RB-{row[0]:06d}",
        "route_version": (batch or {}).get("route_version") or "",
        "work_state": (batch or {}).get("work_state") or row[6],
        "blocked_reason": (batch or {}).get("blocked_reason") or "",
    }


def production_control_active_batch_to_dict(batch, step, employee_names):
    stage = f"{route_step_category(step)} · {step['position']}"
    now = local_now().isoformat()

    return {
        "id": batch["id"],
        "task_kind": "route",
        "product": batch["product_name"],
        "size": batch["product_size"],
        "color": format_color_label(batch["product_color"]),
        "quantity": int(batch["quantity"] or 0),
        "status": batch["status"],
        "status_text": {
            "free": "Свободно",
            "in_work": "В работе",
            "paused": "Пауза",
            "blocked": "Заблокировано",
        }.get(batch.get("work_state") or ("in_work" if batch.get("assigned_employee_id") else "free"), "В работе"),
        "operation": step["operation"],
        "position": step["position"],
        "stage": stage,
        "employee": employee_names.get(batch.get("assigned_employee_id"), ""),
        "good_quantity": int(batch.get("good_quantity") or 0),
        "defect_quantity": int(batch.get("defect_quantity") or 0),
        "priority": batch.get("priority") or "normal",
        "due_date": batch.get("due_date") or "",
        "created_at": batch.get("created_at") or "",
        "assigned_at": batch.get("assigned_at") or "",
        "completed_at": "",
        "cycle_minutes": minutes_between(batch.get("assigned_at"), now),
        "lead_minutes": minutes_between(batch.get("created_at"), now),
        "on_time": None,
        "parent_batch_id": batch.get("parent_batch_id"),
        "trace_code": batch.get("trace_code") or f"RB-{batch['id']:06d}",
        "route_version": batch.get("route_version") or "",
        "work_state": batch.get("work_state") or "",
        "blocked_reason": batch.get("blocked_reason") or "",
        "last_activity_at": batch.get("last_activity_at") or batch.get("updated_at") or "",
        "handover_count": int(batch.get("handover_count") or 0),
    }


def get_production_control_payload(start_date: str, end_date: str):
    route_rows = get_period_route_batch_rows(start_date, end_date)
    defect_rows = get_route_batch_defect_rows(start_date, end_date)
    active_batches = get_active_route_batches()
    warehouse_rows = get_warehouse_stock_rows()
    employee_names = {row[0]: row[1] for row in get_all_employees()}

    planned_rows = [
        row for row in route_rows
        if row[6] != "cancelled" and date_in_period(row[7], start_date, end_date)
    ]
    completed_rows = [row for row in route_rows if date_in_period(row[9], start_date, end_date)]
    plan = sum(int(row[4] or 0) for row in planned_rows)
    good = sum(int(row[14] or 0) for row in completed_rows)
    defect_quantity = sum(int(row[8] or 0) for row in defect_rows)
    first_pass_total = good + defect_quantity
    fpy = round(good * 100 / first_pass_total, 1) if first_pass_total else 0

    cycle_values = [
        value for value in (minutes_between(row[10], row[9]) for row in completed_rows)
        if value is not None
    ]
    lead_values = [
        value for value in (minutes_between(row[7], row[9]) for row in completed_rows)
        if value is not None
    ]
    due_completed = [row for row in completed_rows if row[17]]
    on_time = [row for row in due_completed if (row[9] or "")[:10] <= row[17]]
    schedule_adherence = round(len(on_time) * 100 / len(due_completed), 1) if due_completed else 0

    stage_map = {}
    active_task_details = []
    alerts = []
    today = local_today().isoformat()

    for batch in active_batches:
        step = get_route_step_for_batch(batch)
        if step is None:
            continue
        stage_key = f"{route_step_category(step)} · {step['position']}"
        active_task_details.append(production_control_active_batch_to_dict(batch, step, employee_names))
        stage = stage_map.setdefault(
            stage_key,
            {"stage": stage_key, "tasks": 0, "quantity": 0, "free": 0, "overdue": 0},
        )
        stage["tasks"] += 1
        stage["quantity"] += int(batch["quantity"] or 0)
        if not batch.get("assigned_employee_id"):
            stage["free"] += 1
        work_state = batch.get("work_state") or ("in_work" if batch.get("assigned_employee_id") else "free")
        if work_state == "blocked":
            alerts.append(
                {
                    "type": "blocked",
                    "batch_id": batch["id"],
                    "title": f"Заблокировано задание #{batch['id']}",
                    "detail": batch.get("blocked_reason") or f"{batch['product_name']} · {step['operation']}",
                }
            )
        elif work_state == "paused":
            alerts.append(
                {
                    "type": "paused",
                    "batch_id": batch["id"],
                    "title": f"Пауза по заданию #{batch['id']}",
                    "detail": batch.get("blocked_reason") or f"{batch['product_name']} · {step['operation']}",
                }
            )
        elif batch.get("assigned_employee_id") and (minutes_between(batch.get("last_activity_at"), local_now().isoformat()) or 0) >= 240:
            alerts.append(
                {
                    "type": "stalled",
                    "batch_id": batch["id"],
                    "title": f"Нет движения по заданию #{batch['id']}",
                    "detail": f"{batch['product_name']} · {step['operation']} · более 4 часов",
                }
            )

        if batch.get("due_date") and batch["due_date"] < today:
            stage["overdue"] += 1
            alerts.append(
                {
                    "type": "overdue",
                    "batch_id": batch["id"],
                    "title": f"Просрочено задание #{batch['id']}",
                    "detail": f"{batch['product_name']} · {step['operation']} · срок {batch['due_date']}",
                }
            )
        elif not batch.get("assigned_employee_id"):
            alerts.append(
                {
                    "type": "free",
                    "batch_id": batch["id"],
                    "title": f"Свободное задание #{batch['id']}",
                    "detail": f"{batch['product_name']} · {step['position']} · {batch['quantity']} шт",
                }
            )

    for row in defect_rows[:10]:
        alerts.append(
            {
                "type": "defect",
                "batch_id": row[1],
                "title": f"Брак: {row[8]} шт · {row[3]}",
                "detail": f"{row[6]} · {row[9]} · {row[10]}",
            }
        )

    semifinished_quantity = sum(
        int(row["quantity"] or 0) for row in warehouse_rows if row["item_type"] == "semifinished"
    )
    defect_details = [
        {
            "id": row[0],
            "batch_id": row[1],
            "date": row[2][:10],
            "product": row[3],
            "size": row[4],
            "color": format_color_label(row[5]),
            "stage": row[6],
            "position": row[7],
            "quantity": row[8],
            "reason": row[9],
            "disposition": row[10],
            "comment": row[11] or "",
            "rework_batch_id": row[12],
            "employee": row[14] or "",
        }
        for row in defect_rows
    ]
    planned_task_details = [production_control_route_row_to_dict(row) for row in planned_rows]
    completed_task_details = [production_control_route_row_to_dict(row) for row in completed_rows]
    semifinished_details = [
        {
            "id": row["id"],
            "product": row["product_name"],
            "size": row["product_size"],
            "color": format_color_label(row["product_color"]),
            "stage": row["stage_name"],
            "ready_for": row["ready_for_position"],
            "quantity": int(row["quantity"] or 0),
            "unit": row["unit"] or "шт",
            "updated_at": row["updated_at"] or "",
        }
        for row in warehouse_rows
        if row["item_type"] == "semifinished"
    ]

    return {
        "start_date": start_date,
        "end_date": end_date,
        "plan": plan,
        "fact": good,
        "defect_quantity": defect_quantity,
        "fpy": fpy,
        "completed_tasks": len(completed_rows),
        "average_cycle_minutes": round(sum(cycle_values) / len(cycle_values), 1) if cycle_values else 0,
        "average_lead_minutes": round(sum(lead_values) / len(lead_values), 1) if lead_values else 0,
        "schedule_adherence": schedule_adherence,
        "active_tasks": len(active_batches),
        "active_quantity": sum(int(batch["quantity"] or 0) for batch in active_batches),
        "semifinished_quantity": semifinished_quantity,
        "stages": sorted(stage_map.values(), key=lambda row: (-row["overdue"], -row["quantity"], row["stage"])),
        "alerts": alerts[:20],
        "defects": defect_details,
        "details": {
            "planned_tasks": planned_task_details,
            "completed_tasks": completed_task_details,
            "active_tasks": active_task_details,
            "semifinished": semifinished_details,
            "cycle_tasks": [row for row in completed_task_details if row["cycle_minutes"] is not None],
            "lead_tasks": [row for row in completed_task_details if row["lead_minutes"] is not None],
            "schedule_tasks": [row for row in completed_task_details if row["due_date"]],
            "defects": defect_details,
        },
    }


def get_admin_home_period_payload(
    period_id: str,
    title: str,
    start_date: str,
    end_date: str,
    employees: list[dict],
    open_shifts: list[dict],
):
    report = get_admin_report_payload("period", start_date, end_date)
    control = get_production_control_payload(start_date, end_date)
    performance_rows = get_period_employee_production_performance(start_date, end_date)
    employee_by_name = {employee["full_name"]: employee for employee in employees}
    rows_by_name = {}

    def ensure_employee(full_name: str):
        employee = employee_by_name.get(full_name, {})
        if full_name not in rows_by_name:
            rows_by_name[full_name] = {
                "name": full_name,
                "position": employee.get("position") or "-",
                "plan": 0,
                "plan_text": "0",
                "fact": 0,
                "fact_text": "0",
                "shift_count": 0,
                "on_shift": False,
                "status": "",
                "operations": [],
            }
        return rows_by_name[full_name]

    for shift in report.get("shifts", []):
        employee = ensure_employee(shift.get("employee") or "Сотрудник")
        employee["shift_count"] += 1
        if shift.get("status") == "open":
            employee["on_shift"] = True
        employee["status"] = shift.get("status") or employee["status"]

    if period_id == "today":
        for shift in open_shifts:
            employee = ensure_employee(shift.get("employee") or "Сотрудник")
            employee["on_shift"] = True
            employee["status"] = "open"
            employee["start_time"] = shift.get("start_time") or ""
            employee["date"] = shift.get("date") or start_date

    for operation in report.get("operations", []):
        employee = ensure_employee(operation.get("employee") or "Сотрудник")
        operation_quantity = quantity_as_float(operation.get("quantity"))
        employee["operations"].append(
            {
                "operation": operation.get("operation") or "-",
                "stage": operation.get("group") or "-",
                "date": operation.get("date") or "",
                "size": operation.get("size") or "-",
                "color": operation.get("color") or "-",
                "quantity": operation_quantity,
                "quantity_text": quantity_text(operation_quantity),
                "unit": operation.get("unit") or "шт",
            }
        )

    for _employee_id, full_name, position, plan, fact in performance_rows:
        employee = ensure_employee(full_name)
        if employee["position"] == "-":
            employee["position"] = position or "-"
        employee["plan"] += int(plan or 0)
        employee["fact"] += int(fact or 0)

    for employee in rows_by_name.values():
        employee["fact_text"] = quantity_text(employee["fact"])
        employee["plan_text"] = quantity_text(employee["plan"])
        employee["operations"].sort(key=lambda row: (row["date"], row["operation"], row["size"], row["color"]))

    employee_rows = sorted(
        rows_by_name.values(),
        key=lambda row: (not row["on_shift"], -row["fact"], row["name"]),
    )

    return {
        "id": period_id,
        "title": title,
        "start_date": start_date,
        "end_date": end_date,
        "plan": control["plan"],
        "plan_text": quantity_text(control["plan"]),
        "fact": control["fact"],
        "fact_text": quantity_text(control["fact"]),
        "defect_count": len(control["defects"]),
        "employees": employee_rows,
        "defects": control["defects"],
        "defect_fields": ["Изделие", "Этап", "Причина"],
        "control": control,
    }


def get_admin_home_payload(employees: list[dict], open_shifts: list[dict]):
    today = local_today().isoformat()
    month_start, month_end = month_bounds()
    quarter_start, quarter_end = quarter_bounds()

    return {
        "periods": {
            "today": get_admin_home_period_payload(
                "today",
                "Сегодня",
                today,
                today,
                employees,
                open_shifts,
            ),
            "month": get_admin_home_period_payload(
                "month",
                "Текущий месяц",
                month_start,
                month_end,
                employees,
                open_shifts,
            ),
            "quarter": get_admin_home_period_payload(
                "quarter",
                "Текущий квартал",
                quarter_start,
                quarter_end,
                employees,
                open_shifts,
            ),
        }
    }


def get_admin_dashboard(telegram_id: int):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    employees = get_all_employees()
    user_accounts = get_all_user_accounts()
    web_profiles = get_web_account_profiles_by_telegram_ids(employee[3] for employee in user_accounts)
    employee_rows = [employee_admin_to_dict(employee, web_profiles) for employee in employees]
    open_shift_rows = [open_shift_to_dict(shift) for shift in get_open_shifts()]
    month_start, today = month_bounds()

    return {
        "ok": True,
        "menu": ADMIN_MENU,
        "positions": POSITIONS,
        "period_defaults": {"start_date": month_start, "end_date": today},
        "employees": employee_rows,
        "user_accounts": [
            user_account_admin_to_dict(employee, web_profiles) for employee in user_accounts
        ],
        "active_employees": [
            employee_admin_to_dict(employee, web_profiles) for employee in get_employees_by_status("active")
        ],
        "inactive_employees": [
            employee_admin_to_dict(employee, web_profiles) for employee in get_employees_by_status("inactive")
        ],
        "pending_employees": [
            pending_employee_to_dict(employee, web_profiles) for employee in get_pending_employees()
        ],
        "open_shifts": open_shift_rows,
        "recent_shifts": [recent_shift_to_dict(shift) for shift in get_recent_shifts(20)],
        "operations": [
            operation_admin_to_dict(operation) for operation in get_all_operations()
        ],
        "reports": get_admin_report_payload("period", month_start, today),
        "home": get_admin_home_payload(employee_rows, open_shift_rows),
        "feedback": [
            feedback_row_to_dict(row)
            for row in get_feedback_entries(month_start, today)
        ],
        "production_control": get_production_control_payload(month_start, today),
        "defect_reasons": DEFECT_REASONS,
        "defect_dispositions": DEFECT_DISPOSITIONS,
    }


def get_admin_report_payload(
    report_type: str,
    start_date: str,
    end_date: str,
    employee_id: int | None = None,
):
    if report_type == "today":
        today = local_today().isoformat()
        return {
            "type": "today",
            "title": f"Отчёт за сегодня: {today}",
            "start_date": today,
            "end_date": today,
            "summary": [employee_summary_to_dict(row) for row in get_period_employee_summary(today, today)],
            "shifts": [shift_detail_to_dict(row) for row in get_today_shifts()],
            "operations": [operation_summary_to_dict(row) for row in get_period_operation_rows(today, today)],
        }

    if report_type == "employee" and employee_id:
        summary = get_employee_period_summary(employee_id, start_date, end_date)
        return {
            "type": "employee",
            "title": f"Отчёт по сотруднику: {start_date} — {end_date}",
            "start_date": start_date,
            "end_date": end_date,
            "employee_summary": (
                {
                    "employee_id": summary[0],
                    "full_name": summary[1],
                    "position": summary[2],
                    "shift_count": summary[3],
                    "total_minutes": summary[4],
                    "total_time": format_minutes(summary[4] or 0),
                }
                if summary
                else None
            ),
            "employee_shifts": [
                {
                    "id": row[0],
                    "date": row[1],
                    "start_time": row[2],
                    "end_time": row[3],
                    "total_minutes": row[4],
                    "total_time": format_minutes(row[4] or 0),
                    "status": row[5],
                }
                for row in get_employee_shifts_by_period(employee_id, start_date, end_date)
            ],
            "employee_operations": [
                employee_operation_total_to_dict(row)
                for row in get_employee_period_operation_totals(employee_id, start_date, end_date)
            ],
        }

    title = f"Отчёт за период: {start_date} — {end_date}"
    summary_rows = get_period_employee_summary(start_date, end_date)
    operation_rows = get_period_operation_rows(start_date, end_date)
    shift_rows = get_period_shift_details(start_date, end_date)

    return {
        "type": "period",
        "title": title,
        "start_date": start_date,
        "end_date": end_date,
        "summary": [employee_summary_to_dict(row) for row in summary_rows],
        "shifts": [shift_detail_to_dict(row) for row in shift_rows],
        "operations": [operation_summary_to_dict(row) for row in operation_rows],
    }


def get_admin_report_for_telegram(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    month_start, today = month_bounds()
    report_type = (payload.get("report_type") or "period").strip()
    start_date = clean_date(payload.get("start_date"), month_start)
    end_date = clean_date(payload.get("end_date"), today)

    try:
        employee_id = int(payload.get("employee_id") or 0)
    except (TypeError, ValueError):
        employee_id = 0

    return {
        "ok": True,
        "report": get_admin_report_payload(
            report_type,
            start_date,
            end_date,
            employee_id or None,
        ),
    }


def get_employee_history_for_telegram(telegram_id: int, payload: dict):
    month_start, today = month_bounds()
    start_date = clean_date(payload.get("start_date"), month_start)
    end_date = clean_date(payload.get("end_date"), today)
    employee = get_employee_for_access(telegram_id)

    if employee is None or employee[5] != "active":
        return {
            "ok": False,
            "message": "Нет активного профиля.",
            "start_date": start_date,
            "end_date": end_date,
            "summary": None,
            "shifts": [],
            "operations": [],
        }

    summary = get_employee_period_summary(employee[0], start_date, end_date)

    return {
        "ok": True,
        "message": "",
        "start_date": start_date,
        "end_date": end_date,
        "summary": (
            {
                "employee_id": summary[0],
                "full_name": summary[1],
                "position": summary[2],
                "shift_count": summary[3],
                "total_minutes": summary[4],
                "total_time": format_minutes(summary[4] or 0),
            }
            if summary
            else None
        ),
        "shifts": [
            employee_shift_to_dict(row)
            for row in get_employee_shifts_by_period(employee[0], start_date, end_date)
        ],
        "operations": [
            employee_operation_total_to_dict(row)
            for row in get_employee_period_operation_totals(employee[0], start_date, end_date)
        ],
    }


def get_admin_feedback_for_telegram(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора.", "feedback": []}

    month_start, today = month_bounds()
    start_date = clean_date(payload.get("start_date"), month_start)
    end_date = clean_date(payload.get("end_date"), today)

    return {
        "ok": True,
        "start_date": start_date,
        "end_date": end_date,
        "feedback": [
            feedback_row_to_dict(row)
            for row in get_feedback_entries(start_date, end_date)
        ],
    }


def export_admin_report_for_telegram(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    month_start, today = month_bounds()
    report_type = (payload.get("report_type") or "period").strip()
    start_date = clean_date(payload.get("start_date"), month_start)
    end_date = clean_date(payload.get("end_date"), today)

    try:
        employee_id = int(payload.get("employee_id") or 0)
    except (TypeError, ValueError):
        employee_id = 0

    if report_type == "today":
        start_date = today
        end_date = today
        content = create_period_excel_bytes(start_date, end_date)
        filename = f"report_today_{today}.xlsx"
    elif report_type == "employee":
        if employee_id <= 0:
            return {"ok": False, "message": "Выберите сотрудника для выгрузки."}

        content, employee_name = create_employee_excel_bytes(employee_id, start_date, end_date)

        if content is None:
            return {"ok": False, "message": "Сотрудник не найден."}

        safe_name = "".join(
            char if char.isalnum() or char in {"_", "-"} else "_"
            for char in employee_name
        ).strip("_")
        filename = f"employee_{employee_id}_{safe_name}_{start_date}_{end_date}.xlsx"
    else:
        content = create_period_excel_bytes(start_date, end_date)
        filename = f"report_{start_date}_{end_date}.xlsx"

    add_edit_log(
        telegram_id,
        "admin",
        "Выгрузил отчёт из миниаппа",
        "report",
        None,
        f"{report_type}: {start_date} — {end_date}",
    )

    return {
        "ok": True,
        "filename": filename,
        "content": content,
    }


def set_employee_status_for_admin(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    try:
        employee_id = int(payload.get("employee_id") or 0)
    except (TypeError, ValueError):
        employee_id = 0

    status = (payload.get("status") or "").strip()

    if status not in {"active", "inactive", "pending"}:
        return {"ok": False, "message": "Некорректный статус."}

    current_employee = get_employee_by_id(employee_id)
    if current_employee is None:
        return {"ok": False, "message": "Сотрудник не найден."}
    if current_employee[1] == telegram_id and status != "active":
        return {"ok": False, "message": "Нельзя отключить собственный аккаунт администратора."}
    if status == "active" and current_employee[4] != "admin" and current_employee[3] not in POSITIONS:
        return {"ok": False, "message": "Перед активацией выберите должность сотрудника."}

    access_result = update_employee_access_status(employee_id, status)
    if access_result.get("code") == "last_admin":
        return {"ok": False, "message": "Нельзя отключить последнего активного администратора."}
    employee = access_result.get("employee")

    if employee is None:
        return {"ok": False, "message": "Сотрудник не найден."}

    if status != "active":
        revoke_web_sessions_for_telegram_id(employee[1])

    add_edit_log(
        telegram_id,
        "admin",
        f"Изменил статус сотрудника на {status} из миниаппа",
        "employee",
        employee_id,
        employee[2],
    )
    dashboard = get_admin_dashboard(telegram_id)
    dashboard["message"] = "Статус сотрудника изменён."
    return dashboard


def set_employee_role_for_admin(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    try:
        employee_id = int(payload.get("employee_id") or 0)
    except (TypeError, ValueError):
        employee_id = 0

    role = (payload.get("role") or "").strip()
    position = (payload.get("position") or "").strip()
    current_employee = get_employee_by_id(employee_id)

    if current_employee is None:
        return {"ok": False, "message": "Пользователь не найден."}
    if role not in {"admin", "employee"}:
        return {"ok": False, "message": "Некорректная роль."}
    if current_employee[1] == telegram_id and role != "admin":
        return {"ok": False, "message": "Нельзя снять права администратора у собственного аккаунта."}
    if role == "employee" and position not in POSITIONS:
        return {"ok": False, "message": "Для сотрудника выберите должность из списка."}

    role_result = update_employee_role(employee_id, role, position)
    if role_result.get("code") == "last_admin":
        return {"ok": False, "message": "Нельзя снять права у последнего активного администратора."}
    employee = role_result.get("employee")
    if employee is None:
        return {"ok": False, "message": "Не удалось изменить роль пользователя."}

    role_label = "Администратор" if role == "admin" else "Сотрудник"
    add_edit_log(
        telegram_id,
        "admin",
        f"Изменил роль пользователя на {role_label} из веб-приложения",
        "employee",
        employee_id,
        employee[2],
    )
    dashboard = get_admin_dashboard(telegram_id)
    dashboard["message"] = f"Роль изменена: {role_label}."
    return dashboard


def set_employee_position_for_admin(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    try:
        employee_id = int(payload.get("employee_id") or 0)
    except (TypeError, ValueError):
        employee_id = 0

    position = (payload.get("position") or "").strip()

    if position not in POSITIONS:
        return {"ok": False, "message": "Выберите должность из списка."}

    current_employee = get_employee_by_id(employee_id)
    if current_employee is None:
        return {"ok": False, "message": "Сотрудник не найден."}
    if current_employee[4] == "admin":
        return {"ok": False, "message": "Для администратора измените роль, затем назначьте должность."}

    employee = update_employee_position(employee_id, position)

    if employee is None:
        return {"ok": False, "message": "Сотрудник не найден."}

    add_edit_log(
        telegram_id,
        "admin",
        f"Изменил должность на {position} из миниаппа",
        "employee",
        employee_id,
        employee[2],
    )
    dashboard = get_admin_dashboard(telegram_id)
    dashboard["message"] = "Должность сотрудника изменена."
    return dashboard


def close_shift_for_admin(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    try:
        shift_id = int(payload.get("shift_id") or 0)
    except (TypeError, ValueError):
        shift_id = 0

    end_time = (payload.get("end_time") or "").strip()

    if not end_time:
        return {"ok": False, "message": "Укажите время закрытия в формате 18:00."}

    result = admin_close_shift(shift_id, end_time)

    if result is None:
        return {"ok": False, "message": "Открытая смена с таким ID не найдена."}

    if result == "bad_time":
        return {"ok": False, "message": "Время закрытия меньше времени начала."}

    add_edit_log(
        telegram_id,
        "admin",
        "Закрыл смену сотрудника из миниаппа",
        "shift",
        shift_id,
        f"Окончание: {result['end_time']}, отработано: {format_minutes(result['total_minutes'])}",
    )
    dashboard = get_admin_dashboard(telegram_id)
    dashboard["message"] = "Смена закрыта."
    return dashboard


def delete_shift_for_admin(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    try:
        shift_id = int(payload.get("shift_id") or 0)
    except (TypeError, ValueError):
        shift_id = 0

    shift = delete_shift_by_id(shift_id)

    if shift is None:
        return {"ok": False, "message": "Смена не найдена."}

    add_edit_log(
        telegram_id,
        "admin",
        "Удалил смену из миниаппа",
        "shift",
        shift_id,
        f"{shift[1]} {shift[2]} {shift[3]}-{shift[4] or ''}",
    )
    dashboard = get_admin_dashboard(telegram_id)
    dashboard["message"] = "Смена удалена."
    return dashboard


def operation_action_for_admin(telegram_id: int, payload: dict):
    if not is_admin(telegram_id):
        return {"ok": False, "message": "Нет прав администратора."}

    action = (payload.get("action") or "").strip()

    try:
        number = int(payload.get("number") or 0)
    except (TypeError, ValueError):
        number = 0

    if action == "hide":
        operation = hide_operation(number)
        message = "Операция скрыта."
    elif action == "restore":
        operation = restore_operation(number)
        message = "Операция возвращена."
    elif action == "update":
        operation = update_operation_field(
            number,
            (payload.get("field") or "").strip(),
            (payload.get("value") or "").strip(),
        )
        message = "Операция изменена."
    elif action == "add":
        operation = add_operation(
            (payload.get("name") or "").strip(),
            (payload.get("position") or "").strip(),
            (payload.get("operation_group") or "").strip(),
            (payload.get("folder") or "").strip(),
        )
        message = "Операция добавлена."
    else:
        return {"ok": False, "message": "Неизвестное действие."}

    if operation is None:
        return {"ok": False, "message": "Операция не найдена или данные некорректны."}

    operation_log_id = number

    if not operation_log_id and isinstance(operation, dict):
        operation_log_id = operation.get("id")

    add_edit_log(
        telegram_id,
        "admin",
        message,
        "operation",
        operation_log_id,
        json.dumps(operation, ensure_ascii=False),
    )
    dashboard = get_admin_dashboard(telegram_id)
    dashboard["message"] = message
    return dashboard


def get_app_state(telegram_id: int, message: str = ""):
    shift_state = get_shift_state(telegram_id, message)
    employee = shift_state.get("employee")
    is_admin_user = is_admin(telegram_id)

    return {
        **shift_state,
        "is_admin": is_admin_user,
        "report": get_current_report_for_telegram(telegram_id),
        "history": get_employee_history_for_telegram(telegram_id, {}),
        "routes": get_routes_payload(telegram_id),
        "production": get_production_state_for_telegram(telegram_id),
        "quality": {
            "defect_reasons": DEFECT_REASONS,
            "defect_dispositions": DEFECT_DISPOSITIONS,
        },
        "admin": get_admin_dashboard(telegram_id) if is_admin_user else None,
        "features": {
            "can_work": bool(employee and employee.get("status") == "active"),
            "can_admin": is_admin_user,
            "routes_enabled": ROUTES_MINIAPP_ENABLED,
        },
    }


def can_access_task_attachment(telegram_id: int, task_id: int):
    if is_admin(telegram_id):
        return True

    employee = get_employee_for_access(telegram_id)

    if not employee or employee[5] != "active":
        return False

    task = get_production_task_by_id(task_id)

    if task is None:
        return False

    if employee[3] == "Раскройщик" and task["status"] in {"active", "contours_done", "in_cutting", "formed"}:
        return True

    return False


def make_handler(bot_token: str, debug: bool):
    class MiniAppRequestHandler(BaseHTTPRequestHandler):
        protocol_version = "HTTP/1.1"
        login_attempts = {}
        login_attempts_lock = threading.Lock()
        registration_attempts = {}
        registration_attempts_lock = threading.Lock()

        def client_ip(self):
            if os.getenv("TRUST_PROXY_HEADERS") == "1":
                forwarded = self.headers.get("X-Forwarded-For", "").split(",", 1)[0].strip()
                if forwarded:
                    return forwarded
            return str(self.client_address[0] if self.client_address else "")

        def secure_cookie(self):
            if os.getenv("WEBAPP_COOKIE_SECURE") == "1":
                return True
            if os.getenv("TRUST_PROXY_HEADERS") == "1":
                return self.headers.get("X-Forwarded-Proto", "").lower() == "https"
            return False

        def expected_origin(self):
            configured = os.getenv("WEBAPP_PUBLIC_ORIGIN", "").strip().rstrip("/")
            if configured:
                return configured
            scheme = "https" if self.secure_cookie() else "http"
            return f"{scheme}://{self.headers.get('Host', '').strip()}".rstrip("/")

        def origin_is_valid(self):
            origin = self.headers.get("Origin", "").strip().rstrip("/")
            return bool(origin and hmac.compare_digest(origin, self.expected_origin()))

        def web_session_token(self):
            return session_token_from_cookie(
                self.headers.get("Cookie", ""),
                secure=self.secure_cookie(),
            )

        def authenticate_request(self, payload: dict, *, require_csrf: bool):
            if debug and payload.get("telegram_id"):
                try:
                    return {"id": int(payload["telegram_id"])}
                except (TypeError, ValueError):
                    return None
            session_token = self.web_session_token()
            if session_token:
                session = get_web_session(
                    session_token,
                    self.headers.get("X-CSRF-Token", ""),
                    require_csrf=require_csrf,
                )
                if session:
                    employee = get_employee_for_access(session["telegram_id"])
                    if employee and employee[5] == "active":
                        return {"id": session["telegram_id"], "web_username": session["username"]}
                return None
            return None

        def login_is_rate_limited(self):
            now_epoch = time.time()
            client_ip = self.client_ip()
            with self.login_attempts_lock:
                attempts = [stamp for stamp in self.login_attempts.get(client_ip, []) if now_epoch - stamp < 300]
                self.login_attempts[client_ip] = attempts
                return len(attempts) >= 10

        def record_login_failure(self):
            client_ip = self.client_ip()
            with self.login_attempts_lock:
                self.login_attempts.setdefault(client_ip, []).append(time.time())

        def clear_login_failures(self):
            with self.login_attempts_lock:
                self.login_attempts.pop(self.client_ip(), None)

        def registration_is_rate_limited(self):
            now_epoch = time.time()
            client_ip = self.client_ip()
            with self.registration_attempts_lock:
                attempts = [
                    stamp
                    for stamp in self.registration_attempts.get(client_ip, [])
                    if now_epoch - stamp < 10 * 60
                ]
                if len(attempts) >= 30:
                    self.registration_attempts[client_ip] = attempts
                    return True
                attempts.append(now_epoch)
                self.registration_attempts[client_ip] = attempts
                return False

        def do_GET(self):
            parsed_url = urlparse(self.path)
            path = parsed_url.path

            if send_pwa_resource(
                self,
                path,
                app_shell_revision_token=PWA_SHELL_REVISION,
            ):
                return

            if path == "/favicon.ico":
                self.send_response(204)
                self.send_header("Content-Length", "0")
                self.end_headers()
                return

            if path in {"/", "/app"}:
                self.send_html(PWA_HTML)
                return

            if path == "/health":
                self.send_json({"ok": True})
                return

            if path == "/api/web/session":
                session_token = self.web_session_token()
                session = get_web_session(session_token) if session_token else None
                if session is None:
                    self.send_json({"ok": False, "code": "unauthorized", "message": "Войдите в приложение."}, status=401)
                    return
                employee = get_employee_for_access(session["telegram_id"])
                if employee is None or employee[5] != "active":
                    revoke_web_sessions_for_telegram_id(session["telegram_id"])
                    self.send_json(
                        {
                            "ok": False,
                            "code": "account_disabled",
                            "message": "Доступ к аккаунту отключён. Обратитесь к администратору.",
                        },
                        status=401,
                        extra_headers={"Set-Cookie": build_clear_cookies()},
                    )
                    return
                self.send_json(
                    {
                        "ok": True,
                        "username": session["username"],
                        "telegram_id": session["telegram_id"],
                        "full_name": employee[2] or session.get("full_name") or "",
                        "position": employee[3] or "",
                        "role": employee[4] or "employee",
                        "email": session.get("email") or "",
                        "phone": session.get("phone") or "",
                        "csrf_token": session["csrf_token"],
                        "expires_at": session["expires_at"],
                    }
                )
                return

            if path == "/api/production/task-attachment":
                query = dict(parse_qsl(parsed_url.query, keep_blank_values=True))
                payload = {
                    "telegram_id": query.get("telegram_id", "") or query.get("debug_tg_id", ""),
                }
                user = self.authenticate_request(payload, require_csrf=False)

                if not user or not user.get("id"):
                    self.send_json({"ok": False, "message": "Войдите в веб-приложение или откройте его из Telegram."}, status=401)
                    return

                try:
                    task_id = int(query.get("task_id") or 0)
                except (TypeError, ValueError):
                    task_id = 0

                if task_id <= 0:
                    self.send_json({"ok": False, "message": "Файл не найден."}, status=404)
                    return

                telegram_id = int(user["id"])

                if not can_access_task_attachment(telegram_id, task_id):
                    self.send_json({"ok": False, "message": "Нет доступа к файлу."}, status=403)
                    return

                attachment = get_production_task_attachment(task_id)

                if attachment is None:
                    self.send_json({"ok": False, "message": "Файл не найден."}, status=404)
                    return

                try:
                    content = base64.b64decode(attachment["content_base64"], validate=True)
                except (ValueError, TypeError):
                    self.send_json({"ok": False, "message": "Файл повреждён."}, status=500)
                    return

                disposition = "attachment" if query.get("mode") == "download" else "inline"
                self.send_binary_file(
                    content,
                    attachment["file_name"],
                    attachment["mime_type"],
                    disposition,
                )
                return

            if path in {"/api/routes/qr", "/api/routes/defect-photo"}:
                query = dict(parse_qsl(parsed_url.query, keep_blank_values=True))
                payload = {
                    "telegram_id": query.get("telegram_id", "") or query.get("debug_tg_id", ""),
                }
                user = self.authenticate_request(payload, require_csrf=False)
                if not user or not user.get("id"):
                    self.send_json({"ok": False, "message": "Войдите в веб-приложение или откройте его из Telegram."}, status=401)
                    return
                telegram_id = int(user["id"])

                if path == "/api/routes/defect-photo":
                    try:
                        defect_id = int(query.get("defect_id") or 0)
                    except (TypeError, ValueError):
                        defect_id = 0
                    photo = get_route_batch_defect_photo(defect_id)
                    if photo is None:
                        self.send_json({"ok": False, "message": "Фото не найдено."}, status=404)
                        return
                    batch = get_route_batch_by_id(photo["batch_id"])
                    if not can_access_route_batch(telegram_id, batch):
                        self.send_json({"ok": False, "message": "Нет доступа к фото."}, status=403)
                        return
                    try:
                        content = base64.b64decode(photo["content_base64"], validate=True)
                    except (TypeError, ValueError):
                        self.send_json({"ok": False, "message": "Фото повреждено."}, status=500)
                        return
                    self.send_binary_file(content, photo["file_name"], photo["mime_type"], "inline")
                    return

                try:
                    batch_id = int(query.get("batch_id") or 0)
                except (TypeError, ValueError):
                    batch_id = 0
                batch = get_route_batch_by_id(batch_id)
                if not can_access_route_batch(telegram_id, batch):
                    self.send_json({"ok": False, "message": "Нет доступа к QR партии."}, status=403)
                    return
                try:
                    import qrcode
                    image = qrcode.make(f"TRACE:{batch.get('trace_code') or f'RB-{batch_id:06d}'}")
                    buffer = io.BytesIO()
                    image.save(buffer, format="PNG")
                except (ImportError, OSError):
                    self.send_json({"ok": False, "message": "Не удалось сформировать QR-код."}, status=503)
                    return
                self.send_binary_file(buffer.getvalue(), f"party-{batch_id}.png", "image/png", "inline")
                return

            self.send_json({"ok": False, "message": "Not found"}, status=404)

        def do_HEAD(self):
            path = urlparse(self.path).path
            if send_pwa_resource(
                self,
                path,
                app_shell_revision_token=PWA_SHELL_REVISION,
                head_only=True,
            ):
                return
            self.send_response(404)
            self.send_header("Content-Length", "0")
            self.send_security_headers()
            self.end_headers()

        def do_POST(self):
            path = urlparse(self.path).path

            allowed_paths = {
                "/api/web/login",
                "/api/web/register",
                "/api/web/logout",
                "/api/web/password",
                "/api/app/state",
                "/api/shift/status",
                "/api/shift/open",
                "/api/shift/close",
                "/api/report/history",
                "/api/feedback/send",
                "/api/production/fabric-receipt",
                "/api/production/adjust-stock",
                "/api/production/reject-fabric-rolls",
                "/api/production/create-task",
                "/api/production/create-order-task",
                "/api/production/delete-order-task",
                "/api/production/start-cutting-task",
                "/api/production/submit-contours",
                "/api/production/submit-cutting-stage",
                "/api/routes/create-batch",
                "/api/routes/start",
                "/api/routes/complete",
                "/api/routes/work-action",
                "/api/routes/passport",
                "/api/routes/lookup",
                "/api/admin/dashboard",
                "/api/admin/report",
                "/api/admin/report/export",
                "/api/admin/feedback",
                "/api/admin/employee/status",
                "/api/admin/employee/position",
                "/api/admin/employee/role",
                "/api/admin/shift/close",
                "/api/admin/shift/delete",
                "/api/admin/operation",
            }

            if path not in allowed_paths:
                self.send_json({"ok": False, "message": "Not found"}, status=404)
                return

            if not self.headers.get("Content-Type", "").lower().split(";", 1)[0].strip() == "application/json":
                self.send_json({"ok": False, "message": "Ожидается JSON-запрос."}, status=415)
                return

            payload = self.read_json_body()
            if payload.pop("_request_too_large", False):
                self.send_json({"ok": False, "message": "Запрос или файл слишком большой."}, status=413)
                return

            uses_web_cookie = bool(self.web_session_token())
            if (path in {"/api/web/login", "/api/web/register", "/api/web/logout"} or uses_web_cookie) and not self.origin_is_valid():
                self.send_json({"ok": False, "code": "invalid_origin", "message": "Запрос отклонён."}, status=403)
                return

            if path == "/api/web/login":
                if self.login_is_rate_limited():
                    self.send_json(
                        {"ok": False, "code": "rate_limited", "message": "Слишком много попыток. Повторите вход через несколько минут."},
                        status=429,
                    )
                    return
                account = authenticate_web_credentials(payload.get("username", ""), payload.get("password", ""))
                if account is None:
                    self.record_login_failure()
                    self.send_json(
                        {"ok": False, "code": "invalid_credentials", "message": "Неверный логин или пароль."},
                        status=401,
                    )
                    return
                employee = get_employee_for_access(account["telegram_id"])
                if employee is None:
                    self.send_json(
                        {"ok": False, "code": "invalid_credentials", "message": "Неверный логин или пароль."},
                        status=401,
                    )
                    return
                if employee[5] == "pending":
                    self.send_json(
                        {
                            "ok": False,
                            "code": "account_pending",
                            "message": "Регистрация принята. Дождитесь активации администратором.",
                        },
                        status=403,
                    )
                    return
                if employee[5] != "active":
                    self.send_json(
                        {
                            "ok": False,
                            "code": "account_disabled",
                            "message": "Доступ к аккаунту отключён. Обратитесь к администратору.",
                        },
                        status=403,
                    )
                    return
                self.clear_login_failures()
                session = create_web_session(
                    account,
                    ip_address=self.client_ip(),
                    user_agent=self.headers.get("User-Agent", ""),
                )
                cookie = build_session_cookie(
                    session["session_token"],
                    secure=self.secure_cookie(),
                    max_age=max(0, session["expires_at"] - int(time.time())),
                )
                self.send_json(
                    {
                        "ok": True,
                        "username": session["username"],
                        "telegram_id": session["telegram_id"],
                        "csrf_token": session["csrf_token"],
                        "expires_at": session["expires_at"],
                    },
                    extra_headers={"Set-Cookie": cookie},
                )
                return

            if path == "/api/web/register":
                if self.registration_is_rate_limited():
                    self.send_json(
                        {
                            "ok": False,
                            "code": "rate_limited",
                            "message": "Слишком много регистраций. Повторите попытку позже.",
                        },
                        status=429,
                    )
                    return
                try:
                    registration = register_web_account(
                        payload.get("email", ""),
                        payload.get("phone", ""),
                        payload.get("full_name", ""),
                        payload.get("password", ""),
                    )
                except WebRegistrationError as error:
                    conflict_codes = {"email_exists", "phone_exists", "registration_conflict"}
                    self.send_json(
                        {"ok": False, "code": error.code, "message": str(error)},
                        status=409 if error.code in conflict_codes else 400,
                    )
                    return
                except Exception:
                    logging.exception("Web registration failed")
                    self.send_json(
                        {
                            "ok": False,
                            "code": "registration_unavailable",
                            "message": "Не удалось завершить регистрацию. Повторите попытку позже.",
                        },
                        status=500,
                    )
                    return
                self.send_json(
                    {
                        "ok": True,
                        "status": registration["status"],
                        "message": "Регистрация завершена. Администратор назначит должность и активирует доступ.",
                    },
                    status=201,
                )
                return

            if path == "/api/web/logout":
                session_token = self.web_session_token()
                existing_session = get_web_session(session_token) if session_token else None
                if existing_session is None:
                    self.send_json(
                        {"ok": True, "message": "Вы вышли из приложения."},
                        extra_headers={"Set-Cookie": build_clear_cookies()},
                    )
                    return
                session = get_web_session(
                    session_token,
                    self.headers.get("X-CSRF-Token", ""),
                    require_csrf=True,
                )
                if session is None:
                    self.send_json({"ok": False, "code": "invalid_csrf", "message": "Сессия устарела."}, status=403)
                    return
                revoke_web_session(session_token)
                self.send_json(
                    {"ok": True, "message": "Вы вышли из приложения."},
                    extra_headers={"Set-Cookie": build_clear_cookies()},
                )
                return

            if path == "/api/web/password":
                session_token = self.web_session_token()
                session = get_web_session(
                    session_token,
                    self.headers.get("X-CSRF-Token", ""),
                    require_csrf=True,
                ) if session_token else None
                if session is None:
                    self.send_json({"ok": False, "code": "invalid_csrf", "message": "Сессия устарела."}, status=403)
                    return
                password_result = change_web_password(
                    session["account_id"],
                    payload.get("current_password", ""),
                    payload.get("new_password", ""),
                )
                if not password_result.get("ok"):
                    self.send_json(password_result, status=400)
                    return
                self.send_json(
                    password_result,
                    extra_headers={"Set-Cookie": build_clear_cookies()},
                )
                return

            user = self.authenticate_request(payload, require_csrf=True)

            if not user or not user.get("id"):
                self.send_json(
                    {
                        "ok": False,
                        "code": "unauthorized",
                        "message": "Войдите в веб-приложение или откройте его из Telegram.",
                        "employee": None,
                        "shift": None,
                    },
                    status=401,
                )
                return

            telegram_id = int(user["id"])

            if path == "/api/admin/report/export":
                export_result = export_admin_report_for_telegram(telegram_id, payload)

                if not export_result.get("ok"):
                    self.send_json(export_result, status=400)
                    return

                self.send_file(export_result["content"], export_result["filename"])
                return

            if path == "/api/app/state":
                result = get_app_state(telegram_id, payload.get("message", ""))
            elif path == "/api/report/history":
                result = get_employee_history_for_telegram(telegram_id, payload)
            elif path == "/api/shift/open":
                action_result = open_shift_for_telegram(telegram_id)
                result = get_app_state(telegram_id, action_result.get("message", ""))
            elif path == "/api/shift/close":
                action_result = close_shift_for_telegram(telegram_id)
                result = get_app_state(telegram_id, action_result.get("message", ""))
            elif path == "/api/feedback/send":
                result = submit_feedback_for_telegram(
                    telegram_id,
                    payload.get("category", ""),
                    payload.get("message", ""),
                )
            elif path == "/api/production/fabric-receipt":
                result = add_fabric_receipt_for_telegram(telegram_id, payload)
            elif path == "/api/production/adjust-stock":
                result = adjust_stock_for_telegram(telegram_id, payload)
            elif path == "/api/production/reject-fabric-rolls":
                result = reject_fabric_rolls_for_telegram(telegram_id, payload)
            elif path == "/api/production/create-task":
                result = create_production_task_for_telegram(telegram_id, payload)
            elif path == "/api/production/create-order-task":
                result = create_order_task_for_telegram(telegram_id, payload)
            elif path == "/api/production/delete-order-task":
                result = delete_order_task_for_telegram(telegram_id, payload)
            elif path == "/api/production/start-cutting-task":
                try:
                    task_id = int(payload.get("task_id") or 0)
                except (TypeError, ValueError):
                    task_id = 0
                result = start_cutting_task_for_telegram(telegram_id, task_id)
            elif path == "/api/production/submit-contours":
                result = submit_production_contours_for_telegram(telegram_id, payload)
            elif path == "/api/production/submit-cutting-stage":
                result = submit_cutting_stage_for_telegram(telegram_id, payload)
            elif path == "/api/routes/create-batch":
                result = create_route_batch_for_telegram(telegram_id, payload)
            elif path == "/api/routes/start":
                try:
                    batch_id = int(payload.get("batch_id") or 0)
                except (TypeError, ValueError):
                    batch_id = 0

                result = start_route_task_for_telegram(telegram_id, batch_id)
            elif path == "/api/routes/complete":
                try:
                    batch_id = int(payload.get("batch_id") or 0)
                except (TypeError, ValueError):
                    batch_id = 0

                result = complete_route_task_for_telegram(telegram_id, batch_id, payload)
            elif path == "/api/routes/work-action":
                try:
                    batch_id = int(payload.get("batch_id") or 0)
                except (TypeError, ValueError):
                    batch_id = 0
                result = route_task_work_action_for_telegram(telegram_id, batch_id, payload)
            elif path == "/api/routes/passport":
                try:
                    batch_id = int(payload.get("batch_id") or 0)
                except (TypeError, ValueError):
                    batch_id = 0
                result = get_route_passport_for_telegram(telegram_id, batch_id)
            elif path == "/api/routes/lookup":
                result = lookup_route_trace_for_telegram(telegram_id, str(payload.get("trace_code") or ""))
            elif path == "/api/admin/dashboard":
                result = get_admin_dashboard(telegram_id)
            elif path == "/api/admin/report":
                result = get_admin_report_for_telegram(telegram_id, payload)
            elif path == "/api/admin/feedback":
                result = get_admin_feedback_for_telegram(telegram_id, payload)
            elif path == "/api/admin/employee/status":
                result = set_employee_status_for_admin(telegram_id, payload)
            elif path == "/api/admin/employee/position":
                result = set_employee_position_for_admin(telegram_id, payload)
            elif path == "/api/admin/employee/role":
                result = set_employee_role_for_admin(telegram_id, payload)
            elif path == "/api/admin/shift/close":
                result = close_shift_for_admin(telegram_id, payload)
            elif path == "/api/admin/shift/delete":
                result = delete_shift_for_admin(telegram_id, payload)
            elif path == "/api/admin/operation":
                result = operation_action_for_admin(telegram_id, payload)
            else:
                result = get_app_state(telegram_id)

            self.send_json(result)

        def read_json_body(self):
            try:
                content_length = int(self.headers.get("Content-Length", "0") or "0")
            except ValueError:
                return {}

            if content_length < 0 or content_length > MAX_REQUEST_BODY_BYTES:
                self.close_connection = True
                return {"_request_too_large": True}

            raw_body = self.rfile.read(content_length) if content_length else b"{}"

            try:
                return json.loads(raw_body.decode("utf-8"))
            except json.JSONDecodeError:
                return {}

        def send_security_headers(self):
            self.send_header("X-Content-Type-Options", "nosniff")
            self.send_header("Referrer-Policy", "same-origin")
            self.send_header("Permissions-Policy", "geolocation=(), microphone=(), camera=(self)")
            self.send_header("Content-Security-Policy", CONTENT_SECURITY_POLICY)

        def send_html(self, html_text: str, status: int = 200):
            body = html_text.encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.send_security_headers()
            self.end_headers()
            self.wfile.write(body)

        def send_json(self, payload: dict, status: int = 200, extra_headers: dict | None = None):
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.send_security_headers()
            for header_name, header_value in (extra_headers or {}).items():
                values = header_value if isinstance(header_value, (list, tuple)) else (header_value,)
                for value in values:
                    self.send_header(header_name, value)
            self.end_headers()
            self.wfile.write(body)

        def send_file(self, content: bytes, filename: str):
            self.send_binary_file(
                content,
                filename,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "attachment",
            )

        def send_binary_file(self, content: bytes, filename: str, mime_type: str, disposition: str):
            safe_filename = quote(filename)
            self.send_response(200)
            self.send_header("Content-Type", mime_type or "application/octet-stream")
            self.send_header("Content-Length", str(len(content)))
            self.send_header(
                "Content-Disposition",
                f"{disposition}; filename*=UTF-8''{safe_filename}",
            )
            self.send_header("Cache-Control", "no-store")
            self.send_security_headers()
            self.end_headers()
            self.wfile.write(content)

        def log_message(self, format_string, *args):
            logging.info("Miniapp: " + format_string, *args)

    return MiniAppRequestHandler


class ReusableThreadingHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True
    daemon_threads = True
    block_on_close = False
    request_slots = threading.BoundedSemaphore(MAX_CONCURRENT_REQUESTS)

    def get_request(self):
        request, client_address = super().get_request()
        request.settimeout(REQUEST_SOCKET_TIMEOUT_SECONDS)
        return request, client_address

    def process_request(self, request, client_address):
        if not self.request_slots.acquire(blocking=False):
            self.shutdown_request(request)
            return
        try:
            super().process_request(request, client_address)
        except BaseException:
            self.request_slots.release()
            raise

    def process_request_thread(self, request, client_address):
        try:
            super().process_request_thread(request, client_address)
        finally:
            self.request_slots.release()


def start_miniapp_server(bot_token: str, host: str, port: int, debug: bool = False):
    if os.getenv("MINIAPP_ENABLED", "1") == "0":
        logging.info("Miniapp server disabled")
        return None

    is_loopback = host in {"127.0.0.1", "localhost", "::1"}
    if debug and not is_loopback:
        logging.critical("Miniapp debug impersonation is allowed only on a loopback host")
        return None
    if os.getenv("WEBAPP_ENV", "").strip().lower() == "production":
        public_origin = os.getenv("WEBAPP_PUBLIC_ORIGIN", "").strip()
        if debug or os.getenv("WEBAPP_COOKIE_SECURE") != "1" or not public_origin.startswith("https://"):
            logging.critical("Production webapp requires HTTPS origin, secure cookies, and debug disabled")
            return None

    init_web_auth()

    try:
        server = ReusableThreadingHTTPServer(
            (host, port),
            make_handler(bot_token, debug),
        )
    except OSError:
        logging.exception("Miniapp server failed to bind on %s:%s", host, port)
        return None

    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    logging.info("Miniapp server started on %s:%s", host, port)
    return server
