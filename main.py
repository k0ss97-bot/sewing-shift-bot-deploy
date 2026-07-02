import asyncio
import logging
import os
import re
import shutil
import sqlite3
from datetime import date, datetime

from aiogram import Bot, Dispatcher
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    CallbackQuery,
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
)
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database import (
    DB_NAME,
    add_operation,
    add_shift_operation,
    add_edit_log,
    add_cutting_layout,
    admin_close_shift,
    close_shift,
    create_cutting_contour_batch,
    create_employee,
    create_shift,
    delete_cutting_batch,
    delete_shift_by_id,
    delete_shift_operation,
    get_active_operation_groups,
    get_active_operation_folders,
    get_active_operations,
    get_all_employees,
    get_all_operations,
    get_admin_cutting_batches,
    get_cutting_batch_result_rows,
    get_cutting_batches_for_cutting,
    get_cutting_batches_for_formation,
    get_cutting_batches_for_layout,
    get_database_status,
    get_editable_shift_for_today,
    get_employee_by_telegram_id,
    get_employee_period_operation_totals,
    get_employee_period_summary,
    get_employee_recent_shifts,
    get_employee_shifts_by_period,
    get_employees_by_status,
    get_month_employee_summary,
    get_month_operation_rows,
    get_month_operations_by_employee,
    get_month_shift_details,
    get_open_shift_for_today,
    get_open_shifts,
    get_operation_by_number,
    get_period_employee_summary,
    get_period_operation_rows,
    get_period_operations_by_employee,
    get_period_shift_details,
    get_pending_employees,
    get_product_colors,
    get_product_sizes,
    get_preparation_material_colors,
    get_preparation_operation_sizes,
    get_recent_edit_logs,
    get_recent_shifts,
    get_shift_for_today,
    get_shift_operation_choices,
    get_shift_report,
    get_today_shifts,
    hide_operation,
    init_db,
    is_preparation_operation_folder,
    mark_cutting_batch_formed,
    restore_operation,
    set_shift_operation_quantity,
    update_cutting_batch_progress,
    update_operation_field,
    update_employee_position,
    update_shift_operation_quantity,
    update_employee_status,
    local_now,
    local_today,
)


load_dotenv(".env")

BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_IDS = os.getenv("ADMIN_IDS", "")
LOGS_DIR = "logs"
ERROR_LOG_FILE = os.path.join(LOGS_DIR, "errors.log")

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


class Registration(StatesGroup):
    waiting_for_full_name = State()
    waiting_for_position = State()


class Report(StatesGroup):
    waiting_for_group = State()
    waiting_for_folder = State()
    waiting_for_batch = State()
    waiting_for_progress = State()
    waiting_for_size = State()
    waiting_for_color = State()
    waiting_for_operation = State()
    waiting_for_quantity = State()


class EditReport(StatesGroup):
    waiting_for_operation = State()
    waiting_for_quantity = State()


class DeleteReport(StatesGroup):
    waiting_for_operation = State()


class DateReport(StatesGroup):
    waiting_for_period = State()


class EmployeeAdmin(StatesGroup):
    waiting_for_active_id = State()
    waiting_for_inactive_id = State()
    waiting_for_position_employee_id = State()
    waiting_for_new_position = State()


class AdminPeriod(StatesGroup):
    waiting_for_report_period = State()
    waiting_for_excel_period = State()
    waiting_for_employee_report_id = State()
    waiting_for_employee_report_period = State()


class AdminEditReport(StatesGroup):
    waiting_for_employee = State()
    waiting_for_shift = State()
    waiting_for_operation = State()
    waiting_for_action = State()
    waiting_for_quantity = State()


class AdminCuttingBatch(StatesGroup):
    waiting_for_batch = State()


class OperationAdmin(StatesGroup):
    waiting_for_name = State()
    waiting_for_position = State()
    waiting_for_group = State()
    waiting_for_folder = State()
    waiting_for_edit_number = State()
    waiting_for_edit_field = State()
    waiting_for_edit_value = State()
    waiting_for_hide_number = State()
    waiting_for_restore_number = State()


class DatabaseAdmin(StatesGroup):
    waiting_for_database_file = State()


POSITIONS = {
    "1": "Швея",
    "2": "Упаковщик",
    "3": "Раскройщик",
}

CUTTING_CONTOUR_OPERATION = "Нанесение контуров лекал на ткань"
CUTTING_LAYOUT_OPERATION = "Формирование настила"
CUTTING_CUT_OPERATION = "Раскрой"
CUTTING_FORM_OPERATION = "Формирование готового кроя"
CUTTING_MODE_CONTOURS = "contours"
CUTTING_MODE_LAYOUT = "layout"
CUTTING_MODE_CUT = "cut"
CUTTING_MODE_FORM = "form"
CUTTING_MODE_FULL = "full"


def employee_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="Отчёт"),
                KeyboardButton(text="Закрыть смену"),
            ],
            [
                KeyboardButton(text="Админ-панель"),
            ],
        ],
        resize_keyboard=True,
    )


def report_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="Отправить отчёт"),
                KeyboardButton(text="Текущий отчёт"),
            ],
            [
                KeyboardButton(text="Изменить отчёт"),
                KeyboardButton(text="Удалить операцию"),
            ],
            [
                KeyboardButton(text="Отчёт за даты"),
                KeyboardButton(text="Назад"),
            ],
        ],
        resize_keyboard=True,
    )


def get_cutting_mode(operation_name: str) -> str:
    if operation_name == CUTTING_CONTOUR_OPERATION:
        return CUTTING_MODE_CONTOURS

    if operation_name == CUTTING_LAYOUT_OPERATION:
        return CUTTING_MODE_LAYOUT

    if operation_name == CUTTING_CUT_OPERATION:
        return CUTTING_MODE_CUT

    if operation_name == CUTTING_FORM_OPERATION:
        return CUTTING_MODE_FORM

    return CUTTING_MODE_FULL


def is_packing_preparation_flow(data: dict):
    return (
        data.get("employee_position") == "Упаковщик"
        and is_preparation_operation_folder(data.get("selected_folder", ""))
    )


def admin_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="Отчёты"),
                KeyboardButton(text="Смены"),
            ],
            [
                KeyboardButton(text="Сотрудники"),
                KeyboardButton(text="Операции"),
            ],
            [
                KeyboardButton(text="Файлы"),
                KeyboardButton(text="В меню сотрудника"),
            ],
        ],
        resize_keyboard=True,
    )


def admin_reports_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="Отчёт за сегодня"),
                KeyboardButton(text="Отчёт за месяц"),
            ],
            [
                KeyboardButton(text="Отчёт за период"),
                KeyboardButton(text="Excel за период"),
            ],
            [
                KeyboardButton(text="Отчёт по сотруднику"),
                KeyboardButton(text="Править отчёт"),
            ],
            [
                KeyboardButton(text="Выгрузить отчёт"),
                KeyboardButton(text="Админ меню"),
            ],
        ],
        resize_keyboard=True,
    )


def admin_shifts_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="Открытые смены"),
                KeyboardButton(text="Последние смены"),
            ],
            [
                KeyboardButton(text="Удалить смену"),
                KeyboardButton(text="Админ меню"),
            ],
        ],
        resize_keyboard=True,
    )


def admin_employees_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="Заявки"),
                KeyboardButton(text="Список сотрудников"),
            ],
            [
                KeyboardButton(text="Активные сотрудники"),
                KeyboardButton(text="Неактивные сотрудники"),
            ],
            [
                KeyboardButton(text="Активировать сотрудника"),
                KeyboardButton(text="Отключить сотрудника"),
            ],
            [
                KeyboardButton(text="Сменить должность"),
            ],
            [
                KeyboardButton(text="Админ меню"),
            ],
        ],
        resize_keyboard=True,
    )


def admin_operations_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="Список операций"),
                KeyboardButton(text="Добавить операцию"),
            ],
            [
                KeyboardButton(text="Изменить операцию"),
            ],
            [
                KeyboardButton(text="Скрыть операцию"),
                KeyboardButton(text="Вернуть операцию"),
            ],
            [
                KeyboardButton(text="Админ меню"),
            ],
        ],
        resize_keyboard=True,
    )


def admin_files_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="Журнал"),
                KeyboardButton(text="Скачать базу"),
            ],
            [
                KeyboardButton(text="Проверка базы"),
            ],
            [
                KeyboardButton(text="Загрузить базу"),
            ],
            [
                KeyboardButton(text="Создать копию базы"),
                KeyboardButton(text="Ошибки"),
            ],
            [
                KeyboardButton(text="Админ меню"),
            ],
        ],
        resize_keyboard=True,
    )


def position_list_text():
    return (
        "1. Швея\n"
        "2. Упаковщик\n"
        "3. Раскройщик"
    )


EMPLOYEE_STATUS_NAMES = {
    "active": "активен",
    "inactive": "неактивен",
    "pending": "заявка",
    "rejected": "отклонён",
}


def format_employee_status(status: str):
    return EMPLOYEE_STATUS_NAMES.get(status, status)


def setup_logging():
    os.makedirs(LOGS_DIR, exist_ok=True)

    if logging.getLogger().handlers:
        return

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        handlers=[
            logging.FileHandler(ERROR_LOG_FILE, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


def choice_keyboard(prefix: str, items: dict[str, str | dict], with_navigation: bool = True):
    keyboard = []

    for number, item in items.items():
        if isinstance(item, dict):
            label = item.get("name", number)
        else:
            label = item

        keyboard.append([
            InlineKeyboardButton(
                text=f"{number}. {label}",
                callback_data=f"{prefix}:{number}",
            )
        ])

    if with_navigation:
        keyboard.append([
            InlineKeyboardButton(text="⬅️ Назад", callback_data="nav_back"),
            InlineKeyboardButton(text="❌ Отмена", callback_data="nav_cancel"),
        ])

    return InlineKeyboardMarkup(inline_keyboard=keyboard)


def navigation_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="⬅️ Назад", callback_data="nav_back"),
                InlineKeyboardButton(text="❌ Отмена", callback_data="nav_cancel"),
            ]
        ]
    )


def multi_choice_keyboard(prefix: str, items: dict[str, str], selected_numbers: list[str]):
    keyboard = []
    selected_set = set(selected_numbers)

    for number, label in items.items():
        mark = "☑️" if number in selected_set else "☐"
        keyboard.append([
            InlineKeyboardButton(
                text=f"{mark} {number}. {label}",
                callback_data=f"{prefix}_toggle:{number}",
            )
        ])

    keyboard.append([
        InlineKeyboardButton(text="✅ Далее", callback_data=f"{prefix}_done"),
    ])
    keyboard.append([
        InlineKeyboardButton(text="⬅️ Назад", callback_data="nav_back"),
        InlineKeyboardButton(text="❌ Отмена", callback_data="nav_cancel"),
    ])

    return InlineKeyboardMarkup(inline_keyboard=keyboard)


def progress_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="25%", callback_data="cutting_progress:25"),
                InlineKeyboardButton(text="50%", callback_data="cutting_progress:50"),
            ],
            [
                InlineKeyboardButton(text="75%", callback_data="cutting_progress:75"),
                InlineKeyboardButton(text="100%", callback_data="cutting_progress:100"),
            ],
            [
                InlineKeyboardButton(text="⬅️ Назад", callback_data="nav_back"),
                InlineKeyboardButton(text="❌ Отмена", callback_data="nav_cancel"),
            ],
        ]
    )


async def callback_state_is_current(callback: CallbackQuery, state: FSMContext, expected_state: State):
    if await state.get_state() == expected_state.state:
        return True

    await callback.answer("Это старое меню. Откройте отчёт заново.", show_alert=False)
    return False


def get_admin_ids():
    return [int(admin_id.strip()) for admin_id in ADMIN_IDS.split(",") if admin_id.strip()]


def is_admin(telegram_id: int):
    return telegram_id in get_admin_ids()


def format_minutes(total_minutes: int):
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours} ч. {minutes} мин."


def create_database_backup(kind: str = "manual"):
    backups_dir = "backups"
    os.makedirs(backups_dir, exist_ok=True)

    timestamp = local_now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_path = os.path.join(backups_dir, f"{kind}_{timestamp}_{os.path.basename(DB_NAME)}")

    source = sqlite3.connect(DB_NAME)
    backup = sqlite3.connect(backup_path)

    try:
        source.backup(backup)
    finally:
        backup.close()
        source.close()

    return backup_path


def create_daily_database_backup():
    backups_dir = "backups"
    os.makedirs(backups_dir, exist_ok=True)

    today = local_today().isoformat()
    daily_prefix = f"auto_{today}_"

    if any(file_name.startswith(daily_prefix) for file_name in os.listdir(backups_dir)):
        return None

    return create_database_backup("auto")


def format_operation_line(name: str, product_size: str, product_color: str, quantity: int, unit: str):
    details = []

    if product_size and product_size != "без размера":
        details.append(f"размер {product_size}")

    if product_color and product_color != "без цвета":
        details.append(f"цвет {product_color}")

    details_text = f" ({', '.join(details)})" if details else ""
    return f"{name}{details_text} — {quantity} {unit}"


def is_valid_time(value: str):
    return re.fullmatch(r"([01]\d|2[0-3]):[0-5]\d", value) is not None


async def send_long_text(message: Message, text: str, reply_markup=None):
    max_length = 3500

    while len(text) > max_length:
        split_at = text.rfind("\n", 0, max_length)

        if split_at == -1:
            split_at = max_length

        await message.answer(text[:split_at])
        text = text[split_at:].lstrip()

    await message.answer(text, reply_markup=reply_markup)


def is_valid_date(value: str):
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def parse_period(text: str):
    parts = text.split()

    if len(parts) != 2:
        return None

    start_date, end_date = parts

    if not is_valid_date(start_date) or not is_valid_date(end_date):
        return None

    if start_date > end_date:
        return None

    return start_date, end_date


async def send_report_group_step(message: Message, state: FSMContext):
    data = await state.get_data()
    employee_position = data["employee_position"]
    group_map = data.get("group_map", {})

    text = f"Раздел: {employee_position}\n\nВыберите группу. Напишите номер:\n\n"

    for number, group in group_map.items():
        text += f"{number}. {group}\n"

    await state.set_state(Report.waiting_for_group)
    await message.answer(text, reply_markup=choice_keyboard("report_group", group_map))


async def send_report_folder_step(message: Message, state: FSMContext):
    data = await state.get_data()
    selected_group = data["selected_group"]
    folder_map = data.get("folder_map", {})
    item_word = "раздел подготовки" if data.get("employee_position") == "Упаковщик" and selected_group == "Подготовка" else "изделие"

    text = f"Группа: {selected_group}\n\nВыберите {item_word}. Напишите номер:\n\n"

    for number, folder in folder_map.items():
        text += f"{number}. {folder}\n"

    await state.set_state(Report.waiting_for_folder)
    await message.answer(text, reply_markup=choice_keyboard("report_folder", folder_map))


async def send_report_size_step(message: Message, state: FSMContext):
    data = await state.get_data()
    selected_folder = data["selected_folder"]
    size_map = data.get("size_map", {})
    selected_sizes = data.get("selected_size_numbers", [])

    if not size_map:
        await send_report_folder_step(message, state)
        return

    if data.get("employee_position") == "Раскройщик" or is_packing_preparation_flow(data):
        cutting_mode = data.get("cutting_mode")
        action_text = "Выберите один или несколько размеров"

        if cutting_mode == CUTTING_MODE_CONTOURS:
            action_text = "Выберите размеры, на которые нанесены контуры лекал"

        if is_packing_preparation_flow(data):
            action_text = "Выберите один или несколько размеров для подготовки"

        text = (
            f"Изделие: {selected_folder}\n\n"
            f"{action_text}, затем нажмите «Далее».\n"
            "Можно написать номера через пробел, например: 1 2 3 4\n\n"
        )

        await state.set_state(Report.waiting_for_size)
        await message.answer(
            text,
            reply_markup=multi_choice_keyboard("report_size_multi", size_map, selected_sizes),
        )
        return

    text = f"Изделие: {selected_folder}\n\nВыберите размер. Напишите номер:\n\n"

    for number, product_size in size_map.items():
        text += f"{number}. {product_size}\n"

    await state.set_state(Report.waiting_for_size)
    await message.answer(text, reply_markup=choice_keyboard("report_size", size_map))


async def send_report_color_step(message: Message, state: FSMContext):
    data = await state.get_data()
    selected_folder = data["selected_folder"]
    selected_size = data.get("selected_size", ", ".join(data.get("selected_sizes", [])))
    color_map = data.get("color_map", {})
    selected_colors = data.get("selected_color_numbers", [])

    if not color_map:
        await send_report_size_step(message, state)
        return

    if data.get("employee_position") == "Раскройщик":
        cutting_mode = data.get("cutting_mode")
        size_line = "" if cutting_mode == CUTTING_MODE_LAYOUT else f"Размеры: {selected_size}\n"
        action_text = "Выберите один или несколько цветов"

        if cutting_mode == CUTTING_MODE_LAYOUT:
            action_text = "Выберите цвета настила"

        text = (
            f"Изделие: {selected_folder}\n"
            f"{size_line}\n"
            f"{action_text}, затем нажмите «Далее».\n"
            "Можно написать номера через пробел, например: 1 2 3\n\n"
        )

        await state.set_state(Report.waiting_for_color)
        await message.answer(
            text,
            reply_markup=multi_choice_keyboard("report_color_multi", color_map, selected_colors),
        )
        return

    text = f"Изделие: {selected_folder}\nРазмер: {selected_size}\n\nВыберите цвет. Напишите номер:\n\n"

    for number, color in color_map.items():
        text += f"{number}. {color}\n"

    await state.set_state(Report.waiting_for_color)
    await message.answer(text, reply_markup=choice_keyboard("report_color", color_map))


async def send_report_operation_step(message: Message, state: FSMContext):
    data = await state.get_data()
    selected_folder = data["selected_folder"]
    operation_map = data.get("operation_map", {})

    text = f"Изделие: {selected_folder}\n"

    if data.get("selected_size") or data.get("selected_sizes"):
        text += f"Размер: {data.get('selected_size', ', '.join(data.get('selected_sizes', [])))}\n"

    if data.get("selected_color") or data.get("selected_colors"):
        text += f"Цвет: {data.get('selected_color', ', '.join(data.get('selected_colors', [])))}\n"

    text += "\nВыберите операцию. Напишите номер:\n\n"

    for number, operation in operation_map.items():
        text += f"{number}. {operation['name']}\n"

    await state.set_state(Report.waiting_for_operation)
    await message.answer(text, reply_markup=choice_keyboard("report_operation", operation_map))


def format_cutting_batch_result(rows):
    if not rows:
        return "Расчёта пока нет."

    text = ""
    for product_size, product_color, quantity in rows:
        text += f"- размер {product_size}, цвет {product_color}: {quantity} шт\n"
    return text


async def send_cutting_batch_step(message: Message, state: FSMContext):
    data = await state.get_data()
    selected_folder = data["selected_folder"]
    cutting_mode = data.get("cutting_mode")

    if cutting_mode == CUTTING_MODE_LAYOUT:
        batches = get_cutting_batches_for_layout(selected_folder)
        empty_text = "Для этого изделия нет готовых контуров. Сначала сделайте «Нанесение контуров лекал на ткань»."
        title = "Выберите нанесение контуров для настила:"
    elif cutting_mode == CUTTING_MODE_CUT:
        batches = get_cutting_batches_for_cutting(selected_folder)
        empty_text = "Для этого изделия нет готового настила. Сначала сделайте «Формирование настила»."
        title = "Выберите настил для раскроя:"
    else:
        batches = get_cutting_batches_for_formation(selected_folder)
        empty_text = "Для этого изделия нет раскроя, завершённого на 100%."
        title = "Выберите завершённый раскрой для формирования готового кроя:"

    if not batches:
        await message.answer(empty_text, reply_markup=navigation_keyboard())
        return

    batch_map = {}
    text = f"Изделие: {selected_folder}\n\n{title}\n\n"

    for index, batch in enumerate(batches, start=1):
        if cutting_mode == CUTTING_MODE_LAYOUT:
            batch_id, product_name, contour_date, employee_name, sizes_text = batch
            line = f"{CUTTING_CONTOUR_OPERATION} {contour_date} — {product_name}"
            if sizes_text:
                line += f"\n   размеры: {sizes_text}"
            if employee_name:
                line += f"\n   сотрудник: {employee_name}"
        elif cutting_mode == CUTTING_MODE_CUT:
            batch_id, product_name, contour_date, layout_date, progress, colors_text = batch
            line = (
                f"{CUTTING_LAYOUT_OPERATION} {layout_date} — {product_name}\n"
                f"   контуры: {contour_date}, готовность раскроя: {progress}%"
            )
            if colors_text:
                line += f"\n   цвета: {colors_text}"
        else:
            batch_id, product_name, contour_date, layout_date, progress = batch
            line = (
                f"{CUTTING_CUT_OPERATION} — {product_name}\n"
                f"   контуры: {contour_date}, настил: {layout_date}, готовность: {progress}%"
            )

        batch_map[str(index)] = {
            "id": batch_id,
            "name": line.replace("\n", " "),
        }
        text += f"{index}. {line}\n\n"

    await state.update_data(cutting_batch_map=batch_map)
    await state.set_state(Report.waiting_for_batch)
    await message.answer(text, reply_markup=choice_keyboard("cutting_batch", batch_map))


async def go_back_in_report(message: Message, state: FSMContext, current_state: str):
    if current_state == Report.waiting_for_group.state:
        await state.clear()
        await message.answer("Выберите действие с отчётом:", reply_markup=report_keyboard())
        return

    if current_state == Report.waiting_for_folder.state:
        await send_report_group_step(message, state)
        return

    if current_state == Report.waiting_for_batch.state:
        await send_report_operation_step(message, state)
        return

    if current_state == Report.waiting_for_progress.state:
        await send_cutting_batch_step(message, state)
        return

    if current_state == Report.waiting_for_size.state:
        data = await state.get_data()
        if is_packing_preparation_flow(data):
            await send_report_operation_step(message, state)
            return

        if data.get("employee_position") == "Раскройщик" and data.get("operation_id"):
            await send_report_operation_step(message, state)
            return

        await send_report_folder_step(message, state)
        return

    if current_state == Report.waiting_for_color.state:
        data = await state.get_data()
        if is_packing_preparation_flow(data):
            await send_report_size_step(message, state)
            return

        if data.get("employee_position") == "Раскройщик":
            if data.get("cutting_mode") == CUTTING_MODE_LAYOUT:
                await send_report_operation_step(message, state)
                return

            if data.get("size_map"):
                await send_report_size_step(message, state)
                return

        if data.get("size_map"):
            await send_report_size_step(message, state)
        else:
            await send_report_folder_step(message, state)
        return

    if current_state == Report.waiting_for_operation.state:
        data = await state.get_data()
        if data.get("employee_position") == "Раскройщик":
            await send_report_folder_step(message, state)
            return

        if data.get("color_map"):
            await send_report_color_step(message, state)
        elif data.get("size_map"):
            await send_report_size_step(message, state)
        else:
            await send_report_folder_step(message, state)
        return

    if current_state == Report.waiting_for_quantity.state:
        data = await state.get_data()
        if is_packing_preparation_flow(data):
            await send_report_color_step(message, state)
        elif data.get("employee_position") == "Раскройщик" and data.get("cutting_mode") == CUTTING_MODE_CONTOURS:
            await send_report_size_step(message, state)
        elif data.get("employee_position") == "Раскройщик" and data.get("color_map"):
            await send_report_color_step(message, state)
        else:
            await send_report_operation_step(message, state)
        return

    await state.clear()
    await message.answer("Основное меню:", reply_markup=employee_keyboard())


async def send_admin_edit_employee_step(message: Message, state: FSMContext):
    employees = get_all_employees()

    if not employees:
        await state.clear()
        await message.answer("Сотрудников пока нет.", reply_markup=admin_reports_keyboard())
        return

    employee_map = {}
    text = "Выберите сотрудника, чей отчёт нужно исправить:\n\n"

    for employee in employees:
        employee_id, full_name, position, _, status = employee
        employee_map[str(employee_id)] = {
            "id": employee_id,
            "name": f"{full_name} — {position}, {format_employee_status(status)}",
        }
        text += f"{employee_id}. {full_name} — {position}, {format_employee_status(status)}\n"

    await state.update_data(admin_edit_employee_map=employee_map)
    await state.set_state(AdminEditReport.waiting_for_employee)
    await message.answer(text, reply_markup=choice_keyboard("admin_edit_employee", employee_map))


async def send_admin_edit_shift_step(message: Message, state: FSMContext):
    data = await state.get_data()
    employee_id = data["admin_edit_employee_id"]
    shifts = get_employee_recent_shifts(employee_id, 20)

    if not shifts:
        await message.answer("У этого сотрудника пока нет смен.", reply_markup=admin_reports_keyboard())
        await state.clear()
        return

    shift_map = {}
    text = f"Сотрудник: {data['admin_edit_employee_name']}\n\nВыберите смену:\n\n"

    for index, shift in enumerate(shifts, start=1):
        shift_id, shift_date, start_time, end_time, status, operation_count = shift
        end_text = end_time if end_time else "открыта"
        status_text = "закрыта" if status == "closed" else "открыта"
        line = f"{shift_date}, {start_time}–{end_text}, {status_text}, операций: {operation_count}"
        shift_map[str(index)] = {
            "id": shift_id,
            "name": line,
        }
        text += f"{index}. ID {shift_id} — {line}\n"

    await state.update_data(admin_edit_shift_map=shift_map)
    await state.set_state(AdminEditReport.waiting_for_shift)
    await message.answer(text, reply_markup=choice_keyboard("admin_edit_shift", shift_map))


async def send_admin_edit_operation_step(message: Message, state: FSMContext):
    data = await state.get_data()
    shift_id = data["admin_edit_shift_id"]
    operations = get_shift_operation_choices(shift_id)

    if not operations:
        await message.answer("В этой смене нет операций для правки.", reply_markup=admin_reports_keyboard())
        await state.clear()
        return

    operation_map = {}
    text = (
        f"Сотрудник: {data['admin_edit_employee_name']}\n"
        f"Смена ID {shift_id}: {data['admin_edit_shift_name']}\n\n"
        "Выберите строку отчёта:\n\n"
    )

    for index, operation in enumerate(operations, start=1):
        shift_operation_id, name, product_size, product_color, quantity, unit = operation
        operation_line = format_operation_line(name, product_size, product_color, quantity, unit)
        operation_map[str(index)] = {
            "id": shift_operation_id,
            "name": operation_line,
        }
        text += f"{index}. {operation_line}\n"

    await state.update_data(admin_edit_operation_map=operation_map)
    await state.set_state(AdminEditReport.waiting_for_operation)
    await message.answer(text, reply_markup=choice_keyboard("admin_edit_operation", operation_map))


async def send_admin_edit_action_step(message: Message, state: FSMContext):
    data = await state.get_data()
    action_map = {
        "1": {"id": "quantity", "name": "Изменить количество"},
        "2": {"id": "delete", "name": "Удалить строку"},
    }

    await state.update_data(admin_edit_action_map=action_map)
    await state.set_state(AdminEditReport.waiting_for_action)
    await message.answer(
        f"Строка отчёта:\n{data['admin_edit_operation_name']}\n\nЧто сделать?",
        reply_markup=choice_keyboard("admin_edit_action", action_map),
    )


def format_cutting_batch_status(status: str):
    statuses = {
        "contours_done": "контуры нанесены",
        "layout_done": "настил сформирован",
        "cutting_in_progress": "раскрой в работе",
        "cutting_done": "раскрой 100%",
        "formed": "готовый крой сформирован",
    }
    return statuses.get(status, status)


async def send_admin_cutting_batches_step(message: Message, state: FSMContext):
    batches = get_admin_cutting_batches()

    if not batches:
        await state.clear()
        await message.answer("Партий раскроя пока нет.", reply_markup=admin_reports_keyboard())
        return

    batch_map = {}
    text = "Партии раскроя.\nВыберите тестовую/ошибочную партию для удаления:\n\n"

    for index, batch in enumerate(batches, start=1):
        batch_id, product_name, status, contour_date, layout_date, progress, employee_name, sizes_text, colors_text = batch
        line = (
            f"ID {batch_id} — {product_name}, {format_cutting_batch_status(status)}\n"
            f"   контуры: {contour_date or '-'}, настил: {layout_date or '-'}, готовность: {progress}%"
        )

        if employee_name:
            line += f"\n   сотрудник: {employee_name}"

        if sizes_text:
            line += f"\n   размеры: {sizes_text}"

        if colors_text:
            line += f"\n   цвета: {colors_text}"

        batch_map[str(index)] = {
            "id": batch_id,
            "name": f"ID {batch_id} — {product_name}, {format_cutting_batch_status(status)}",
        }
        text += f"{index}. {line}\n\n"

    await state.update_data(admin_cutting_batch_map=batch_map)
    await state.set_state(AdminCuttingBatch.waiting_for_batch)
    await send_long_text(message, text, reply_markup=choice_keyboard("admin_cutting_batch", batch_map))


async def go_back_in_admin_edit_report(message: Message, state: FSMContext, current_state: str):
    if current_state == AdminEditReport.waiting_for_employee.state:
        await state.clear()
        await message.answer("Раздел отчётов:", reply_markup=admin_reports_keyboard())
        return

    if current_state == AdminEditReport.waiting_for_shift.state:
        await send_admin_edit_employee_step(message, state)
        return

    if current_state == AdminEditReport.waiting_for_operation.state:
        await send_admin_edit_shift_step(message, state)
        return

    if current_state == AdminEditReport.waiting_for_action.state:
        await send_admin_edit_operation_step(message, state)
        return

    if current_state == AdminEditReport.waiting_for_quantity.state:
        await send_admin_edit_action_step(message, state)
        return

    await state.clear()
    await message.answer("Раздел отчётов:", reply_markup=admin_reports_keyboard())


async def reset_state_if_command(message: Message, state: FSMContext):
    current_state = await state.get_state()

    if message.text in {"Отмена", "Отменить"}:
        await state.clear()
        reply_markup = admin_reports_keyboard() if current_state and current_state.startswith("AdminEditReport:") else employee_keyboard()
        await message.answer("Действие отменено.", reply_markup=reply_markup)
        return True

    if message.text == "Админ меню" and is_admin(message.from_user.id):
        await state.clear()
        await message.answer("Админ-панель:", reply_markup=admin_keyboard())
        return True

    if message.text == "Админ-панель" and is_admin(message.from_user.id):
        await state.clear()
        await message.answer("Админ-панель:", reply_markup=admin_keyboard())
        return True

    if message.text == "Отчёт":
        await state.clear()
        await message.answer("Выберите действие с отчётом:", reply_markup=report_keyboard())
        return True

    if message.text == "Назад":
        if current_state and current_state.startswith("Report:"):
            await go_back_in_report(message, state, current_state)
        else:
            await state.clear()
            await message.answer("Основное меню:", reply_markup=employee_keyboard())
        return True

    if message.text and message.text.startswith("/"):
        await state.clear()
        await message.answer(
            "Предыдущее действие отменено. Повторите команду ещё раз."
        )
        return True

    return False


@dp.callback_query(lambda callback: callback.data in {"nav_back", "nav_cancel"})
async def process_navigation_button(callback: CallbackQuery, state: FSMContext):
    current_state = await state.get_state()

    if callback.data == "nav_cancel":
        await state.clear()
        await callback.answer("Отменено")
        reply_markup = admin_reports_keyboard() if current_state and current_state.startswith("AdminEditReport:") else employee_keyboard()
        await callback.message.answer("Действие отменено.", reply_markup=reply_markup)
        return

    if current_state and current_state.startswith("Report:"):
        await callback.answer()
        await go_back_in_report(callback.message, state, current_state)
        return

    if current_state and current_state.startswith("AdminEditReport:"):
        await callback.answer()
        await go_back_in_admin_edit_report(callback.message, state, current_state)
        return

    if current_state and current_state.startswith("AdminCuttingBatch:"):
        await state.clear()
        await callback.answer("Назад")
        await callback.message.answer("Раздел отчётов:", reply_markup=admin_reports_keyboard())
        return

    await state.clear()
    await callback.answer("Назад")
    await callback.message.answer("Основное меню:", reply_markup=employee_keyboard())


@dp.message(Command("admin"))
async def admin_menu(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await message.answer("Админ-панель:", reply_markup=admin_keyboard())


@dp.errors()
async def error_handler(event):
    logging.exception("Bot error", exc_info=event.exception)

    update = getattr(event, "update", None)
    message = getattr(update, "message", None)
    callback_query = getattr(update, "callback_query", None)

    if message is not None:
        await message.answer("Произошла ошибка. Администратор сможет посмотреть лог в разделе «Файлы».")
    elif callback_query is not None:
        await callback_query.answer("Произошла ошибка. Попробуйте ещё раз.", show_alert=True)

    return True


def minutes_to_excel_time(total_minutes):
    if total_minutes is None:
        return ""

    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours}:{minutes:02d}"


def style_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    if sheet.max_row == 0:
        return

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = min(max_length + 3, 45)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def append_total_row(sheet, label: str, value, label_column: int = 1, value_column: int = 2):
    row_number = sheet.max_row + 1
    sheet.cell(row=row_number, column=label_column, value=label)
    sheet.cell(row=row_number, column=value_column, value=value)

    for cell in sheet[row_number]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF6EE")


def append_operation_rows(sheet, operation_rows):
    sheet.append(["Дата", "Сотрудник", "Группа", "Операция", "Размер", "Цвет", "Количество", "Ед."])

    total_quantity = 0

    for row in operation_rows:
        shift_date, full_name, work_group, operation_name, product_size, product_color, quantity, unit = row
        total_quantity += quantity or 0
        sheet.append([
            shift_date,
            full_name,
            work_group,
            operation_name,
            product_size,
            product_color,
            quantity,
            unit,
        ])

    append_total_row(sheet, "Итого количество", total_quantity, label_column=6, value_column=7)


def create_excel_report(start_date: str | None = None, end_date: str | None = None):
    today = local_today()
    month_name = today.strftime("%Y-%m")
    is_period_report = start_date is not None and end_date is not None
    exports_dir = "exports"
    os.makedirs(exports_dir, exist_ok=True)

    if is_period_report:
        file_path = os.path.join(exports_dir, f"report_{start_date}_{end_date}.xlsx")
        employee_summary = get_period_employee_summary(start_date, end_date)
        operation_rows = get_period_operation_rows(start_date, end_date)
        shift_details = get_period_shift_details(start_date, end_date)
    else:
        file_path = os.path.join(exports_dir, f"report_{month_name}.xlsx")
        employee_summary = get_month_employee_summary()
        operation_rows = get_month_operation_rows()
        shift_details = get_month_shift_details()

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    summary_sheet.append(["Сотрудник", "Смен", "Часы"])
    total_shifts = 0
    total_minutes = 0

    for employee in employee_summary:
        _, full_name, shift_count, employee_minutes = employee
        total_shifts += shift_count or 0
        total_minutes += employee_minutes or 0
        summary_sheet.append([full_name, shift_count, minutes_to_excel_time(employee_minutes)])

    append_total_row(summary_sheet, "Итого", total_shifts, label_column=1, value_column=2)
    summary_sheet.cell(row=summary_sheet.max_row, column=3, value=minutes_to_excel_time(total_minutes))

    operations_sheet = workbook.create_sheet("Операции")
    append_operation_rows(operations_sheet, operation_rows)

    group_sheet_names = {
        "Пошив": "Пошив",
        "Упаковка": "Упаковка",
        "Раскрой": "Раскрой",
    }

    for group_name, sheet_name in group_sheet_names.items():
        group_rows = [row for row in operation_rows if row[2] == group_name]

        if not group_rows:
            continue

        group_sheet = workbook.create_sheet(sheet_name)
        append_operation_rows(group_sheet, group_rows)

    days_sheet = workbook.create_sheet("Дни")
    days_sheet.append(["Дата", "Сотрудник", "Начало", "Окончание", "Часы", "Статус"])

    for row in shift_details:
        shift_date, full_name, start_time, end_time, total_minutes, status = row
        days_sheet.append([
            shift_date,
            full_name,
            start_time,
            end_time or "",
            minutes_to_excel_time(total_minutes),
            status,
        ])

    for sheet in workbook.worksheets:
        style_sheet(sheet)

    workbook.save(file_path)
    return file_path


def create_employee_excel_report(employee_id: int, start_date: str, end_date: str):
    exports_dir = "exports"
    os.makedirs(exports_dir, exist_ok=True)

    employee_summary = get_employee_period_summary(employee_id, start_date, end_date)

    if employee_summary is None:
        return None

    _, full_name, position, shift_count, total_minutes = employee_summary
    operations = get_employee_period_operation_totals(employee_id, start_date, end_date)
    safe_name = re.sub(r"[^0-9A-Za-zА-Яа-я_-]+", "_", full_name).strip("_")
    file_path = os.path.join(exports_dir, f"employee_{employee_id}_{safe_name}_{start_date}_{end_date}.xlsx")

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Итог"
    summary_sheet.append(["Показатель", "Значение"])
    summary_sheet.append(["Сотрудник", full_name])
    summary_sheet.append(["Должность", position or ""])
    summary_sheet.append(["Период", f"{start_date} — {end_date}"])
    summary_sheet.append(["Отработано смен", shift_count])
    summary_sheet.append(["Отработано часов", minutes_to_excel_time(total_minutes)])

    operations_sheet = workbook.create_sheet("Операции")
    operations_sheet.append(["Операция", "Количество", "Ед."])
    total_quantity = 0

    for operation_name, quantity, unit in operations:
        total_quantity += quantity or 0
        operations_sheet.append([operation_name, quantity, unit])

    append_total_row(operations_sheet, "Итого количество", total_quantity, label_column=1, value_column=2)

    for sheet in workbook.worksheets:
        style_sheet(sheet)

    workbook.save(file_path)
    return file_path


@dp.message(Command("pending"))
async def pending_employees(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    employees = get_pending_employees()

    if not employees:
        await message.answer("Новых заявок нет.", reply_markup=admin_employees_keyboard())
        return

    text = "Заявки на регистрацию:\n\n"

    for employee in employees:
        employee_id, full_name, position, telegram_id, registered_at = employee
        text += (
            f"ID: {employee_id}\n"
            f"ФИО: {full_name}\n"
            f"Должность: {position}\n"
            f"Telegram ID: {telegram_id}\n"
            f"Дата заявки: {registered_at}\n\n"
            f"Подтвердить: /approve {employee_id}\n"
            f"Отклонить: /reject {employee_id}\n\n"
        )

    await message.answer(text, reply_markup=admin_employees_keyboard())


@dp.message(Command("approve"))
async def approve_employee(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    parts = message.text.split()

    if len(parts) != 2 or not parts[1].isdigit():
        await message.answer("Используйте команду так: /approve 1")
        return

    employee_id = int(parts[1])
    employee = update_employee_status(employee_id, "active")

    if employee is None:
        await message.answer("Сотрудник не найден.")
        return

    _, telegram_id, full_name, _ = employee

    add_edit_log(
        message.from_user.id,
        "admin",
        "Подтвердил сотрудника",
        "employee",
        employee_id,
        full_name,
    )

    await message.answer(
        f"Сотрудник подтверждён: {full_name}",
        reply_markup=admin_employees_keyboard(),
    )

    await bot.send_message(
        telegram_id,
        "Ваша регистрация подтверждена.\n\n"
        "Теперь вы можете открыть смену командой /start.",
        reply_markup=employee_keyboard(),
    )


@dp.message(Command("reject"))
async def reject_employee(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    parts = message.text.split()

    if len(parts) != 2 or not parts[1].isdigit():
        await message.answer("Используйте команду так: /reject 1")
        return

    employee_id = int(parts[1])
    employee = update_employee_status(employee_id, "rejected")

    if employee is None:
        await message.answer("Сотрудник не найден.")
        return

    _, telegram_id, full_name, _ = employee

    add_edit_log(
        message.from_user.id,
        "admin",
        "Отклонил сотрудника",
        "employee",
        employee_id,
        full_name,
    )

    await message.answer(
        f"Заявка отклонена: {full_name}",
        reply_markup=admin_employees_keyboard(),
    )

    await bot.send_message(
        telegram_id,
        "Ваша регистрация отклонена. Обратитесь к администратору."
    )


async def change_employee_status(message: Message, employee_id: int, status: str):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    employee = update_employee_status(employee_id, status)

    if employee is None:
        await message.answer("Сотрудник не найден.")
        return

    _, telegram_id, full_name, new_status = employee
    status_text = "активен" if new_status == "active" else "неактивен"

    add_edit_log(
        message.from_user.id,
        "admin",
        "Изменил статус сотрудника",
        "employee",
        employee_id,
        f"{full_name}: {new_status}",
    )

    await message.answer(
        f"Статус сотрудника изменён.\n\n"
        f"{full_name}: {status_text}",
        reply_markup=admin_employees_keyboard(),
    )

    if new_status == "active":
        await bot.send_message(
            telegram_id,
            "Ваш профиль снова активен.\n\n"
            "Можно открыть смену командой /start.",
            reply_markup=employee_keyboard(),
        )
    elif new_status == "inactive":
        await bot.send_message(
            telegram_id,
            "Ваш профиль временно отключён. Обратитесь к администратору."
        )


async def change_employee_position(message: Message, employee_id: int, position: str):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    employee = update_employee_position(employee_id, position)

    if employee is None:
        await message.answer("Сотрудник не найден.", reply_markup=admin_employees_keyboard())
        return

    _, telegram_id, full_name, new_position, status = employee

    add_edit_log(
        message.from_user.id,
        "admin",
        "Изменил должность сотрудника",
        "employee",
        employee_id,
        f"{full_name}: {new_position}",
    )

    await message.answer(
        "Должность сотрудника изменена.\n\n"
        f"{full_name}: {new_position}",
        reply_markup=admin_employees_keyboard(),
    )

    if status == "active":
        await bot.send_message(
            telegram_id,
            "Ваша должность изменена.\n\n"
            f"Новая должность: {new_position}",
            reply_markup=employee_keyboard(),
        )


@dp.message(Command("set_active"))
async def set_employee_active(message: Message):
    parts = message.text.split()

    if len(parts) != 2 or not parts[1].isdigit():
        await message.answer("Используйте команду так: /set_active 1")
        return

    await change_employee_status(message, int(parts[1]), "active")


@dp.message(Command("set_inactive"))
async def set_employee_inactive(message: Message):
    parts = message.text.split()

    if len(parts) != 2 or not parts[1].isdigit():
        await message.answer("Используйте команду так: /set_inactive 1")
        return

    await change_employee_status(message, int(parts[1]), "inactive")


@dp.message(Command("set_position"))
async def set_employee_position(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    parts = message.text.split(maxsplit=2)

    if len(parts) != 3 or not parts[1].isdigit():
        await message.answer(
            "Используйте команду так:\n"
            "/set_position 1 Швея\n\n"
            f"{position_list_text()}"
        )
        return

    position = parts[2].strip()

    if position not in POSITIONS.values():
        await message.answer(
            "Такой должности нет. Выберите одну из списка:\n\n"
            f"{position_list_text()}"
        )
        return

    await change_employee_position(message, int(parts[1]), position)


@dp.message(lambda message: message.text == "Активировать сотрудника")
async def activate_employee_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await state.set_state(EmployeeAdmin.waiting_for_active_id)
    await message.answer("Введите ID сотрудника, которого нужно активировать.")


@dp.message(lambda message: message.text == "Отключить сотрудника")
async def deactivate_employee_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await state.set_state(EmployeeAdmin.waiting_for_inactive_id)
    await message.answer("Введите ID сотрудника, которого нужно отключить.")


@dp.message(lambda message: message.text == "Сменить должность")
async def change_employee_position_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    employees = get_all_employees()

    if not employees:
        await message.answer("Сотрудников пока нет.", reply_markup=admin_employees_keyboard())
        return

    text = "Введите ID сотрудника, которому нужно сменить должность:\n\n"

    for employee in employees:
        employee_id, full_name, position, _, status = employee
        text += f"{employee_id}. {full_name} — {position}, {format_employee_status(status)}\n"

    await state.set_state(EmployeeAdmin.waiting_for_position_employee_id)
    await message.answer(text)


@dp.message(EmployeeAdmin.waiting_for_active_id)
async def process_active_employee_id(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit():
        await message.answer("Введите ID числом, например: 1")
        return

    await state.clear()
    await change_employee_status(message, int(message.text), "active")


@dp.message(EmployeeAdmin.waiting_for_inactive_id)
async def process_inactive_employee_id(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit():
        await message.answer("Введите ID числом, например: 1")
        return

    await state.clear()
    await change_employee_status(message, int(message.text), "inactive")


@dp.message(EmployeeAdmin.waiting_for_position_employee_id)
async def process_position_employee_id(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit():
        await message.answer("Введите ID числом, например: 1")
        return

    await state.update_data(employee_id=int(message.text))
    await state.set_state(EmployeeAdmin.waiting_for_new_position)
    await message.answer(
        "Выберите новую должность. Напишите номер:\n\n"
        f"{position_list_text()}"
    )


@dp.message(EmployeeAdmin.waiting_for_new_position)
async def process_new_employee_position(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    selected_position = message.text.strip()

    if selected_position not in POSITIONS:
        await message.answer("Введите номер должности: 1, 2 или 3.")
        return

    data = await state.get_data()
    await state.clear()
    await change_employee_position(
        message,
        data["employee_id"],
        POSITIONS[selected_position],
    )


@dp.message(Command("open_shifts"))
async def open_shifts(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    shifts = get_open_shifts()

    if not shifts:
        await message.answer("Открытых смен нет.", reply_markup=admin_shifts_keyboard())
        return

    text = "Открытые смены:\n\n"

    for shift in shifts:
        shift_id, full_name, shift_date, start_time = shift
        text += (
            f"ID смены: {shift_id}\n"
            f"Сотрудник: {full_name}\n"
            f"Дата: {shift_date}\n"
            f"Начало: {start_time}\n"
            f"Закрыть: /admin_close {shift_id} 17:00\n\n"
        )

    await message.answer(text, reply_markup=admin_shifts_keyboard())


@dp.message(Command("shift_list"))
async def shift_list(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    shifts = get_recent_shifts()

    if not shifts:
        await message.answer("Смен пока нет.", reply_markup=admin_shifts_keyboard())
        return

    text = "Последние смены:\n\n"

    for shift in shifts:
        shift_id, full_name, shift_date, start_time, end_time, status, operation_count = shift
        status_text = "открыта" if status == "open" else "закрыта"
        end_text = end_time if end_time else "смена открыта"
        text += (
            f"ID смены: {shift_id}\n"
            f"Сотрудник: {full_name}\n"
            f"Дата: {shift_date}\n"
            f"Время: {start_time}–{end_text}\n"
            f"Статус: {status_text}\n"
            f"Операций: {operation_count}\n"
            f"Удалить: /delete_shift {shift_id}\n\n"
        )

    await message.answer(text, reply_markup=admin_shifts_keyboard())


@dp.message(Command("delete_shift"))
async def delete_shift(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    parts = message.text.split()

    if len(parts) != 2 or not parts[1].isdigit():
        await message.answer("Используйте команду так: /delete_shift 12")
        return

    shift_id = int(parts[1])
    deleted_shift = delete_shift_by_id(shift_id)

    if deleted_shift is None:
        await message.answer("Смена с таким ID не найдена.", reply_markup=admin_shifts_keyboard())
        return

    _, full_name, shift_date, start_time, end_time, status = deleted_shift
    status_text = "открыта" if status == "open" else "закрыта"
    end_text = end_time if end_time else "смена открыта"

    add_edit_log(
        message.from_user.id,
        "admin",
        "Удалил смену",
        "shift",
        shift_id,
        f"{full_name}, {shift_date}, {start_time}–{end_text}, {status_text}",
    )

    await message.answer(
        "Смена удалена.\n\n"
        f"ID смены: {shift_id}\n"
        f"Сотрудник: {full_name}\n"
        f"Дата: {shift_date}\n"
        f"Время: {start_time}–{end_text}\n"
        f"Статус: {status_text}",
        reply_markup=admin_shifts_keyboard(),
    )


@dp.message(Command("admin_close"))
async def admin_close(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    parts = message.text.split()

    if len(parts) != 3 or not parts[1].isdigit() or not is_valid_time(parts[2]):
        await message.answer("Используйте команду так: /admin_close 3 17:00")
        return

    shift_id = int(parts[1])
    end_time = parts[2]

    result = admin_close_shift(shift_id, end_time)

    if result is None:
        await message.answer("Открытая смена с таким ID не найдена.")
        return

    if result == "bad_time":
        await message.answer("Время окончания не может быть раньше начала смены.")
        return

    add_edit_log(
        message.from_user.id,
        "admin",
        "Закрыл смену",
        "shift",
        shift_id,
        f"Окончание: {result['end_time']}, отработано: {format_minutes(result['total_minutes'])}",
    )

    await message.answer(
        "Смена закрыта администратором.\n\n"
        f"ID смены: {shift_id}\n"
        f"Окончание: {result['end_time']}\n"
        f"Отработано: {format_minutes(result['total_minutes'])}",
        reply_markup=admin_shifts_keyboard(),
    )


@dp.message(Command("today"))
async def today_report(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    shifts = get_today_shifts()

    if not shifts:
        await message.answer("Сегодня смен ещё нет.", reply_markup=admin_reports_keyboard())
        return

    text = "Отчёт за сегодня:\n\n"

    for shift in shifts:
        shift_id, full_name, shift_date, start_time, end_time, total_minutes, status = shift
        report = get_shift_report(shift_id)

        if status == "open":
            end_text = "смена открыта"
            worked_text = "ещё не рассчитано"
        else:
            end_text = end_time
            worked_text = format_minutes(total_minutes)

        text += (
            f"{full_name}\n"
            f"Смена: {start_time}–{end_text}\n"
            f"Отработано: {worked_text}\n"
        )

        if not report:
            text += "Операции: не добавлены\n\n"
        else:
            text += "Операции:\n"
            for row in report:
                name, product_size, product_color, qty, unit = row
                text += f"- {format_operation_line(name, product_size, product_color, qty, unit)}\n"
            text += "\n"

    await send_long_text(message, text, reply_markup=admin_reports_keyboard())


@dp.message(Command("month"))
async def month_report(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    employees = get_month_employee_summary()

    if not employees:
        await message.answer("За текущий месяц закрытых смен ещё нет.", reply_markup=admin_reports_keyboard())
        return

    text = "Отчёт за месяц:\n\n"

    for employee in employees:
        employee_id, full_name, shift_count, total_minutes = employee
        operations = get_month_operations_by_employee(employee_id)

        text += (
            f"{full_name}\n"
            f"Смен: {shift_count}\n"
            f"Часов: {format_minutes(total_minutes)}\n"
        )

        if not operations:
            text += "Операции: не добавлены\n\n"
        else:
            text += "Операции:\n"
            for operation in operations:
                name, product_size, product_color, quantity, unit = operation
                text += f"- {format_operation_line(name, product_size, product_color, quantity, unit)}\n"
            text += "\n"

    await send_long_text(message, text, reply_markup=admin_reports_keyboard())


async def send_period_report(message: Message, start_date: str, end_date: str):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    shifts = get_period_shift_details(start_date, end_date)

    if not shifts:
        await message.answer(
            f"За период {start_date} — {end_date} смен нет.",
            reply_markup=admin_reports_keyboard(),
        )
        return

    employees = get_period_employee_summary(start_date, end_date)
    text = f"Отчёт за период:\n{start_date} — {end_date}\n\n"

    if not employees:
        text += "Закрытых смен за период нет.\n\n"
    else:
        for employee in employees:
            employee_id, full_name, shift_count, total_minutes = employee
            operations = get_period_operations_by_employee(employee_id, start_date, end_date)

            text += (
                f"{full_name}\n"
                f"Смен: {shift_count}\n"
                f"Часов: {format_minutes(total_minutes)}\n"
            )

            if not operations:
                text += "Операции: не добавлены\n\n"
            else:
                text += "Операции:\n"
                for operation in operations:
                    name, product_size, product_color, quantity, unit = operation
                    text += f"- {format_operation_line(name, product_size, product_color, quantity, unit)}\n"
                text += "\n"

    text += "Смены:\n"

    for shift in shifts:
        shift_date, full_name, start_time, end_time, total_minutes, status = shift
        end_text = end_time if end_time else "смена открыта"
        worked_text = format_minutes(total_minutes) if total_minutes else "ещё не рассчитано"
        status_text = "закрыта" if status == "closed" else "открыта"

        text += (
            f"- {shift_date}: {full_name}, "
            f"{start_time}–{end_text}, {worked_text}, {status_text}\n"
        )

    await send_long_text(message, text, reply_markup=admin_reports_keyboard())


@dp.message(Command("period"))
async def period_report(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    period = parse_period(message.text.replace("/period", "", 1).strip())

    if period is None:
        await message.answer("Используйте команду так: /period 2026-06-01 2026-06-25")
        return

    await send_period_report(message, period[0], period[1])


@dp.message(Command("excel"))
async def excel_report(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    file_path = create_excel_report()
    document = FSInputFile(file_path)

    await message.answer_document(
        document,
        caption="Excel-отчёт за текущий месяц."
    )
    await message.answer("Готово.", reply_markup=admin_reports_keyboard())


async def send_period_excel(message: Message, start_date: str, end_date: str):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    shifts = get_period_shift_details(start_date, end_date)

    if not shifts:
        await message.answer(
            f"За период {start_date} — {end_date} смен нет.",
            reply_markup=admin_reports_keyboard(),
        )
        return

    file_path = create_excel_report(start_date, end_date)
    document = FSInputFile(file_path)

    await message.answer_document(
        document,
        caption=f"Excel-отчёт за период {start_date} — {end_date}."
    )
    await message.answer("Готово.", reply_markup=admin_reports_keyboard())


async def send_employee_excel_report(message: Message, employee_id: int, start_date: str, end_date: str):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    file_path = create_employee_excel_report(employee_id, start_date, end_date)

    if file_path is None:
        await message.answer("Сотрудник с таким ID не найден.", reply_markup=admin_reports_keyboard())
        return

    document = FSInputFile(file_path)

    await message.answer_document(
        document,
        caption=f"Отчёт по сотруднику за период {start_date} — {end_date}."
    )
    await message.answer("Готово.", reply_markup=admin_reports_keyboard())


@dp.message(Command("excel_period"))
async def excel_period_report(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    period = parse_period(message.text.replace("/excel_period", "", 1).strip())

    if period is None:
        await message.answer("Используйте команду так: /excel_period 2026-06-01 2026-06-25")
        return

    await send_period_excel(message, period[0], period[1])


@dp.message(Command("employee_report"))
async def employee_report_command(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    parts = message.text.split()

    if len(parts) != 4 or not parts[1].isdigit():
        await message.answer(
            "Используйте команду так:\n"
            "/employee_report 1 2026-06-01 2026-06-25"
        )
        return

    employee_id = int(parts[1])
    start_date = parts[2]
    end_date = parts[3]

    if not is_valid_date(start_date) or not is_valid_date(end_date) or start_date > end_date:
        await message.answer("Введите даты в формате: 2026-06-01 2026-06-25")
        return

    await send_employee_excel_report(message, employee_id, start_date, end_date)


@dp.message(Command("logs"))
async def logs_report(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    logs = get_recent_edit_logs()

    if not logs:
        await message.answer("Журнал изменений пока пуст.", reply_markup=admin_files_keyboard())
        return

    text = "Последние изменения:\n\n"

    for log in logs:
        changed_at, changed_by, role, action, entity_type, entity_id, details = log
        text += (
            f"{changed_at}\n"
            f"Кто: {changed_by} ({role})\n"
            f"Действие: {action}\n"
            f"Объект: {entity_type} #{entity_id}\n"
            f"Детали: {details}\n\n"
        )

    await message.answer(text, reply_markup=admin_files_keyboard())


@dp.message(Command("errors"))
async def errors_report(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    if not os.path.exists(ERROR_LOG_FILE) or os.path.getsize(ERROR_LOG_FILE) == 0:
        await message.answer("Лог ошибок пока пуст.", reply_markup=admin_files_keyboard())
        return

    document = FSInputFile(ERROR_LOG_FILE)

    await message.answer_document(
        document,
        caption="Лог ошибок бота."
    )
    await message.answer("Готово.", reply_markup=admin_files_keyboard())


@dp.message(Command("operations"))
async def operations_report(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    operations = get_all_operations()

    if not operations:
        await message.answer("Операций пока нет.", reply_markup=admin_operations_keyboard())
        return

    text = "Список операций:\n\n"

    for operation in operations:
        _, number, name, position, operation_group, folder, unit, is_active = operation
        status = "активна" if is_active else "скрыта"
        group_text = operation_group if operation_group else "без группы"
        folder_text = folder if folder else "без папки"
        text += f"{number}. {group_text} / {folder_text}: {name} ({unit}) — {position}, {status}\n"

    text += "\nДобавить: /add_operation\n"
    text += "Скрыть: /hide_operation"

    await send_long_text(message, text, reply_markup=admin_operations_keyboard())


@dp.message(Command("add_operation"))
async def add_operation_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await state.set_state(OperationAdmin.waiting_for_name)
    await message.answer("Введите название новой операции:")


@dp.message(OperationAdmin.waiting_for_name)
async def add_operation_name(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    name = message.text.strip()

    if len(name) < 3:
        await message.answer("Название операции слишком короткое.")
        return

    await state.update_data(operation_name=name)
    await state.set_state(OperationAdmin.waiting_for_position)

    await message.answer(
        "Выберите раздел операции. Напишите номер:\n\n"
        f"{position_list_text()}"
    )


@dp.message(OperationAdmin.waiting_for_position)
async def add_operation_position(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    selected_position = message.text.strip()

    if selected_position not in POSITIONS:
        await message.answer("Введите номер раздела: 1, 2 или 3.")
        return

    position = POSITIONS[selected_position]

    await state.update_data(operation_position=position)
    await state.set_state(OperationAdmin.waiting_for_group)

    await message.answer(
        "Введите группу операции.\n\n"
        "Например: Брюки / низ"
    )


@dp.message(OperationAdmin.waiting_for_group)
async def add_operation_group(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    operation_group = message.text.strip()

    if len(operation_group) < 3:
        await message.answer("Название группы слишком короткое.")
        return

    await state.update_data(operation_group=operation_group)
    await state.set_state(OperationAdmin.waiting_for_folder)

    await message.answer(
        "Введите изделие или папку для операции.\n\n"
        "Например: Брюки со стрелками детские"
    )


@dp.message(OperationAdmin.waiting_for_folder)
async def add_operation_folder(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    folder = message.text.strip()

    if len(folder) < 3:
        await message.answer("Название папки слишком короткое.")
        return

    data = await state.get_data()
    name = data["operation_name"]
    position = data["operation_position"]
    operation_group = data["operation_group"]
    operation = add_operation(name, position, operation_group, folder)

    add_edit_log(
        message.from_user.id,
        "admin",
        "Добавил операцию",
        "operation",
        operation["id"],
        f"{operation['number']}. {operation['name']} — {operation['position']}",
    )

    await state.clear()
    await message.answer(
        f"Операция добавлена:\n{operation['number']}. {operation['name']}\n"
        f"Раздел: {operation['position']}\n"
        f"Группа: {operation['operation_group']}\n"
        f"Папка: {operation['folder']}",
        reply_markup=admin_operations_keyboard(),
    )


OPERATION_EDIT_FIELDS = {
    "1": ("name", "название"),
    "2": ("position", "должность"),
    "3": ("operation_group", "группа"),
    "4": ("folder", "изделие/папка"),
}


def operation_edit_fields_text():
    return (
        "1. Название операции\n"
        "2. Должность\n"
        "3. Группа\n"
        "4. Изделие/папка"
    )


@dp.message(Command("edit_operation"))
async def edit_operation_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    operations = get_all_operations()

    if not operations:
        await message.answer("Операций пока нет.", reply_markup=admin_operations_keyboard())
        return

    text = "Какую операцию изменить? Напишите номер:\n\n"

    for operation in operations:
        _, number, name, position, operation_group, folder, unit, is_active = operation
        status = "активна" if is_active else "скрыта"
        group_text = operation_group if operation_group else "без группы"
        folder_text = folder if folder else "без папки"
        text += f"{number}. {group_text} / {folder_text}: {name} ({unit}) — {position}, {status}\n"

    await state.set_state(OperationAdmin.waiting_for_edit_number)
    await send_long_text(message, text)


@dp.message(OperationAdmin.waiting_for_edit_number)
async def edit_operation_number(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit():
        await message.answer("Введите номер операции числом.")
        return

    operation = get_operation_by_number(int(message.text))

    if operation is None:
        await message.answer("Операция с таким номером не найдена или она скрыта.")
        return

    operation_id, number, name, unit = operation

    await state.update_data(operation_number=number, operation_id=operation_id, operation_name=name)
    await state.set_state(OperationAdmin.waiting_for_edit_field)
    await message.answer(
        f"Операция: {number}. {name} ({unit})\n\n"
        "Что изменить? Напишите номер:\n\n"
        f"{operation_edit_fields_text()}"
    )


@dp.message(OperationAdmin.waiting_for_edit_field)
async def edit_operation_field(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    selected_field = message.text.strip()

    if selected_field not in OPERATION_EDIT_FIELDS:
        await message.answer(
            "Введите номер поля из списка:\n\n"
            f"{operation_edit_fields_text()}"
        )
        return

    field, label = OPERATION_EDIT_FIELDS[selected_field]

    await state.update_data(operation_field=field, operation_field_label=label)
    await state.set_state(OperationAdmin.waiting_for_edit_value)

    if field == "position":
        await message.answer(
            "Выберите новую должность операции. Напишите номер:\n\n"
            f"{position_list_text()}"
        )
        return

    await message.answer(f"Введите новое значение для поля «{label}»:")


@dp.message(OperationAdmin.waiting_for_edit_value)
async def edit_operation_value(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    data = await state.get_data()
    field = data["operation_field"]
    value = message.text.strip()

    if field == "position":
        if value not in POSITIONS:
            await message.answer("Введите номер должности: 1, 2 или 3.")
            return

        value = POSITIONS[value]
    elif len(value) < 3:
        await message.answer("Значение слишком короткое.")
        return

    result = update_operation_field(data["operation_number"], field, value)

    if result is None:
        await message.answer("Не удалось изменить операцию.")
        await state.clear()
        return

    old_id, old_number, old_name, old_position, old_group, old_folder = result["old"]
    _, new_number, new_name, new_position, new_group, new_folder = result["new"]

    add_edit_log(
        message.from_user.id,
        "admin",
        "Изменил операцию",
        "operation",
        old_id,
        (
            f"{old_number}. {old_group}: {old_folder}: {old_name} — {old_position} -> "
            f"{new_number}. {new_group}: {new_folder}: {new_name} — {new_position}"
        ),
    )

    await state.clear()
    await message.answer(
        "Операция изменена.\n\n"
        f"Было: {old_number}. {old_group}: {old_folder}: {old_name} — {old_position}\n"
        f"Стало: {new_number}. {new_group}: {new_folder}: {new_name} — {new_position}",
        reply_markup=admin_operations_keyboard(),
    )


@dp.message(Command("hide_operation"))
async def hide_operation_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    operations = get_all_operations()

    if not operations:
        await message.answer("Операций пока нет.", reply_markup=admin_operations_keyboard())
        return

    text = "Какую операцию скрыть? Напишите номер:\n\n"

    for operation in operations:
        _, number, name, position, operation_group, folder, unit, is_active = operation
        status = "активна" if is_active else "скрыта"
        group_text = operation_group if operation_group else "без группы"
        folder_text = folder if folder else "без папки"
        text += f"{number}. {group_text} / {folder_text}: {name} ({unit}) — {position}, {status}\n"

    await state.set_state(OperationAdmin.waiting_for_hide_number)
    await send_long_text(message, text)


@dp.message(OperationAdmin.waiting_for_hide_number)
async def hide_operation_number(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit():
        await message.answer("Введите номер операции числом.")
        return

    operation = hide_operation(int(message.text))

    if operation is None:
        await message.answer("Операция с таким номером не найдена.")
        return

    operation_id, number, name, position, operation_group, folder = operation

    add_edit_log(
        message.from_user.id,
        "admin",
        "Скрыл операцию",
        "operation",
        operation_id,
        f"{number}. {operation_group}: {folder}: {name} — {position}",
    )

    await state.clear()
    await message.answer(
        f"Операция скрыта:\n{number}. {operation_group}: {folder}: {name}",
        reply_markup=admin_operations_keyboard(),
    )


@dp.message(Command("restore_operation"))
async def restore_operation_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    operations = get_all_operations()
    hidden_operations = [operation for operation in operations if not operation[7]]

    if not hidden_operations:
        await message.answer("Скрытых операций нет.", reply_markup=admin_operations_keyboard())
        return

    text = "Какую операцию вернуть? Напишите номер:\n\n"

    for operation in hidden_operations:
        _, number, name, position, operation_group, folder, unit, _ = operation
        group_text = operation_group if operation_group else "без группы"
        folder_text = folder if folder else "без папки"
        text += f"{number}. {group_text} / {folder_text}: {name} ({unit}) — {position}, скрыта\n"

    await state.set_state(OperationAdmin.waiting_for_restore_number)
    await send_long_text(message, text)


@dp.message(OperationAdmin.waiting_for_restore_number)
async def restore_operation_number(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit():
        await message.answer("Введите номер операции числом.")
        return

    operation = restore_operation(int(message.text))

    if operation is None:
        await message.answer("Операция с таким номером не найдена.")
        return

    operation_id, number, name, position, operation_group, folder = operation

    add_edit_log(
        message.from_user.id,
        "admin",
        "Вернул операцию",
        "operation",
        operation_id,
        f"{number}. {operation_group}: {folder}: {name} — {position}",
    )

    await state.clear()
    await message.answer(
        f"Операция снова активна:\n{number}. {operation_group}: {folder}: {name}",
        reply_markup=admin_operations_keyboard(),
    )


@dp.message(Command("employees"))
async def employees_report(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await send_employees_list(message, "Сотрудники", get_all_employees())


async def send_employees_list(message: Message, title: str, employees):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    if not employees:
        await message.answer(f"{title}: список пуст.", reply_markup=admin_employees_keyboard())
        return

    text = f"{title}:\n\n"

    for employee in employees:
        employee_id, full_name, position, telegram_id, status = employee
        text += (
            f"{employee_id}. {full_name}\n"
            f"Должность: {position}\n"
            f"Telegram ID: {telegram_id}\n"
            f"Статус: {format_employee_status(status)}\n\n"
            f"Активировать: /set_active {employee_id}\n"
            f"Отключить: /set_inactive {employee_id}\n\n"
            f"Сменить должность: /set_position {employee_id} Швея\n\n"
        )

    await send_long_text(message, text, reply_markup=admin_employees_keyboard())


@dp.message(Command("active_employees"))
async def active_employees_report(message: Message):
    await send_employees_list(message, "Активные сотрудники", get_employees_by_status("active"))


@dp.message(Command("inactive_employees"))
async def inactive_employees_report(message: Message):
    await send_employees_list(message, "Неактивные сотрудники", get_employees_by_status("inactive"))


@dp.message(Command("backup"))
async def backup_database(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    document = FSInputFile(DB_NAME)

    await message.answer_document(
        document,
        caption="Текущая база данных."
    )
    await message.answer("База отправлена.", reply_markup=admin_files_keyboard())


@dp.message(Command("db_status"))
async def database_status(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    status = get_database_status()
    text = (
        "Проверка базы:\n\n"
        f"Путь: {status['path']}\n"
        f"Файл найден: {'да' if status['exists'] else 'нет'}\n"
        f"Размер: {status['size']} байт\n"
        f"Сотрудники: {status['employees']}\n"
        f"Смены: {status['shifts']}\n"
        f"Операции в отчётах: {status['shift_operations']}\n"
        f"Справочник операций: {status['operations']}"
    )

    if "error" in status:
        text += f"\nОшибка: {status['error']}"

    await message.answer(text, reply_markup=admin_files_keyboard())


def read_database_counts(db_path: str):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    counts = {}

    for table in ["employees", "shifts", "shift_operations", "operations"]:
        cursor.execute(f"SELECT COUNT(*) FROM {table}")
        counts[table] = cursor.fetchone()[0]

    conn.close()
    return counts


@dp.message(Command("restore_db"))
async def restore_database_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await state.set_state(DatabaseAdmin.waiting_for_database_file)
    await message.answer(
        "Отправьте файл рабочей базы данных bot.db сюда в Telegram.\n\n"
        "Бот проверит файл, сделает копию текущей базы и заменит активную базу.",
        reply_markup=admin_files_keyboard(),
    )


@dp.message(DatabaseAdmin.waiting_for_database_file)
async def restore_database_file(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    if not message.document:
        await message.answer("Нужно отправить файл bot.db как документ.")
        return

    file_name = message.document.file_name or ""

    if not file_name.endswith(".db"):
        await message.answer("Нужен файл базы с расширением .db, например bot.db.")
        return

    db_dir = os.path.dirname(DB_NAME)

    if db_dir:
        os.makedirs(db_dir, exist_ok=True)

    upload_path = f"{DB_NAME}.upload"

    await bot.download(message.document, destination=upload_path)

    try:
        counts = read_database_counts(upload_path)
    except sqlite3.Error as error:
        if os.path.exists(upload_path):
            os.remove(upload_path)

        await message.answer(f"Это не подходит как база данных.\nОшибка: {error}")
        return

    if counts["employees"] == 0:
        os.remove(upload_path)
        await message.answer(
            "В этой базе нет сотрудников. Похоже, это пустой файл.\n"
            "Выберите рабочий bot.db из папки sewing_shift_bot."
        )
        return

    backup_path = None

    if os.path.exists(DB_NAME):
        backups_dir = "backups"
        os.makedirs(backups_dir, exist_ok=True)
        timestamp = local_now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_path = os.path.join(backups_dir, f"before_restore_{timestamp}_{os.path.basename(DB_NAME)}")
        shutil.copy2(DB_NAME, backup_path)

    shutil.move(upload_path, DB_NAME)
    await state.clear()

    text = (
        "База заменена.\n\n"
        f"Путь: {DB_NAME}\n"
        f"Сотрудники: {counts['employees']}\n"
        f"Смены: {counts['shifts']}\n"
        f"Операции в отчётах: {counts['shift_operations']}\n"
        f"Справочник операций: {counts['operations']}"
    )

    if backup_path:
        text += f"\n\nСтарая база сохранена:\n{backup_path}"

    await message.answer(text, reply_markup=admin_files_keyboard())


@dp.message(Command("create_backup"))
async def create_backup_command(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    backup_path = create_database_backup("manual")
    document = FSInputFile(backup_path)

    add_edit_log(
        message.from_user.id,
        "admin",
        "Создал резервную копию базы",
        "backup",
        None,
        backup_path,
    )

    await message.answer_document(
        document,
        caption=f"Резервная копия создана:\n{backup_path}"
    )
    await message.answer("Готово.", reply_markup=admin_files_keyboard())


@dp.message(lambda message: message.text == "Заявки")
async def pending_button(message: Message):
    await pending_employees(message)


@dp.message(lambda message: message.text == "Отчёты")
async def admin_reports_button(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await message.answer("Раздел отчётов:", reply_markup=admin_reports_keyboard())


@dp.message(lambda message: message.text == "Смены")
async def admin_shifts_button(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await message.answer("Раздел смен:", reply_markup=admin_shifts_keyboard())


@dp.message(lambda message: message.text == "Сотрудники")
async def admin_employees_section_button(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await message.answer("Раздел сотрудников:", reply_markup=admin_employees_keyboard())


@dp.message(lambda message: message.text == "Операции")
async def admin_operations_section_button(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await message.answer("Раздел операций:", reply_markup=admin_operations_keyboard())


@dp.message(lambda message: message.text == "Файлы")
async def admin_files_button(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await message.answer("Файлы и журнал:", reply_markup=admin_files_keyboard())


@dp.message(lambda message: message.text == "Админ меню")
async def admin_back_button(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await message.answer("Админ-панель:", reply_markup=admin_keyboard())


@dp.message(lambda message: message.text == "В меню сотрудника")
async def employee_menu_from_admin_button(message: Message):
    await message.answer("Основное меню:", reply_markup=employee_keyboard())


@dp.message(lambda message: message.text == "Открытые смены")
async def open_shifts_button(message: Message):
    await open_shifts(message)


@dp.message(lambda message: message.text == "Последние смены")
async def shift_list_button(message: Message):
    await shift_list(message)


@dp.message(lambda message: message.text == "Удалить смену")
async def delete_shift_button(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await message.answer(
        "Чтобы удалить ошибочную смену, сначала посмотрите её ID через «Последние смены».\n\n"
        "Потом отправьте команду:\n"
        "/delete_shift ID\n\n"
        "Например: /delete_shift 12",
        reply_markup=admin_shifts_keyboard(),
    )


@dp.message(lambda message: message.text == "Отчёт за сегодня")
async def today_button(message: Message):
    await today_report(message)


@dp.message(lambda message: message.text == "Отчёт за месяц")
async def month_button(message: Message):
    await month_report(message)


@dp.message(lambda message: message.text == "Отчёт за период")
async def period_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await state.set_state(AdminPeriod.waiting_for_report_period)
    await message.answer(
        "Введите даты отчёта через пробел.\n\n"
        "Пример: 2026-06-01 2026-06-25"
    )


@dp.message(lambda message: message.text == "Excel за период")
async def excel_period_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await state.set_state(AdminPeriod.waiting_for_excel_period)
    await message.answer(
        "Введите даты для Excel через пробел.\n\n"
        "Пример: 2026-06-01 2026-06-25"
    )


@dp.message(lambda message: message.text == "Отчёт по сотруднику")
async def employee_report_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    employees = get_all_employees()

    if not employees:
        await message.answer("Сотрудников пока нет.", reply_markup=admin_reports_keyboard())
        return

    text = "Выберите сотрудника. Напишите ID:\n\n"

    for employee in employees:
        employee_id, full_name, position, _, status = employee
        text += f"{employee_id}. {full_name} — {position}, {status}\n"

    await state.set_state(AdminPeriod.waiting_for_employee_report_id)
    await message.answer(text)


@dp.message(lambda message: message.text == "Править отчёт")
async def admin_edit_report_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await send_admin_edit_employee_step(message, state)


@dp.message(lambda message: message.text == "Партии раскроя")
async def admin_cutting_batches_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("У вас нет прав администратора.")
        return

    await send_admin_cutting_batches_step(message, state)


async def select_admin_cutting_batch(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    batch_map = data.get("admin_cutting_batch_map", {})

    if selected_number not in batch_map:
        await message.answer("Выберите номер партии из списка.")
        return

    batch = batch_map[selected_number]
    result = delete_cutting_batch(batch["id"])

    if result is None:
        await state.clear()
        await message.answer("Не удалось удалить партию. Возможно, она уже удалена.", reply_markup=admin_reports_keyboard())
        return

    batch_id, product_name, status, contour_date, employee_name = result

    add_edit_log(
        message.from_user.id,
        "admin",
        "Удалил партию раскроя",
        "cutting_batch",
        batch_id,
        f"{product_name}, статус {status}, контуры {contour_date}, сотрудник {employee_name or '-'}",
    )

    await state.clear()
    await message.answer(
        "Партия раскроя удалена.\n\n"
        f"ID: {batch_id}\n"
        f"Изделие: {product_name}\n"
        f"Статус: {format_cutting_batch_status(status)}\n"
        f"Дата контуров: {contour_date or '-'}\n\n"
        "Она больше не будет появляться в выборе следующих операций.",
        reply_markup=admin_reports_keyboard(),
    )


@dp.message(AdminCuttingBatch.waiting_for_batch)
async def process_admin_cutting_batch(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_admin_cutting_batch(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("admin_cutting_batch:"))
async def process_admin_cutting_batch_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, AdminCuttingBatch.waiting_for_batch):
        return

    await callback.answer()
    await select_admin_cutting_batch(callback.message, state, callback.data.rsplit(":", 1)[1])


async def select_admin_edit_employee(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    employee_map = data.get("admin_edit_employee_map", {})

    if selected_number not in employee_map:
        await message.answer("Введите ID сотрудника из списка.")
        return

    employee = employee_map[selected_number]
    await state.update_data(
        admin_edit_employee_id=employee["id"],
        admin_edit_employee_name=employee["name"],
    )
    await send_admin_edit_shift_step(message, state)


@dp.message(AdminEditReport.waiting_for_employee)
async def process_admin_edit_employee(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_admin_edit_employee(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("admin_edit_employee:"))
async def process_admin_edit_employee_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, AdminEditReport.waiting_for_employee):
        return

    await callback.answer()
    await select_admin_edit_employee(callback.message, state, callback.data.rsplit(":", 1)[1])


async def select_admin_edit_shift(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    shift_map = data.get("admin_edit_shift_map", {})

    if selected_number not in shift_map:
        await message.answer("Введите номер смены из списка.")
        return

    shift = shift_map[selected_number]
    await state.update_data(
        admin_edit_shift_id=shift["id"],
        admin_edit_shift_name=shift["name"],
    )
    await send_admin_edit_operation_step(message, state)


@dp.message(AdminEditReport.waiting_for_shift)
async def process_admin_edit_shift(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_admin_edit_shift(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("admin_edit_shift:"))
async def process_admin_edit_shift_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, AdminEditReport.waiting_for_shift):
        return

    await callback.answer()
    await select_admin_edit_shift(callback.message, state, callback.data.rsplit(":", 1)[1])


async def select_admin_edit_operation(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    operation_map = data.get("admin_edit_operation_map", {})

    if selected_number not in operation_map:
        await message.answer("Введите номер строки из списка.")
        return

    operation = operation_map[selected_number]
    await state.update_data(
        admin_edit_shift_operation_id=operation["id"],
        admin_edit_operation_name=operation["name"],
    )
    await send_admin_edit_action_step(message, state)


@dp.message(AdminEditReport.waiting_for_operation)
async def process_admin_edit_operation(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_admin_edit_operation(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("admin_edit_operation:"))
async def process_admin_edit_operation_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, AdminEditReport.waiting_for_operation):
        return

    await callback.answer()
    await select_admin_edit_operation(callback.message, state, callback.data.rsplit(":", 1)[1])


async def select_admin_edit_action(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    action_map = data.get("admin_edit_action_map", {})

    if selected_number not in action_map:
        await message.answer("Выберите действие из списка.")
        return

    action = action_map[selected_number]["id"]

    if action == "quantity":
        await state.set_state(AdminEditReport.waiting_for_quantity)
        await message.answer("Введите новое количество:", reply_markup=navigation_keyboard())
        return

    result = delete_shift_operation(data["admin_edit_shift_operation_id"])

    if result is None:
        await state.clear()
        await message.answer("Не удалось удалить строку отчёта.", reply_markup=admin_reports_keyboard())
        return

    add_edit_log(
        message.from_user.id,
        "admin",
        "Удалил строку отчёта сотрудника",
        "shift",
        result["shift_id"],
        (
            f"{data['admin_edit_employee_name']}: "
            f"{result['operation_name']}, размер {result['product_size']}, "
            f"цвет {result['product_color']} — {result['quantity']} {result['unit']}"
        ),
    )

    report = get_shift_report(result["shift_id"])
    text = "Строка удалена.\n\nОтчёт после правки:\n\n"

    if not report:
        text += "Операции ещё не добавлены."
    else:
        for index, row in enumerate(report, start=1):
            name, product_size, product_color, qty, unit = row
            text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

    await state.clear()
    await send_long_text(message, text, reply_markup=admin_reports_keyboard())


@dp.message(AdminEditReport.waiting_for_action)
async def process_admin_edit_action(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_admin_edit_action(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("admin_edit_action:"))
async def process_admin_edit_action_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, AdminEditReport.waiting_for_action):
        return

    await callback.answer()
    await select_admin_edit_action(callback.message, state, callback.data.rsplit(":", 1)[1])


@dp.message(AdminEditReport.waiting_for_quantity)
async def process_admin_edit_quantity(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit() or int(message.text) <= 0:
        await message.answer("Введите количество положительным числом, например: 280")
        return

    data = await state.get_data()
    quantity = int(message.text)
    result = update_shift_operation_quantity(data["admin_edit_shift_operation_id"], quantity)

    if result is None:
        await state.clear()
        await message.answer("Не удалось изменить строку отчёта.", reply_markup=admin_reports_keyboard())
        return

    add_edit_log(
        message.from_user.id,
        "admin",
        "Изменил отчёт сотрудника",
        "shift",
        result["shift_id"],
        (
            f"{data['admin_edit_employee_name']}: {result['operation_name']}: "
            f"размер {result['product_size']}, цвет {result['product_color']}: "
            f"{result['old_quantity']} -> {result['new_quantity']} {result['unit']}"
        ),
    )

    report = get_shift_report(result["shift_id"])
    text = "Количество изменено.\n\nОтчёт после правки:\n\n"

    for index, row in enumerate(report, start=1):
        name, product_size, product_color, qty, unit = row
        text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

    await state.clear()
    await send_long_text(message, text, reply_markup=admin_reports_keyboard())


@dp.message(AdminPeriod.waiting_for_report_period)
async def process_period_button(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    period = parse_period(message.text)

    if period is None:
        await message.answer("Введите две даты в формате: 2026-06-01 2026-06-25")
        return

    await state.clear()
    await send_period_report(message, period[0], period[1])


@dp.message(AdminPeriod.waiting_for_excel_period)
async def process_excel_period_button(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    period = parse_period(message.text)

    if period is None:
        await message.answer("Введите две даты в формате: 2026-06-01 2026-06-25")
        return

    await state.clear()
    await send_period_excel(message, period[0], period[1])


@dp.message(AdminPeriod.waiting_for_employee_report_id)
async def process_employee_report_id(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit():
        await message.answer("Введите ID сотрудника числом, например: 1")
        return

    await state.update_data(employee_id=int(message.text))
    await state.set_state(AdminPeriod.waiting_for_employee_report_period)
    await message.answer(
        "Введите период отчёта через пробел.\n\n"
        "Пример: 2026-06-01 2026-06-25"
    )


@dp.message(AdminPeriod.waiting_for_employee_report_period)
async def process_employee_report_period(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    period = parse_period(message.text)

    if period is None:
        await message.answer("Введите две даты в формате: 2026-06-01 2026-06-25")
        return

    data = await state.get_data()
    await state.clear()
    await send_employee_excel_report(message, data["employee_id"], period[0], period[1])


@dp.message(lambda message: message.text == "Excel")
async def excel_button(message: Message):
    await excel_report(message)


@dp.message(lambda message: message.text == "Выгрузить отчёт")
async def export_report_button(message: Message):
    await excel_report(message)


@dp.message(lambda message: message.text == "Журнал")
async def logs_button(message: Message):
    await logs_report(message)


@dp.message(lambda message: message.text == "Ошибки")
async def errors_button(message: Message):
    await errors_report(message)


@dp.message(lambda message: message.text == "Список операций")
async def operations_button(message: Message):
    await operations_report(message)


@dp.message(lambda message: message.text == "Добавить операцию")
async def add_operation_button(message: Message, state: FSMContext):
    await add_operation_start(message, state)


@dp.message(lambda message: message.text == "Изменить операцию")
async def edit_operation_button(message: Message, state: FSMContext):
    await edit_operation_start(message, state)


@dp.message(lambda message: message.text == "Скрыть операцию")
async def hide_operation_button(message: Message, state: FSMContext):
    await hide_operation_start(message, state)


@dp.message(lambda message: message.text == "Вернуть операцию")
async def restore_operation_button(message: Message, state: FSMContext):
    await restore_operation_start(message, state)


@dp.message(lambda message: message.text == "Список сотрудников")
async def employees_button(message: Message):
    await employees_report(message)


@dp.message(lambda message: message.text == "Активные сотрудники")
async def active_employees_button(message: Message):
    await active_employees_report(message)


@dp.message(lambda message: message.text == "Неактивные сотрудники")
async def inactive_employees_button(message: Message):
    await inactive_employees_report(message)


@dp.message(lambda message: message.text == "Скачать базу")
async def backup_button(message: Message):
    await backup_database(message)


@dp.message(lambda message: message.text == "Проверка базы")
async def database_status_button(message: Message):
    await database_status(message)


@dp.message(lambda message: message.text == "Загрузить базу")
async def restore_database_button(message: Message, state: FSMContext):
    await restore_database_start(message, state)


@dp.message(lambda message: message.text == "Создать копию базы")
async def create_backup_button(message: Message):
    await create_backup_command(message)


@dp.message(Command("report"))
async def start_report(message: Message, state: FSMContext):
    employee = get_employee_by_telegram_id(message.from_user.id)

    if employee is None or employee[5] != "active":
        await message.answer("Сначала нужно зарегистрироваться и дождаться подтверждения.")
        return

    shift = get_open_shift_for_today(employee[0])

    if shift is None:
        await message.answer("У вас нет открытой смены. Нажмите /start, чтобы открыть смену.")
        return

    employee_position = employee[3]
    groups = get_active_operation_groups(employee_position)

    if not groups:
        await message.answer(
            f"Для должности «{employee_position}» пока нет активных операций.\n"
            "Обратитесь к администратору."
        )
        return

    group_map = {}
    text = f"Раздел: {employee_position}\n\nВыберите группу. Напишите номер:\n\n"

    for index, group in enumerate(groups, start=1):
        group_map[str(index)] = group
        text += f"{index}. {group}\n"

    await state.update_data(
        shift_id=shift[0],
        employee_id=employee[0],
        employee_position=employee_position,
        group_map=group_map,
    )
    await state.set_state(Report.waiting_for_group)
    await message.answer(text, reply_markup=choice_keyboard("report_group", group_map))


async def select_operation_group(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    group_map = data.get("group_map", {})

    if selected_number not in group_map:
        await message.answer("Введите номер группы из списка.")
        return

    selected_group = group_map[selected_number]
    employee_position = data["employee_position"]
    folders = get_active_operation_folders(employee_position, selected_group)

    if not folders:
        await message.answer("В этой группе нет активных изделий.")
        return

    folder_map = {}
    item_word = "раздел подготовки" if employee_position == "Упаковщик" and selected_group == "Подготовка" else "изделие"
    text = f"Группа: {selected_group}\n\nВыберите {item_word}. Напишите номер:\n\n"

    for index, folder in enumerate(folders, start=1):
        folder_map[str(index)] = folder
        text += f"{index}. {folder}\n"

    await state.update_data(selected_group=selected_group, folder_map=folder_map)
    await state.set_state(Report.waiting_for_folder)
    await message.answer(text, reply_markup=choice_keyboard("report_folder", folder_map))


@dp.message(Report.waiting_for_group)
async def process_operation_group(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_operation_group(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("report_group:"))
async def process_operation_group_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_group):
        return

    await callback.answer()
    await select_operation_group(callback.message, state, callback.data.rsplit(":", 1)[1])


async def select_operation_folder(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    folder_map = data.get("folder_map", {})

    if selected_number not in folder_map:
        await message.answer("Введите номер папки из списка.")
        return

    selected_folder = folder_map[selected_number]
    sizes = get_product_sizes(selected_folder)

    if data.get("employee_position") == "Упаковщик" and is_preparation_operation_folder(selected_folder):
        await state.update_data(
            selected_folder=selected_folder,
            selected_size="",
            selected_color="",
            selected_sizes=[],
            selected_colors=[],
            selected_size_numbers=[],
            selected_color_numbers=[],
            size_map={},
            color_map={},
        )
        await ask_report_operation(message, state)
        return

    if data.get("employee_position") == "Раскройщик":
        size_map = {}
        color_map = {}

        for index, product_size in enumerate(sizes, start=1):
            size_map[str(index)] = product_size

        for index, color in enumerate(get_product_colors(selected_folder), start=1):
            color_map[str(index)] = color

        await state.update_data(
            selected_folder=selected_folder,
            selected_size="",
            selected_color="",
            selected_sizes=[],
            selected_colors=[],
            selected_size_numbers=[],
            selected_color_numbers=[],
            size_map=size_map,
            color_map=color_map,
            operation_id=None,
            operation_name="",
            cutting_mode="",
            cutting_quantity_index=0,
            cutting_quantities={},
        )
        await ask_report_operation(message, state)
        return

    if sizes:
        size_map = {}

        for index, product_size in enumerate(sizes, start=1):
            size_map[str(index)] = product_size

        await state.update_data(
            selected_folder=selected_folder,
            size_map=size_map,
            selected_size_numbers=[],
            selected_sizes=[],
        )
        await state.set_state(Report.waiting_for_size)

        if data.get("employee_position") == "Раскройщик":
            await send_report_size_step(message, state)
            return

        text = f"Изделие: {selected_folder}\n\nВыберите размер. Напишите номер:\n\n"

        for number, product_size in size_map.items():
            text += f"{number}. {product_size}\n"

        await message.answer(text, reply_markup=choice_keyboard("report_size", size_map))
        return

    await state.update_data(selected_folder=selected_folder, selected_size="без размера")
    await ask_report_color(message, state)


@dp.message(Report.waiting_for_folder)
async def process_operation_folder(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_operation_folder(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("report_folder:"))
async def process_operation_folder_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_folder):
        return

    await callback.answer()
    await select_operation_folder(callback.message, state, callback.data.rsplit(":", 1)[1])


async def ask_report_color(message: Message, state: FSMContext):
    data = await state.get_data()
    selected_folder = data["selected_folder"]
    colors = get_product_colors(selected_folder)

    if is_packing_preparation_flow(data):
        colors = get_preparation_material_colors()

    if colors:
        color_map = {}

        for index, color in enumerate(colors, start=1):
            color_map[str(index)] = color

        await state.update_data(color_map=color_map, selected_color_numbers=[], selected_colors=[])
        await state.set_state(Report.waiting_for_color)

        if data.get("employee_position") == "Раскройщик":
            await send_report_color_step(message, state)
            return

        text = f"Изделие: {selected_folder}\nРазмер: {data['selected_size']}\n\nВыберите цвет. Напишите номер:\n\n"

        for number, color in color_map.items():
            text += f"{number}. {color}\n"

        await message.answer(text, reply_markup=choice_keyboard("report_color", color_map))
        return

    if data.get("employee_position") == "Раскройщик":
        await state.update_data(selected_colors=["без цвета"], selected_color="без цвета")
    else:
        await state.update_data(selected_color="без цвета")

    await ask_report_operation(message, state)


async def ask_report_operation(message: Message, state: FSMContext):
    data = await state.get_data()
    selected_folder = data["selected_folder"]
    employee_position = data["employee_position"]
    selected_group = data["selected_group"]
    operations = get_active_operations(employee_position, selected_folder, selected_group)

    if not operations:
        await message.answer("В этой папке нет активных операций.")
        return

    operation_map = {}
    size_text = data.get("selected_size", ", ".join(data.get("selected_sizes", [])))
    color_text = data.get("selected_color", ", ".join(data.get("selected_colors", [])))
    text = f"Изделие: {selected_folder}\n"

    if size_text:
        text += f"Размер: {size_text}\n"

    if color_text:
        text += f"Цвет: {color_text}\n"

    text += "\nВыберите операцию. Напишите номер:\n\n"

    for index, operation in enumerate(operations, start=1):
        operation_id, _, name, unit = operation
        operation_map[str(index)] = {
            "id": operation_id,
            "name": name,
            "unit": unit,
        }
        text += f"{index}. {name} ({unit})\n"

    await state.update_data(selected_folder=selected_folder, operation_map=operation_map)
    await state.set_state(Report.waiting_for_operation)
    await message.answer(text, reply_markup=choice_keyboard("report_operation", operation_map))


def parse_multiple_numbers(text: str, available_numbers: set[str]):
    numbers = re.findall(r"\d+", text)
    unique_numbers = []

    for number in numbers:
        if number in available_numbers and number not in unique_numbers:
            unique_numbers.append(number)

    return unique_numbers


async def finish_multi_size_selection(message: Message, state: FSMContext):
    data = await state.get_data()
    size_map = data.get("size_map", {})
    selected_numbers = data.get("selected_size_numbers", [])

    if not selected_numbers:
        await message.answer("Выберите хотя бы один размер.")
        return

    selected_sizes = [size_map[number] for number in selected_numbers]
    await state.update_data(
        selected_size=", ".join(selected_sizes),
        selected_sizes=selected_sizes,
    )

    if data.get("employee_position") == "Раскройщик" and data.get("cutting_mode") == CUTTING_MODE_CONTOURS:
        await state.update_data(cutting_quantity_index=0, cutting_quantities={})
        await state.set_state(Report.waiting_for_quantity)
        await ask_cutting_size_quantity(message, state)
        return

    if is_packing_preparation_flow({**data, "selected_folder": data.get("selected_folder", "")}):
        await ask_report_color(message, state)
        return

    await ask_report_color(message, state)


async def finish_multi_color_selection(message: Message, state: FSMContext):
    data = await state.get_data()
    color_map = data.get("color_map", {})
    selected_numbers = data.get("selected_color_numbers", [])

    if not selected_numbers:
        await message.answer("Выберите хотя бы один цвет.")
        return

    selected_colors = [color_map[number] for number in selected_numbers]
    await state.update_data(
        selected_color=", ".join(selected_colors),
        selected_colors=selected_colors,
    )

    if data.get("employee_position") == "Раскройщик" and data.get("operation_id"):
        await state.update_data(cutting_quantity_index=0, cutting_quantities={})
        await state.set_state(Report.waiting_for_quantity)
        await ask_cutting_color_quantity(message, state)
        return

    await ask_report_operation(message, state)


async def select_report_size(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    size_map = data.get("size_map", {})

    if selected_number not in size_map:
        await message.answer("Введите номер размера из списка.")
        return

    await state.update_data(selected_size=size_map[selected_number])
    await ask_report_color(message, state)


@dp.message(Report.waiting_for_size)
async def process_report_size(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    data = await state.get_data()

    if data.get("employee_position") == "Раскройщик" or is_packing_preparation_flow(data):
        size_map = data.get("size_map", {})
        selected_numbers = parse_multiple_numbers(message.text, set(size_map.keys()))

        if not selected_numbers:
            await message.answer("Напишите номера размеров через пробел или выберите кнопками.")
            return

        await state.update_data(selected_size_numbers=selected_numbers)
        await finish_multi_size_selection(message, state)
        return

    await select_report_size(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("report_size_multi_toggle:"))
async def process_multi_report_size_toggle(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_size):
        return

    data = await state.get_data()
    size_map = data.get("size_map", {})
    selected_numbers = data.get("selected_size_numbers", [])
    selected_number = callback.data.rsplit(":", 1)[1]

    if selected_number not in size_map:
        await callback.answer("Такого размера нет")
        return

    if selected_number in selected_numbers:
        selected_numbers = [number for number in selected_numbers if number != selected_number]
    else:
        selected_numbers.append(selected_number)

    await state.update_data(selected_size_numbers=selected_numbers)
    await callback.answer()
    await callback.message.edit_reply_markup(
        reply_markup=multi_choice_keyboard("report_size_multi", size_map, selected_numbers)
    )


@dp.callback_query(lambda callback: callback.data == "report_size_multi_done")
async def process_multi_report_size_done(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_size):
        return

    await callback.answer()
    await finish_multi_size_selection(callback.message, state)


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("report_size:"))
async def process_report_size_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_size):
        return

    await callback.answer()
    await select_report_size(callback.message, state, callback.data.rsplit(":", 1)[1])


async def select_report_color(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    color_map = data.get("color_map", {})

    if selected_number not in color_map:
        await message.answer("Введите номер цвета из списка.")
        return

    await state.update_data(selected_color=color_map[selected_number])
    await ask_report_operation(message, state)


@dp.message(Report.waiting_for_color)
async def process_report_color(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    data = await state.get_data()

    if data.get("employee_position") == "Раскройщик":
        color_map = data.get("color_map", {})
        selected_numbers = parse_multiple_numbers(message.text, set(color_map.keys()))

        if not selected_numbers:
            await message.answer("Напишите номера цветов через пробел или выберите кнопками.")
            return

        await state.update_data(selected_color_numbers=selected_numbers)
        await finish_multi_color_selection(message, state)
        return

    await select_report_color(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("report_color_multi_toggle:"))
async def process_multi_report_color_toggle(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_color):
        return

    data = await state.get_data()
    color_map = data.get("color_map", {})
    selected_numbers = data.get("selected_color_numbers", [])
    selected_number = callback.data.rsplit(":", 1)[1]

    if selected_number not in color_map:
        await callback.answer("Такого цвета нет")
        return

    if selected_number in selected_numbers:
        selected_numbers = [number for number in selected_numbers if number != selected_number]
    else:
        selected_numbers.append(selected_number)

    await state.update_data(selected_color_numbers=selected_numbers)
    await callback.answer()
    await callback.message.edit_reply_markup(
        reply_markup=multi_choice_keyboard("report_color_multi", color_map, selected_numbers)
    )


@dp.callback_query(lambda callback: callback.data == "report_color_multi_done")
async def process_multi_report_color_done(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_color):
        return

    await callback.answer()
    await finish_multi_color_selection(callback.message, state)


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("report_color:"))
async def process_report_color_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_color):
        return

    await callback.answer()
    await select_report_color(callback.message, state, callback.data.rsplit(":", 1)[1])


async def select_cutting_batch(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    batch_map = data.get("cutting_batch_map", {})

    if selected_number not in batch_map:
        await message.answer("Выберите номер партии из списка.")
        return

    batch = batch_map[selected_number]
    cutting_mode = data.get("cutting_mode")

    await state.update_data(
        cutting_batch_id=batch["id"],
        cutting_batch_name=batch["name"],
    )

    if cutting_mode == CUTTING_MODE_LAYOUT:
        await state.update_data(
            selected_size="без размера",
            selected_sizes=["без размера"],
            selected_color_numbers=[],
            selected_colors=[],
        )
        await send_report_color_step(message, state)
        return

    if cutting_mode == CUTTING_MODE_CUT:
        await state.set_state(Report.waiting_for_progress)
        await message.answer(
            f"Партия:\n{batch['name']}\n\nВыберите процент готовности раскроя:",
            reply_markup=progress_keyboard(),
        )
        return

    if cutting_mode == CUTTING_MODE_FORM:
        result_rows = get_cutting_batch_result_rows(batch["id"])

        if not result_rows:
            await message.answer("По этой партии нет расчёта размеров и цветов.", reply_markup=report_keyboard())
            await state.clear()
            return

        marked = mark_cutting_batch_formed(
            batch["id"],
            data["shift_id"],
            data["employee_id"],
            data["operation_id"],
        )

        if not marked:
            await message.answer("Эту партию уже сформировали или она недоступна.", reply_markup=report_keyboard())
            await state.clear()
            return

        added_count = 0
        for product_size, product_color, quantity in result_rows:
            add_shift_operation(
                data["shift_id"],
                data["employee_id"],
                data["operation_id"],
                product_size,
                product_color,
                quantity,
            )
            added_count += 1

        add_edit_log(
            message.from_user.id,
            "employee",
            "Сформировал готовый крой",
            "shift",
            data["shift_id"],
            f"{data['operation_name']}, партия {batch['id']}, строк добавлено: {added_count}",
        )

        report = get_shift_report(data["shift_id"])
        text = (
            f"Готовый крой сформирован. Строк добавлено: {added_count}\n\n"
            "Расчёт партии:\n"
            f"{format_cutting_batch_result(result_rows)}\n"
            "Текущий отчёт за смену:\n\n"
        )

        for index, row in enumerate(report, start=1):
            name, product_size, product_color, qty, unit = row
            text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

        await state.clear()
        await send_long_text(message, text, reply_markup=report_keyboard())


@dp.message(Report.waiting_for_batch)
async def process_cutting_batch(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_cutting_batch(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("cutting_batch:"))
async def process_cutting_batch_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_batch):
        return

    await callback.answer()
    await select_cutting_batch(callback.message, state, callback.data.rsplit(":", 1)[1])


async def select_operation(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    operation_map = data.get("operation_map", {})

    if selected_number not in operation_map:
        await message.answer("Такой операции нет. Введите номер из списка.")
        return

    operation = operation_map[selected_number]
    operation_name = f"{data['selected_folder']}: {operation['name']}"
    cutting_mode = get_cutting_mode(operation["name"])

    update_payload = {
        "operation_id": operation["id"],
        "operation_name": operation_name,
        "operation_unit": operation.get("unit", "шт"),
    }

    if data.get("employee_position") == "Раскройщик":
        update_payload.update(
            {
                "cutting_mode": cutting_mode,
                "cutting_quantity_index": 0,
                "cutting_quantities": {},
            }
        )

    await state.update_data(**update_payload)

    if data.get("employee_position") == "Раскройщик":
        if cutting_mode == CUTTING_MODE_LAYOUT:
            await send_cutting_batch_step(message, state)
            return

        if cutting_mode in {CUTTING_MODE_CUT, CUTTING_MODE_FORM}:
            await send_cutting_batch_step(message, state)
            return

        if cutting_mode == CUTTING_MODE_CONTOURS:
            await state.update_data(
                selected_color="без цвета",
                selected_colors=["без цвета"],
                selected_size_numbers=[],
                selected_sizes=[],
            )
            await send_report_size_step(message, state)
            return

        await state.update_data(
            selected_color="",
            selected_colors=[],
            selected_size_numbers=[],
            selected_sizes=[],
        )
        await send_report_size_step(message, state)
        return

    if is_packing_preparation_flow({**data, **update_payload}):
        preparation_sizes = get_preparation_operation_sizes(operation["name"])
        size_map = {
            str(index): product_size
            for index, product_size in enumerate(preparation_sizes, start=1)
        }
        color_map = {
            str(index): color
            for index, color in enumerate(get_preparation_material_colors(), start=1)
        }

        if not size_map:
            await message.answer("Для этой подготовки пока нет размеров в справочнике.")
            return

        await state.update_data(
            size_map=size_map,
            color_map=color_map,
            selected_size_numbers=[],
            selected_sizes=[],
            selected_color="",
            selected_colors=[],
        )
        await send_report_size_step(message, state)
        return

    await state.set_state(Report.waiting_for_quantity)
    quantity_prompt = "Введите время в минутах:" if operation.get("unit") == "мин" else "Введите количество:"
    await message.answer(
        f"Операция: {operation_name}\n"
        f"Размер: {data.get('selected_size', ', '.join(data.get('selected_sizes', [])))}\n"
        f"Цвет: {data.get('selected_color', ', '.join(data.get('selected_colors', [])))}\n\n"
        f"{quantity_prompt}",
        reply_markup=navigation_keyboard(),
    )


async def ask_cutting_color_quantity(message: Message, state: FSMContext):
    data = await state.get_data()
    selected_colors = data.get("selected_colors") or [data["selected_color"]]
    selected_sizes = data.get("selected_sizes") or [data["selected_size"]]
    color_index = data.get("cutting_quantity_index", 0)

    if color_index >= len(selected_colors):
        return

    current_color = selected_colors[color_index]
    is_layout = data.get("cutting_mode") == CUTTING_MODE_LAYOUT
    size_line = "" if is_layout else f"Размеры: {', '.join(selected_sizes)}\n"
    unit_text = "количество слоёв" if is_layout else "количество"

    await message.answer(
        f"Операция: {data['operation_name']}\n"
        f"{size_line}"
        f"Цвет {color_index + 1} из {len(selected_colors)}: {current_color}\n\n"
        f"Введите {unit_text} для цвета «{current_color}»:",
        reply_markup=navigation_keyboard(),
    )


async def ask_cutting_size_quantity(message: Message, state: FSMContext):
    data = await state.get_data()
    selected_sizes = data.get("selected_sizes") or [data["selected_size"]]
    size_index = data.get("cutting_quantity_index", 0)

    if size_index >= len(selected_sizes):
        return

    current_size = selected_sizes[size_index]

    await message.answer(
        f"Операция: {data['operation_name']}\n"
        f"Размер {size_index + 1} из {len(selected_sizes)}: {current_size}\n\n"
        f"Введите количество штук для размера {current_size}:",
        reply_markup=navigation_keyboard(),
    )


@dp.message(Report.waiting_for_operation)
async def process_operation(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_operation(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("report_operation:"))
async def process_operation_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_operation):
        return

    await callback.answer()
    await select_operation(callback.message, state, callback.data.rsplit(":", 1)[1])


async def finish_cutting_progress(message: Message, state: FSMContext, progress: int):
    if progress not in {25, 50, 75, 100}:
        await message.answer("Выберите процент: 25, 50, 75 или 100.")
        return

    data = await state.get_data()
    updated = update_cutting_batch_progress(
        data["cutting_batch_id"],
        data["shift_id"],
        data["employee_id"],
        data["operation_id"],
        progress,
    )

    if not updated:
        await state.clear()
        await message.answer("Партия недоступна для раскроя.", reply_markup=report_keyboard())
        return

    set_shift_operation_quantity(
        data["shift_id"],
        data["employee_id"],
        data["operation_id"],
        "готовность",
        "без цвета",
        progress,
    )

    add_edit_log(
        message.from_user.id,
        "employee",
        "Обновил процент раскроя",
        "shift",
        data["shift_id"],
        f"{data['operation_name']}, партия {data['cutting_batch_id']}: {progress}%",
    )

    text = f"Готовность раскроя сохранена: {progress}%.\n"

    if progress == 100:
        result_rows = get_cutting_batch_result_rows(data["cutting_batch_id"])
        text += (
            "\nРаскрой завершён. Расчёт для сверки:\n"
            f"{format_cutting_batch_result(result_rows)}\n"
            "Теперь партия доступна в операции «Формирование готового кроя»."
        )
    else:
        text += "\nПартия останется доступной в операции «Раскрой» для продолжения."

    await state.clear()
    await send_long_text(message, text, reply_markup=report_keyboard())


@dp.message(Report.waiting_for_progress)
async def process_cutting_progress(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit():
        await message.answer("Введите процент: 25, 50, 75 или 100.")
        return

    await finish_cutting_progress(message, state, int(message.text))


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("cutting_progress:"))
async def process_cutting_progress_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, Report.waiting_for_progress):
        return

    await callback.answer()
    await finish_cutting_progress(callback.message, state, int(callback.data.rsplit(":", 1)[1]))


@dp.message(Report.waiting_for_quantity)
async def process_quantity(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    data = await state.get_data()
    quantity_error_text = (
        "Введите время положительным числом минут, например: 15"
        if data.get("operation_unit") == "мин"
        else "Введите количество положительным числом, например: 200"
    )

    if not message.text.isdigit() or int(message.text) <= 0:
        await message.answer(quantity_error_text)
        return

    quantity = int(message.text)

    if data.get("employee_position") == "Раскройщик":
        if data.get("cutting_mode") == CUTTING_MODE_CONTOURS:
            selected_sizes = data.get("selected_sizes") or [data["selected_size"]]
            size_index = data.get("cutting_quantity_index", 0)
            cutting_quantities = data.get("cutting_quantities", {})
            current_size = selected_sizes[size_index]
            cutting_quantities[current_size] = quantity
            next_index = size_index + 1

            if next_index < len(selected_sizes):
                await state.update_data(
                    cutting_quantity_index=next_index,
                    cutting_quantities=cutting_quantities,
                )
                await ask_cutting_size_quantity(message, state)
                return

            batch_id = create_cutting_contour_batch(
                data["selected_folder"],
                data["shift_id"],
                data["employee_id"],
                data["operation_id"],
                cutting_quantities,
            )

            added_count = 0
            for selected_size in selected_sizes:
                add_shift_operation(
                    data["shift_id"],
                    data["employee_id"],
                    data["operation_id"],
                    selected_size,
                    "без цвета",
                    cutting_quantities[selected_size],
                )
                added_count += 1

            quantity_text = ", ".join(f"{size}: {cutting_quantities[size]}" for size in selected_sizes)

            add_edit_log(
                message.from_user.id,
                "employee",
                "Создал партию раскроя",
                "shift",
                data["shift_id"],
                f"{data['operation_name']}, партия {batch_id}, контуры: {quantity_text}",
            )

            report = get_shift_report(data["shift_id"])
            text = (
                f"Контуры сохранены. Партия ID {batch_id}\n\n"
                "Количество по размерам:\n"
                f"{quantity_text}\n\n"
                "Теперь эту партию можно выбрать в операции «Формирование настила».\n\n"
                "Текущий отчёт за смену:\n\n"
            )

            for index, row in enumerate(report, start=1):
                name, product_size, product_color, qty, unit = row
                text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

            await state.clear()
            await send_long_text(message, text, reply_markup=report_keyboard())
            return

        selected_sizes = data.get("selected_sizes") or [data["selected_size"]]
        selected_colors = data.get("selected_colors") or [data["selected_color"]]
        color_index = data.get("cutting_quantity_index", 0)
        cutting_quantities = data.get("cutting_quantities", {})

        current_color = selected_colors[color_index]
        cutting_quantities[current_color] = quantity
        next_index = color_index + 1

        if next_index < len(selected_colors):
            await state.update_data(
                cutting_quantity_index=next_index,
                cutting_quantities=cutting_quantities,
            )
            await ask_cutting_color_quantity(message, state)
            return

        added_count = 0

        if data.get("cutting_mode") == CUTTING_MODE_LAYOUT:
            layout_saved = add_cutting_layout(
                data["cutting_batch_id"],
                data["shift_id"],
                data["employee_id"],
                data["operation_id"],
                cutting_quantities,
            )

            if not layout_saved:
                await state.clear()
                await message.answer("Не удалось сохранить настил. Возможно, партия уже использована.", reply_markup=report_keyboard())
                return

        for selected_color in selected_colors:
            color_quantity = cutting_quantities[selected_color]

            for selected_size in selected_sizes:
                add_shift_operation(
                    data["shift_id"],
                    data["employee_id"],
                    data["operation_id"],
                    selected_size,
                    selected_color,
                    color_quantity,
                )
                added_count += 1

        is_layout = data.get("cutting_mode") == CUTTING_MODE_LAYOUT
        quantity_text = ", ".join(
            f"{color}: {cutting_quantities[color]}" for color in selected_colors
        )
        quantity_label = "слоёв по цветам" if is_layout else "количество по цветам"

        add_edit_log(
            message.from_user.id,
            "employee",
            "Добавил операцию",
            "shift",
            data["shift_id"],
            (
                f"{data['operation_name']}, размеры {', '.join(selected_sizes)}, "
                f"{quantity_label}: {quantity_text}, строк добавлено: {added_count}"
            ),
        )

        report = get_shift_report(data["shift_id"])

        text = (
            f"Операция добавлена. Строк добавлено: {added_count}\n\n"
            f"{quantity_label.capitalize()}:\n"
            f"{quantity_text}\n\n"
            "Текущий отчёт за смену:\n\n"
        )

        for index, row in enumerate(report, start=1):
            name, product_size, product_color, qty, unit = row
            text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

        text += "\nВыберите следующее действие в меню ниже."

        await state.clear()
        await send_long_text(message, text, reply_markup=report_keyboard())
        return

    selected_sizes = data.get("selected_sizes") or [data["selected_size"]]
    selected_colors = data.get("selected_colors") or [data["selected_color"]]
    added_count = 0

    for selected_size in selected_sizes:
        for selected_color in selected_colors:
            add_shift_operation(
                data["shift_id"],
                data["employee_id"],
                data["operation_id"],
                selected_size,
                selected_color,
                quantity,
            )
            added_count += 1

    add_edit_log(
        message.from_user.id,
        "employee",
        "Добавил операцию",
        "shift",
        data["shift_id"],
        (
            f"{data['operation_name']}, размеры {', '.join(selected_sizes)}, "
            f"цвета {', '.join(selected_colors)} — {quantity}, строк добавлено: {added_count}"
        ),
    )

    report = get_shift_report(data["shift_id"])

    text = (
        f"Операция добавлена. Строк добавлено: {added_count}\n\n"
        "Текущий отчёт за смену:\n\n"
    )

    for index, row in enumerate(report, start=1):
        name, product_size, product_color, qty, unit = row
        text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

    text += "\nВыберите следующее действие в меню ниже."

    await state.clear()
    await send_long_text(message, text, reply_markup=report_keyboard())


@dp.message(Command("current"))
async def current_report(message: Message):
    employee = get_employee_by_telegram_id(message.from_user.id)

    if employee is None or employee[5] != "active":
        await message.answer("Сначала нужно зарегистрироваться и дождаться подтверждения.")
        return

    shift = get_shift_for_today(employee[0])

    if shift is None:
        await message.answer("За сегодня отчёта нет. Нажмите /start, чтобы открыть смену.")
        return

    report = get_shift_report(shift[0])
    end_time = shift[4] if shift[4] else "смена открыта"
    status_text = "закрыта" if shift[5] == "closed" else "открыта"

    text = (
        "Отчёт за сегодня:\n\n"
        f"Дата: {shift[2]}\n"
        f"Начало смены: {shift[3]}\n"
        f"Окончание: {end_time}\n"
        f"Статус: {status_text}\n\n"
    )

    if not report:
        text += "Операции ещё не добавлены."
    else:
        text += "Операции:\n"
        for index, row in enumerate(report, start=1):
            name, product_size, product_color, qty, unit = row
            text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

    if shift[5] == "open":
        text += "\nВыберите следующее действие в меню отчёта."
    else:
        text += "\nМожно посмотреть отчёт за другой период через кнопку «Отчёт за даты»."

    await send_long_text(message, text, reply_markup=report_keyboard())


async def send_employee_period_report(message: Message, start_date: str, end_date: str):
    employee = get_employee_by_telegram_id(message.from_user.id)

    if employee is None or employee[5] != "active":
        await message.answer("Сначала нужно зарегистрироваться и дождаться подтверждения.")
        return

    if not is_valid_date(start_date) or not is_valid_date(end_date):
        await message.answer("Введите даты в формате: 2026-06-01 2026-06-25")
        return

    if start_date > end_date:
        await message.answer("Первая дата должна быть раньше второй или равна ей.")
        return

    shifts = get_employee_shifts_by_period(employee[0], start_date, end_date)

    if not shifts:
        await message.answer(
            f"За период {start_date} — {end_date} отчётов нет.",
            reply_markup=report_keyboard(),
        )
        return

    text = (
        f"Ваш отчёт за период:\n"
        f"{start_date} — {end_date}\n\n"
    )
    total_minutes = 0

    for shift in shifts:
        shift_id, shift_date, start_time, end_time, minutes, status = shift
        status_text = "закрыта" if status == "closed" else "открыта"
        end_text = end_time if end_time else "смена открыта"

        if minutes:
            total_minutes += minutes

        text += (
            f"{shift_date}\n"
            f"Начало: {start_time}\n"
            f"Окончание: {end_text}\n"
            f"Статус: {status_text}\n"
        )

        if minutes:
            text += f"Отработано: {format_minutes(minutes)}\n"

        report = get_shift_report(shift_id)

        if not report:
            text += "Операции не добавлены.\n\n"
            continue

        text += "Операции:\n"
        for index, row in enumerate(report, start=1):
            name, product_size, product_color, qty, unit = row
            text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

        text += "\n"

    if total_minutes:
        text += f"Итого времени: {format_minutes(total_minutes)}"

    await send_long_text(message, text, reply_markup=report_keyboard())


@dp.message(Command("my_report"))
async def my_period_report(message: Message, state: FSMContext):
    parts = message.text.split()

    if len(parts) == 3 and parts[0] == "/my_report":
        await send_employee_period_report(message, parts[1], parts[2])
        return

    await state.set_state(DateReport.waiting_for_period)
    await message.answer(
        "Введите даты отчёта через пробел.\n\n"
        "Пример: 2026-06-01 2026-06-25",
        reply_markup=report_keyboard(),
    )


@dp.message(DateReport.waiting_for_period)
async def process_period_report(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    parts = message.text.split()

    if len(parts) != 2:
        await message.answer("Введите две даты через пробел, например: 2026-06-01 2026-06-25")
        return

    await state.clear()
    await send_employee_period_report(message, parts[0], parts[1])


@dp.message(Command("edit"))
async def edit_report(message: Message, state: FSMContext):
    employee = get_employee_by_telegram_id(message.from_user.id)

    if employee is None or employee[5] != "active":
        await message.answer("Сначала нужно зарегистрироваться и дождаться подтверждения.")
        return

    shift = get_editable_shift_for_today(employee[0])

    if shift is None:
        await message.answer(
            "Редактирование недоступно.\n\n"
            "Можно редактировать открытую смену или закрытую смену в течение 1 часа."
        )
        return

    operations = get_shift_operation_choices(shift[0])

    if not operations:
        await message.answer("В отчёте пока нет операций для изменения.")
        return

    operation_map = {}
    text = "Что изменить? Напишите номер строки:\n\n"

    for index, operation in enumerate(operations, start=1):
        shift_operation_id, name, product_size, product_color, quantity, unit = operation
        operation_line = format_operation_line(name, product_size, product_color, quantity, unit)
        operation_map[str(index)] = {
            "id": shift_operation_id,
            "name": operation_line,
        }
        text += f"{index}. {operation_line}\n"

    await state.update_data(shift_id=shift[0], operation_map=operation_map)
    await state.set_state(EditReport.waiting_for_operation)
    await message.answer(text, reply_markup=choice_keyboard("edit_operation", operation_map))


async def select_edit_operation(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    operation_map = data.get("operation_map", {})

    if selected_number not in operation_map:
        await message.answer("Введите номер строки из списка.")
        return

    await state.update_data(shift_operation_id=operation_map[selected_number]["id"])
    await state.set_state(EditReport.waiting_for_quantity)
    await message.answer("Введите новое количество:")


@dp.message(EditReport.waiting_for_operation)
async def edit_operation_selected(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_edit_operation(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("edit_operation:"))
async def edit_operation_selected_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, EditReport.waiting_for_operation):
        return

    await callback.answer()
    await select_edit_operation(callback.message, state, callback.data.rsplit(":", 1)[1])


@dp.message(EditReport.waiting_for_quantity)
async def edit_quantity_entered(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    if not message.text.isdigit() or int(message.text) <= 0:
        await message.answer("Введите количество положительным числом, например: 280")
        return

    data = await state.get_data()
    quantity = int(message.text)

    result = update_shift_operation_quantity(data["shift_operation_id"], quantity)

    if result is None:
        await message.answer("Не удалось изменить операцию.")
        await state.clear()
        return

    add_edit_log(
        message.from_user.id,
        "employee",
        "Изменил операцию",
        "shift",
        result["shift_id"],
        (
            f"{result['operation_name']}: "
            f"размер {result['product_size']}, цвет {result['product_color']}: "
            f"{result['old_quantity']} -> {result['new_quantity']} {result['unit']}"
        ),
    )

    report = get_shift_report(result["shift_id"])

    text = "Количество изменено.\n\nТекущий отчёт:\n\n"

    for index, row in enumerate(report, start=1):
        name, product_size, product_color, qty, unit = row
        text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

    await state.clear()
    await send_long_text(message, text, reply_markup=report_keyboard())


@dp.message(Command("delete"))
async def delete_report_operation(message: Message, state: FSMContext):
    employee = get_employee_by_telegram_id(message.from_user.id)

    if employee is None or employee[5] != "active":
        await message.answer("Сначала нужно зарегистрироваться и дождаться подтверждения.")
        return

    shift = get_editable_shift_for_today(employee[0])

    if shift is None:
        await message.answer(
            "Удаление недоступно.\n\n"
            "Можно удалять операции в открытой смене или в течение 1 часа после закрытия."
        )
        return

    operations = get_shift_operation_choices(shift[0])

    if not operations:
        await message.answer("В отчёте пока нет операций для удаления.")
        return

    operation_map = {}
    text = "Что удалить? Напишите номер строки:\n\n"

    for index, operation in enumerate(operations, start=1):
        shift_operation_id, name, product_size, product_color, quantity, unit = operation
        operation_line = format_operation_line(name, product_size, product_color, quantity, unit)
        operation_map[str(index)] = {
            "id": shift_operation_id,
            "name": operation_line,
        }
        text += f"{index}. {operation_line}\n"

    await state.update_data(shift_id=shift[0], operation_map=operation_map)
    await state.set_state(DeleteReport.waiting_for_operation)
    await message.answer(text, reply_markup=choice_keyboard("delete_operation", operation_map))


async def select_delete_operation(message: Message, state: FSMContext, selected_number: str):
    data = await state.get_data()
    operation_map = data.get("operation_map", {})

    if selected_number not in operation_map:
        await message.answer("Введите номер строки из списка.")
        return

    result = delete_shift_operation(operation_map[selected_number]["id"])

    if result is None:
        await message.answer("Не удалось удалить операцию.")
        await state.clear()
        return

    add_edit_log(
        message.from_user.id,
        "employee",
        "Удалил операцию",
        "shift",
        result["shift_id"],
        f"{result['operation_name']}, размер {result['product_size']}, цвет {result['product_color']} — {result['quantity']} {result['unit']}",
    )

    report = get_shift_report(result["shift_id"])

    text = "Операция удалена.\n\nТекущий отчёт:\n\n"

    if not report:
        text += "Операции ещё не добавлены."
    else:
        for index, row in enumerate(report, start=1):
            name, product_size, product_color, qty, unit = row
            text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

    await state.clear()
    await send_long_text(message, text, reply_markup=report_keyboard())


@dp.message(DeleteReport.waiting_for_operation)
async def delete_operation_selected(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    await select_delete_operation(message, state, message.text.strip())


@dp.callback_query(lambda callback: callback.data and callback.data.startswith("delete_operation:"))
async def delete_operation_selected_button(callback: CallbackQuery, state: FSMContext):
    if not await callback_state_is_current(callback, state, DeleteReport.waiting_for_operation):
        return

    await callback.answer()
    await select_delete_operation(callback.message, state, callback.data.rsplit(":", 1)[1])


@dp.message(Command("close"))
async def close_report_preview(message: Message):
    employee = get_employee_by_telegram_id(message.from_user.id)

    if employee is None or employee[5] != "active":
        await message.answer("Сначала нужно зарегистрироваться и дождаться подтверждения.")
        return

    shift = get_open_shift_for_today(employee[0])

    if shift is None:
        await message.answer("У вас нет открытой смены.")
        return

    report = get_shift_report(shift[0])

    if not report:
        await message.answer("Нельзя закрыть смену без добавленных операций.")
        return

    text = (
        "Проверьте итоговый отчёт:\n\n"
        f"Дата: {shift[2]}\n"
        f"Начало смены: {shift[3]}\n\n"
        "Операции:\n"
    )

    for index, row in enumerate(report, start=1):
        name, product_size, product_color, qty, unit = row
        text += f"{index}. {format_operation_line(name, product_size, product_color, qty, unit)}\n"

    text += "\nПодтвердить закрытие смены: /confirm_close"

    await message.answer(text, reply_markup=employee_keyboard())


@dp.message(Command("confirm_close"))
async def confirm_close_shift(message: Message):
    employee = get_employee_by_telegram_id(message.from_user.id)

    if employee is None or employee[5] != "active":
        await message.answer("Сначала нужно зарегистрироваться и дождаться подтверждения.")
        return

    shift = get_open_shift_for_today(employee[0])

    if shift is None:
        await message.answer("У вас нет открытой смены.")
        return

    result = close_shift(shift[0])

    if result is None:
        await message.answer("Не удалось закрыть смену.")
        return

    add_edit_log(
        message.from_user.id,
        "employee",
        "Закрыл смену",
        "shift",
        shift[0],
        f"Окончание: {result['end_time']}, отработано: {format_minutes(result['total_minutes'])}",
    )

    await message.answer(
        "Отчёт отправлен.\n\n"
        "Смена закрыта.\n"
        f"Окончание смены: {result['end_time']}\n"
        f"Отработано: {format_minutes(result['total_minutes'])}\n\n"
        "Редактирование сотрудником доступно в течение 1 часа."
    )


@dp.message(lambda message: message.text == "Отчёт")
async def report_menu_button(message: Message):
    await message.answer("Выберите действие с отчётом:", reply_markup=report_keyboard())


@dp.message(lambda message: message.text == "Отправить отчёт")
async def report_button(message: Message, state: FSMContext):
    await start_report(message, state)


@dp.message(lambda message: message.text == "Текущий отчёт")
async def current_button(message: Message):
    await current_report(message)


@dp.message(lambda message: message.text == "Отчёт за даты")
async def my_period_report_button(message: Message, state: FSMContext):
    await state.set_state(DateReport.waiting_for_period)
    await message.answer(
        "Введите даты отчёта через пробел.\n\n"
        "Пример: 2026-06-01 2026-06-25",
        reply_markup=report_keyboard(),
    )


@dp.message(lambda message: message.text == "Закрыть смену")
async def close_button(message: Message):
    await confirm_close_shift(message)


@dp.message(lambda message: message.text == "Изменить отчёт")
async def edit_button(message: Message, state: FSMContext):
    await edit_report(message, state)


@dp.message(lambda message: message.text == "Удалить операцию")
async def delete_button(message: Message, state: FSMContext):
    await delete_report_operation(message, state)


@dp.message(lambda message: message.text == "Назад")
async def back_to_employee_menu(message: Message):
    await message.answer("Основное меню:", reply_markup=employee_keyboard())


@dp.message(lambda message: message.text == "Админ-панель")
async def admin_panel_button(message: Message):
    await admin_menu(message)


@dp.message(CommandStart())
async def start(message: Message, state: FSMContext):
    employee = get_employee_by_telegram_id(message.from_user.id)

    if employee is None:
        await message.answer(
            "Здравствуйте. Вы ещё не зарегистрированы.\n\n"
            "Введите ваше ФИО:"
        )
        await state.set_state(Registration.waiting_for_full_name)
        return

    full_name = employee[2]
    status = employee[5]

    if status == "pending":
        await message.answer(
            f"{full_name}, ваша заявка ожидает подтверждения администратора."
        )
        return

    if status == "active":
        employee_id = employee[0]
        existing_shift = get_shift_for_today(employee_id)

        if existing_shift is not None and existing_shift[5] == "closed":
            await message.answer(
                f"Здравствуйте, {full_name}.\n\n"
                "Ваша смена за сегодня уже закрыта.\n"
                f"Дата: {existing_shift[2]}\n"
                f"Начало: {existing_shift[3]}\n"
                f"Окончание: {existing_shift[4]}\n\n"
                "Новая смена сегодня уже не создаётся.\n"
                "Посмотреть отчёт можно через кнопку «Отчёт».",
                reply_markup=employee_keyboard(),
            )
            return

        shift = get_open_shift_for_today(employee_id)

        if shift is None:
            shift = create_shift(employee_id)

            add_edit_log(
                message.from_user.id,
                "employee",
                "Открыл смену",
                "shift",
                shift["id"],
                f"Дата: {shift['shift_date']}, начало: {shift['start_time']}",
            )

            await message.answer(
                f"Здравствуйте, {full_name}.\n\n"
                "Смена открыта.\n"
                f"Дата: {shift['shift_date']}\n"
                f"Время начала: {shift['start_time']}\n\n"
                "Выберите действие в меню ниже.\n"
                "Все действия с отчётом находятся в кнопке «Отчёт».",
                reply_markup=employee_keyboard(),
            )
            return

        await message.answer(
            f"Здравствуйте, {full_name}.\n\n"
            "Ваша смена уже открыта.\n"
            f"Дата: {shift[2]}\n"
            f"Время начала: {shift[3]}\n\n"
            "Выберите действие в меню ниже.\n"
            "Все действия с отчётом находятся в кнопке «Отчёт».",
            reply_markup=employee_keyboard(),
        )
        return

    await message.answer("Ваш профиль не активен. Обратитесь к администратору.")


@dp.message(Registration.waiting_for_full_name)
async def process_full_name(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    full_name = message.text.strip()

    if len(full_name.split()) < 2:
        await message.answer("Введите ФИО полностью, например: Иванова Мария Сергеевна")
        return

    await state.update_data(full_name=full_name)
    await state.set_state(Registration.waiting_for_position)

    await message.answer(
        "Выберите должность. Напишите номер:\n\n"
        "1. Швея\n"
        "2. Упаковщик\n"
        "3. Раскройщик"
    )


@dp.message(Registration.waiting_for_position)
async def process_position(message: Message, state: FSMContext):
    if await reset_state_if_command(message, state):
        return

    selected_position = message.text.strip()

    if selected_position not in POSITIONS:
        await message.answer("Введите номер должности: 1, 2 или 3.")
        return

    data = await state.get_data()
    full_name = data["full_name"]
    position = POSITIONS[selected_position]

    create_employee(message.from_user.id, full_name, position)

    await message.answer(
        "Регистрация принята.\n\n"
        f"ФИО: {full_name}\n"
        f"Должность: {position}\n\n"
        "Ваша заявка отправлена администратору. "
        "После подтверждения вы сможете открывать смену."
    )

    for admin_id in get_admin_ids():
        await bot.send_message(
            admin_id,
            "Новая заявка на регистрацию:\n\n"
            f"ФИО: {full_name}\n"
            f"Должность: {position}\n"
            f"Telegram ID: {message.from_user.id}\n\n"
            "Проверить заявки: /pending"
        )

    await state.clear()


async def main():
    setup_logging()
    init_db()
    create_daily_database_backup()
    logging.info("Bot started")
    await dp.start_polling(bot)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception:
        setup_logging()
        logging.exception("Bot stopped with critical error")
        raise
    
