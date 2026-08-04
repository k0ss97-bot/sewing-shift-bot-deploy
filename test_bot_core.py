import importlib
import base64
import io
import json
import os
import sqlite3
import sys
import tempfile
import unittest
from concurrent.futures import ThreadPoolExecutor
from datetime import timedelta
from pathlib import Path
from unittest.mock import patch


PROJECT_DIR = Path(__file__).resolve().parent


class IsolatedDatabaseTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.old_db_dir = os.environ.get("DB_DIR")
        self.old_admin_ids = os.environ.get("ADMIN_IDS")
        self.old_wms_auto_receipt = os.environ.get("WMS_AUTO_RECEIPT_ENABLED")
        self.old_cwd = os.getcwd()
        os.environ["DB_DIR"] = self.temp_dir.name
        os.environ["WMS_AUTO_RECEIPT_ENABLED"] = "0"
        os.chdir(self.temp_dir.name)

        for module_name in ["database", "catalog", "route_maps", "miniapp_server", "webapp_auth"]:
            sys.modules.pop(module_name, None)

        sys.path.insert(0, str(PROJECT_DIR))
        self.database = importlib.import_module("database")
        self.database.init_db()

    def tearDown(self):
        if self.old_db_dir is None:
            os.environ.pop("DB_DIR", None)
        else:
            os.environ["DB_DIR"] = self.old_db_dir

        if self.old_admin_ids is None:
            os.environ.pop("ADMIN_IDS", None)
        else:
            os.environ["ADMIN_IDS"] = self.old_admin_ids

        if self.old_wms_auto_receipt is None:
            os.environ.pop("WMS_AUTO_RECEIPT_ENABLED", None)
        else:
            os.environ["WMS_AUTO_RECEIPT_ENABLED"] = self.old_wms_auto_receipt

        if str(PROJECT_DIR) in sys.path:
            sys.path.remove(str(PROJECT_DIR))

        os.chdir(self.old_cwd)
        for module_name in ["database", "catalog", "route_maps", "miniapp_server", "webapp_auth"]:
            sys.modules.pop(module_name, None)

        self.temp_dir.cleanup()

    def test_product_packing_options_are_isolated(self):
        route_maps = importlib.import_module("route_maps")

        def option_ids(product_name: str):
            packing_step = next(
                route_step
                for route_step in route_maps.PRODUCT_ROUTE_MAPS[product_name]
                if route_step["operation"] == "Упаковка"
            )
            return packing_step, {option["id"] for option in packing_step["packing_options"]}

        tshirt_step, tshirt_options = option_ids("Футболки")
        leggings_step, leggings_options = option_ids("Легинсы")
        joggers_step, joggers_options = option_ids("Брюки-джоггеры")
        pants_step, pants_options = option_ids("Брюки со стрелками детские")
        cardigan_step, cardigan_options = option_ids("Кардиган")
        shorts_step, shorts_options = option_ids("Шорты")

        self.assertEqual(tshirt_options, {"individual", "tshirt_3"})
        self.assertEqual(leggings_options, {"individual", "leggings_3"})
        self.assertEqual(joggers_options, {"individual", "joggers_2", "joggers_sweatshirt", "suit"})
        self.assertEqual(pants_options, {"individual", "suit"})
        self.assertEqual(cardigan_options, {"individual", "suit"})
        self.assertEqual(shorts_options, {"individual"})
        self.assertEqual(len({id(step) for step in (tshirt_step, leggings_step, joggers_step, pants_step, cardigan_step, shorts_step)}), 6)

    def test_packaging_is_the_final_production_step(self):
        route_maps = importlib.import_module("route_maps")

        for product_name, steps in route_maps.PRODUCT_ROUTE_MAPS.items():
            self.assertEqual(steps[-1]["operation"], "Упаковка", product_name)
            self.assertNotIn("Размещение на склад", [step["operation"] for step in steps])

    def test_legacy_route_snapshot_skips_production_putaway_after_packaging(self):
        miniapp_server = importlib.import_module("miniapp_server")
        legacy_steps = [
            {"position": "Упаковщик", "operation": "Упаковка", "status_after": "Упаковано"},
            {"position": "Упаковщик", "operation": "Размещение на склад", "status_after": "На складе"},
        ]
        batch = {
            "product_name": "Легинсы",
            "route_step_index": 0,
            "route_snapshot": json.dumps(legacy_steps, ensure_ascii=False),
        }

        next_index, next_step = miniapp_server.get_route_flow_next_step(batch)

        self.assertEqual(next_index, 2)
        self.assertIsNone(next_step)
        self.assertEqual(
            miniapp_server.get_route_previous_status(
                {**batch, "status": "done", "route_step_index": 1}
            ),
            "Упаковано",
        )

    def test_opening_shift_returns_close_shift_reminder_once(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(4120, "Тест Сотрудник", "Швея")
        employee = self.database.get_employee_by_telegram_id(4120)
        self.database.update_employee_status(employee[0], "active")

        opened = miniapp_server.open_shift_for_telegram(4120)
        repeated = miniapp_server.open_shift_for_telegram(4120)

        self.assertTrue(opened["has_open_shift"])
        self.assertIn("обязательно закройте", opened["shift_close_reminder"])
        self.assertNotIn("shift_close_reminder", repeated)

    def test_admin_web_push_subscription_test_and_disable(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(4121, "Тест Администратор Push", "Швея")
        admin = self.database.get_employee_by_telegram_id(4121)
        self.database.update_employee_status(admin[0], "active")
        self.assertTrue(self.database.update_employee_role(admin[0], "admin")["ok"])
        subscription = {
            "endpoint": "https://push.example.test/subscription-4121",
            "keys": {"p256dh": "synthetic-p256dh", "auth": "synthetic-auth"},
        }

        with patch.object(
            miniapp_server,
            "get_public_web_push_config",
            return_value={"configured": True, "public_key": "synthetic-public-key"},
        ):
            enabled = miniapp_server.subscribe_web_push_for_admin(
                4121, {"subscription": subscription}, "Synthetic Test Browser"
            )
            with patch.object(miniapp_server, "send_web_push") as sender:
                delivered = miniapp_server.send_test_web_push_for_admin(4121)

        self.assertTrue(enabled["ok"], enabled)
        self.assertTrue(enabled["subscription"]["active"])
        self.assertTrue(delivered["ok"], delivered)
        sender.assert_called_once()
        disabled = miniapp_server.unsubscribe_web_push_for_admin(
            4121, {"subscription": subscription}
        )
        self.assertTrue(disabled["ok"], disabled)
        self.assertFalse(disabled["subscription"]["active"])

    def test_critical_notification_delivery_is_idempotent(self):
        self.database.create_employee(4122, "Тест Администратор Alerts", "Швея")
        admin = self.database.get_employee_by_telegram_id(4122)
        self.database.update_employee_status(admin[0], "active")
        self.assertTrue(self.database.update_employee_role(admin[0], "admin")["ok"])
        notification_id = self.database.create_or_refresh_operational_notification(
            "synthetic-critical-alert",
            "Тестовое критичное уведомление",
            "Только изолированная проверка.",
            severity="critical",
        )

        self.assertTrue(
            self.database.claim_notification_delivery(notification_id, admin[0], "webpush:test")
        )
        self.assertTrue(
            self.database.finish_notification_delivery(
                notification_id, admin[0], "webpush:test", sent=True
            )
        )
        self.assertFalse(
            self.database.claim_notification_delivery(notification_id, admin[0], "webpush:test")
        )

    def test_only_selected_active_employee_or_admin_can_access_wms(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(4101, "Тест Упаковщик WMS", "Упаковщик")
        storekeeper = self.database.get_employee_by_telegram_id(4101)
        self.database.update_employee_status(storekeeper[0], "active")

        self.database.create_employee(4102, "Тест Упаковщик без WMS", "Упаковщик")
        second_packer = self.database.get_employee_by_telegram_id(4102)
        self.database.update_employee_status(second_packer[0], "active")

        self.assertTrue(self.database.update_employee_wms_access(storekeeper[0], True)["ok"])
        self.assertTrue(miniapp_server.can_access_wms(4101))
        self.assertFalse(miniapp_server.can_access_wms(4102))
        self.assertEqual(storekeeper[3], "Упаковщик")
        self.database.update_employee_status(storekeeper[0], "inactive")
        self.assertFalse(miniapp_server.can_access_wms(4101))

    def test_admin_can_grant_and_revoke_wms_without_changing_position(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(4110, "Тест Администратор", "Швея")
        admin = self.database.get_employee_by_telegram_id(4110)
        self.assertTrue(self.database.update_employee_role(admin[0], "admin")["ok"])

        self.database.create_employee(4111, "Тест Упаковщик", "Упаковщик")
        packer = self.database.get_employee_by_telegram_id(4111)
        self.database.update_employee_status(packer[0], "active")

        granted = miniapp_server.set_employee_wms_access_for_admin(
            4110, {"employee_id": packer[0], "enabled": True}
        )
        self.assertTrue(granted["ok"])
        granted_row = next(row for row in granted["user_accounts"] if row["id"] == packer[0])
        self.assertTrue(granted_row["can_access_wms"])
        self.assertEqual(granted_row["position"], "Упаковщик")
        self.assertTrue(miniapp_server.can_access_wms(4111))

        revoked = miniapp_server.set_employee_wms_access_for_admin(
            4110, {"employee_id": packer[0], "enabled": False}
        )
        self.assertTrue(revoked["ok"])
        self.assertFalse(miniapp_server.can_access_wms(4111))

    def test_employee_wms_permission_column_is_additive_and_defaults_off(self):
        conn = self.database.get_db_connection()
        columns = {row[1] for row in conn.execute("PRAGMA table_info(employees)").fetchall()}
        self.assertIn("can_access_wms", columns)
        conn.close()

        self.database.create_employee(4112, "Тест Без Склада", "Упаковщик")
        employee = self.database.get_employee_by_telegram_id(4112)
        access_map = self.database.get_employee_wms_access_map([employee[0]])
        self.assertFalse(access_map[employee[0]])
        self.assertEqual(employee[3], "Упаковщик")

    def route_step_index(self, product_name: str, operation_name: str, position: str | None = None):
        route_maps = importlib.import_module("route_maps")

        for index, route_step in enumerate(route_maps.PRODUCT_ROUTE_MAPS[product_name]):
            if route_step["operation"] == operation_name and (position is None or route_step["position"] == position):
                return index

        self.fail(f"Route step not found: {product_name} / {operation_name}")

    def test_database_uses_isolated_path_and_indexes(self):
        self.assertEqual(Path(self.database.DB_NAME).parent, Path(self.temp_dir.name))
        self.assertEqual(self.database.get_backup_dir(), str(Path(self.temp_dir.name) / "backups"))

        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type = 'index' AND name LIKE 'idx_%'"
        )
        indexes = {row[0] for row in cursor.fetchall()}
        conn.close()

        self.assertIn("idx_shifts_employee_date", indexes)
        self.assertIn("idx_shifts_one_open", indexes)
        self.assertIn("idx_operations_navigation", indexes)
        self.assertIn("idx_cutting_batches_product_status", indexes)
        self.assertIn("idx_cutting_batches_active_task", indexes)
        self.assertIn("idx_route_batch_history_step", indexes)
        self.assertIn("idx_feedback_entries_date", indexes)
        self.assertIn("idx_feedback_entries_employee_date", indexes)
        self.assertIn("idx_route_batches_status_step", indexes)
        self.assertIn("idx_route_batch_history_batch", indexes)
        self.assertIn("idx_route_batch_inputs_batch", indexes)
        self.assertIn("idx_route_batch_inputs_stock", indexes)
        self.assertIn("idx_cutting_batches_task", indexes)
        self.assertIn("idx_cutting_batch_matrix_batch", indexes)
        self.assertIn("idx_cutting_batch_arbitrary_batch", indexes)
        self.assertIn("idx_fabric_stock_color", indexes)
        self.assertIn("idx_production_tasks_status", indexes)
        self.assertIn("idx_production_task_items_task", indexes)
        self.assertIn("idx_production_task_fabric_rolls_task", indexes)
        self.assertIn("idx_warehouse_stock_type_position", indexes)
        self.assertIn("idx_warehouse_stock_product", indexes)
        self.assertIn("idx_route_batches_trace_code", indexes)
        self.assertIn("idx_production_tasks_trace_code", indexes)
        self.assertIn("idx_route_batch_input_lots_batch", indexes)
        self.assertIn("idx_production_task_fabric_lots_task", indexes)
        self.assertIn("idx_trace_events_batch", indexes)

    def test_init_db_repairs_legacy_duplicates_before_unique_indexes(self):
        self.database.create_employee(1099, "Тест Старые Дубли", "Раскройщик")
        employee = self.database.get_employee_by_telegram_id(1099)
        self.database.update_employee_status(employee[0], "active")
        first_shift = self.database.create_shift(employee[0])
        task = self.database.create_production_task("Легинсы", ["86"], ["Бежевый"], None)
        route_batch = self.database.create_route_batch("Легинсы", "86", "Бежевый", 5, None, 4)
        now = self.database.local_now().isoformat()

        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("DROP INDEX idx_shifts_one_open")
        cursor.execute("DROP INDEX idx_cutting_batches_active_task")
        cursor.execute("DROP INDEX idx_route_batch_history_step")
        cursor.execute(
            """
            INSERT INTO shifts (employee_id, shift_date, start_time, status, created_at)
            VALUES (?, ?, '23:59', 'open', ?)
            """,
            (employee[0], self.database.local_today().isoformat(), now),
        )
        cursor.execute(
            """
            INSERT INTO cutting_batches (product_name, production_task_id, status, created_at, updated_at)
            VALUES ('Легинсы', ?, 'contours_done', ?, ?)
            """,
            (task["id"], now, now),
        )
        cursor.execute(
            """
            INSERT INTO cutting_batches (product_name, production_task_id, status, created_at, updated_at)
            VALUES ('Легинсы', ?, 'layout_done', ?, ?)
            """,
            (task["id"], now, now),
        )
        for quantity in (4, 5):
            cursor.execute(
                """
                INSERT INTO route_batch_history (
                    batch_id, step_index, operation_name, position,
                    employee_id, quantity, completed_at
                )
                VALUES (?, 4, 'Тест', 'Швея', ?, ?, ?)
                """,
                (route_batch["id"], employee[0], quantity, now),
            )
        conn.commit()
        conn.close()

        self.database.init_db()

        conn = sqlite3.connect(self.database.DB_NAME)
        open_count = conn.execute(
            "SELECT COUNT(*) FROM shifts WHERE employee_id = ? AND status = 'open'",
            (employee[0],),
        ).fetchone()[0]
        active_cutting_count = conn.execute(
            "SELECT COUNT(*) FROM cutting_batches WHERE production_task_id = ? AND status != 'cancelled'",
            (task["id"],),
        ).fetchone()[0]
        history_count = conn.execute(
            "SELECT COUNT(*) FROM route_batch_history WHERE batch_id = ? AND step_index = 4",
            (route_batch["id"],),
        ).fetchone()[0]
        closed_first_shift = conn.execute(
            "SELECT status FROM shifts WHERE id = ?",
            (first_shift["id"],),
        ).fetchone()[0]
        repair_count = conn.execute(
            "SELECT COUNT(*) FROM data_repairs WHERE repair_key LIKE 'deduplicate-%'",
        ).fetchone()[0]
        conn.close()

        self.assertEqual(open_count, 1)
        self.assertEqual(active_cutting_count, 1)
        self.assertEqual(history_count, 1)
        self.assertEqual(closed_first_shift, "closed")
        self.assertGreaterEqual(repair_count, 3)

    def test_catalog_seed_contains_expected_operations(self):
        operations = self.database.get_active_operations(position="Упаковщик", folder="Нарезание резинки")
        operation_names = {operation[2] for operation in operations}

        self.assertIn("Ползунки — резинка 25 мм", operation_names)
        self.assertIn("Шорты — резинка 25 мм", operation_names)
        self.assertIn("80 (43 см)", self.database.get_preparation_operation_sizes("Шорты — резинка 25 мм"))

        dublerin_operations = self.database.get_active_operations(
            position="Упаковщик",
            folder="Нарезание дублерина",
        )
        self.assertIn("Кардиганы — дублерин 25 мм", {operation[2] for operation in dublerin_operations})
        self.assertEqual(
            self.database.get_preparation_operation_sizes("Кардиганы — дублерин 25 мм"),
            [
                "92 (100 см)",
                "98 (100 см)",
                "104 (100 см)",
                "110 (100 см)",
                "116 (100 см)",
                "122 (100 см)",
                "128 (100 см)",
                "134 (100 см)",
                "140 (100 см)",
                "146 (100 см)",
                "152 (100 см)",
                "158 (100 см)",
                "164 (100 см)",
            ],
        )
        self.assertEqual(
            self.database.get_preparation_operation_colors("Кардиганы — дублерин 25 мм"),
            ["Черный", "Белый"],
        )

        for operation_name in ["Установка петель", "Установка пуговиц"]:
            preparation_operations = self.database.get_active_operations(
                position="Упаковщик",
                folder=operation_name,
            )
            self.assertEqual([operation[2] for operation in preparation_operations], [operation_name])
            self.assertTrue(self.database.is_preparation_operation_folder(operation_name))
            self.assertEqual(
                self.database.get_preparation_operation_sizes(operation_name),
                ["92", "98", "104", "110", "116", "122", "128", "134", "140", "146", "152", "158", "164"],
            )
            self.assertEqual(
                self.database.get_preparation_operation_colors(operation_name),
                ["Голубой", "Бежевый", "Капучино", "Темно-синий", "Брауни", "Светло-серый", "Черный"],
            )

    def test_route_maps_cover_catalog_products_and_operations(self):
        catalog = importlib.import_module("catalog")
        route_maps = importlib.import_module("route_maps")
        known_operations = {
            *catalog.CUTTING_OPERATIONS,
            *catalog.PACKING_OPERATIONS,
            *(operation[2] for operation in catalog.PRODUCTION_OPERATIONS),
        }

        self.assertEqual(set(route_maps.PRODUCT_ROUTE_MAPS), set(catalog.CUTTING_PRODUCTS))

        for product_name, steps in route_maps.PRODUCT_ROUTE_MAPS.items():
            self.assertGreater(len(steps), 0, product_name)

            for route_step in steps:
                self.assertIn(route_step["operation"], known_operations, product_name)
                self.assertIn(route_step["position"], {"Раскройщик", "Упаковщик", "Швея"}, product_name)
                self.assertTrue(route_step["status_after"], product_name)

    def test_legacy_cardigan_route_names_keep_catalog_options(self):
        miniapp_server = importlib.import_module("miniapp_server")

        catalog = {
            item["product_name"]: item
            for item in miniapp_server.production_catalog_to_dict()
        }

        self.assertEqual(
            catalog["Кардиган детский"]["sizes"],
            catalog["Кардиган"]["sizes"],
        )
        self.assertEqual(
            catalog["Кардиган детский"]["colors"],
            catalog["Кардиган"]["colors"],
        )
        self.assertTrue(catalog["Брюки со стрелками детские"]["sizes"])
        self.assertTrue(catalog["Брюки со стрелками детские"]["colors"])

        route_catalog = {
            item["product_name"]: item
            for item in miniapp_server.get_route_catalog()
        }
        self.assertEqual(
            route_catalog["Кардиган подростковый"]["sizes"],
            catalog["Кардиган"]["sizes"],
        )

    def test_route_batch_flow_advances_and_writes_history(self):
        route_maps = importlib.import_module("route_maps")
        self.database.create_employee(4004, "Тест Маршрут", "Раскройщик")
        employee = self.database.get_employee_by_telegram_id(4004)
        employee_id = employee[0]
        self.database.update_employee_status(employee_id, "active")

        batch = self.database.create_route_batch(
            "Шорты",
            "110",
            "Голубой",
            12,
            employee_id,
        )
        first_step = route_maps.PRODUCT_ROUTE_MAPS["Шорты"][0]

        self.assertIsNotNone(batch)
        self.assertEqual(batch["route_step_index"], 0)
        self.assertEqual(batch["status"], "active")

        updated = self.database.complete_route_batch_step(
            batch["id"],
            employee_id,
            first_step["operation"],
            first_step["position"],
            1,
            "active",
        )

        self.assertEqual(updated["route_step_index"], 1)
        self.assertEqual(updated["status"], "active")

        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT operation_name, position, quantity FROM route_batch_history WHERE batch_id = ?",
            (batch["id"],),
        )
        self.assertEqual(cursor.fetchone(), (first_step["operation"], first_step["position"], 12))
        conn.close()

    def test_period_employee_performance_includes_every_production_position(self):
        today = self.database.local_today().isoformat()
        employees = {}
        for telegram_id, full_name, position in (
            (4101, "Тест Раскрой Отчёт", "Раскройщик"),
            (4102, "Тест Швея Отчёт", "Швея"),
            (4103, "Тест Упаковка Отчёт", "Упаковщик"),
        ):
            self.database.create_employee(telegram_id, full_name, position)
            employee = self.database.get_employee_by_telegram_id(telegram_id)
            self.database.update_employee_status(employee[0], "active")
            employees[position] = employee[0]

        cutter_shift = self.database.create_shift(employees["Раскройщик"])
        task = self.database.create_production_task("Легинсы", ["86"], ["Бежевый"], None)
        batch = self.database.create_route_batch("Легинсы", "86", "Бежевый", 10, None, 4)
        now = self.database.local_now().isoformat()

        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute(
            """
            UPDATE production_task_items
            SET contour_quantity = 8, formed_quantity = 8
            WHERE task_id = ?
            """,
            (task["id"],),
        )
        cursor.execute(
            """
            INSERT INTO cutting_batches (
                product_name, production_task_id, status,
                contour_shift_id, contour_employee_id, contour_date,
                layout_shift_id, layout_employee_id, layout_date,
                cutting_shift_id, cutting_employee_id, cutting_progress,
                formed_shift_id, formed_employee_id, formed_date,
                created_at, updated_at
            )
            VALUES (?, ?, 'formed', ?, ?, ?, ?, ?, ?, ?, ?, 100, ?, ?, ?, ?, ?)
            """,
            (
                "Легинсы",
                task["id"],
                cutter_shift["id"], employees["Раскройщик"], today,
                cutter_shift["id"], employees["Раскройщик"], today,
                cutter_shift["id"], employees["Раскройщик"],
                cutter_shift["id"], employees["Раскройщик"], today,
                now, now,
            ),
        )
        cursor.executemany(
            """
            INSERT INTO route_batch_history (
                batch_id, step_index, operation_name, position,
                employee_id, quantity, completed_at
            )
            VALUES (?, ?, ?, ?, ?, 9, ?)
            """,
            (
                (batch["id"], 4, "Тестовая швейная операция", "Швея", employees["Швея"], now),
                (batch["id"], 5, "Тестовая упаковочная операция", "Упаковщик", employees["Упаковщик"], now),
            ),
        )
        cursor.execute("UPDATE route_batches SET status = 'done' WHERE id = ?", (batch["id"],))
        conn.commit()
        conn.close()

        performance = {
            position: (plan, fact)
            for _employee_id, _full_name, position, plan, fact
            in self.database.get_period_employee_production_performance(today, today)
        }

        self.assertEqual(performance["Раскройщик"], (32, 32))
        self.assertEqual(performance["Швея"], (10, 9))
        self.assertEqual(performance["Упаковщик"], (10, 9))

    def test_cutting_batch_flow_multiplies_sizes_by_layers(self):
        batch_id = self._create_layout_done_batch()

        self.assertTrue(
            self.database.update_cutting_batch_progress(
                batch_id,
                self.shift_id,
                self.employee_id,
                self.operation_id,
                100,
            )
        )

        rows = self.database.get_cutting_batch_result_rows(batch_id)

        self.assertEqual(
            rows,
            [
                ("80", "Бежевый", 20),
                ("80", "Синий", 10),
                ("86", "Бежевый", 30),
                ("86", "Синий", 15),
            ],
        )

    def test_cutting_progress_is_atomic_idempotent_and_backfilled(self):
        batch_id = self._create_layout_done_batch()
        cutting_operation = next(
            operation
            for operation in self.database.get_active_operations(
                position="Раскройщик",
                folder="Брюки-ползунки",
            )
            if operation[2] == "Раскрой"
        )
        cutting_operation_id = cutting_operation[0]

        self.assertEqual(cutting_operation[3], "%")
        self.assertTrue(
            self.database.update_cutting_batch_progress(
                batch_id,
                self.shift_id,
                self.employee_id,
                cutting_operation_id,
                100,
            )
        )
        self.assertTrue(
            self.database.update_cutting_batch_progress(
                batch_id,
                self.shift_id,
                self.employee_id,
                cutting_operation_id,
                100,
            )
        )
        self.assertIn(
            ("Брюки-ползунки: Раскрой", f"партия #{batch_id}", "без цвета", 100, "%"),
            self.database.get_shift_report(self.shift_id),
        )

        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute(
            """
            DELETE FROM shift_operations
            WHERE shift_id = ?
              AND operation_id = ?
              AND product_size = ?
              AND product_color = 'без цвета'
            """,
            (self.shift_id, cutting_operation_id, f"партия #{batch_id}"),
        )
        conn.commit()
        conn.close()

        self.database.init_db()
        self.assertIn(
            ("Брюки-ползунки: Раскрой", f"партия #{batch_id}", "без цвета", 100, "%"),
            self.database.get_shift_report(self.shift_id),
        )

    def test_cutting_batch_at_75_percent_stays_visible_for_cutter(self):
        miniapp_server = importlib.import_module("miniapp_server")
        batch_id = self._create_layout_done_batch()

        self.assertTrue(
            self.database.update_cutting_batch_progress(
                batch_id,
                self.shift_id,
                self.employee_id,
                self.operation_id,
                75,
            )
        )

        state = miniapp_server.get_production_state_for_telegram(1001)
        cutting_tasks = [
            task
            for task in state["cutting_tasks"]
            if task["id"] == batch_id and task["stage"] == "cutting"
        ]

        self.assertEqual(len(cutting_tasks), 1)
        self.assertEqual(cutting_tasks[0]["progress"], 75)
        self.assertEqual(cutting_tasks[0]["status_text"], "В работе")
        self.assertEqual(cutting_tasks[0]["process_status_text"], "готовность 75%")
        self.assertTrue(cutting_tasks[0]["is_assigned_to_me"])

    def test_incomplete_cutting_survives_closed_shift_and_can_be_reopened_once(self):
        batch_id = self._create_layout_done_batch()
        shift_date = "2026-01-01"
        self.database.close_shift(self.shift_id)

        self.assertEqual(
            [row[0] for row in self.database.get_cutting_batches_for_cutting("Брюки-ползунки")],
            [batch_id],
        )

        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE shifts SET shift_date = ? WHERE id = ?",
            (shift_date, self.shift_id),
        )
        restored = self.database.restore_incomplete_cutting_shift(
            cursor,
            repair_key="test_restore_incomplete_cutting",
            batch_id=batch_id,
            shift_id=self.shift_id,
            expected_shift_date=shift_date,
        )
        restored_again = self.database.restore_incomplete_cutting_shift(
            cursor,
            repair_key="test_restore_incomplete_cutting",
            batch_id=batch_id,
            shift_id=self.shift_id,
            expected_shift_date=shift_date,
        )
        conn.commit()
        cursor.execute(
            "SELECT status, end_time, total_minutes, closed_at FROM shifts WHERE id = ?",
            (self.shift_id,),
        )
        shift = cursor.fetchone()
        conn.close()

        self.assertTrue(restored)
        self.assertFalse(restored_again)
        self.assertEqual(shift, ("open", None, None, None))
        self.assertEqual(
            self.database.get_open_shift_for_today(self.employee_id)[0],
            self.shift_id,
        )

    def test_production_task_creates_matrix_cutting_batch(self):
        task = self.database.create_production_task(
            "Шорты",
            ["80", "92"],
            ["Бежевый", "Синий"],
            None,
        )
        self.assertIsNotNone(task)

        batch_id = self.database.create_cutting_contour_batch_for_task(
            task["id"],
            "Шорты",
            1,
            1,
            1,
            {
                ("80", "Бежевый"): 2,
                ("92", "Бежевый"): 3,
                ("80", "Синий"): 4,
                ("92", "Синий"): 0,
            },
        )

        self.assertIsNotNone(batch_id)
        self.assertEqual(self.database.get_production_task_by_id(task["id"])["status"], "contours_done")
        self.assertEqual(
            self.database.get_cutting_batch_task_options(batch_id),
            (["80", "92"], ["Бежевый", "Синий"]),
        )
        self.assertEqual(
            self.database.get_cutting_batch_contour_matrix(batch_id),
            [
                {"size": "80", "color": "Бежевый", "quantity": 2},
                {"size": "80", "color": "Синий", "quantity": 4},
                {"size": "92", "color": "Бежевый", "quantity": 3},
            ],
        )
        self.assertEqual(len(self.database.get_cutting_batches_for_layout("Шорты")[0]), 7)

        self.assertTrue(
            self.database.add_cutting_layout(
                batch_id,
                1,
                1,
                1,
                {"Бежевый": 3, "Синий": 2},
            )
        )
        self.assertEqual(self.database.get_production_task_by_id(task["id"])["status"], "in_cutting")
        preparation_batches = self.database.get_active_route_batches()
        elastic_step_index = self.route_step_index("Шорты", "Шорты — резинка 25 мм", "Упаковщик")
        self.assertEqual({batch["route_step_index"] for batch in preparation_batches}, {elastic_step_index})
        self.assertEqual(
            sorted((batch["product_size"], batch["product_color"], batch["quantity"]) for batch in preparation_batches),
            [
                ("80 (43 см)", "Черный", 14),
                ("92 (45 см)", "Черный", 9),
            ],
        )

        self.assertTrue(self.database.update_cutting_batch_progress(batch_id, 1, 1, 1, 100))
        self.assertEqual(
            self.database.get_cutting_batch_result_rows(batch_id),
            [
                ("80", "Бежевый", 6),
                ("80", "Синий", 8),
                ("92", "Бежевый", 9),
            ],
        )

        self.assertTrue(self.database.mark_cutting_batch_formed(batch_id, 1, 1, 1))
        self.assertEqual(self.database.get_production_task_by_id(task["id"])["status"], "formed")

        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT product_size, product_color, contour_quantity, formed_quantity
            FROM production_task_items
            WHERE task_id = ?
            ORDER BY product_color, CAST(product_size AS INTEGER)
            """,
            (task["id"],),
        )
        self.assertEqual(
            cursor.fetchall(),
            [
                ("80", "Бежевый", 2, 6),
                ("92", "Бежевый", 3, 9),
                ("80", "Синий", 4, 8),
                ("92", "Синий", 0, 0),
            ],
        )
        conn.close()
        warehouse_rows = self.database.get_warehouse_stock_rows()
        self.assertEqual(
            {
                (row["product_size"], row["product_color"], row["quantity"], row["stage_name"], row["ready_for_position"])
                for row in warehouse_rows
            },
            {
                ("80", "Бежевый", 6, "Раскроенные", "Швея"),
                ("92", "Бежевый", 9, "Раскроенные", "Швея"),
                ("80", "Синий", 8, "Раскроенные", "Швея"),
            },
        )
        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT product_size, product_color, quantity, movement_type, source_type, source_id
            FROM warehouse_stock_movements
            ORDER BY product_color, CAST(product_size AS INTEGER)
            """
        )
        self.assertEqual(
            cursor.fetchall(),
            [
                ("80", "Бежевый", 6, "receipt", "cutting_batch", batch_id),
                ("92", "Бежевый", 9, "receipt", "cutting_batch", batch_id),
                ("80", "Синий", 8, "receipt", "cutting_batch", batch_id),
            ],
        )
        conn.close()

    def test_arbitrary_cutting_operation_adds_layers_to_common_output(self):
        task = self.database.create_production_task(
            "Шорты",
            ["80", "92"],
            ["Бежевый", "Синий"],
            None,
        )
        operations = self.database.get_active_operations("Раскройщик", "Шорты", "Раскрой изделий")
        operation_ids = {row[2]: row[0] for row in operations}
        self.assertIn("Нанесение контуров лекал на ткань", operation_ids)
        self.assertIn("Формирование настила", operation_ids)
        self.assertIn("Произвольная операция", operation_ids)

        batch_id = self.database.create_cutting_contour_batch_for_task(
            task["id"],
            "Шорты",
            1,
            1,
            operation_ids["Нанесение контуров лекал на ткань"],
            {
                ("80", "Бежевый"): 2,
                ("92", "Бежевый"): 3,
                ("80", "Синий"): 1,
            },
        )

        self.assertTrue(self.database.add_cutting_layout(
            batch_id,
            1,
            1,
            operation_ids["Формирование настила"],
            {"Бежевый": 3, "Синий": 2},
            arbitrary_operations=[{
                "product_size": "92",
                "product_color": "Бежевый",
                "layers": 5,
                "parts_count": 2,
            }],
            arbitrary_operation_id=operation_ids["Произвольная операция"],
        ))

        arbitrary_rows = self.database.get_cutting_batch_arbitrary_operations(batch_id)
        self.assertEqual(len(arbitrary_rows), 1)
        self.assertEqual(
            {key: arbitrary_rows[0][key] for key in ("product_size", "product_color", "layers", "parts_count", "quantity")},
            {"product_size": "92", "product_color": "Бежевый", "layers": 5, "parts_count": 2, "quantity": 5},
        )
        self.assertEqual(
            self.database.get_cutting_batch_result_rows(batch_id),
            [("80", "Бежевый", 6), ("80", "Синий", 2), ("92", "Бежевый", 14)],
        )

        conn = sqlite3.connect(self.database.DB_NAME)
        arbitrary_shift_quantity = conn.execute(
            "SELECT quantity FROM shift_operations WHERE operation_id = ?",
            (operation_ids["Произвольная операция"],),
        ).fetchone()[0]
        conn.close()
        self.assertEqual(arbitrary_shift_quantity, 5)

        self.assertTrue(self.database.update_cutting_batch_progress(batch_id, 1, 1, 1, 100))
        self.assertTrue(self.database.mark_cutting_batch_formed(batch_id, 1, 1, 1))
        conn = sqlite3.connect(self.database.DB_NAME)
        formed = conn.execute(
            "SELECT formed_quantity FROM production_task_items WHERE task_id = ? AND product_size = '92' AND product_color = 'Бежевый'",
            (task["id"],),
        ).fetchone()[0]
        conn.close()
        self.assertEqual(formed, 14)

    def test_layout_creates_only_dublerin_before_dubling(self):
        task = self.database.create_production_task(
            "Жакет для девочек",
            ["98"],
            ["Черный"],
            None,
        )
        batch_id = self.database.create_cutting_contour_batch_for_task(
            task["id"],
            "Жакет для девочек",
            1,
            1,
            1,
            {("98", "Черный"): 5},
        )

        self.assertTrue(
            self.database.add_cutting_layout(
                batch_id,
                1,
                1,
                1,
                {"Черный": 2},
            )
        )

        dublerin_index = self.route_step_index("Жакет для девочек", "Жакеты — дублерин 80 мм", "Упаковщик")
        dubling_index = self.route_step_index("Жакет для девочек", "Жакет для девочек — Дублирование", "Упаковщик")
        preparation_batches = self.database.get_active_route_batches()

        self.assertEqual(
            sorted((batch["route_step_index"], batch["product_size"], batch["product_color"], batch["quantity"]) for batch in preparation_batches),
            [
                (dublerin_index, "98 (28,5 см)", "Черный", 10),
            ],
        )

    def test_cardigan_layout_creates_two_dublerin_pieces_per_garment(self):
        expected = []
        for product_name, size in (("Кардиган", "98"), ("Кардиган", "134")):
            task = self.database.create_production_task(product_name, [size], ["Черный"], None)
            batch_id = self.database.create_cutting_contour_batch_for_task(
                task["id"], product_name, 1, 1, 1, {(size, "Черный"): 5}
            )
            self.assertTrue(self.database.add_cutting_layout(batch_id, 1, 1, 1, {"Черный": 2}))
            dublerin_index = self.route_step_index(product_name, "Кардиганы — дублерин 25 мм", "Упаковщик")
            expected.extend([
                (product_name, dublerin_index, f"{size} (100 см)", "Черный", 20),
            ])

        actual = sorted(
            (batch["product_name"], batch["route_step_index"], batch["product_size"], batch["product_color"], batch["quantity"])
            for batch in self.database.get_active_route_batches()
        )
        self.assertEqual(actual, sorted(expected))

    def test_dublerin_tasks_group_garment_colors_by_material_color(self):
        task = self.database.create_production_task(
            "Кардиган",
            ["98"],
            ["Брауни", "Черный"],
            None,
        )
        batch_id = self.database.create_cutting_contour_batch_for_task(
            task["id"],
            "Кардиган",
            1,
            1,
            1,
            {("98", "Брауни"): 2, ("98", "Черный"): 2},
        )
        self.assertTrue(self.database.add_cutting_layout(batch_id, 1, 1, 1, {"Брауни": 2, "Черный": 2}))

        dublerin_index = self.route_step_index("Кардиган", "Кардиганы — дублерин 25 мм", "Упаковщик")
        batches = self.database.get_active_route_batches()
        dublerin_rows = [row for row in batches if row["route_step_index"] == dublerin_index]
        self.assertEqual([(row["product_size"], row["product_color"], row["quantity"]) for row in dublerin_rows], [("98 (100 см)", "Черный", 16)])
        self.assertFalse(any(row["route_step_index"] == dublerin_index + 1 for row in batches))

    def test_miniapp_production_creates_task_and_submits_contours(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")

        self.database.create_employee(9002, "Тест Раскройщик Миниапп", "Раскройщик")
        cutter = self.database.get_employee_by_telegram_id(9002)
        cutter_id = cutter[0]
        self.database.update_employee_status(cutter_id, "active")
        shift = self.database.create_shift(cutter_id)

        fabric_result = miniapp_server.add_fabric_receipt_for_telegram(
            9001,
            {
                "material_name": "Ткань",
                "product_color": "Бежевый",
                "quantity": "12",
            },
        )
        self.assertTrue(fabric_result["ok"], fabric_result)
        self.assertEqual(len(fabric_result["production"]["fabric_stock"]), 1)

        task_result = miniapp_server.create_production_task_for_telegram(
            9001,
            {
                "product_name": "Шорты",
                "sizes": ["80", "92"],
                "colors": ["Бежевый"],
                "fabric_rolls": {"Бежевый": "2"},
                "attachment": {
                    "file_name": "layout.pdf",
                    "mime_type": "",
                    "content_base64": "ZmlsZQ==",
                },
            },
        )
        self.assertTrue(task_result["ok"], task_result)
        self.assertEqual(task_result["production"]["fabric_stock"][0]["quantity"], 10)

        task_id = self.database.get_active_production_tasks()[0][0]
        cutter_state = miniapp_server.get_production_state_for_telegram(9002)
        self.assertTrue(cutter_state["can_contours"])
        self.assertEqual(cutter_state["contour_tasks"][0]["id"], task_id)
        self.assertEqual(cutter_state["contour_tasks"][0]["attachment"]["mime_type"], "application/pdf")
        self.assertEqual(cutter_state["contour_tasks"][0]["fabric_rolls"][0]["rolls"], 2)
        started = miniapp_server.start_cutting_task_for_telegram(9002, task_id)
        self.assertTrue(started["ok"], started)

        contour_result = miniapp_server.submit_production_contours_for_telegram(
            9002,
            {
                "task_id": task_id,
                "quantities": {
                    "Шорты|80|Бежевый": "4",
                    "Шорты|92|Бежевый": "6",
                },
            },
        )
        self.assertTrue(contour_result["ok"], contour_result)
        self.assertEqual(self.database.get_production_task_by_id(task_id)["status"], "contours_done")
        self.assertEqual(len(self.database.get_shift_report(shift["id"])), 2)

    def test_miniapp_admin_creates_order_tasks(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.add_fabric_receipt("Ткань", "Бежевый", 8, None)
        self.database.add_fabric_receipt("Ткань", "Фуксия", 12, None)

        cutting_result = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Шорты",
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["80", "92"],
                "colors": ["Бежевый", "Фуксия"],
                "fabric_rolls": {"Бежевый": "1", "Фуксия": "2"},
                "priority": "high",
                "due_date": "2026-07-20",
                "attachment": {
                    "file_name": "cutting.xlsx",
                    "mime_type": "application/octet-stream",
                    "content_base64": "ZmlsZQ==",
                },
            },
        )

        self.assertTrue(cutting_result["ok"], cutting_result)
        cutting_task = self.database.get_production_task_by_id(1)
        self.assertEqual(cutting_task["priority"], "high")
        self.assertEqual(cutting_task["due_date"], "2026-07-20")
        self.assertEqual(len(self.database.get_active_production_tasks()), 1)
        self.assertEqual(set(self.database.get_production_task_options(1)[1]), {"Бежевый", "Фуксия"})
        self.assertEqual(
            {(row["product_color"], row["unit"]): row["quantity"] for row in cutting_result["production"]["fabric_stock"]},
            {("Бежевый", "рул"): 7, ("Фуксия", "рул"): 10},
        )
        self.assertEqual(
            {(row["product_color"], row["rolls"]) for row in cutting_result["production"]["tasks"][0]["fabric_rolls"]},
            {("Бежевый", 1), ("Фуксия", 2)},
        )
        self.assertEqual(
            cutting_result["production"]["tasks"][0]["attachment"]["mime_type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        sewing_step_index = self.route_step_index("Легинсы", "Сборка на оверлоке", "Швея")
        route_result = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Легинсы",
                "task_type": "route",
                "route_step_index": sewing_step_index,
                "priority": "urgent",
                "due_date": "2026-07-21",
            },
        )

        self.assertFalse(route_result["ok"], route_result)
        self.assertIn("Администратор создаёт только задание на раскрой", route_result["message"])
        self.assertEqual(self.database.get_active_route_batches(), [])

    def test_miniapp_admin_creates_shared_nastil_for_multiple_products(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.add_fabric_receipt("Ткань", "Черный", 4, None)

        result = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Брюки со стрелками детские",
                "product_names": ["Брюки со стрелками детские", "Жакет для девочек"],
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["98"],
                "colors": ["Черный"],
                "fabric_rolls": {"Черный": "1"},
            },
        )

        self.assertTrue(result["ok"], result)
        self.assertIn("Общий настил", result["message"])
        tasks = self.database.get_active_production_tasks()
        self.assertEqual(len(tasks), 2)
        self.assertEqual({task[1] for task in tasks}, {"Брюки со стрелками детские", "Жакет для девочек"})
        roll_counts = {
            task[1]: self.database.get_production_task_fabric_rolls(task[0])
            for task in tasks
        }
        self.assertEqual(len(roll_counts["Брюки со стрелками детские"]), 1)
        self.assertEqual(roll_counts["Жакет для девочек"], [])
        self.assertEqual(self.database.get_fabric_stock_rows()[0][2], 3)

    def test_reset_cutting_task_returns_employee_to_contour_entry(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9010, "Курасова Наталья Валерьевна", "Раскройщик")
        cutter = self.database.get_employee_by_telegram_id(9010)
        self.database.update_employee_status(cutter[0], "active")
        self.database.create_shift(cutter[0])
        self.database.add_fabric_receipt("Ткань", "Капучино", 2, None)
        created = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_names": ["Брюки со стрелками детские", "Кардиган детский"],
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["98"],
                "colors": ["Капучино"],
                "fabric_rolls": {"Капучино": "1"},
            },
        )
        self.assertTrue(created["ok"], created)
        tasks = self.database.get_active_production_tasks()
        for task in tasks:
            self.assertIsNotNone(self.database.assign_production_task(task[0], cutter[0]))
            submitted = miniapp_server.submit_production_contours_for_telegram(
                9010,
                {"task_id": task[0], "quantities": {f"{task[1]}|98|Капучино": 3}},
            )
            self.assertTrue(submitted["ok"], submitted)

        reset = self.database.reset_cutting_tasks_to_contours_entry(
            "Курасова Наталья Валерьевна",
            ["Кардиган детский", "Брюки со стрелками детские"],
            replacement_color="Брауни",
        )
        self.assertEqual(len(reset), 2)
        contour_tasks = self.database.get_active_production_tasks_for_contours()
        self.assertEqual({row[1] for row in contour_tasks}, {"Кардиган детский", "Брюки со стрелками детские"})
        for row in contour_tasks:
            task = self.database.get_production_task_by_id(row[0])
            self.assertEqual(task["assigned_employee_id"], cutter[0])
            conn = self.database.get_db_connection()
            colors = [item[0] for item in conn.execute(
                "SELECT product_color FROM production_task_colors WHERE task_id = ?", (row[0],)
            ).fetchall()]
            conn.close()
            self.assertEqual(colors, ["Брауни"])

    def test_kurasova_migration_accepts_database_spelling_and_one_remaining_task(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9011, "Курасова Наталия Валерьевна", "Раскройщик")
        cutter = self.database.get_employee_by_telegram_id(9011)
        self.database.update_employee_status(cutter[0], "active")
        self.database.create_shift(cutter[0])
        self.database.add_fabric_receipt("Ткань", "Капучино", 2, None)
        created = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Кардиган детский",
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["98"],
                "colors": ["Капучино"],
                "fabric_rolls": {"Капучино": "1"},
            },
        )
        self.assertTrue(created["ok"], created)
        task_id = self.database.get_active_production_tasks()[0][0]
        self.assertIsNotNone(self.database.assign_production_task(task_id, cutter[0]))
        submitted = miniapp_server.submit_production_contours_for_telegram(
            9011,
            {"task_id": task_id, "quantities": {"Кардиган детский|98|Капучино": 3}},
        )
        self.assertTrue(submitted["ok"], submitted)

        self.assertTrue(self.database.apply_kurasova_brownie_contour_migration())
        task = self.database.get_production_task_by_id(task_id)
        self.assertEqual(task["status"], "active")
        conn = self.database.get_db_connection()
        colors = [row[0] for row in conn.execute(
            "SELECT product_color FROM production_task_colors WHERE task_id = ?", (task_id,)
        ).fetchall()]
        fabric_colors = [row[0] for row in conn.execute(
            "SELECT product_color FROM fabric_stock ORDER BY product_color"
        ).fetchall()]
        conn.close()
        self.assertEqual(colors, ["Брауни"])
        self.assertEqual(fabric_colors, ["Брауни"])

    def test_contour_reset_can_target_an_exact_task_when_assignment_is_missing(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9012, "Курасова Наталия Валерьевна", "Раскройщик")
        cutter = self.database.get_employee_by_telegram_id(9012)
        self.database.update_employee_status(cutter[0], "active")
        self.database.create_shift(cutter[0])
        self.database.add_fabric_receipt("Ткань", "Капучино", 2, None)
        created = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Брюки со стрелками детские",
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["104"],
                "colors": ["Капучино"],
                "fabric_rolls": {"Капучино": "1"},
            },
        )
        self.assertTrue(created["ok"], created)
        task_id = self.database.get_active_production_tasks()[0][0]
        self.assertIsNotNone(self.database.assign_production_task(task_id, cutter[0]))
        submitted = miniapp_server.submit_production_contours_for_telegram(
            9012,
            {"task_id": task_id, "quantities": {"Брюки со стрелками детские|104|Капучино": 3}},
        )
        self.assertTrue(submitted["ok"], submitted)
        conn = self.database.get_db_connection()
        conn.execute("UPDATE production_tasks SET assigned_employee_id = NULL WHERE id = ?", (task_id,))
        conn.commit()
        conn.close()

        reset = self.database.reset_cutting_tasks_to_contours_entry(
            "Курасова Наталия Валерьевна",
            ["Брюки со стрелками детские"],
            replacement_color="Брауни",
            task_ids=[task_id],
        )

        self.assertEqual([row["task_id"] for row in reset], [task_id])
        task = self.database.get_production_task_by_id(task_id)
        self.assertEqual(task["status"], "active")
        self.assertIsNone(task["assigned_employee_id"])
        conn = self.database.get_db_connection()
        colors = [row[0] for row in conn.execute(
            "SELECT product_color FROM production_task_colors WHERE task_id = ?", (task_id,)
        ).fetchall()]
        conn.close()
        self.assertEqual(colors, ["Брауни"])

    def test_contour_reset_cancels_zero_output_completed_downstream_tasks(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9013, "Курасова Наталия Валерьевна", "Раскройщик")
        cutter = self.database.get_employee_by_telegram_id(9013)
        self.database.update_employee_status(cutter[0], "active")
        self.database.create_shift(cutter[0])
        self.database.add_fabric_receipt("Ткань", "Капучино", 2, None)
        created = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Кардиган детский",
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["104"],
                "colors": ["Капучино"],
                "fabric_rolls": {"Капучино": "1"},
            },
        )
        self.assertTrue(created["ok"], created)
        task_id = self.database.get_active_production_tasks()[0][0]
        self.assertIsNotNone(self.database.assign_production_task(task_id, cutter[0]))
        submitted = miniapp_server.submit_production_contours_for_telegram(
            9013,
            {"task_id": task_id, "quantities": {"Кардиган детский|104|Капучино": 3}},
        )
        self.assertTrue(submitted["ok"], submitted)
        conn = self.database.get_db_connection()
        cutting_batch_id = conn.execute(
            "SELECT id FROM cutting_batches WHERE production_task_id = ?", (task_id,)
        ).fetchone()[0]
        now = self.database.local_now().isoformat()
        cursor = conn.execute(
            """
            INSERT INTO route_batches (
                product_name, product_size, product_color, quantity,
                route_step_index, status, good_quantity, defect_quantity,
                source_cutting_batch_id, work_state, created_at, updated_at, completed_at
            ) VALUES (?, ?, ?, 0, 0, 'done', 0, 0, ?, 'completed', ?, ?, ?)
            """,
            ("Кардиган детский", "104", "Капучино", cutting_batch_id, now, now, now),
        )
        downstream_id = cursor.lastrowid
        conn.commit()
        conn.close()

        reset = self.database.reset_cutting_tasks_to_contours_entry(
            "Курасова Наталия Валерьевна",
            ["Кардиган детский"],
            replacement_color="Брауни",
            task_ids=[task_id],
        )

        self.assertEqual([row["task_id"] for row in reset], [task_id])
        conn = self.database.get_db_connection()
        downstream = conn.execute(
            "SELECT status, work_state, blocked_reason FROM route_batches WHERE id = ?", (downstream_id,)
        ).fetchone()
        conn.close()
        self.assertEqual(downstream, ("cancelled", "cancelled", "Откат задания к нанесению контуров"))

    def test_kurasova_migration_retries_after_old_partial_marker_and_resets_cutting_progress(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9014, "Курасова Наталия Валерьевна", "Раскройщик")
        cutter = self.database.get_employee_by_telegram_id(9014)
        self.database.update_employee_status(cutter[0], "active")
        self.database.create_shift(cutter[0])
        self.database.add_fabric_receipt("Ткань", "Капучино", 3, None)
        created = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Кардиган детский",
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["104"],
                "colors": ["Капучино"],
                "fabric_rolls": {"Капучино": "1"},
            },
        )
        self.assertTrue(created["ok"], created)
        task_id = self.database.get_active_production_tasks()[0][0]
        self.assertTrue(miniapp_server.start_cutting_task_for_telegram(9014, task_id)["ok"])
        contours = miniapp_server.submit_cutting_stage_for_telegram(
            9014,
            {"stage": "contours", "task_id": task_id, "quantities": {"104|Капучино": "5"}},
        )
        self.assertTrue(contours["ok"], contours)
        batch_id = self.database.get_cutting_batches_for_layout("Кардиган детский")[0][0]
        layout = miniapp_server.submit_cutting_stage_for_telegram(
            9014,
            {"stage": "layout", "batch_id": batch_id, "color_layers": {"Капучино": "2"}},
        )
        self.assertTrue(layout["ok"], layout)
        cutting = miniapp_server.submit_cutting_stage_for_telegram(
            9014,
            {"stage": "cutting", "batch_id": batch_id, "progress": "75"},
        )
        self.assertTrue(cutting["ok"], cutting)
        conn = self.database.get_db_connection()
        conn.execute(
            "CREATE TABLE IF NOT EXISTS business_data_migrations (migration_key TEXT PRIMARY KEY, applied_at TEXT NOT NULL)"
        )
        conn.execute(
            "INSERT INTO business_data_migrations (migration_key, applied_at) VALUES (?, ?)",
            ("2026-08-04-kurasova-brownie-contours-v5", self.database.local_now().isoformat()),
        )
        conn.commit()
        conn.close()

        self.assertTrue(self.database.apply_kurasova_brownie_contour_migration())

        task = self.database.get_production_task_by_id(task_id)
        self.assertEqual(task["status"], "active")
        self.assertEqual(self.database.get_cutting_batches_for_cutting("Кардиган детский"), [])
        conn = self.database.get_db_connection()
        colors = [row[0] for row in conn.execute(
            "SELECT product_color FROM production_task_colors WHERE task_id = ?", (task_id,)
        ).fetchall()]
        migration = conn.execute(
            "SELECT 1 FROM business_data_migrations WHERE migration_key = ?",
            ("2026-08-04-kurasova-brownie-contours-v6",),
        ).fetchone()
        conn.close()
        self.assertEqual(colors, ["Брауни"])
        self.assertIsNotNone(migration)

    def test_admin_can_edit_writeoff_and_delete_empty_material_card(self):
        row = self.database.add_fabric_receipt("Ткань", "Брауни", 3, None)
        edited = self.database.update_fabric_stock_card(row[0], "Ткань костюмная", "Брауни", "рул", None, "Уточнение")
        self.assertEqual(edited[1], "Ткань костюмная")
        written_off = self.database.write_off_fabric_stock(row[0], 3, None, "Повреждение")
        self.assertEqual(int(written_off[3]), 0)
        deleted = self.database.delete_zero_fabric_stock(row[0], None, "Карточка больше не используется")
        self.assertEqual(deleted["id"], row[0])
        self.assertIsNone(self.database.get_fabric_stock_by_id(row[0]))

    def test_elastic_preparation_completion_creates_sewing_task(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9003, "Тест Упаковщик", "Упаковщик")
        self.database.create_employee(9004, "Тест Швея", "Швея")
        packer = self.database.get_employee_by_telegram_id(9003)
        seamstress = self.database.get_employee_by_telegram_id(9004)
        self.database.update_employee_status(packer[0], "active")
        self.database.update_employee_status(seamstress[0], "active")
        self.database.create_shift(packer[0])
        elastic_step_index = self.route_step_index("Шорты", "Шорты — резинка 25 мм", "Упаковщик")

        batch = self.database.create_route_batch(
            "Шорты",
            "80 (43 см)",
            "Черный",
            14,
            packer[0],
            route_step_index=elastic_step_index,
        )
        self.assertTrue(miniapp_server.start_route_task_for_telegram(9003, batch["id"])["ok"])
        result = miniapp_server.complete_route_task_for_telegram(9003, batch["id"])

        self.assertTrue(result["ok"], result)
        self.assertIn("Следующее задание создано", result["message"])
        active_batches = self.database.get_active_route_batches()
        self.assertEqual(len(active_batches), 1)

        sewing_batch = active_batches[0]
        sewing_step = miniapp_server.get_route_step("Шорты", sewing_batch["route_step_index"])
        self.assertEqual(sewing_step["operation"], "Сшивание резинок в кольцо")
        self.assertEqual(sewing_batch["quantity"], 14)
        self.assertIsNotNone(sewing_batch["source_stock_id"])
        sewing_stock = self.database.get_warehouse_stock_by_id(sewing_batch["source_stock_id"])
        self.assertEqual(sewing_stock["quantity"], 14)
        self.assertEqual(sewing_stock["reserved_quantity"], 14)

        seamstress_tasks = miniapp_server.get_route_tasks_for_telegram(9004)
        self.assertTrue(seamstress_tasks["ok"], seamstress_tasks)
        self.assertEqual(seamstress_tasks["tasks"][0]["category"], "Подготовка")

    def test_route_start_reserves_input_and_completion_consumes_it(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9120, "Тест Швея Резерв", "Швея")
        employee = self.database.get_employee_by_telegram_id(9120)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        step_index = self.route_step_index("Легинсы", "Сборка на оверлоке", "Швея")
        stock = self.database.add_warehouse_stock(
            "semifinished", "Легинсы", "86", "Бежевый", "Раскроенные", "Швея", 10, None,
        )
        batch = self.database.create_route_batch("Легинсы", "86", "Бежевый", 10, None, step_index)

        self.assertTrue(miniapp_server.start_route_task_for_telegram(9120, batch["id"])["ok"])
        reserved = self.database.get_warehouse_stock_by_id(stock["id"])
        self.assertEqual((reserved["quantity"], reserved["reserved_quantity"]), (10, 10))
        self.assertEqual(self.database.get_route_batch_inputs(batch["id"])[0]["quantity"], 10)

        self.assertTrue(miniapp_server.complete_route_task_for_telegram(
            9120, batch["id"], {"good_quantity": 10, "defect_quantity": 0},
        )["ok"])
        consumed = self.database.get_warehouse_stock_by_id(stock["id"])
        self.assertEqual((consumed["quantity"], consumed["reserved_quantity"]), (0, 0))

    def test_stock_shortage_alert_does_not_block_route_task(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9121, "Тест Швея Дефицит", "Швея")
        employee = self.database.get_employee_by_telegram_id(9121)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        step_index = self.route_step_index("Легинсы", "Сборка на оверлоке", "Швея")
        batch = self.database.create_route_batch("Легинсы", "86", "Бежевый", 6, None, step_index)

        self.assertTrue(miniapp_server.start_route_task_for_telegram(9121, batch["id"])["ok"])
        notifications = self.database.get_open_critical_notifications()
        self.assertEqual(len(notifications), 1)
        self.assertEqual(notifications[0]["batch_id"], batch["id"])
        self.assertEqual(notifications[0]["severity"], "critical")
        self.assertTrue(miniapp_server.complete_route_task_for_telegram(
            9121, batch["id"], {"good_quantity": 6, "defect_quantity": 0},
        )["ok"])

    def test_employee_starts_and_completes_route_task_with_defects(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9013, "Тест Упаковщик", "Упаковщик")
        packer = self.database.get_employee_by_telegram_id(9013)
        self.database.update_employee_status(packer[0], "active")
        self.database.create_shift(packer[0])
        elastic_step_index = self.route_step_index("Шорты", "Шорты — резинка 25 мм", "Упаковщик")
        batch = self.database.create_route_batch(
            "Шорты",
            "80 (43 см)",
            "Черный",
            10,
            None,
            route_step_index=elastic_step_index,
            priority="high",
            due_date=self.database.local_today().isoformat(),
        )

        free_tasks = miniapp_server.get_route_tasks_for_telegram(9013)
        self.assertTrue(free_tasks["ok"], free_tasks)
        self.assertEqual(free_tasks["tasks"][0]["status_text"], "Свободно")
        self.assertTrue(free_tasks["tasks"][0]["can_take"])

        started = miniapp_server.start_route_task_for_telegram(9013, batch["id"])
        self.assertTrue(started["ok"], started)
        started_task = started["tasks"][0]
        self.assertEqual(started_task["status_text"], "В работе")
        self.assertEqual(started_task["assigned_employee_name"], "Тест Упаковщик")
        self.assertTrue(started_task["can_complete"])

        missing_reason = miniapp_server.complete_route_task_for_telegram(
            9013,
            batch["id"],
            {"good_quantity": 9, "defect_quantity": 1},
        )
        self.assertFalse(missing_reason["ok"], missing_reason)

        completed = miniapp_server.complete_route_task_for_telegram(
            9013,
            batch["id"],
            {
                "good_quantity": 9,
                "defect_quantity": 1,
                "defect_reason": "Повреждение фурнитуры",
                "defect_disposition": "На переделку",
                "defect_comment": "Проверить установку",
            },
        )
        self.assertTrue(completed["ok"], completed)
        completed_batch = self.database.get_route_batch_by_id(batch["id"])
        self.assertEqual(completed_batch["status"], "done")
        self.assertEqual(completed_batch["assigned_employee_id"], packer[0])
        self.assertEqual(completed_batch["good_quantity"], 9)
        self.assertEqual(completed_batch["defect_quantity"], 1)
        self.assertEqual(completed["completed_tasks"][0]["status_text"], "Завершено")
        self.assertEqual(completed["completed_tasks"][0]["good_quantity"], 9)
        self.assertEqual(completed["completed_tasks"][0]["defect_quantity"], 1)
        defect_rows = self.database.get_route_batch_defects(batch["id"])
        self.assertEqual(defect_rows[0]["reason"], "Повреждение фурнитуры")
        self.assertEqual(defect_rows[0]["disposition"], "На переделку")
        self.assertIsNotNone(defect_rows[0]["rework_batch_id"])
        rework = self.database.get_route_batch_by_id(defect_rows[0]["rework_batch_id"])
        self.assertEqual(rework["parent_batch_id"], batch["id"])
        self.assertEqual(rework["priority"], "urgent")
        self.assertEqual(rework["quantity"], 1)
        control = miniapp_server.get_production_control_payload(
            self.database.local_today().isoformat(),
            self.database.local_today().isoformat(),
        )
        self.assertEqual(control["plan"], 20)
        self.assertEqual(control["fact"], 9)
        self.assertEqual(control["defect_quantity"], 1)
        self.assertEqual(control["fpy"], 90)
        self.assertEqual(control["schedule_adherence"], 100)
        self.assertEqual(control["active_tasks"], 2)
        self.assertTrue(any(alert["type"] == "defect" for alert in control["alerts"]))
        self.assertTrue(control["details"]["planned_tasks"])
        self.assertEqual(control["details"]["completed_tasks"][0]["id"], batch["id"])
        self.assertEqual(control["details"]["completed_tasks"][0]["employee"], "Тест Упаковщик")
        self.assertEqual(control["details"]["defects"][0]["batch_id"], batch["id"])
        self.assertTrue(all(task["stage"] for task in control["details"]["active_tasks"]))
        self.assertTrue(all("batch_id" in alert for alert in control["alerts"]))

    def test_open_shift_survives_midnight_and_remains_current(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9040, "Тест Ночная Смена", "Швея")
        employee = self.database.get_employee_by_telegram_id(9040)
        self.database.update_employee_status(employee[0], "active")
        shift = self.database.create_shift(employee[0])
        yesterday = (self.database.local_today() - timedelta(days=1)).isoformat()

        conn = sqlite3.connect(self.database.DB_NAME)
        conn.execute("UPDATE shifts SET shift_date = ? WHERE id = ?", (yesterday, shift["id"]))
        conn.commit()
        conn.close()

        open_shift = self.database.get_open_shift_for_today(employee[0])
        self.assertEqual(open_shift[0], shift["id"])
        self.assertEqual(open_shift[2], yesterday)

        reopened = miniapp_server.open_shift_for_telegram(9040)
        self.assertTrue(reopened["ok"], reopened)
        self.assertEqual(reopened["message"], "Смена уже открыта.")
        self.assertEqual(reopened["shift"]["id"], shift["id"])

        report = miniapp_server.get_current_report_for_telegram(9040)
        self.assertEqual(report["shift"]["id"], shift["id"])

        conn = sqlite3.connect(self.database.DB_NAME)
        self.assertEqual(conn.execute("SELECT COUNT(*) FROM shifts WHERE employee_id = ?", (employee[0],)).fetchone()[0], 1)
        conn.close()

    def test_miniapp_excel_export_contains_detailed_production_sheets(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9041, "Тест Выгрузка", "Упаковщик")
        employee = self.database.get_employee_by_telegram_id(9041)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        self.database.add_fabric_receipt("Ткань", "Черный", 3, None, "Тестовая приёмка")
        stock = self.database.add_warehouse_stock(
            "semifinished", "Шорты", "80", "Черный", "Раскроенные", "Упаковщик", 10, None,
        )
        step_index = self.route_step_index("Шорты", "Шорты — резинка 25 мм", "Упаковщик")
        batch = self.database.create_route_batch_with_inputs(
            "Шорты", "80", "Черный", 10, None, step_index,
            [{"stock_id": stock["id"], "quantity": 10, "input_role": "main"}],
        )
        self.assertIsNotNone(batch)
        self.assertTrue(miniapp_server.start_route_task_for_telegram(9041, batch["id"])["ok"])
        completed = miniapp_server.complete_route_task_for_telegram(
            9041,
            batch["id"],
            {
                "good_quantity": 9,
                "defect_quantity": 1,
                "defect_reason": "Повреждение ткани",
                "defect_disposition": "Списать",
            },
        )
        self.assertTrue(completed["ok"], completed)

        today = self.database.local_today().isoformat()
        export = miniapp_server.export_admin_report_for_telegram(
            9001, {"report_type": "period", "start_date": today, "end_date": today},
        )
        self.assertTrue(export["ok"], export)

        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(export["content"]), data_only=False)
        expected_sheets = {
            "Сводка", "Смены по дням", "Операции", "Раскрой", "Пошив", "Упаковка",
            "Задания раскроя", "Маршрутные задания", "Входы заданий", "Брак",
            "Партии компонентов", "Хронология партий",
            "Движения склада", "Движения материалов", "Остатки склада", "Остатки материалов",
            "WIP по этапам", "Отклонения",
        }
        self.assertTrue(expected_sheets.issubset(set(workbook.sheetnames)))

        route_sheet = workbook["Маршрутные задания"]
        route_values = list(route_sheet.values)
        self.assertEqual(route_values[1][9:14], (10, 9, 1, 90, 10))
        self.assertEqual(route_values[1][15], "Тест Выгрузка")
        self.assertTrue(route_values[1][19].startswith("RB-"))
        self.assertTrue(route_values[1][20])
        defect_values = list(workbook["Брак"].values)
        self.assertEqual(defect_values[1][8], 1)
        self.assertEqual(defect_values[1][9], "Повреждение ткани")
        self.assertEqual(defect_values[1][10], "Списать")
        self.assertGreaterEqual(workbook["Движения склада"].max_row, 3)
        self.assertGreaterEqual(workbook["Движения материалов"].max_row, 2)
        self.assertGreaterEqual(workbook["Партии компонентов"].max_row, 2)
        self.assertGreaterEqual(workbook["Хронология партий"].max_row, 4)

    def test_miniapp_full_production_cycle_is_strictly_sequential(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        route_maps = importlib.import_module("route_maps")

        employees = {
            "Раскройщик": (9002, "Тест Раскройщик Цикл"),
            "Швея": (9003, "Тест Швея Цикл"),
            "Упаковщик": (9004, "Тест Упаковщик Цикл"),
        }
        employee_rows = {}

        for position, (telegram_id, full_name) in employees.items():
            self.database.create_employee(telegram_id, full_name, position)
            employee = self.database.get_employee_by_telegram_id(telegram_id)
            self.database.update_employee_status(employee[0], "active")
            employee_rows[position] = employee
            self.database.create_shift(employee[0])

        receipt = miniapp_server.add_fabric_receipt_for_telegram(
            9001,
            {"material_name": "Ткань", "product_color": "Бежевый", "quantity": "3"},
        )
        self.assertTrue(receipt["ok"], receipt)

        cutting_task = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Футболки",
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["86"],
                "colors": ["Бежевый"],
                "fabric_rolls": {"Бежевый": "1"},
            },
        )
        self.assertTrue(cutting_task["ok"], cutting_task)
        self.assertEqual(cutting_task["production"]["fabric_stock"][0]["quantity"], 2)
        task_id = self.database.get_active_production_tasks()[0][0]
        started_cutting = miniapp_server.start_cutting_task_for_telegram(9002, task_id)
        self.assertTrue(started_cutting["ok"], started_cutting)
        admin_released_cutting = miniapp_server.release_cutting_task_for_telegram(
            9001,
            task_id,
            "Администратор передаёт задание",
        )
        self.assertTrue(admin_released_cutting["ok"], admin_released_cutting)
        self.assertIsNone(self.database.get_production_task_by_id(task_id)["assigned_employee_id"])
        started_cutting = miniapp_server.start_cutting_task_for_telegram(9002, task_id)
        self.assertTrue(started_cutting["ok"], started_cutting)
        released_cutting = miniapp_server.release_cutting_task_for_telegram(
            9002,
            task_id,
            "Проверка возврата задания",
        )
        self.assertTrue(released_cutting["ok"], released_cutting)
        self.assertIsNone(self.database.get_production_task_by_id(task_id)["assigned_employee_id"])
        started_cutting = miniapp_server.start_cutting_task_for_telegram(9002, task_id)
        self.assertTrue(started_cutting["ok"], started_cutting)

        blocked_layout = miniapp_server.submit_cutting_stage_for_telegram(
            9002,
            {"stage": "layout", "batch_id": task_id, "color_layers": {"Бежевый": "2"}},
        )
        self.assertFalse(blocked_layout["ok"], blocked_layout)

        contours = miniapp_server.submit_cutting_stage_for_telegram(
            9002,
            {"stage": "contours", "task_id": task_id, "quantities": {"86|Бежевый": "5"}},
        )
        self.assertTrue(contours["ok"], contours)
        batch_id = self.database.get_cutting_batches_for_layout("Футболки")[0][0]

        blocked_cutting = miniapp_server.submit_cutting_stage_for_telegram(
            9002,
            {"stage": "cutting", "batch_id": batch_id, "progress": "100"},
        )
        self.assertFalse(blocked_cutting["ok"], blocked_cutting)

        layout = miniapp_server.submit_cutting_stage_for_telegram(
            9002,
            {"stage": "layout", "batch_id": batch_id, "color_layers": {"Бежевый": "2"}},
        )
        self.assertTrue(layout["ok"], layout)

        partial_cutting = miniapp_server.submit_cutting_stage_for_telegram(
            9002,
            {"stage": "cutting", "batch_id": batch_id, "progress": "50"},
        )
        self.assertTrue(partial_cutting["ok"], partial_cutting)
        blocked_formation = miniapp_server.submit_cutting_stage_for_telegram(
            9002,
            {"stage": "formation", "batch_id": batch_id},
        )
        self.assertFalse(blocked_formation["ok"], blocked_formation)

        completed_cutting = miniapp_server.submit_cutting_stage_for_telegram(
            9002,
            {"stage": "cutting", "batch_id": batch_id, "progress": "100"},
        )
        self.assertTrue(completed_cutting["ok"], completed_cutting)
        formation = miniapp_server.submit_cutting_stage_for_telegram(
            9002,
            {"stage": "formation", "batch_id": batch_id},
        )
        self.assertTrue(formation["ok"], formation)

        stock = next(row for row in self.database.get_warehouse_stock_rows() if row["quantity"] > 0)
        self.assertEqual(
            (stock["item_type"], stock["stage_name"], stock["ready_for_position"], stock["quantity"]),
            ("semifinished", "Раскроенные", "Швея", 10),
        )

        route_steps = route_maps.PRODUCT_ROUTE_MAPS["Футболки"]
        first_step_index = len(route_maps.CUTTING_ROUTE)
        manual_route = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Футболки",
                "task_type": "route",
                "route_step_index": first_step_index + 1,
                "stock_items": [{"stock_id": stock["id"], "quantity": "10"}],
            },
        )
        self.assertFalse(manual_route["ok"], manual_route)
        self.assertEqual(self.database.get_warehouse_stock_by_id(stock["id"])["quantity"], 10)

        position_telegram_ids = {position: data[0] for position, data in employees.items()}

        for step_index in range(first_step_index, len(route_steps)):
            active_batches = self.database.get_active_route_batches()
            self.assertEqual(len(active_batches), 1, (step_index, active_batches))
            active_batch = active_batches[0]
            self.assertEqual(active_batch["route_step_index"], step_index)
            telegram_id = position_telegram_ids[route_steps[step_index]["position"]]
            started = miniapp_server.start_route_task_for_telegram(telegram_id, active_batch["id"])
            self.assertTrue(started["ok"], (step_index, started))
            completed = miniapp_server.complete_route_task_for_telegram(
                telegram_id,
                active_batch["id"],
                {"good_quantity": 10, "defect_quantity": 0},
            )
            self.assertTrue(completed["ok"], (step_index, completed))

        final_stock = [row for row in self.database.get_warehouse_stock_rows() if row["quantity"] > 0]
        self.assertEqual(len(final_stock), 1)
        self.assertEqual(
            (
                final_stock[0]["item_type"],
                final_stock[0]["stage_name"],
                final_stock[0]["ready_for_position"],
                final_stock[0]["quantity"],
            ),
            ("finished", "Упаковано", "Склад", 10),
        )
        self.assertEqual(self.database.get_active_route_batches(), [])

    def test_miniapp_route_task_combines_main_and_component_inputs(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9014, "Тест Швея Компоненты", "Швея")
        seamstress = self.database.get_employee_by_telegram_id(9014)
        self.database.update_employee_status(seamstress[0], "active")

        main_stock = self.database.add_warehouse_stock(
            "semifinished",
            "Шорты",
            "80",
            "Бежевый",
            "Раскроенные",
            "Швея",
            10,
            None,
        )
        second_main_stock = self.database.add_warehouse_stock(
            "semifinished",
            "Шорты",
            "86",
            "Голубой",
            "Раскроенные",
            "Швея",
            15,
            None,
        )
        elastic_stock = self.database.add_warehouse_stock(
            "semifinished",
            "Шорты",
            "80, 86 (43 см)",
            "Черный",
            "Резинка сшита в кольцо",
            "Швея",
            25,
            None,
        )
        assembly_step_index = self.route_step_index("Шорты", "Сборка шорт на оверлоке", "Швея")

        manual_route = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Шорты",
                "task_type": "route",
                "route_step_index": assembly_step_index,
                "stock_items": [{"stock_id": main_stock["id"], "quantity": "10"}],
            },
        )
        self.assertFalse(manual_route["ok"], manual_route)
        self.assertIn("создаёт система", manual_route["message"])
        self.assertEqual(self.database.get_warehouse_stock_by_id(main_stock["id"])["quantity"], 10)
        self.assertEqual(self.database.get_warehouse_stock_by_id(second_main_stock["id"])["quantity"], 15)
        self.assertEqual(self.database.get_warehouse_stock_by_id(elastic_stock["id"])["quantity"], 25)
        self.assertEqual(self.database.get_active_route_batches(), [])

    def test_miniapp_admin_deletes_order_tasks(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.add_fabric_receipt("Ткань", "Бежевый", 5, None)

        cutting_result = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Шорты",
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["80"],
                "colors": ["Бежевый"],
                "fabric_rolls": {"Бежевый": "2"},
            },
        )
        self.assertTrue(cutting_result["ok"], cutting_result)
        self.assertEqual(cutting_result["production"]["fabric_stock"][0]["quantity"], 3)
        sewing_step_index = self.route_step_index("Легинсы", "Сборка на оверлоке", "Швея")
        route_batch = self.database.create_route_batch(
            "Легинсы", "86", "Бежевый", 3, None, sewing_step_index,
        )
        self.assertIsNotNone(route_batch)

        production_id = self.database.get_active_production_tasks()[0][0]
        route_id = route_batch["id"]

        delete_production = miniapp_server.delete_order_task_for_telegram(
            9001,
            {"task_kind": "production", "task_id": production_id},
        )
        self.assertTrue(delete_production["ok"], delete_production)
        self.assertEqual(self.database.get_production_task_by_id(production_id)["status"], "cancelled")
        self.assertEqual(self.database.get_active_production_tasks(), [])
        self.assertEqual(self.database.get_fabric_stock_rows()[0][2], 5)

        delete_route = miniapp_server.delete_order_task_for_telegram(
            9001,
            {"task_kind": "route", "task_id": route_id},
        )
        self.assertTrue(delete_route["ok"], delete_route)
        self.assertEqual(self.database.get_route_batch_by_id(route_id)["status"], "cancelled")
        self.assertEqual(self.database.get_active_route_batches(), [])

    def test_admin_cancels_cutting_in_progress_and_restores_fabric(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9020, "Тест Раскройщик Отмена", "Раскройщик")
        cutter = self.database.get_employee_by_telegram_id(9020)
        self.database.update_employee_status(cutter[0], "active")
        self.database.create_shift(cutter[0])
        self.database.add_fabric_receipt("Ткань", "Брауни", 3, None)

        created = miniapp_server.create_order_task_for_telegram(
            9001,
            {
                "product_name": "Футболки",
                "task_type": "cutting",
                "material_name": "Ткань",
                "sizes": ["86"],
                "colors": ["Брауни"],
                "fabric_rolls": {"Брауни": "1"},
            },
        )
        self.assertTrue(created["ok"], created)
        task_id = self.database.get_active_production_tasks()[0][0]
        self.assertTrue(miniapp_server.start_cutting_task_for_telegram(9020, task_id)["ok"])
        contours = miniapp_server.submit_cutting_stage_for_telegram(
            9020,
            {"stage": "contours", "task_id": task_id, "quantities": {"86|Брауни": "5"}},
        )
        self.assertTrue(contours["ok"], contours)
        batch_id = self.database.get_cutting_batches_for_layout("Футболки")[0][0]
        layout = miniapp_server.submit_cutting_stage_for_telegram(
            9020,
            {"stage": "layout", "batch_id": batch_id, "color_layers": {"Брауни": "2"}},
        )
        self.assertTrue(layout["ok"], layout)
        cutting = miniapp_server.submit_cutting_stage_for_telegram(
            9020,
            {"stage": "cutting", "batch_id": batch_id, "progress": "50"},
        )
        self.assertTrue(cutting["ok"], cutting)

        deleted = miniapp_server.delete_order_task_for_telegram(
            9001, {"task_kind": "production", "task_id": task_id}
        )
        self.assertTrue(deleted["ok"], deleted)
        self.assertEqual(self.database.get_production_task_by_id(task_id)["status"], "cancelled")
        self.assertEqual(self.database.get_fabric_stock_rows()[0][2], 3)
        self.assertEqual(self.database.get_active_production_tasks(), [])

    def test_zero_quantity_is_not_saved_or_reported(self):
        self.database.create_employee(2002, "Тест Швея", "Швея")
        employee = self.database.get_employee_by_telegram_id(2002)
        employee_id = employee[0]
        self.database.update_employee_status(employee_id, "active")
        shift = self.database.create_shift(employee_id)

        operation = self.database.get_active_operations(
            position="Швея",
            folder="Брюки-джоггеры",
        )[0]
        operation_id = operation[0]

        self.assertFalse(
            self.database.add_shift_operation(
                shift["id"],
                employee_id,
                operation_id,
                "86",
                "Капучино",
                0,
            )
        )
        self.assertEqual(self.database.get_shift_report(shift["id"]), [])
        self.assertEqual(self.database.get_shift_operation_choices(shift["id"]), [])

        self.assertTrue(
            self.database.add_shift_operation(
                shift["id"],
                employee_id,
                operation_id,
                "86",
                "Капучино",
                5,
            )
        )
        self.assertEqual(len(self.database.get_shift_report(shift["id"])), 1)

        self.assertFalse(
            self.database.set_shift_operation_quantity(
                shift["id"],
                employee_id,
                operation_id,
                "86",
                "Капучино",
                0,
            )
        )
        self.assertEqual(self.database.get_shift_report(shift["id"]), [])
        self.assertEqual(self.database.get_shift_operation_choices(shift["id"]), [])

    def test_feedback_entries_are_saved_and_listed(self):
        self.database.create_employee(3003, "Тест Сотрудник", "Упаковщик")
        employee = self.database.get_employee_by_telegram_id(3003)
        employee_id = employee[0]
        self.database.update_employee_status(employee_id, "active")
        shift = self.database.create_shift(employee_id)

        today = self.database.local_today().isoformat()
        feedback_id = self.database.add_feedback_entry(
            employee_id,
            shift["id"],
            "Производство",
            "Заканчивается резинка 25 мм",
        )

        self.assertIsNotNone(feedback_id)
        self.assertIsNone(self.database.add_feedback_entry(employee_id, shift["id"], "Бытовое", "   "))

        period_rows = self.database.get_feedback_entries(today, today)
        shift_rows = self.database.get_feedback_entries_by_shift(shift["id"])
        employee_rows = self.database.get_feedback_entries(today, today, employee_id=employee_id)
        status = self.database.get_database_status()

        self.assertEqual(period_rows, shift_rows)
        self.assertEqual(period_rows, employee_rows)
        self.assertEqual(len(period_rows), 1)
        self.assertEqual(period_rows[0][2], "Тест Сотрудник")
        self.assertEqual(period_rows[0][4], "Производство")
        self.assertEqual(period_rows[0][5], "Заканчивается резинка 25 мм")
        self.assertEqual(period_rows[0][6], shift["id"])
        self.assertEqual(status["feedback_entries"], 1)

    def test_admins_are_excluded_from_employee_reports(self):
        self.database.create_employee(5001, "Тест Рабочий", "Швея")
        employee = self.database.get_employee_by_telegram_id(5001)
        employee_id = employee[0]
        self.database.update_employee_status(employee_id, "active")
        admin = self.database.ensure_admin_employee(9001)
        admin_id = admin[0]

        employee_shift = self.database.create_shift(employee_id)
        admin_shift = self.database.create_shift(admin_id)
        today = self.database.local_today().isoformat()

        open_shift_names = {row[1] for row in self.database.get_open_shifts()}
        self.assertIn("Тест Рабочий", open_shift_names)
        self.assertNotIn(admin[2], open_shift_names)

        operation = self.database.get_active_operations(
            position="Швея",
            folder="Брюки-джоггеры",
        )[0]
        operation_id = operation[0]

        self.database.add_shift_operation(employee_shift["id"], employee_id, operation_id, "86", "Капучино", 5)
        self.database.add_shift_operation(admin_shift["id"], admin_id, operation_id, "86", "Капучино", 99)
        self.database.add_feedback_entry(employee_id, employee_shift["id"], "Производство", "Рабочее сообщение")
        self.database.add_feedback_entry(admin_id, admin_shift["id"], "Производство", "Админское сообщение")
        self.database.close_shift(employee_shift["id"])
        self.database.close_shift(admin_shift["id"])

        self.assertEqual([row[1] for row in self.database.get_all_employees()], ["Тест Рабочий"])
        self.assertEqual([row[1] for row in self.database.get_employees_by_status("active")], ["Тест Рабочий"])
        self.assertEqual([row[1] for row in self.database.get_period_employee_summary(today, today)], ["Тест Рабочий"])
        self.assertEqual([row[1] for row in self.database.get_period_shift_details(today, today)], ["Тест Рабочий"])
        self.assertEqual({row[1] for row in self.database.get_period_operation_rows(today, today)}, {"Тест Рабочий"})
        self.assertEqual([row[2] for row in self.database.get_feedback_entries(today, today)], ["Тест Рабочий"])
        self.assertEqual([row[1] for row in self.database.get_recent_shifts(10)], ["Тест Рабочий"])
        self.assertIsNone(self.database.get_employee_period_summary(admin_id, today, today))
        self.assertEqual(self.database.get_employee_period_operation_totals(admin_id, today, today), [])

    def test_admin_can_update_and_rollback_cutting_batch(self):
        batch_id = self._create_layout_done_batch()

        progress_result = self.database.admin_update_cutting_batch_progress(batch_id, 50)
        self.assertEqual(progress_result["new_status"], "cutting_in_progress")
        self.assertEqual(progress_result["new_progress"], 50)

        rollback_to_layout = self.database.rollback_cutting_batch(batch_id, "layout_done")
        self.assertEqual(rollback_to_layout["new_status"], "layout_done")

        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT status, cutting_progress FROM cutting_batches WHERE id = ?", (batch_id,))
        self.assertEqual(cursor.fetchone(), ("layout_done", 0))
        cursor.execute(
            "SELECT COUNT(*) FROM route_batches WHERE source_cutting_batch_id = ? AND status = 'active'",
            (batch_id,),
        )
        self.assertGreater(cursor.fetchone()[0], 0)
        conn.close()

        rollback_to_contours = self.database.rollback_cutting_batch(batch_id, "contours_done")
        self.assertEqual(rollback_to_contours["new_status"], "contours_done")

        conn = sqlite3.connect(self.database.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM cutting_batch_colors WHERE batch_id = ?", (batch_id,))
        self.assertEqual(cursor.fetchone()[0], 0)
        cursor.execute("SELECT status, layout_date, cutting_progress FROM cutting_batches WHERE id = ?", (batch_id,))
        self.assertEqual(cursor.fetchone(), ("contours_done", None, 0))
        cursor.execute(
            "SELECT COUNT(*) FROM route_batches WHERE source_cutting_batch_id = ? AND status = 'active'",
            (batch_id,),
        )
        self.assertEqual(cursor.fetchone()[0], 0)
        conn.close()

    def test_concurrent_shift_open_returns_single_open_shift(self):
        self.database.create_employee(1101, "Тест Одна Смена", "Швея")
        employee = self.database.get_employee_by_telegram_id(1101)
        self.database.update_employee_status(employee[0], "active")

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(lambda _index: self.database.create_shift(employee[0]), range(2)))

        self.assertEqual(results[0]["id"], results[1]["id"])
        conn = sqlite3.connect(self.database.DB_NAME)
        count = conn.execute(
            "SELECT COUNT(*) FROM shifts WHERE employee_id = ? AND status = 'open'",
            (employee[0],),
        ).fetchone()[0]
        conn.close()
        self.assertEqual(count, 1)

    def test_concurrent_shift_close_is_applied_once(self):
        self.database.create_employee(1105, "Тест Закрытие Смены", "Швея")
        employee = self.database.get_employee_by_telegram_id(1105)
        self.database.update_employee_status(employee[0], "active")
        shift = self.database.create_shift(employee[0])

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(lambda _index: self.database.close_shift(shift["id"]), range(2)))

        self.assertEqual(sum(result is not None for result in results), 1)
        conn = sqlite3.connect(self.database.DB_NAME)
        row = conn.execute("SELECT status, end_time FROM shifts WHERE id = ?", (shift["id"],)).fetchone()
        conn.close()
        self.assertEqual(row[0], "closed")
        self.assertTrue(row[1])

    def test_concurrent_contour_submission_creates_one_batch(self):
        self.database.create_employee(1102, "Тест Контуры Один Раз", "Раскройщик")
        employee = self.database.get_employee_by_telegram_id(1102)
        self.database.update_employee_status(employee[0], "active")
        shift = self.database.create_shift(employee[0])
        operation = self.database.get_active_operations(position="Раскройщик")[0]
        task = self.database.create_production_task(
            "Легинсы",
            ["86"],
            ["Бежевый"],
            employee[0],
        )

        def submit(_index):
            return self.database.create_cutting_contour_batch_for_task(
                task["id"],
                "Легинсы",
                shift["id"],
                employee[0],
                operation[0],
                {("86", "Бежевый"): 7},
            )

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(submit, range(2)))

        self.assertEqual(sum(result is not None for result in results), 1)
        conn = sqlite3.connect(self.database.DB_NAME)
        batch_count = conn.execute(
            "SELECT COUNT(*) FROM cutting_batches WHERE production_task_id = ? AND status != 'cancelled'",
            (task["id"],),
        ).fetchone()[0]
        reported = conn.execute(
            "SELECT quantity FROM shift_operations WHERE shift_id = ? AND operation_id = ?",
            (shift["id"], operation[0]),
        ).fetchone()[0]
        conn.close()
        self.assertEqual(batch_count, 1)
        self.assertEqual(reported, 7)

    def test_concurrent_route_completion_posts_stock_once(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        route_maps = importlib.import_module("route_maps")
        last_step_index = len(route_maps.PRODUCT_ROUTE_MAPS["Легинсы"]) - 1
        last_step = route_maps.PRODUCT_ROUTE_MAPS["Легинсы"][last_step_index]
        telegram_id = 1103
        self.database.create_employee(telegram_id, "Тест Завершение Один Раз", last_step["position"])
        employee = self.database.get_employee_by_telegram_id(telegram_id)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        batch = self.database.create_route_batch(
            "Легинсы",
            "86",
            "Бежевый",
            10,
            employee[0],
            route_step_index=last_step_index,
        )
        self.assertTrue(miniapp_server.start_route_task_for_telegram(telegram_id, batch["id"])["ok"])

        def complete(_index):
            return miniapp_server.complete_route_task_for_telegram(
                telegram_id,
                batch["id"],
                {"good_quantity": 10, "defect_quantity": 0},
            )

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(complete, range(2)))

        self.assertEqual(sum(result["ok"] for result in results), 1)
        finished = [row for row in self.database.get_warehouse_stock_rows() if row["item_type"] == "finished"]
        self.assertEqual(sum(row["quantity"] for row in finished), 10)
        conn = sqlite3.connect(self.database.DB_NAME)
        history_count = conn.execute(
            "SELECT COUNT(*) FROM route_batch_history WHERE batch_id = ?",
            (batch["id"],),
        ).fetchone()[0]
        conn.close()
        self.assertEqual(history_count, 1)

    def test_final_packaging_receives_finished_goods_in_wms_once(self):
        from wms.models import OperationResult

        os.environ["WMS_AUTO_RECEIPT_ENABLED"] = "1"
        miniapp_server = importlib.import_module("miniapp_server")
        route_maps = importlib.import_module("route_maps")
        step_index = len(route_maps.PRODUCT_ROUTE_MAPS["Легинсы"]) - 1
        self.assertEqual(route_maps.PRODUCT_ROUTE_MAPS["Легинсы"][step_index]["operation"], "Упаковка")
        telegram_id = 1104
        self.database.create_employee(telegram_id, "Тест Автоприёмка", "Упаковщик")
        employee = self.database.get_employee_by_telegram_id(telegram_id)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        batch = self.database.create_route_batch(
            "Легинсы", "86", "Бежевый", 6, employee[0], route_step_index=step_index
        )
        self.assertTrue(miniapp_server.start_route_task_for_telegram(telegram_id, batch["id"])["ok"])

        with patch.object(
            miniapp_server.wms_operations,
            "receive_from_production",
            return_value=OperationResult(ok=True, movement_id=73),
        ) as receive:
            completed = miniapp_server.complete_route_task_for_telegram(
                telegram_id,
                batch["id"],
                {"request_id": "pack-auto-receipt-1", "good_quantity": 6, "defect_quantity": 0},
            )
            replayed = miniapp_server.complete_route_task_for_telegram(
                telegram_id,
                batch["id"],
                {"request_id": "pack-auto-receipt-1", "good_quantity": 6, "defect_quantity": 0},
            )

        self.assertTrue(completed["ok"], completed)
        self.assertTrue(replayed["ok"], replayed)
        self.assertIn("зону приёмки", completed["message"])
        receive.assert_called_once()
        product_key, quantity = receive.call_args.args
        self.assertEqual(quantity, 6)
        self.assertEqual(
            (
                product_key.item_type,
                product_key.product_name,
                product_key.product_size,
                product_key.product_color,
                product_key.stage_name,
                product_key.ready_for_position,
            ),
            ("finished", "Легинсы", "86", "Бежевый", "Упаковано", "Склад"),
        )
        self.assertEqual(receive.call_args.kwargs["source_id"], batch["id"])
        self.assertEqual(
            receive.call_args.kwargs["request_key"],
            f"production:route-batch:{batch['id']}",
        )
        outbox = self.database.get_wms_receipt_outbox_by_batch_id(batch["id"])
        self.assertEqual((outbox["status"], outbox["quantity"], outbox["attempts"]), ("sent", 6, 1))
        self.assertEqual(self.database.get_active_route_batches(), [])

    def test_preparation_without_packaging_does_not_enter_wms_receiving(self):
        os.environ["WMS_AUTO_RECEIPT_ENABLED"] = "1"
        miniapp_server = importlib.import_module("miniapp_server")
        route_maps = importlib.import_module("route_maps")
        step_index = len(route_maps.PRODUCT_ROUTE_MAPS["Легинсы"]) - 2
        self.assertEqual(
            route_maps.PRODUCT_ROUTE_MAPS["Легинсы"][step_index]["operation"],
            "Подготовка к упаковке",
        )
        telegram_id = 1106
        self.database.create_employee(telegram_id, "Тест До Упаковки", "Упаковщик")
        employee = self.database.get_employee_by_telegram_id(telegram_id)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        batch = self.database.create_route_batch(
            "Легинсы", "98", "Синий", 3, employee[0], route_step_index=step_index
        )
        self.assertTrue(miniapp_server.start_route_task_for_telegram(telegram_id, batch["id"])["ok"])

        with patch.object(miniapp_server.wms_operations, "receive_from_production") as receive:
            completed = miniapp_server.complete_route_task_for_telegram(
                telegram_id,
                batch["id"],
                {"good_quantity": 3, "defect_quantity": 0},
            )

        self.assertTrue(completed["ok"], completed)
        receive.assert_not_called()
        self.assertIsNone(self.database.get_wms_receipt_outbox_by_batch_id(batch["id"]))
        next_batches = self.database.get_active_route_batches()
        self.assertEqual(len(next_batches), 1)
        self.assertEqual(
            route_maps.PRODUCT_ROUTE_MAPS["Легинсы"][next_batches[0]["route_step_index"]]["operation"],
            "Упаковка",
        )

    def test_failed_wms_receipt_stays_queued_and_retries(self):
        from wms.models import OperationResult

        os.environ["WMS_AUTO_RECEIPT_ENABLED"] = "1"
        miniapp_server = importlib.import_module("miniapp_server")
        route_maps = importlib.import_module("route_maps")
        step_index = len(route_maps.PRODUCT_ROUTE_MAPS["Легинсы"]) - 1
        telegram_id = 1105
        self.database.create_employee(telegram_id, "Тест Очередь WMS", "Упаковщик")
        employee = self.database.get_employee_by_telegram_id(telegram_id)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        batch = self.database.create_route_batch(
            "Легинсы", "92", "Черный", 4, employee[0], route_step_index=step_index
        )
        self.assertTrue(miniapp_server.start_route_task_for_telegram(telegram_id, batch["id"])["ok"])

        with patch.object(
            miniapp_server.wms_operations,
            "receive_from_production",
            side_effect=RuntimeError("synthetic WMS outage"),
        ):
            completed = miniapp_server.complete_route_task_for_telegram(
                telegram_id,
                batch["id"],
                {"good_quantity": 4, "defect_quantity": 0},
            )

        self.assertTrue(completed["ok"], completed)
        self.assertIn("поставлена в очередь", completed["message"])
        failed = self.database.get_wms_receipt_outbox_by_batch_id(batch["id"])
        self.assertEqual((failed["status"], failed["attempts"]), ("failed", 1))

        with patch.object(
            miniapp_server.wms_operations,
            "receive_from_production",
            return_value=OperationResult(ok=True, movement_id=74),
        ) as receive:
            summary = miniapp_server.flush_pending_wms_receipts()

        self.assertEqual(summary["sent"], 1)
        receive.assert_called_once()
        sent = self.database.get_wms_receipt_outbox_by_batch_id(batch["id"])
        self.assertEqual((sent["status"], sent["attempts"]), ("sent", 2))

    def test_concurrent_orders_cannot_consume_same_semifinished_stock(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        route_maps = importlib.import_module("route_maps")
        step_index = len(route_maps.CUTTING_ROUTE)
        step = route_maps.PRODUCT_ROUTE_MAPS["Легинсы"][step_index]
        stock = self.database.add_warehouse_stock(
            "semifinished",
            "Легинсы",
            "86",
            "Бежевый",
            "Раскроенные",
            step["position"],
            10,
            None,
        )
        payload = {
            "product_name": "Легинсы",
            "task_type": "route",
            "route_step_index": step_index,
            "stock_items": [{"stock_id": stock["id"], "quantity": "10"}],
        }

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(lambda _index: miniapp_server.create_order_task_for_telegram(9001, payload), range(2)))

        self.assertEqual(sum(result["ok"] for result in results), 0)
        self.assertEqual(self.database.get_warehouse_stock_by_id(stock["id"])["quantity"], 10)
        self.assertEqual(len(self.database.get_active_route_batches()), 0)

    def test_concurrent_route_cancellation_restores_stock_once(self):
        os.environ["ADMIN_IDS"] = "9001"
        miniapp_server = importlib.import_module("miniapp_server")
        stock = self.database.add_warehouse_stock(
            "semifinished", "Легинсы", "86", "Бежевый", "Раскроенные", "Швея", 10, None,
        )
        batch = self.database.create_route_batch_with_inputs(
            "Легинсы",
            "86",
            "Бежевый",
            10,
            None,
            4,
            [{"stock_id": stock["id"], "quantity": 10, "input_role": "main"}],
        )
        payload = {"task_kind": "route", "task_id": batch["id"]}

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(lambda _index: miniapp_server.delete_order_task_for_telegram(9001, payload), range(2)))

        self.assertEqual(sum(result["ok"] for result in results), 1)
        self.assertEqual(self.database.get_warehouse_stock_by_id(stock["id"])["quantity"], 10)

    def test_mass_route_creation_rolls_back_every_batch_on_shortage(self):
        stock = self.database.add_warehouse_stock(
            "semifinished", "Легинсы", "86, 92", "Бежевый", "Раскроенные", "Швея", 10, None,
        )
        specs = [
            {
                "product_size": product_size,
                "product_color": "Бежевый",
                "quantity": 6,
                "input_items": [{"stock_id": stock["id"], "quantity": 6, "input_role": "main"}],
            }
            for product_size in ("86", "92")
        ]

        result = self.database.create_route_batches_with_inputs(
            "Легинсы", None, 4, specs,
        )

        self.assertIsNone(result)
        self.assertEqual(self.database.get_warehouse_stock_by_id(stock["id"])["quantity"], 10)
        self.assertEqual(self.database.get_active_route_batches(), [])

    def test_concurrent_cutting_claim_has_one_owner_and_survives_new_shift(self):
        miniapp_server = importlib.import_module("miniapp_server")
        cutters = []
        for telegram_id, full_name in ((1201, "Первый Раскройщик"), (1202, "Второй Раскройщик")):
            self.database.create_employee(telegram_id, full_name, "Раскройщик")
            employee = self.database.get_employee_by_telegram_id(telegram_id)
            self.database.update_employee_status(employee[0], "active")
            self.database.create_shift(employee[0])
            cutters.append((telegram_id, employee[0]))

        task = self.database.create_production_task("Легинсы", ["86"], ["Бежевый"], None)

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(
                lambda cutter: miniapp_server.start_cutting_task_for_telegram(cutter[0], task["id"]),
                cutters,
            ))

        self.assertEqual(sum(result["ok"] for result in results), 1)
        assigned_task = self.database.get_production_task_by_id(task["id"])
        self.assertIn(assigned_task["assigned_employee_id"], {employee_id for _, employee_id in cutters})

        owner_telegram_id = next(
            telegram_id for telegram_id, employee_id in cutters
            if employee_id == assigned_task["assigned_employee_id"]
        )
        owner_shift = self.database.get_open_shift_for_today(assigned_task["assigned_employee_id"])
        self.database.close_shift(owner_shift[0])
        self.database.create_shift(assigned_task["assigned_employee_id"])
        state = miniapp_server.get_production_state_for_telegram(owner_telegram_id)
        self.assertTrue(state["cutting_tasks"][0]["is_assigned_to_me"])
        self.assertEqual(state["cutting_tasks"][0]["status_text"], "В работе")

    def test_cutting_progress_cannot_move_backwards(self):
        batch_id = self._create_layout_done_batch()
        self.assertTrue(self.database.update_cutting_batch_progress(
            batch_id, self.shift_id, self.employee_id, self.operation_id, 75,
        ))
        self.assertFalse(self.database.update_cutting_batch_progress(
            batch_id, self.shift_id, self.employee_id, self.operation_id, 25,
        ))
        row = next(row for row in self.database.get_cutting_batches_for_cutting("Брюки-ползунки") if row[0] == batch_id)
        self.assertEqual(row[4], 75)

    def test_started_cutting_task_cannot_be_cancelled_by_employee(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.add_fabric_receipt("Ткань", "Бежевый", 2, None)
        task = self.database.create_production_task(
            "Легинсы", ["86"], ["Бежевый"], None,
            fabric_rolls={"Бежевый": 1},
        )
        self.database.create_employee(1203, "Раскройщик Отмена", "Раскройщик")
        employee = self.database.get_employee_by_telegram_id(1203)
        self.database.update_employee_status(employee[0], "active")
        shift = self.database.create_shift(employee[0])
        self.database.assign_production_task(task["id"], employee[0])
        operation = self.database.get_active_operations(position="Раскройщик")[0]
        batch_id = self.database.create_cutting_contour_batch_for_task(
            task["id"], "Легинсы", shift["id"], employee[0], operation[0], {("86", "Бежевый"): 5},
        )

        self.assertIsNotNone(batch_id)
        result = miniapp_server.delete_order_task_for_telegram(
            1203,
            {"task_kind": "production", "task_id": task["id"]},
        )
        self.assertFalse(result["ok"])
        self.assertIn("администратора", result["message"].lower())
        self.assertEqual(self.database.get_fabric_stock_rows()[0][2], 1)
        self.assertEqual(self.database.get_production_task_by_id(task["id"])["status"], "contours_done")

    def test_route_completion_requires_assignment_and_open_shift(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(1204, "Швея Проверка", "Швея")
        employee = self.database.get_employee_by_telegram_id(1204)
        self.database.update_employee_status(employee[0], "active")
        step_index = self.route_step_index("Легинсы", "Сборка на оверлоке", "Швея")
        batch = self.database.create_route_batch("Легинсы", "86", "Бежевый", 5, None, step_index)

        no_shift = miniapp_server.complete_route_task_for_telegram(1204, batch["id"])
        self.assertFalse(no_shift["ok"])
        self.database.create_shift(employee[0])
        not_assigned = miniapp_server.complete_route_task_for_telegram(1204, batch["id"])
        self.assertFalse(not_assigned["ok"])
        self.assertIn("выберите задание", not_assigned["message"].lower())

    def test_fabric_lots_are_reserved_fifo_and_restored_exactly(self):
        self.database.add_fabric_receipt("Ткань", "Бежевый", 2, None, comment="Первая партия")
        self.database.add_fabric_receipt("Ткань", "Бежевый", 2, None, comment="Вторая партия")
        task = self.database.create_production_task(
            "Легинсы",
            ["86"],
            ["Бежевый"],
            None,
            fabric_rolls={"Бежевый": 3},
        )

        self.assertTrue(task["trace_code"].startswith("CUT-"))
        self.assertTrue(task["route_version"])
        conn = sqlite3.connect(self.database.DB_NAME)
        allocations = conn.execute(
            """
            SELECT fabric_stock_lots.lot_code, production_task_fabric_lots.rolls
            FROM production_task_fabric_lots
            JOIN fabric_stock_lots ON fabric_stock_lots.id = production_task_fabric_lots.lot_id
            WHERE production_task_fabric_lots.task_id = ?
            ORDER BY fabric_stock_lots.id
            """,
            (task["id"],),
        ).fetchall()
        available_before_cancel = conn.execute(
            "SELECT SUM(rolls_available) FROM fabric_stock_lots",
        ).fetchone()[0]
        conn.close()

        self.assertEqual([row[1] for row in allocations], [2, 1])
        self.assertEqual(available_before_cancel, 1)
        self.assertIsNotNone(self.database.cancel_production_task(task["id"]))

        conn = sqlite3.connect(self.database.DB_NAME)
        available_after_cancel = conn.execute(
            "SELECT SUM(rolls_available) FROM fabric_stock_lots",
        ).fetchone()[0]
        stock_quantity = conn.execute("SELECT quantity FROM fabric_stock").fetchone()[0]
        conn.close()
        self.assertEqual(available_after_cancel, 4)
        self.assertEqual(stock_quantity, 4)

    def test_route_passport_connects_output_lot_to_previous_operation(self):
        miniapp_server = importlib.import_module("miniapp_server")
        step_index = self.route_step_index("Легинсы", "Сборка на оверлоке", "Швея")
        stock = self.database.add_warehouse_stock(
            "semifinished", "Легинсы", "86", "Бежевый", "Раскроенные", "Швея", 10, None,
        )
        parent = self.database.create_route_batch_with_inputs(
            "Легинсы", "86", "Бежевый", 10, None, step_index,
            [{"stock_id": stock["id"], "quantity": 10, "input_role": "main"}],
        )
        self.database.create_employee(1301, "Тест Швея Партия", "Швея")
        employee = self.database.get_employee_by_telegram_id(1301)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        self.assertTrue(miniapp_server.start_route_task_for_telegram(1301, parent["id"])["ok"])
        self.assertTrue(miniapp_server.complete_route_task_for_telegram(
            1301, parent["id"], {"good_quantity": 10, "defect_quantity": 0},
        )["ok"])

        child = next(
            row for row in self.database.get_active_route_batches()
            if row["route_step_index"] == step_index + 1
        )
        passport = self.database.get_route_batch_passport(child["id"])

        self.assertTrue({parent["id"], child["id"]}.issubset({row["id"] for row in passport["batches"]}))
        child_lots = passport["inputs"][str(child["id"])]
        self.assertEqual(child_lots[0]["source_type"], "route_batch")
        self.assertEqual(child_lots[0]["source_id"], parent["id"])
        self.assertTrue(any(event["event_type"] == "operation_completed" for event in passport["events"]))

    def test_route_snapshot_is_frozen_when_catalog_changes(self):
        route_maps = importlib.import_module("route_maps")
        miniapp_server = importlib.import_module("miniapp_server")
        step_index = self.route_step_index("Легинсы", "Сборка на оверлоке", "Швея")
        batch = self.database.create_route_batch("Легинсы", "86", "Бежевый", 5, None, step_index)
        original_operation = miniapp_server.get_route_step_for_batch(batch)["operation"]
        saved_operation = route_maps.PRODUCT_ROUTE_MAPS["Легинсы"][step_index]["operation"]
        try:
            route_maps.PRODUCT_ROUTE_MAPS["Легинсы"][step_index]["operation"] = "Новая версия операции"
            self.assertEqual(miniapp_server.get_route_step_for_batch(batch)["operation"], original_operation)
            self.assertNotEqual(original_operation, "Новая версия операции")
        finally:
            route_maps.PRODUCT_ROUTE_MAPS["Легинсы"][step_index]["operation"] = saved_operation

    def test_legacy_cardigan_product_alias_uses_current_dublerin_route(self):
        snapshot = self.database.current_route_snapshot("Кардиган детский")
        steps = self.database.route_steps_from_snapshot(snapshot, "Кардиган детский")
        operations = [step["operation"] for step in steps]
        self.assertIn("Кардиганы — дублерин 25 мм", operations)
        self.assertIn("Кардиган — Дублирование", operations)

    def test_route_task_pause_block_and_handover_persist(self):
        self.database.create_employee(1302, "Тест Передача", "Швея")
        employee = self.database.get_employee_by_telegram_id(1302)
        self.database.update_employee_status(employee[0], "active")
        step_index = self.route_step_index("Легинсы", "Сборка на оверлоке", "Швея")
        batch = self.database.create_route_batch("Легинсы", "86", "Бежевый", 5, None, step_index)
        self.assertIsNotNone(self.database.assign_route_batch(batch["id"], employee[0]))

        paused = self.database.set_route_batch_work_state(batch["id"], employee[0], "pause", "Перерыв")
        self.assertEqual(paused["work_state"], "paused")
        blocked = self.database.set_route_batch_work_state(batch["id"], employee[0], "block", "Нет комплектующей")
        self.assertEqual(blocked["work_state"], "blocked")
        resumed = self.database.set_route_batch_work_state(batch["id"], employee[0], "resume")
        self.assertEqual(resumed["work_state"], "in_work")
        released = self.database.set_route_batch_work_state(
            batch["id"], employee[0], "release", "Передача следующей смене",
        )
        self.assertIsNone(released["assigned_employee_id"])
        self.assertEqual(released["work_state"], "free")
        self.assertEqual(released["handover_count"], 1)

        conn = sqlite3.connect(self.database.DB_NAME)
        handoff = conn.execute(
            "SELECT from_employee_id, reason FROM route_batch_handoffs WHERE batch_id = ?",
            (batch["id"],),
        ).fetchone()
        event_types = {
            row[0] for row in conn.execute(
                "SELECT event_type FROM production_trace_events WHERE batch_id = ?",
                (batch["id"],),
            )
        }
        conn.close()
        self.assertEqual(handoff, (employee[0], "Передача следующей смене"))
        self.assertTrue({"task_paused", "task_blocked", "task_resumed", "task_released"}.issubset(event_types))

    def test_completion_request_is_idempotent_and_saves_defect_photo(self):
        miniapp_server = importlib.import_module("miniapp_server")
        route_maps = importlib.import_module("route_maps")
        step_index = len(route_maps.PRODUCT_ROUTE_MAPS["Легинсы"]) - 1
        position = route_maps.PRODUCT_ROUTE_MAPS["Легинсы"][step_index]["position"]
        self.database.create_employee(1303, "Тест Офлайн", position)
        employee = self.database.get_employee_by_telegram_id(1303)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        batch = self.database.create_route_batch("Легинсы", "86", "Бежевый", 5, None, step_index)
        self.assertTrue(miniapp_server.start_route_task_for_telegram(1303, batch["id"])["ok"])
        payload = {
            "request_id": "offline-request-1",
            "good_quantity": 4,
            "defect_quantity": 1,
            "defect_reason": "Дефект строчки",
            "defect_disposition": "Списать",
            "defect_comment": "Тест",
            "defect_photo": {
                "file_name": "defect.png",
                "mime_type": "image/png",
                "content_base64": base64.b64encode(b"test-image").decode("ascii"),
            },
        }

        first = miniapp_server.complete_route_task_for_telegram(1303, batch["id"], payload)
        second = miniapp_server.complete_route_task_for_telegram(1303, batch["id"], payload)
        self.assertTrue(first["ok"])
        self.assertTrue(second["ok"])
        self.assertIn("синхронизировано", second["message"].lower())
        finished_quantity = sum(
            row["quantity"] for row in self.database.get_warehouse_stock_rows()
            if row["item_type"] == "finished"
        )
        self.assertEqual(finished_quantity, 4)
        defect = self.database.get_route_batch_defects(batch["id"])[0]
        self.assertTrue(defect["has_photo"])
        photo = self.database.get_route_batch_defect_photo(defect["id"])
        self.assertEqual(photo["file_name"], "defect.png")
        self.assertEqual(base64.b64decode(photo["content_base64"]), b"test-image")

    def test_admin_adjusts_fabric_and_warehouse_balances_with_reason(self):
        os.environ["ADMIN_IDS"] = "9401"
        miniapp_server = importlib.import_module("miniapp_server")
        fabric = self.database.add_fabric_receipt("Ткань", "Бежевый", 5, None)
        warehouse = self.database.add_warehouse_stock(
            "semifinished", "Легинсы", "86", "Бежевый", "Раскроенные", "Швея", 10, None,
        )

        fabric_result = miniapp_server.adjust_stock_for_telegram(
            9401,
            {"stock_kind": "fabric", "stock_id": fabric[0], "quantity": 3, "reason": "Инвентаризация ткани"},
        )
        warehouse_result = miniapp_server.adjust_stock_for_telegram(
            9401,
            {"stock_kind": "warehouse", "stock_id": warehouse["id"], "quantity": 6, "reason": "Пересчёт кроя"},
        )
        self.assertTrue(fabric_result["ok"], fabric_result)
        self.assertTrue(warehouse_result["ok"], warehouse_result)
        self.assertEqual(self.database.get_fabric_stock_by_id(fabric[0])[3], 3)
        self.assertEqual(self.database.get_warehouse_stock_by_id(warehouse["id"])["quantity"], 6)

        conn = sqlite3.connect(self.database.DB_NAME)
        fabric_lots = conn.execute(
            "SELECT SUM(rolls_available) FROM fabric_stock_lots WHERE stock_id = ?", (fabric[0],),
        ).fetchone()[0]
        warehouse_lots = conn.execute(
            "SELECT SUM(quantity_available) FROM warehouse_stock_lots WHERE stock_id = ?", (warehouse["id"],),
        ).fetchone()[0]
        fabric_movement = conn.execute(
            "SELECT quantity, movement_type, comment FROM fabric_stock_movements ORDER BY id DESC LIMIT 1",
        ).fetchone()
        warehouse_movement = conn.execute(
            "SELECT quantity, movement_type, comment FROM warehouse_stock_movements ORDER BY id DESC LIMIT 1",
        ).fetchone()
        conn.close()
        self.assertEqual(fabric_lots, 3)
        self.assertEqual(warehouse_lots, 6)
        self.assertEqual(fabric_movement, (-2, "adjustment", "Инвентаризация ткани"))
        self.assertEqual(warehouse_movement, (-4, "adjustment", "Пересчёт кроя"))

    def test_cutter_rejects_assigned_fabric_rolls_and_cancel_restores_only_good_rolls(self):
        miniapp_server = importlib.import_module("miniapp_server")
        self.database.create_employee(9402, "Тест Раскройщик Брак", "Раскройщик")
        employee = self.database.get_employee_by_telegram_id(9402)
        self.database.update_employee_status(employee[0], "active")
        self.database.create_shift(employee[0])
        self.database.add_fabric_receipt("Ткань", "Бежевый", 3, None)
        task = self.database.create_production_task(
            "Легинсы", ["86"], ["Бежевый"], None,
            fabric_rolls={"Бежевый": 3},
        )
        self.assertIsNotNone(self.database.assign_production_task(task["id"], employee[0]))

        rejected = miniapp_server.reject_fabric_rolls_for_telegram(
            9402,
            {
                "task_id": task["id"],
                "product_color": "Бежевый",
                "quantity": 1,
                "comment": "Повреждение полотна",
            },
        )
        repeated_too_large = miniapp_server.reject_fabric_rolls_for_telegram(
            9402,
            {
                "task_id": task["id"],
                "product_color": "Бежевый",
                "quantity": 3,
                "comment": "Повторное списание",
            },
        )
        self.assertTrue(rejected["ok"], rejected)
        self.assertFalse(repeated_too_large["ok"])
        roll = rejected["production"]["cutting_tasks"][0]["fabric_rolls"][0]
        self.assertEqual((roll["rolls"], roll["rejected_rolls"], roll["available_rolls"]), (3, 1, 2))
        self.assertEqual(roll["defects"][0]["comment"], "Повреждение полотна")

        cancelled = self.database.cancel_production_task(task["id"], employee[0])
        self.assertIsNotNone(cancelled)
        self.assertEqual(self.database.get_fabric_stock_rows()[0][2], 2)
        conn = sqlite3.connect(self.database.DB_NAME)
        available_lots = conn.execute("SELECT SUM(rolls_available) FROM fabric_stock_lots").fetchone()[0]
        conn.close()
        self.assertEqual(available_lots, 2)

    def _create_layout_done_batch(self):
        self.database.create_employee(1001, "Тест Раскройщик", "Раскройщик")
        employee = self.database.get_employee_by_telegram_id(1001)
        self.employee_id = employee[0]
        self.database.update_employee_status(self.employee_id, "active")
        shift = self.database.create_shift(self.employee_id)
        self.shift_id = shift["id"]

        contour_operation = self.database.get_active_operations(
            position="Раскройщик",
            folder="Брюки-ползунки",
        )[0]
        self.operation_id = contour_operation[0]

        batch_id = self.database.create_cutting_contour_batch(
            "Брюки-ползунки",
            self.shift_id,
            self.employee_id,
            self.operation_id,
            {"80": 2, "86": 3},
        )

        self.assertTrue(
            self.database.add_cutting_layout(
                batch_id,
                self.shift_id,
                self.employee_id,
                self.operation_id,
                {"Синий": 5, "Бежевый": 10},
            )
        )

        return batch_id


if __name__ == "__main__":
    unittest.main()
