"""Read-only shipment views and document exports for the addressed WMS."""

from __future__ import annotations

import io
import re
from collections import defaultdict
from datetime import datetime

from .connection import get_pg_connection


_NUMBER_RE = re.compile(r"(?:ТЕСТОВАЯ )?ОТГРУЗКА\s+([A-Z0-9-]+)", re.IGNORECASE)


def _number(reason: str) -> str:
    match = _NUMBER_RE.search(str(reason or ""))
    return match.group(1) if match else "Без номера"


def shipment_detail(number: str) -> dict | None:
    number = str(number or "").strip().upper()
    if not number:
        return None
    conn = get_pg_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT m.id,m.product_key,m.quantity,m.reason,m.occurred_at,
                          l.code AS from_location_code,l.name_ru AS from_location_name
                     FROM wms_movements m
                LEFT JOIN wms_locations l ON l.id=m.from_location_id
                    WHERE (m.movement_type='ship' OR m.source_type='shipment')
                 ORDER BY m.occurred_at DESC,m.id DESC"""
            )
            rows = cur.fetchall()
        lines = []
        reason = ""
        occurred_at = ""
        for row in rows:
            if _number(row[3]) != number:
                continue
            product = row[1] if isinstance(row[1], dict) else {}
            lines.append({
                "product_name": str(product.get("product_name") or "Товар"),
                "product_size": str(product.get("product_size") or "—"),
                "product_color": str(product.get("product_color") or "—"),
                "quantity": int(row[2] or 0),
                "from_location_code": str(row[5] or "—"),
                "from_location_name": str(row[6] or ""),
            })
            reason = str(row[3] or reason)
            occurred_at = str(row[4] or occurred_at)
        if not lines:
            return None
        lines.sort(key=lambda item: (item["from_location_code"], item["product_name"], item["product_size"], item["product_color"]))
        return {
            "number": number,
            "reason": reason,
            "occurred_at": occurred_at,
            "lines": lines,
            "total": sum(item["quantity"] for item in lines),
            "locations": len({item["from_location_code"] for item in lines}),
        }
    finally:
        conn.rollback()


def shipment_excel_bytes(shipment: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отгрузка"
    sheet.append(["Отгрузка", shipment["number"]])
    sheet.append(["Дата", shipment.get("occurred_at") or "—"])
    sheet.append(["Всего", shipment["total"]])
    sheet.append([])
    sheet.append(["№", "Товар", "Размер", "Цвет", "Количество", "Ячейка", "Название ячейки"])
    for index, line in enumerate(shipment["lines"], start=1):
        sheet.append([index, line["product_name"], line["product_size"], line["product_color"], line["quantity"], line["from_location_code"], line["from_location_name"]])
    sheet.append([])
    sheet.append(["ИТОГО", "", "", "", shipment["total"], "", ""])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
    for cell in sheet[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
    for column, width in {"A": 8, "B": 42, "C": 12, "D": 18, "E": 14, "F": 18, "G": 26}.items():
        sheet.column_dimensions[column].width = width
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def shipment_pdf_bytes(shipment: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    bold_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    pdfmetrics.registerFont(TTFont("DejaVu", font_path))
    pdfmetrics.registerFont(TTFont("DejaVu-Bold", bold_path))
    styles = getSampleStyleSheet()
    title = styles["Title"].clone("shipment-title")
    title.fontName = "DejaVu-Bold"
    title.fontSize = 18
    normal = styles["Normal"].clone("shipment-normal")
    normal.fontName = "DejaVu"
    normal.fontSize = 8
    normal.leading = 10
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    data = [["№", "Товар", "Размер", "Цвет", "Кол-во", "Ячейка"]]
    for index, line in enumerate(shipment["lines"], start=1):
        data.append([str(index), Paragraph(line["product_name"], normal), line["product_size"], line["product_color"], str(line["quantity"]), line["from_location_code"]])
    data.append(["", Paragraph("ИТОГО", normal), "", "", str(shipment["total"]), ""])
    table = Table(data, colWidths=[8*mm, 72*mm, 18*mm, 28*mm, 18*mm, 30*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "DejaVu"),
        ("FONTNAME", (0, 0), (-1, 0), "DejaVu-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (4, 1), (4, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
        ("TOPPADDING", (0, 0), (-1, 0), 7),
    ]))
    document.build([
        Paragraph(f"Отгрузка {shipment['number']}", title),
        Spacer(1, 5*mm),
        Paragraph(f"Дата: {shipment.get('occurred_at') or '—'} &nbsp;&nbsp; Позиций: {len(shipment['lines'])} &nbsp;&nbsp; Всего: {shipment['total']} шт.", normal),
        Spacer(1, 5*mm),
        table,
    ])
    return output.getvalue()
